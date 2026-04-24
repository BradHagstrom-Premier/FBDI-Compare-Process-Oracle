"""Stage 8 post-run verification for fbdi-compare-release.

- Runs `python -m fbdi diagnose --release <ver>` and parses the Diagnostic
  xlsx output for NO_HEADER regressions.
- Reads FBDI_Master_Catalog.xlsx Issues tab, filters by release, and flags
  catalog Issues-tab regression if:
      new_count > 2 * prior_count   OR   new_count - prior_count > 50

Never blocks. Exit 0 = clean, 1 = regression detected.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from collections import Counter
from contextlib import closing
from pathlib import Path

from openpyxl import load_workbook

ISSUE_MULTIPLIER_THRESHOLD = 2.0
ISSUE_ABSOLUTE_THRESHOLD = 50


def check_catalog_issues(catalog_path: Path, release: str) -> dict:
    """Read Issues tab, group by release, compare release against most-recent prior."""
    release = release.upper()
    with closing(load_workbook(catalog_path, read_only=True, data_only=True)) as wb:
        if "Issues" not in wb.sheetnames:
            return {
                "release_issue_count": 0,
                "prior_release": None,
                "prior_issue_count": 0,
                "regression": False,
                "threshold": {"multiplier": ISSUE_MULTIPLIER_THRESHOLD,
                              "absolute": ISSUE_ABSOLUTE_THRESHOLD},
            }
        ws = wb["Issues"]
        counter: Counter[str] = Counter()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or not row[0]:
                continue
            counter[str(row[0]).upper()] += 1

    release_count = counter.get(release, 0)
    priors = sorted(r for r in counter if r < release)
    prior = priors[-1] if priors else None
    prior_count = counter.get(prior, 0) if prior else 0

    regression = False
    if prior and prior_count > 0:
        if release_count > ISSUE_MULTIPLIER_THRESHOLD * prior_count:
            regression = True
        if release_count - prior_count > ISSUE_ABSOLUTE_THRESHOLD:
            regression = True

    return {
        "release_issue_count": release_count,
        "prior_release": prior,
        "prior_issue_count": prior_count,
        "regression": regression,
        "threshold": {"multiplier": ISSUE_MULTIPLIER_THRESHOLD,
                      "absolute": ISSUE_ABSOLUTE_THRESHOLD},
    }


def run_diagnose(release: str, repo_root: Path) -> dict:
    """Invoke `python -m fbdi diagnose --release <release>` and parse its output xlsx."""
    diagnostic_path = repo_root / f"Diagnostic_Report_{release.upper()}.xlsx"
    if diagnostic_path.is_file():
        diagnostic_path.unlink()

    proc = subprocess.run(
        [sys.executable, "-m", "fbdi", "diagnose", "--release", release],
        cwd=repo_root,
        capture_output=True, text=True,
    )
    if proc.returncode != 0 or not diagnostic_path.is_file():
        return {
            "no_header_count": None,
            "file_error_count": None,
            "regression": False,
            "error": f"diagnose invocation failed: {proc.stderr[:500]}",
        }

    no_header = 0
    file_error = 0
    with closing(load_workbook(diagnostic_path, read_only=True, data_only=True)) as wb:
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row:
                continue
            result = row[2]  # "Detection Result" column
            if result == "NO_HEADER":
                no_header += 1
            elif result == "FILE_ERROR":
                file_error += 1

    return {
        "no_header_count": no_header,
        "file_error_count": file_error,
        "regression": no_header > 0,
        "diagnostic_path": str(diagnostic_path),
    }


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Stage 8 post-run verification")
    parser.add_argument("--release", required=True)
    parser.add_argument(
        "--catalog", type=Path, default=Path("FBDI_Master_Catalog.xlsx"),
    )
    parser.add_argument(
        "--repo-root", type=Path, default=Path.cwd(),
        help="Repo root for fbdi diagnose invocation",
    )
    parser.add_argument(
        "--skip-diagnose", action="store_true",
        help="Skip the diagnose subprocess (for unit tests / quick runs)",
    )
    args = parser.parse_args(argv)

    release = args.release.upper()

    diag = {"skipped": True}
    if not args.skip_diagnose:
        diag = run_diagnose(release, args.repo_root)

    cat = check_catalog_issues(args.catalog, release)

    overall = bool(diag.get("regression")) or bool(cat.get("regression"))
    payload = {
        "release": release,
        "diagnose": diag,
        "catalog_issues": cat,
        "overall_regression": overall,
    }
    print(json.dumps(payload, indent=2))
    return 1 if overall else 0


if __name__ == "__main__":
    sys.exit(main())
