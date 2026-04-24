"""Stage 7 summary for fbdi-compare-release.

Reads Comparison_Report_<OLD>_<NEW>.xlsx and prints a JSON summary:
total changes, distinct files with changes, top-5 most-changed files.
Accepts Stage 4 timeout filenames via --timeouts for inclusion in the
summary (so the user knows which blanks files need manual clearing).
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def summarize(report_path: Path) -> dict:
    report_path = Path(report_path)
    wb = load_workbook(report_path, read_only=True, data_only=True)
    ws = wb.active
    counter: Counter[str] = Counter()
    total = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if row and row[0]:
            counter[row[0]] += 1
            total += 1
    wb.close()

    top = [{"file": name, "changes": n} for name, n in counter.most_common(5)]
    return {
        "total_changes": total,
        "files_with_changes": len(counter),
        "top_files": top,
    }


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Stage 7 summary")
    parser.add_argument(
        "--report", type=Path, required=True,
        help="Path to Comparison_Report_<OLD>_<NEW>.xlsx",
    )
    parser.add_argument(
        "--catalog", type=Path, default=Path("FBDI_Master_Catalog.xlsx"),
        help="Path to FBDI_Master_Catalog.xlsx",
    )
    parser.add_argument(
        "--timeouts", type=str, default="",
        help="Comma-separated Stage 4 timeout filenames (optional)",
    )
    args = parser.parse_args(argv)

    summary = summarize(args.report)
    payload = {
        "report_path": str(args.report),
        "catalog_path": str(args.catalog),
        **summary,
        "stage4_timeouts": [t for t in args.timeouts.split(",") if t.strip()],
    }
    print(json.dumps(payload, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
