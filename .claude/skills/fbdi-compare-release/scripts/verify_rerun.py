"""Stage 8 macro-signal validator for fbdi-compare-release.

Adds checks not already covered by verify_run.py:
- Catalog row count delta (post-rerun vs pre-rerun catalog)
- Compare-report changes count vs expected baseline
- Module column population % in the working mapping spreadsheet

Never blocks. Exit 0 = clean, 1 = regression detected.
"""

from __future__ import annotations

import argparse
import json
import sys
from contextlib import closing
from pathlib import Path

from openpyxl import load_workbook


# Tunable thresholds — bump in future quarters as macro signals shift
CATALOG_DELTA_PCT_THRESHOLD = 5.0          # ±5% on per-release row count
COMPARE_CHANGES_DELTA_THRESHOLD = 50       # absolute delta around expected
DEFAULT_EXPECTED_COMPARE_CHANGES = 706     # baseline 26A→26B ground truth
MODULE_PCT_THRESHOLD = 95.0                # ≥95% rows with col A populated


def _count_release_rows(catalog_path: Path, release: str) -> int:
    """Return the number of data rows on the per-release sheet of the catalog."""
    with closing(load_workbook(catalog_path, read_only=True, data_only=True)) as wb:
        if release not in wb.sheetnames:
            return 0
        ws = wb[release]
        return max((ws.max_row or 1) - 1, 0)


def _count_compare_changes(report_path: Path) -> int:
    with closing(load_workbook(report_path, read_only=True, data_only=True)) as wb:
        ws = wb.active
        return max((ws.max_row or 1) - 1, 0)


def _module_pct(mapping_path: Path) -> tuple[float, int, int]:
    """Compute Module column population %: rows with col A non-blank
    are the denominator; rows with col F non-blank are the numerator.
    Returns (pct, populated, total)."""
    with closing(load_workbook(mapping_path, read_only=True, data_only=True)) as wb:
        if "FBDI Mapping" not in wb.sheetnames:
            return 0.0, 0, 0
        ws = wb["FBDI Mapping"]
        total = 0
        populated = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row[0]:  # col A non-blank
                total += 1
                if len(row) >= 6 and row[5]:
                    populated += 1
    pct = (populated / total * 100.0) if total > 0 else 0.0
    return pct, populated, total


def run_checks(
    new_catalog: Path,
    baseline_catalog: Path,
    compare_report: Path,
    mapping: Path,
    release: str,
    expected_compare_changes: int = DEFAULT_EXPECTED_COMPARE_CHANGES,
) -> dict:
    """Run all macro checks. Returns a JSON-serializable dict."""
    regressions: list[str] = []

    # Catalog row delta — only if baseline exists
    delta_pct = None
    new_rows = _count_release_rows(new_catalog, release)
    if baseline_catalog.is_file():
        baseline_rows = _count_release_rows(baseline_catalog, release)
        if baseline_rows > 0:
            delta_pct = (new_rows - baseline_rows) / baseline_rows * 100.0
            if abs(delta_pct) > CATALOG_DELTA_PCT_THRESHOLD:
                regressions.append(
                    f"Catalog row count for {release}: {new_rows} vs baseline "
                    f"{baseline_rows} ({delta_pct:+.1f}%, threshold ±{CATALOG_DELTA_PCT_THRESHOLD}%)"
                )

    # Compare changes delta
    changes = _count_compare_changes(compare_report) if compare_report.is_file() else None
    if changes is not None:
        delta = abs(changes - expected_compare_changes)
        if delta > COMPARE_CHANGES_DELTA_THRESHOLD:
            regressions.append(
                f"Compare report changes: {changes} vs expected ~{expected_compare_changes} "
                f"(±{COMPARE_CHANGES_DELTA_THRESHOLD})"
            )

    # Module pct populated
    module_pct, populated, total = _module_pct(mapping) if mapping.is_file() else (None, 0, 0)
    if module_pct is not None and module_pct < MODULE_PCT_THRESHOLD:
        regressions.append(
            f"Module column populated: {module_pct:.1f}% ({populated}/{total}) "
            f"vs threshold ≥{MODULE_PCT_THRESHOLD}%"
        )

    return {
        "release": release,
        "catalog_rows_new": new_rows,
        "catalog_delta_pct": delta_pct,
        "compare_changes": changes,
        "expected_compare_changes": expected_compare_changes,
        "module_pct_populated": module_pct,
        "regressions": regressions,
    }


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Stage 8 macro-signal validator")
    parser.add_argument("--release", required=True)
    parser.add_argument("--new-catalog", type=Path,
                        default=Path("FBDI_Master_Catalog.xlsx"))
    parser.add_argument("--baseline-catalog", type=Path,
                        default=Path("FBDI_Master_Catalog.bak.xlsx"),
                        help="Pre-rerun catalog snapshot for delta check")
    parser.add_argument("--compare-report", type=Path,
                        help="e.g. Comparison_Report_26A_26B.xlsx")
    parser.add_argument("--mapping", type=Path,
                        default=Path("FBDI_to_ApplaudTables_Mapping.xlsx"))
    parser.add_argument("--expected-compare-changes", type=int,
                        default=DEFAULT_EXPECTED_COMPARE_CHANGES)
    args = parser.parse_args(argv)

    report_path = args.compare_report or Path(f"Comparison_Report_*_{args.release}.xlsx")

    result = run_checks(
        new_catalog=args.new_catalog,
        baseline_catalog=args.baseline_catalog,
        compare_report=report_path,
        mapping=args.mapping,
        release=args.release.upper(),
        expected_compare_changes=args.expected_compare_changes,
    )
    print(json.dumps(result, indent=2))
    return 1 if result["regressions"] else 0


if __name__ == "__main__":
    sys.exit(main())
