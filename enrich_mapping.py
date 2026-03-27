"""Enrich fbdi_applaud_mapping.xlsx with Applaud table/prefix data from AP5TBFLD.CSV."""

import re
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd

# --- Constants ---
SCRIPT_DIR = Path(__file__).resolve().parent
CSV_PATH = SCRIPT_DIR / "AP5TBFLD.CSV"
MAPPING_PATH = SCRIPT_DIR / "fbdi_applaud_mapping.xlsx"
SHEET_NAME = "FBDI Mapping"
VALID_DD_TYPES = {"X", "N", "D", "U"}
ORACLE_TAB_PATTERN = re.compile(r"^[A-Z][A-Z0-9_]+$")
PREFIX_PATTERN = re.compile(r"^([A-Z][A-Z0-9]{1,2})([A-Z_].*)")

# Column indices (1-based) in the mapping sheet
COL_FBDI_FILE = 1
COL_FBDI_TAB = 2
COL_APPLAUD_TABLE = 3
COL_PREFIX = 4
COL_IN_SCOPE = 5
COL_NOTES = 7


def extract_majority_prefix(fields: list[str]) -> str:
    """Extract the most common prefix from a list of Applaud field names."""
    prefixes = []
    for f in fields:
        m = PREFIX_PATTERN.match(f)
        if m:
            prefixes.append(m.group(1))
    if not prefixes:
        return ""
    return Counter(prefixes).most_common(1)[0][0]


def build_csv_lookup(csv_path: Path) -> dict[str, list[str]]:
    """Load CSV, filter to valid DD_Type rows, return {DB_Name: [DD_Name, ...]}."""
    df = pd.read_csv(csv_path)
    clean = df[df["DD_Type"].isin(VALID_DD_TYPES)].copy()
    clean["DB_Name"] = clean["DB_Name"].str.strip()
    lookup: dict[str, list[str]] = {}
    for db_name, dd_name in zip(clean["DB_Name"], clean["DD_Name"]):
        lookup.setdefault(db_name, []).append(dd_name)
    return lookup


def enrich() -> None:
    """Main enrichment logic."""
    # Step 1: Build CSV lookup
    csv_lookup = build_csv_lookup(CSV_PATH)
    csv_tables = set(csv_lookup.keys())

    # Step 2: Load mapping workbook
    wb = openpyxl.load_workbook(MAPPING_PATH)
    ws = wb[SHEET_NAME]

    updated_rows: list[tuple[str, str, str, str]] = []  # (file, tab, table, prefix)
    rows_skipped_no_match = 0
    rows_skipped_filled = 0

    # Step 3: Iterate data rows
    for row_idx in range(2, ws.max_row + 1):
        in_scope = ws.cell(row_idx, COL_IN_SCOPE).value
        if in_scope != "TBD":
            rows_skipped_filled += 1
            continue

        tab = ws.cell(row_idx, COL_FBDI_TAB).value
        if tab is None or not ORACLE_TAB_PATTERN.match(str(tab).strip()):
            rows_skipped_no_match += 1
            continue

        tab_clean = str(tab).strip()
        candidate = "T_" + tab_clean

        if candidate not in csv_tables:
            rows_skipped_no_match += 1
            continue

        # Confirmed match — extract prefix
        fields = csv_lookup[candidate]
        prefix = extract_majority_prefix(fields)

        # Update cells
        ws.cell(row_idx, COL_APPLAUD_TABLE).value = candidate
        ws.cell(row_idx, COL_PREFIX).value = prefix if prefix else ""
        ws.cell(row_idx, COL_IN_SCOPE).value = "YES"

        # Handle notes
        if not prefix:
            note = "No prefix found in CSV fields"
            existing_notes = ws.cell(row_idx, COL_NOTES).value
            if existing_notes:
                ws.cell(row_idx, COL_NOTES).value = str(existing_notes) + "; " + note
            else:
                ws.cell(row_idx, COL_NOTES).value = note

        # T_GL_INTERFACE edge case: note the secondary prefix
        if candidate == "T_GL_INTERFACE":
            prefix_counts = Counter()
            for f in fields:
                m = PREFIX_PATTERN.match(f)
                if m:
                    prefix_counts[m.group(1)] += 1
            if len(prefix_counts) > 1:
                secondary = [
                    (p, c) for p, c in prefix_counts.most_common() if p != prefix
                ]
                if secondary:
                    sec_prefix, sec_count = secondary[0]
                    note = f"Secondary prefix {sec_prefix} ({sec_count} fields) also present"
                    existing_notes = ws.cell(row_idx, COL_NOTES).value
                    if existing_notes:
                        ws.cell(row_idx, COL_NOTES).value = (
                            str(existing_notes) + "; " + note
                        )
                    else:
                        ws.cell(row_idx, COL_NOTES).value = note

        fbdi_file = ws.cell(row_idx, COL_FBDI_FILE).value or ""
        updated_rows.append((fbdi_file, tab_clean, candidate, prefix))

    # Step 4: Save (try in-place first, fall back to _enriched copy if locked)
    try:
        wb.save(MAPPING_PATH)
        save_path = MAPPING_PATH
    except PermissionError:
        save_path = MAPPING_PATH.with_stem(MAPPING_PATH.stem + "_enriched")
        wb.save(save_path)
        print(f"  NOTE: Original file was locked. Saved to: {save_path.name}")
        print("  Close Excel and rename to replace the original.")

    # Step 5: Console output
    print("Enrichment complete.")
    print(f"  Rows updated:   {len(updated_rows)}")
    print(f"  Rows skipped (no match): {rows_skipped_no_match}")
    print(f"  Rows skipped (already filled): {rows_skipped_filled}")
    print()
    print("Updated rows:")
    for fbdi_file, tab, table, prefix in sorted(updated_rows):
        print(f"  {fbdi_file} | {tab} -> {table} (prefix={prefix})")


if __name__ == "__main__":
    enrich()
