"""Complete the FBDI–Applaud mapping: exact matches, inferred matches, and NO_MATCH."""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

SCRIPT_DIR = Path(__file__).resolve().parent
MAPPING_PATH = SCRIPT_DIR / "fbdi_applaud_mapping.xlsx"
COVERAGE_PATH = SCRIPT_DIR / "applaud_table_coverage.csv"
CSV_PATH = SCRIPT_DIR / "AP5TBFLD.CSV"
SHEET_NAME = "FBDI Mapping"

# Column indices (1-based) in mapping sheet
COL_FBDI_FILE = 1
COL_FBDI_TAB = 2
COL_APPLAUD_TABLE = 3
COL_PREFIX = 4
COL_IN_SCOPE = 5
COL_MODULE = 6
COL_NOTES = 7
COL_MATCH_TYPE = 8
COL_CONFIDENCE = 9

# Fills and fonts for formatting
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_BOLD = Font(bold=True)
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_LOW = Font(bold=True, color="C00000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def build_applaud_lookup() -> dict[str, str]:
    """Return {applaud_table: prefix} from coverage CSV."""
    cov = pd.read_csv(COVERAGE_PATH)
    lookup = {}
    for _, row in cov.iterrows():
        t = str(row["applaud_table"]).strip()
        p = str(row["prefix"]).strip() if pd.notna(row["prefix"]) else ""
        lookup[t] = p
    return lookup


def build_semantic_map() -> dict[tuple[str, str], tuple[str, str, str]]:
    """
    Build manual semantic mapping: (fbdi_file, fbdi_tab) -> (applaud_table, confidence, note).
    Only for human-readable tabs that can't be matched mechanically.
    """
    m: dict[tuple[str, str], tuple[str, str, str]] = {}

    def add(file: str, tab: str, table: str, conf: str, note: str) -> None:
        m[(file, tab)] = (table, conf, note)

    # --- Awards ---
    add("ImportAwards", "Awards", "T_AWARDS", "HIGH", "Semantic: Awards tab maps to T_AWARDS")
    add("ImportAwards", "Award Budget Periods", "T_AWARD_BUDGET_PERIODS", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Certifications", "T_AWARD_CERTIFICATIONS", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Assistance Listing Number", "T_AWARD_FED_DOM_ASSIST_PRG", "MEDIUM", "Semantic: Assistance Listing = Federal Domestic Assistance Program")
    add("ImportAwards", "Award Funding", "T_AWARD_FUNDING", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Funding Allocations", "T_AWARD_FUNDING_ALLOCATIONS", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Funding Sources", "T_AWARD_FUNDING_SOURCE", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Keywords", "T_AWARD_KEYWORDS", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Organization Credits", "T_AWARD_ORGANIZATION_CREDITS", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Personnel", "T_AWARD_PERSONNEL", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Prj Task Burden Schedules", "T_AWARD_PRJ_TASK_BURDEN_SCHED", "HIGH", "Semantic: direct name match (truncated)")
    add("ImportAwards", "Award Project Funding Sources", "T_AWARD_PROJECT_FUNDING_SOURCE", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Projects", "T_AWARD_PROJECTS", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award References", "T_AWARD_REFERENCES", "HIGH", "Semantic: direct name match")
    add("ImportAwards", "Award Terms and Conditions", "T_AWARD_TERMS_AND_CONDITIONS", "HIGH", "Semantic: direct name match")

    # --- Grants Personnel / Sponsors ---
    add("ImportGrantsPersonnel", "Grants Personnel", "T_GMS_PERSONNEL_INT", "HIGH", "Semantic: GMS = Grants Management System personnel interface")
    add("ImportGrantsPersonnel", "Grants Personnel Keywords", "T_AWARD_KEYWORDS", "MEDIUM", "Semantic: keywords shared with Awards module")
    add("ImportFundingSources", "Sponsors", "T_GMS_SPONSORS_INT", "HIGH", "Semantic: GMS sponsors interface")

    # --- Work Orders (discrete) ---
    add("WorkOrderTemplate", "Work Order Batches", "T_WO_BATCHES", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Header", "T_WO_HEADER", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Operations", "T_WO_OPERATIONS", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Operation Materials", "T_WO_OPERATION_MATERIALS", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Operation Resources", "T_WO_OPERATION_RESOURCES", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Opn Res Instances", "T_WO_OPN_RES_INSTANCES", "HIGH", "Semantic: direct name match (truncated)")
    add("WorkOrderTemplate", "Work Order Assembly Component", "T_WO_ASSEMBLY_COMPONENT", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Material Lot Numbers", "T_WO_MATERIAL_LOT_NUMBERS", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Matl Serial Numbers", "T_WO_MATL_SERIAL_NUMBER", "HIGH", "Semantic: direct name match (truncated)")
    add("WorkOrderTemplate", "Work Order Operation Outputs", "T_WO_OPERATIONS", "LOW", "Semantic: Operation Outputs may map to WO_OPERATIONS or separate table not in Applaud")
    add("WorkOrderTemplate", "Work Order Product Lot Numbers", "T_WO_PRODUCT_LOT_NUMBERS", "HIGH", "Semantic: direct name match")
    add("WorkOrderTemplate", "Work Order Serial Numbers", "T_WO_SERIAL_NUMBERS", "HIGH", "Semantic: direct name match")

    # --- Process Work Orders ---
    add("ProcessWorkOrderTemplate", "Work Order Batches", "T_WO_BATCHES", "HIGH", "Semantic: process WO uses same batch table")
    add("ProcessWorkOrderTemplate", "Work Order Header", "T_WO_HEADER", "HIGH", "Semantic: process WO uses same header table")
    add("ProcessWorkOrderTemplate", "Work Order Operations", "T_WO_OPERATIONS", "HIGH", "Semantic: process WO uses same ops table")
    add("ProcessWorkOrderTemplate", "Work Order Operation Materials", "T_WO_OPERATION_MATERIALS", "HIGH", "Semantic: process WO uses same materials table")
    add("ProcessWorkOrderTemplate", "Work Order Operation Resources", "T_WO_OPERATION_RESOURCES", "HIGH", "Semantic: process WO uses same resources table")
    add("ProcessWorkOrderTemplate", "Work Order Opn Res Instances", "T_WO_OPN_RES_INSTANCES", "HIGH", "Semantic: process WO uses same resource instances table")
    add("ProcessWorkOrderTemplate", "Work Order Operation Outputs", "T_WO_OPERATIONS", "LOW", "Semantic: Operation Outputs may map to WO_OPERATIONS")
    add("ProcessWorkOrderTemplate", "Work Order Product Lots", "T_WO_PRODUCT_LOT_NUMBERS", "HIGH", "Semantic: direct name match")

    # --- Work Definitions ---
    add("WorkDefinitionTemplate", "Work Definition Headers", "T_WORK_DEFINITION_AND_OPS", "HIGH", "Semantic: WD headers stored in T_WORK_DEFINITION_AND_OPS")
    add("WorkDefinitionTemplate", "Work Definition Operations", "T_WORK_DEFINITION_AND_OPS", "HIGH", "Semantic: WD ops stored in same table")
    add("WorkDefinitionTemplate", "Operation Items - Standard", "T_WIS_WD_DETAILS_INT-STNDCMP", "HIGH", "Semantic: Standard component items")
    add("WorkDefinitionTemplate", "Operation Items - ATO Model", "T_WIS_WD_DETAILS_INT-ATO-CMP", "HIGH", "Semantic: ATO model component items")
    add("WorkDefinitionTemplate", "Operation Resources", "T_OPERATION_RESOURCES", "HIGH", "Semantic: direct name match")
    add("WorkDefinitionTemplate", "Operation Outputs", "T_OPERATION_ITEMS", "MEDIUM", "Semantic: operation outputs stored in T_OPERATION_ITEMS")
    add("WorkDefinitionTemplate", "Operation Alternate Resources", "T_WIS_WD_DETAILS_INT-ARES", "HIGH", "Semantic: ARES = Alternate Resources")
    add("WorkDefinitionTemplate", "Import Batches", "T_WIE_INT_BATCHES_VL", "HIGH", "Semantic: WIE batch interface table")

    # --- Process Work Definitions ---
    add("ProcessWorkDefinitionTemplate", "Work Definition Operations", "T_WIS_WD_DETAILS_INT-OPS", "HIGH", "Semantic: process WD operations")
    add("ProcessWorkDefinitionTemplate", "Operation Resources", "T_WIS_WD_DETAILS_INT-RSC", "HIGH", "Semantic: process WD resources (RSC)")
    add("ProcessWorkDefinitionTemplate", "Operation Items", "T_OPERATION_ITEMS", "HIGH", "Semantic: direct name match")
    add("ProcessWorkDefinitionTemplate", "Operation Outputs", "T_OPERATION_ITEMS", "MEDIUM", "Semantic: operation outputs stored in T_OPERATION_ITEMS")
    add("ProcessWorkDefinitionTemplate", "Operation Alternate Resources", "T_WIS_WD_DETAILS_INT-ARES", "HIGH", "Semantic: ARES = Alternate Resources")
    add("ProcessWorkDefinitionTemplate", "Import Batches", "T_WIE_INT_BATCHES_VL", "HIGH", "Semantic: WIE batch interface table")

    # --- Maintenance Work Definitions ---
    add("MaintenanceWorkDefinitionTemplate", "Work Definitions", "T_WORK_DEFINITION_AND_OPS", "HIGH", "Semantic: maintenance WD uses same table")
    add("MaintenanceWorkDefinitionTemplate", "Work Definition Operations", "T_WORK_DEFINITION_AND_OPS", "HIGH", "Semantic: WD ops stored in same table")
    add("MaintenanceWorkDefinitionTemplate", "Operations Resources", "T_OPERATION_RESOURCES", "HIGH", "Semantic: direct name match")
    add("MaintenanceWorkDefinitionTemplate", "Operations Materials", "T_OPERATION_ITEMS", "MEDIUM", "Semantic: operation materials stored in T_OPERATION_ITEMS")
    add("MaintenanceWorkDefinitionTemplate", "Import Batches", "T_WIE_INT_BATCHES_VL", "HIGH", "Semantic: WIE batch interface table")

    # --- Maintenance Work Orders ---
    add("MaintenanceWorkOrderTemplate", "Work orders", "T_WO_HEADER", "HIGH", "Semantic: work order header table")
    add("MaintenanceWorkOrderTemplate", "Work Order Operations", "T_WO_OPERATIONS", "HIGH", "Semantic: direct name match")
    add("MaintenanceWorkOrderTemplate", "Operation Materials", "T_WO_OPERATION_MATERIALS", "HIGH", "Semantic: direct name match")
    add("MaintenanceWorkOrderTemplate", "Operation resources", "T_WO_OPERATION_RESOURCES", "HIGH", "Semantic: direct name match")
    add("MaintenanceWorkOrderTemplate", "Operation Resource instances", "T_WO_OPN_RES_INSTANCES", "HIGH", "Semantic: resource instances table")
    add("MaintenanceWorkOrderTemplate", "Work order Asset", "T_CSE_ASSETS_INT", "MEDIUM", "Semantic: maintenance asset interface")
    add("MaintenanceWorkOrderTemplate", "Import Batches", "T_WO_BATCHES", "HIGH", "Semantic: WO batch table")

    # --- Projects ---
    add("ProjectImportTemplate", "Projects", "T_PROJECTS", "HIGH", "Semantic: direct name match")
    add("ProjectImportTemplate", "Project Classifications", "T_PROJECT_CLASSIFICATIONS", "HIGH", "Semantic: direct name match")
    add("ProjectImportTemplate", "Project Team Members", "T_PROJECT_TEAM_MEMBERS", "HIGH", "Semantic: direct name match")

    # --- Project Billing Events ---
    add("CreateBillingEventsTemplate", "Project Billing Events", "T_PJB_BILLING_EVENTS_INT", "HIGH", "Semantic: PJB = Project Billing events interface")

    # --- Supplier Address / Contact ---
    add("UploadCustomersTemplate", "Customers", "T_CUSTOMERS", "HIGH", "Semantic: direct name match")
    add("UploadCustomersTemplate", "Contacts", "T_CONTACTS", "HIGH", "Semantic: direct name match")
    add("UploadCustomersTemplate", "Reference Accounts", "T_REFERENCE_ACCOUNTS", "HIGH", "Semantic: direct name match")
    add("UploadCustomersTemplate", "Customer Bank Accounts", "T_IBY_TEMP_EXT_BANK_ACCT", "MEDIUM", "Semantic: customer bank accounts stored in IBY external bank account table")

    # --- DOO Order additional tabs ---
    # (DOO_ORDER_CHARGE_COMPS, DOO_ORDER_HDRS_ALL_EFF_B, DOO_ORDER_LINES_EFF_B are Oracle-style, handled by exact match)

    # --- Supply Chain Planning (Scp*) tabs ---
    add("ScpSalesOrderImportTemplate", "SalesOrder_", "T_SCP_SALESORDER", "HIGH", "Semantic: SCP sales order staging table")
    add("ScpExternalForecastImportTemplate", "ExternalForecast_", "T_SCP_EXTERNALFORECAST", "HIGH", "Semantic: SCP external forecast staging table")
    add("ScpSafetyStockLevelImportTemplate", "SafetyStockLevel_", "T_SAFETYSTOCKLEVEL", "HIGH", "Semantic: direct name match")
    add("ScpUOMImportTemplate", "UOMConversion_", "T_SCP_UOM_CONVERSION", "HIGH", "Semantic: SCP UOM conversion table")
    add("ScpBookingHistoryImportTemplate", "BookingHistory_", "T_MSC_ST_MEASURE_DATA_BOOKINGS", "HIGH", "Semantic: MSC booking history measure data")
    add("ScpSourcingImportTemplate", "SourcingRules_", "T_MSC_ST_SOURCING_RULES", "HIGH", "Semantic: MSC sourcing rules staging table")
    add("ScpSourcingImportTemplate", "AssignmentSets_", "T_MSC_ST_ASSIGNMENT_SETS", "HIGH", "Semantic: MSC assignment sets staging table")
    add("ScpForecastMeasureImportTemplate", "DPForecasts_", "T_MSC_ST_MEASURE_DATA", "MEDIUM", "Semantic: demand planning forecasts -> MSC measure data")

    # --- Transfer Orders ---
    add("DosSupplyOrderImportTemplate", "Transfer Order Lines", "T_TRANSFERORDERLINES", "HIGH", "Semantic: direct name match")
    add("DosSupplyOrderImportTemplate", "Additional Transfer Order Costs", "T_ADDITIONALTRANSFERORDERCOST", "HIGH", "Semantic: direct name match")

    # --- CSE (Installed Base) ---
    add("CseGenealogyBulkImport", "Import Batches", "T_CSE_INT_BATCHES_B", "HIGH", "Semantic: CSE batch interface table")

    # --- Standard Cost ---
    add("StandardCostImportTemplate", "CST_INTERFACE_STD_COST_DETAILS", "T_CST_INTR_STD_COST_DETAIL", "HIGH", "Semantic: CST standard cost detail (name truncated in Applaud)")
    add("StandardCostImportTemplate", "CST_INTERFACE_STD_COST_HEADERS", "T_CST_INTR_STD_COST_HEADERS", "HIGH", "Semantic: CST standard cost headers (name truncated in Applaud)")

    # --- RA Interface Distributions ---
    # This is an Oracle-style tab that doesn't match because Applaud table drops _ALL suffix

    # --- GL Budgets ---
    # GL_BUDGETS_INTERFACE is Oracle-style, handled by exact match

    # --- Supplier sites ---
    # POZ_SUPPLIER_SITES_INT — check if there's an Oracle-style tab for it

    # --- Banks ---
    add("RapidImplementationForCashManagement", "Bank Account", "T_BANKS", "MEDIUM", "Semantic: bank account setup -> T_BANKS")

    # --- Supplier Address/Contact/Sites (Oracle-style tabs with non-exact Applaud names) ---
    add("SupplierAddressImportTemplate", "POZ_SUPPLIER_ADDRESSES_INT", "T_POZ_SUP_ADDRESSES_INT", "HIGH", "Name variant: POZ_SUPPLIER_ADDRESSES_INT -> T_POZ_SUP_ADDRESSES_INT (abbreviated)")
    add("SupplierContactImportTemplate", "POZ_SUP_CONTACTS", "T_POZ_SUP_CONTACTS_INT", "HIGH", "Name variant: POZ_SUP_CONTACTS -> T_POZ_SUP_CONTACTS_INT (suffix added)")
    add("SupplierContactImportTemplate", "POZ_SUPP_CONTACT_ADDRESSES_INT", "T_POZ_SUP_CONTACT_ADDRESS_INT", "HIGH", "Name variant: POZ_SUPP_CONTACT_ADDRESSES_INT -> T_POZ_SUP_CONTACT_ADDRESS_INT (abbreviated)")

    # --- Customer / HZ tabs with naming variants ---
    add("CustomerImportTemplate", "HZ_IMP_PARTYSITES_T", "T_HZ_IMP_PARTY_SITES_T", "HIGH", "Name variant: HZ_IMP_PARTYSITES_T -> T_HZ_IMP_PARTY_SITES_T (underscore added)")

    # --- Supplier Bank Accounts ---
    add("SupplierBankAccountImportTemplate", "IBY_TEMP_EXT_BANK_ACCTS", "T_IBY_TEMP_EXT_BANK_ACCT", "HIGH", "Name variant: IBY_TEMP_EXT_BANK_ACCTS -> T_IBY_TEMP_EXT_BANK_ACCT (plural dropped)")

    # --- Item tabs ---
    add("ItemImportTemplate", "EGP_ITEM_CATEGORIES_INTERFACE", "T_EGP_ITEM_CATEGORIES_INT", "HIGH", "Name variant: INTERFACE truncated to INT in Applaud")
    add("ItemImportTemplate", "EGP_ITEM_REVISIONS_INTERFACE", "T_EGP_ITEM_REVISION_INT", "HIGH", "Name variant: REVISIONS->REVISION, INTERFACE->INT")
    add("ItemImportTemplate", "EGP_TRADING_PARTNER_ITEMS_INTF", "T_EGP_TRADING_PARTNER_ITEMS", "HIGH", "Name variant: _INTF suffix dropped in Applaud")

    # --- RA Interface Distributions ---
    add("AutoInvoiceImportTemplate", "RA_INTERFACE_DISTRIBUTIONS_ALL", "T_RA_INTERFACE_DISTRIBUTIONS", "HIGH", "Name variant: _ALL suffix dropped in Applaud")

    # --- GL Budgets ---
    add("GeneralLedgerBudgetBalanceImportTemplate", "GL_BUDGET_INTERFACE", "T_GL_BUDGETS_INTERFACE", "HIGH", "Name variant: GL_BUDGET -> GL_BUDGETS (plural)")

    # --- GL Segment Values ---
    add("ChartofAccountsSegmentValuesandHierarchiesImportTemplate", "GL_SEGMENT_VALUES_INTERFACE", "T_GL_SEGMENT_VALUES_INT", "HIGH", "Name variant: INTERFACE->INT")

    # --- FA tabs ---
    add("FixedAssetMassAdjustmentsImportTemplate", "FA_ADJUSTMENTS_T", "T_FA_ADJUSTMENTS", "HIGH", "Name variant: FA_ADJUSTMENTS_T -> T_FA_ADJUSTMENTS (_T suffix position changed)")
    add("FixedAssetMassRevaluationsImportTemplate", "FA_ADJUSTMENTS_T", "T_FA_ADJUSTMENTS", "HIGH", "Name variant: FA_ADJUSTMENTS_T -> T_FA_ADJUSTMENTS (revaluations use same table)")

    # --- INV Serial Numbers (non-exact variant) ---
    add("InventoryTransactionImportTemplate", "INV_TRANSACTION_LOTS_INTERFACE", "T_INV_TRANSACT_LOTS_INTERFACE", "HIGH", "Name variant: TRANSACTION->TRANSACT (truncated in Applaud)")
    add("InterfacedPickTransactionsImportTemplate", "INV_TRANSACTION_LOTS_INTERFACE", "T_INV_TRANSACT_LOTS_INTERFACE", "HIGH", "Name variant: TRANSACTION->TRANSACT (truncated in Applaud)")
    add("PerformShippingTransactionImportTemplate", "INV_TRANSACTION_LOTS_INTERFACE", "T_INV_TRANSACT_LOTS_INTERFACE", "HIGH", "Name variant: TRANSACTION->TRANSACT")
    add("ReceivingReceiptImportTemplate", "INV_TRANSACTION_LOTS_INTERFACE", "T_INV_TRANSACT_LOTS_INTERFACE", "HIGH", "Name variant: TRANSACTION->TRANSACT")

    # --- BPA PO tabs ---
    add("POBlanketPurchaseAgreementImportTemplate", "PO_GA_ORG_ASSIGN_INTERFACE", "T_BPA_PO_GA_ORG_ASSIGN_INTERFA", "MEDIUM", "Semantic: BPA org assignment -> T_BPA variant (truncated)")
    add("POContractPurchaseAgreementImportTemplate", "PO_GA_ORG_ASSIGN_INTERFACE", "T_BPA_PO_GA_ORG_ASSIGN_INTERFA", "MEDIUM", "Semantic: Contract PA org assignment -> same BPA table")

    # --- Project tabs ---
    add("ProjectEnterpriseExpenseResourcesImportTemplate", "PJT_PRJ_ENT_RES_INTERFACE", "T_PROJ_ENT_RES_INTERFACE", "HIGH", "Name variant: PJT_PRJ -> PROJ (abbreviated)")
    add("ProjectEnterpriseResourcesImportTemplate", "PJT_PRJ_ENT_RES_INTERFACE", "T_PROJ_ENT_RES_INTERFACE", "HIGH", "Name variant: PJT_PRJ -> PROJ (abbreviated)")
    add("ProjectResourceRequestImportTemplate", "PJR_RES_REQ_INTERFACE", "T_PROJ_RES_REQ_INTERFACE", "HIGH", "Name variant: PJR -> PROJ (abbreviated)")

    # --- Supplier sites ---
    # POZ_SUPPLIER_SITES_INT doesn't have a direct FBDI tab match — it's populated via SupplierSiteImportTemplate which has different tab names

    # --- WSH Transit Times ---
    add("ShipmentRequestImportTemplate", "WSH_TRANSACTIONS_HISTORY", "T_WSH_TRANSIT_TIMES", "LOW", "Speculative: WSH transactions history may map to transit times staging")

    # --- Trading Partners ---
    add("ScpSupplierImportTemplate", "Supplier_", "T_TRADING_PARTNERS_IMPORT", "MEDIUM", "Semantic: SCP supplier -> trading partners import")

    # --- Worker/Person/Assignment (HR templates) ---
    # These are human-readable tabs from HCM Worker Import templates — need specific file names
    # Leaving as NO_MATCH since we don't have HCM FBDI template files in scope

    # --- SO Attachments ---
    add("SourceSalesOrderImportTemplate", "DOO_ORDER_DOC_REFERENCES_INT", "T_SO_ATTACHMENTS", "LOW", "Speculative: order doc references may map to SO_ATTACHMENTS")

    return m


def try_exact_match(tab: str, applaud_lookup: dict[str, str]) -> str | None:
    """Try exact matching with variant suffixes. Returns applaud_table or None."""
    tab_upper = tab.upper().strip()
    candidates = [
        "T_" + tab_upper,
        "T_" + tab_upper + "_T",
    ]
    # Also try stripping trailing _INT, _ALL, _V
    for suffix in ("_INT", "_ALL", "_V"):
        if tab_upper.endswith(suffix):
            stripped = tab_upper[: -len(suffix)]
            candidates.append("T_" + stripped)
            candidates.append("T_" + stripped + "_T")

    # Case-insensitive check
    upper_to_real = {k.upper(): k for k in applaud_lookup}
    for c in candidates:
        if c.upper() in upper_to_real:
            return upper_to_real[c.upper()]
    return None


def run() -> None:
    applaud_lookup = build_applaud_lookup()
    semantic_map = build_semantic_map()

    wb = openpyxl.load_workbook(MAPPING_PATH)
    ws = wb[SHEET_NAME]

    # Add headers for new columns
    ws.cell(1, COL_MATCH_TYPE).value = "match_type"
    ws.cell(1, COL_CONFIDENCE).value = "confidence"

    # Track stats
    exact_count = 0
    inferred_count = 0
    no_match_count = 0
    updated_applaud_tables: dict[str, list[str]] = {}  # applaud_table -> [fbdi descriptions]

    oracle_tab_pattern = re.compile(r"^[A-Z][A-Z0-9_]+$")

    for row_idx in range(2, ws.max_row + 1):
        in_scope = ws.cell(row_idx, COL_IN_SCOPE).value
        if in_scope != "TBD":
            # Preserve existing match_type if already set; only back-fill for
            # rows that were YES before this script existed (original 87)
            if in_scope == "YES" and not ws.cell(row_idx, COL_MATCH_TYPE).value:
                ws.cell(row_idx, COL_MATCH_TYPE).value = "EXACT"
                ws.cell(row_idx, COL_CONFIDENCE).value = "HIGH"
            continue

        fbdi_file = ws.cell(row_idx, COL_FBDI_FILE).value or ""
        fbdi_tab = ws.cell(row_idx, COL_FBDI_TAB).value or ""
        fbdi_tab_clean = fbdi_tab.strip()

        matched = False

        # --- Pass 1: Exact match (Oracle-style tab names) ---
        if oracle_tab_pattern.match(fbdi_tab_clean):
            applaud_table = try_exact_match(fbdi_tab_clean, applaud_lookup)
            if applaud_table:
                prefix = applaud_lookup.get(applaud_table, "")
                ws.cell(row_idx, COL_APPLAUD_TABLE).value = applaud_table
                ws.cell(row_idx, COL_PREFIX).value = prefix
                ws.cell(row_idx, COL_IN_SCOPE).value = "YES"
                ws.cell(row_idx, COL_MATCH_TYPE).value = "EXACT"
                ws.cell(row_idx, COL_CONFIDENCE).value = "HIGH"
                exact_count += 1
                matched = True
                updated_applaud_tables.setdefault(applaud_table, []).append(
                    f"{fbdi_file} / {fbdi_tab_clean}"
                )

        # --- Pass 2: Semantic/inferred match ---
        if not matched:
            key = (fbdi_file.strip(), fbdi_tab_clean)
            if key in semantic_map:
                applaud_table, confidence, note = semantic_map[key]
                if applaud_table in applaud_lookup:
                    prefix = applaud_lookup[applaud_table]
                    ws.cell(row_idx, COL_APPLAUD_TABLE).value = applaud_table
                    ws.cell(row_idx, COL_PREFIX).value = prefix
                    ws.cell(row_idx, COL_IN_SCOPE).value = "YES"
                    ws.cell(row_idx, COL_MATCH_TYPE).value = "INFERRED"
                    ws.cell(row_idx, COL_CONFIDENCE).value = confidence
                    # Append note
                    existing = ws.cell(row_idx, COL_NOTES).value
                    if existing:
                        ws.cell(row_idx, COL_NOTES).value = str(existing) + "; " + note
                    else:
                        ws.cell(row_idx, COL_NOTES).value = note
                    inferred_count += 1
                    matched = True
                    updated_applaud_tables.setdefault(applaud_table, []).append(
                        f"{fbdi_file} / {fbdi_tab_clean}"
                    )

        # --- Pass 3: NO_MATCH ---
        if not matched:
            ws.cell(row_idx, COL_MATCH_TYPE).value = "NO_MATCH"
            # Leave confidence blank, leave in_scope as TBD
            no_match_count += 1

    # --- Check for prefix conflict: T_POZ_SUP_THIRDPARTY_INT vs T_AWARDS both TE1 ---
    for row_idx in range(2, ws.max_row + 1):
        at = ws.cell(row_idx, COL_APPLAUD_TABLE).value
        if at in ("T_POZ_SUP_THIRDPARTY_INT", "T_AWARDS"):
            existing = ws.cell(row_idx, COL_NOTES).value
            conflict_note = "Prefix TE1 shared with T_AWARDS" if at == "T_POZ_SUP_THIRDPARTY_INT" else "Prefix TE1 shared with T_POZ_SUP_THIRDPARTY_INT"
            if existing:
                if conflict_note not in str(existing):
                    ws.cell(row_idx, COL_NOTES).value = str(existing) + "; " + conflict_note
            else:
                ws.cell(row_idx, COL_NOTES).value = conflict_note

    # --- Apply formatting ---
    # Header row
    for col in range(1, COL_CONFIDENCE + 1):
        cell = ws.cell(1, col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx in range(2, ws.max_row + 1):
        match_type = ws.cell(row_idx, COL_MATCH_TYPE).value
        confidence = ws.cell(row_idx, COL_CONFIDENCE).value

        # Apply row fill based on match_type
        if match_type == "EXACT":
            for col in range(1, COL_CONFIDENCE + 1):
                ws.cell(row_idx, col).fill = FILL_GREEN
                ws.cell(row_idx, col).border = THIN_BORDER
        elif match_type == "INFERRED":
            for col in range(1, COL_CONFIDENCE + 1):
                ws.cell(row_idx, col).fill = FILL_YELLOW
                ws.cell(row_idx, col).border = THIN_BORDER
        else:
            for col in range(1, COL_CONFIDENCE + 1):
                ws.cell(row_idx, col).border = THIN_BORDER

        # Bold font for LOW confidence
        if confidence == "LOW":
            for col in range(1, COL_CONFIDENCE + 1):
                ws.cell(row_idx, col).font = FONT_LOW

    # Set column widths for readability
    col_widths = {
        COL_FBDI_FILE: 45,
        COL_FBDI_TAB: 40,
        COL_APPLAUD_TABLE: 42,
        COL_PREFIX: 10,
        COL_IN_SCOPE: 14,
        COL_MODULE: 20,
        COL_NOTES: 60,
        COL_MATCH_TYPE: 14,
        COL_CONFIDENCE: 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(COL_CONFIDENCE)}{ws.max_row}"

    # Save
    try:
        wb.save(MAPPING_PATH)
        print(f"Saved: {MAPPING_PATH.name}")
    except PermissionError:
        fallback = MAPPING_PATH.with_stem(MAPPING_PATH.stem + "_completed")
        wb.save(fallback)
        print(f"Original locked. Saved: {fallback.name}")
        print("Close Excel and rename to replace the original.")

    # --- Update coverage CSV ---
    # Reload to get all mapped tables (including pre-existing YES)
    wb2 = openpyxl.load_workbook(MAPPING_PATH if MAPPING_PATH.exists() else fallback)
    ws2 = wb2[SHEET_NAME]
    all_mapped: dict[str, list[str]] = {}
    for r in range(2, ws2.max_row + 1):
        at = ws2.cell(r, COL_APPLAUD_TABLE).value
        scope = ws2.cell(r, COL_IN_SCOPE).value
        if at and scope == "YES":
            f = ws2.cell(r, COL_FBDI_FILE).value or ""
            t = ws2.cell(r, COL_FBDI_TAB).value or ""
            all_mapped.setdefault(at.strip(), []).append(f"{f} / {t}")

    cov = pd.read_csv(COVERAGE_PATH)
    for idx, row in cov.iterrows():
        t = str(row["applaud_table"]).strip()
        if t in all_mapped:
            cov.at[idx, "status"] = "MAPPED"
            cov.at[idx, "fbdi_mappings"] = "; ".join(all_mapped[t])
    cov.to_csv(COVERAGE_PATH, index=False)
    print(f"Updated: {COVERAGE_PATH.name}")

    # --- Summary ---
    mapped_count = len([t for t in cov["status"] if t == "MAPPED"])
    unmapped_count = len([t for t in cov["status"] if t == "UNMAPPED"])

    print(f"\n{'='*60}")
    print(f"MAPPING COMPLETION SUMMARY")
    print(f"{'='*60}")
    print(f"  New EXACT matches:    {exact_count}")
    print(f"  New INFERRED matches: {inferred_count}")
    print(f"  NO_MATCH rows:        {no_match_count}")
    print(f"  Applaud tables now MAPPED:   {mapped_count} / {len(cov)}")
    print(f"  Applaud tables still UNMAPPED: {unmapped_count}")
    print(f"{'='*60}")


if __name__ == "__main__":
    run()
