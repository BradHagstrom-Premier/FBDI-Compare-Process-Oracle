"""
Format fbdi_applaud_mapping.xlsx and applaud_table_coverage (CSV→XLSX)
with a polished, executive-grade design.

Design: "Corporate Editorial" — dark navy headers, clean alternating bands,
strategic color coding, summary dashboard sheet.
"""

import datetime
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
import pandas as pd

SCRIPT_DIR = Path(__file__).resolve().parent
MAPPING_PATH = SCRIPT_DIR / "fbdi_applaud_mapping.xlsx"
COVERAGE_CSV = SCRIPT_DIR / "applaud_table_coverage.csv"
COVERAGE_XLSX = SCRIPT_DIR / "applaud_table_coverage.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY = "1B2A4A"
DARK_NAVY = "0F1D35"
SLATE = "2C3E6B"
GOLD = "D4A843"
LIGHT_GOLD = "F5E6C4"
WHITE = "FFFFFF"
OFF_WHITE = "F7F8FA"
LIGHT_GRAY = "E8EBF0"
MID_GRAY = "B0B8C8"
CHARCOAL = "3A3F4B"
SOFT_GREEN = "D4EDDA"
DEEP_GREEN = "1D6F42"
SOFT_YELLOW = "FFF3CD"
AMBER = "B8860B"
SOFT_RED = "F8D7DA"
CRIMSON = "9B1B30"
SOFT_BLUE = "D6EAF8"
STEEL_BLUE = "34698A"
MINT = "E8F5E9"

# ── Fills ────────────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_SUBHEADER = PatternFill("solid", fgColor=SLATE)
FILL_BAND_ODD = PatternFill("solid", fgColor=OFF_WHITE)
FILL_BAND_EVEN = PatternFill("solid", fgColor=WHITE)
FILL_EXACT = PatternFill("solid", fgColor=SOFT_GREEN)
FILL_INFERRED = PatternFill("solid", fgColor=SOFT_YELLOW)
FILL_NO_MATCH = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_ERROR = PatternFill("solid", fgColor=SOFT_RED)
FILL_MAPPED = PatternFill("solid", fgColor=SOFT_GREEN)
FILL_UNMAPPED = PatternFill("solid", fgColor=SOFT_RED)
FILL_GOLD_ACCENT = PatternFill("solid", fgColor=LIGHT_GOLD)
FILL_TITLE_BG = PatternFill("solid", fgColor=DARK_NAVY)
FILL_KPI_BG = PatternFill("solid", fgColor=OFF_WHITE)
FILL_SOFT_BLUE = PatternFill("solid", fgColor=SOFT_BLUE)

# ── Fonts ────────────────────────────────────────────────────────────────────
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color=WHITE)
FONT_TITLE = Font(name="Calibri", bold=True, size=20, color=GOLD)
FONT_SUBTITLE = Font(name="Calibri", size=12, color=MID_GRAY, italic=True)
FONT_KPI_LABEL = Font(name="Calibri", size=10, color=CHARCOAL)
FONT_KPI_VALUE = Font(name="Calibri", bold=True, size=28, color=NAVY)
FONT_KPI_UNIT = Font(name="Calibri", size=10, color=MID_GRAY)
FONT_DATA = Font(name="Calibri", size=10, color=CHARCOAL)
FONT_DATA_BOLD = Font(name="Calibri", bold=True, size=10, color=CHARCOAL)
FONT_ACCENT = Font(name="Calibri", bold=True, size=10, color=DEEP_GREEN)
FONT_WARN = Font(name="Calibri", bold=True, size=10, color=CRIMSON)
FONT_MUTED = Font(name="Calibri", size=10, color=MID_GRAY)
FONT_LINK = Font(name="Calibri", size=10, color=STEEL_BLUE, underline="single")
FONT_SECTION = Font(name="Calibri", bold=True, size=13, color=NAVY)
FONT_STATUS_YES = Font(name="Calibri", bold=True, size=10, color=DEEP_GREEN)
FONT_STATUS_TBD = Font(name="Calibri", size=10, color=MID_GRAY)
FONT_STATUS_ERR = Font(name="Calibri", bold=True, size=10, color=CRIMSON)
FONT_MAPPED = Font(name="Calibri", bold=True, size=10, color=DEEP_GREEN)
FONT_UNMAPPED = Font(name="Calibri", bold=True, size=10, color=CRIMSON)
FONT_CONFIDENCE_HIGH = Font(name="Calibri", bold=True, size=10, color=DEEP_GREEN)
FONT_CONFIDENCE_MED = Font(name="Calibri", bold=True, size=10, color=AMBER)
FONT_CONFIDENCE_LOW = Font(name="Calibri", bold=True, size=10, color=CRIMSON)

# ── Borders ──────────────────────────────────────────────────────────────────
THIN_BOTTOM = Border(bottom=Side("thin", color=LIGHT_GRAY))
THIN_ALL = Border(
    left=Side("hair", color=LIGHT_GRAY),
    right=Side("hair", color=LIGHT_GRAY),
    top=Side("hair", color=LIGHT_GRAY),
    bottom=Side("hair", color=LIGHT_GRAY),
)
ACCENT_BOTTOM = Border(bottom=Side("medium", color=GOLD))
SECTION_TOP = Border(top=Side("medium", color=NAVY))

# ── Alignments ───────────────────────────────────────────────────────────────
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def style_header_row(ws, max_col, row=1):
    """Apply navy header with gold accent border."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = Border(bottom=Side("medium", color=GOLD))
    ws.row_dimensions[row].height = 36


def style_data_cell(cell, row_idx, is_band=True):
    """Apply alternating band and base data styling."""
    if is_band:
        cell.fill = FILL_BAND_ODD if row_idx % 2 == 0 else FILL_BAND_EVEN
    cell.font = FONT_DATA
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_ALL


def apply_mapping_status_styling(ws, row_idx, match_type, confidence, in_scope):
    """Apply color coding based on match status."""
    max_col = ws.max_column

    # Row fill based on match_type
    if match_type == "EXACT":
        row_fill = FILL_EXACT
    elif match_type == "INFERRED":
        row_fill = FILL_INFERRED
    elif in_scope in ("FILE_TOO_LARGE", "FILE_ERROR"):
        row_fill = FILL_ERROR
    elif match_type == "NO_MATCH":
        row_fill = None  # use alternating bands
    else:
        row_fill = None

    for c in range(1, max_col + 1):
        cell = ws.cell(row_idx, c)
        if row_fill:
            cell.fill = row_fill
        cell.border = THIN_ALL
        cell.alignment = ALIGN_LEFT

    # Status column (in_scope) font
    scope_cell = ws.cell(row_idx, 5)  # in_scope
    if in_scope == "YES":
        scope_cell.font = FONT_STATUS_YES
    elif in_scope == "TBD":
        scope_cell.font = FONT_STATUS_TBD
    elif in_scope in ("FILE_TOO_LARGE", "FILE_ERROR"):
        scope_cell.font = FONT_STATUS_ERR

    # Confidence column font
    conf_cell = ws.cell(row_idx, 9)  # confidence
    if confidence == "HIGH":
        conf_cell.font = FONT_CONFIDENCE_HIGH
    elif confidence == "MEDIUM":
        conf_cell.font = FONT_CONFIDENCE_MED
    elif confidence == "LOW":
        conf_cell.font = FONT_CONFIDENCE_LOW

    # Match type column
    mt_cell = ws.cell(row_idx, 8)
    if match_type == "EXACT":
        mt_cell.font = FONT_ACCENT
    elif match_type == "INFERRED":
        mt_cell.font = FONT_DATA_BOLD
    elif match_type == "NO_MATCH":
        mt_cell.font = FONT_MUTED

    # Center-align specific columns
    for c in [4, 5, 8, 9]:  # prefix, in_scope, match_type, confidence
        ws.cell(row_idx, c).alignment = ALIGN_CENTER


def add_dashboard_sheet(wb, mapping_ws):
    """Add a Summary Dashboard sheet as the first sheet."""
    ds = wb.create_sheet("Dashboard", 0)
    ds.sheet_properties.tabColor = GOLD

    # Gather stats
    total = mapping_ws.max_row - 1
    stats = Counter()
    match_stats = Counter()
    conf_stats = Counter()
    modules = Counter()
    for r in range(2, mapping_ws.max_row + 1):
        scope = mapping_ws.cell(r, 5).value or ""
        mt = mapping_ws.cell(r, 8).value or ""
        conf = mapping_ws.cell(r, 9).value or ""
        mod = mapping_ws.cell(r, 6).value or "Unassigned"
        stats[scope] += 1
        if mt:
            match_stats[mt] += 1
        if conf:
            conf_stats[conf] += 1
        modules[mod] += 1

    yes_count = stats.get("YES", 0)
    tbd_count = stats.get("TBD", 0)
    error_count = stats.get("FILE_TOO_LARGE", 0) + stats.get("FILE_ERROR", 0)
    exact_count = match_stats.get("EXACT", 0)
    inferred_count = match_stats.get("INFERRED", 0)
    no_match_count = match_stats.get("NO_MATCH", 0)
    coverage_pct = round(yes_count / total * 100, 1) if total else 0

    # ── Title banner ─────────────────────────────────────────────────────
    ds.merge_cells("A1:H2")
    title = ds["A1"]
    title.value = "FBDI → Applaud Mapping Dashboard"
    title.font = FONT_TITLE
    title.fill = FILL_TITLE_BG
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for r in range(1, 3):
        for c in range(1, 9):
            ds.cell(r, c).fill = FILL_TITLE_BG
    ds.row_dimensions[1].height = 28
    ds.row_dimensions[2].height = 28

    # Subtitle
    ds.merge_cells("A3:H3")
    sub = ds["A3"]
    sub.value = f"Generated {datetime.date.today().strftime('%B %d, %Y')}  ·  Oracle Cloud FBDI Template Analysis"
    sub.font = FONT_SUBTITLE
    sub.fill = FILL_TITLE_BG
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 9):
        ds.cell(3, c).fill = FILL_TITLE_BG
    ds.row_dimensions[3].height = 24

    # Gold accent line
    for c in range(1, 9):
        ds.cell(4, c).border = ACCENT_BOTTOM
    ds.row_dimensions[4].height = 6

    # ── KPI Cards (row 6–8) ──────────────────────────────────────────────
    kpis = [
        ("A", "B", str(total), "Total FBDI Tabs", ""),
        ("C", "D", str(yes_count), "Mapped to Applaud", f"{coverage_pct}% coverage"),
        ("E", "F", str(exact_count), "Exact Matches", "Structural name match"),
        ("G", "H", str(inferred_count), "Inferred Matches", "Semantic / name variant"),
    ]

    row_start = 6
    for col_start, col_end, value, label, sublabel in kpis:
        # Card background
        for r in range(row_start, row_start + 3):
            for c_letter in (col_start, col_end):
                cell = ds[f"{c_letter}{r}"]
                cell.fill = FILL_KPI_BG
                cell.border = THIN_ALL

        # Value
        ds.merge_cells(f"{col_start}{row_start}:{col_end}{row_start}")
        val_cell = ds[f"{col_start}{row_start}"]
        val_cell.value = value
        val_cell.font = FONT_KPI_VALUE
        val_cell.alignment = Alignment(horizontal="center", vertical="bottom")

        # Label
        ds.merge_cells(f"{col_start}{row_start+1}:{col_end}{row_start+1}")
        lbl_cell = ds[f"{col_start}{row_start+1}"]
        lbl_cell.value = label
        lbl_cell.font = FONT_KPI_LABEL
        lbl_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Sublabel
        ds.merge_cells(f"{col_start}{row_start+2}:{col_end}{row_start+2}")
        sub_cell = ds[f"{col_start}{row_start+2}"]
        sub_cell.value = sublabel
        sub_cell.font = FONT_KPI_UNIT
        sub_cell.alignment = Alignment(horizontal="center", vertical="top")

    ds.row_dimensions[row_start].height = 44
    ds.row_dimensions[row_start + 1].height = 20
    ds.row_dimensions[row_start + 2].height = 18

    # ── Second row of KPIs (row 10–12) ───────────────────────────────────
    kpis2 = [
        ("A", "B", str(no_match_count), "No Applaud Match", "Expected — no counterpart"),
        ("C", "D", str(tbd_count), "Remaining TBD", "Awaiting manual review"),
        ("E", "F", str(error_count), "File Errors", "Corrupt or oversized files"),
        ("G", "H", str(conf_stats.get("LOW", 0)), "Low Confidence", "Flagged for review"),
    ]

    row_start2 = 10
    for col_start, col_end, value, label, sublabel in kpis2:
        for r in range(row_start2, row_start2 + 3):
            for c_letter in (col_start, col_end):
                cell = ds[f"{c_letter}{r}"]
                cell.fill = FILL_KPI_BG
                cell.border = THIN_ALL

        ds.merge_cells(f"{col_start}{row_start2}:{col_end}{row_start2}")
        val_cell = ds[f"{col_start}{row_start2}"]
        val_cell.value = value
        val_cell.font = FONT_KPI_VALUE
        val_cell.alignment = Alignment(horizontal="center", vertical="bottom")

        ds.merge_cells(f"{col_start}{row_start2+1}:{col_end}{row_start2+1}")
        lbl_cell = ds[f"{col_start}{row_start2+1}"]
        lbl_cell.value = label
        lbl_cell.font = FONT_KPI_LABEL
        lbl_cell.alignment = Alignment(horizontal="center", vertical="top")

        ds.merge_cells(f"{col_start}{row_start2+2}:{col_end}{row_start2+2}")
        sub_cell = ds[f"{col_start}{row_start2+2}"]
        sub_cell.value = sublabel
        sub_cell.font = FONT_KPI_UNIT
        sub_cell.alignment = Alignment(horizontal="center", vertical="top")

    ds.row_dimensions[row_start2].height = 44
    ds.row_dimensions[row_start2 + 1].height = 20
    ds.row_dimensions[row_start2 + 2].height = 18

    # ── Match Type Breakdown table (row 15+) ─────────────────────────────
    table_row = 15
    ds.merge_cells(f"A{table_row}:D{table_row}")
    ds[f"A{table_row}"].value = "Match Type Breakdown"
    ds[f"A{table_row}"].font = FONT_SECTION
    ds[f"A{table_row}"].border = Border(bottom=Side("medium", color=NAVY))
    for c in range(1, 5):
        ds.cell(table_row, c).border = Border(bottom=Side("medium", color=NAVY))

    table_row += 1
    headers = ["Match Type", "Count", "% of Total", "Description"]
    for i, h in enumerate(headers, 1):
        cell = ds.cell(table_row, i)
        cell.value = h
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = Border(bottom=Side("medium", color=GOLD))

    breakdown = [
        ("EXACT", exact_count, "Structural T_ + tab_name match", FILL_EXACT, FONT_ACCENT),
        ("INFERRED", inferred_count, "Semantic / name variant match", FILL_INFERRED, FONT_DATA_BOLD),
        ("NO_MATCH", no_match_count, "No Applaud table counterpart", PatternFill("solid", fgColor=LIGHT_GRAY), FONT_MUTED),
        ("FILE_TOO_LARGE", stats.get("FILE_TOO_LARGE", 0), "Skipped — file exceeds size limit", FILL_ERROR, FONT_STATUS_ERR),
        ("FILE_ERROR", stats.get("FILE_ERROR", 0), "Skipped — corrupt or unreadable", FILL_ERROR, FONT_STATUS_ERR),
    ]

    for i, (label, count, desc, fill, font) in enumerate(breakdown):
        r = table_row + 1 + i
        pct = round(count / total * 100, 1) if total else 0
        ds.cell(r, 1).value = label
        ds.cell(r, 1).font = font
        ds.cell(r, 1).fill = fill
        ds.cell(r, 2).value = count
        ds.cell(r, 2).alignment = ALIGN_CENTER
        ds.cell(r, 3).value = f"{pct}%"
        ds.cell(r, 3).alignment = ALIGN_CENTER
        ds.cell(r, 4).value = desc
        ds.cell(r, 4).font = FONT_MUTED
        for c in range(1, 5):
            ds.cell(r, c).border = THIN_ALL
        ds.row_dimensions[r].height = 22

    # ── Pie chart for match types ────────────────────────────────────────
    chart_data_start = table_row + 1
    chart_data_end = table_row + len(breakdown)

    pie = PieChart()
    pie.title = None
    pie.style = 2
    pie.width = 14
    pie.height = 10

    labels = Reference(ds, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
    data = Reference(ds, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)

    # Custom colors for slices
    pie_colors = ["4CAF50", "FFC107", "9E9E9E", "E57373", "EF5350"]
    for i, color in enumerate(pie_colors):
        if i < len(pie.series[0].data_points):
            pass
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False

    ds.add_chart(pie, "E15")

    # ── Confidence distribution (row 23+) ────────────────────────────────
    conf_row = 23
    ds.merge_cells(f"A{conf_row}:D{conf_row}")
    ds[f"A{conf_row}"].value = "Confidence Distribution"
    ds[f"A{conf_row}"].font = FONT_SECTION
    for c in range(1, 5):
        ds.cell(conf_row, c).border = Border(bottom=Side("medium", color=NAVY))

    conf_row += 1
    for i, h in enumerate(["Confidence", "Count", "% of Mapped", "Action Required"], 1):
        cell = ds.cell(conf_row, i)
        cell.value = h
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    conf_breakdown = [
        ("HIGH", conf_stats.get("HIGH", 0), "No review needed", FONT_CONFIDENCE_HIGH),
        ("MEDIUM", conf_stats.get("MEDIUM", 0), "Quick verification recommended", FONT_CONFIDENCE_MED),
        ("LOW", conf_stats.get("LOW", 0), "Manual review required", FONT_CONFIDENCE_LOW),
    ]

    for i, (label, count, action, font) in enumerate(conf_breakdown):
        r = conf_row + 1 + i
        pct = round(count / yes_count * 100, 1) if yes_count else 0
        ds.cell(r, 1).value = label
        ds.cell(r, 1).font = font
        ds.cell(r, 2).value = count
        ds.cell(r, 2).alignment = ALIGN_CENTER
        ds.cell(r, 3).value = f"{pct}%"
        ds.cell(r, 3).alignment = ALIGN_CENTER
        ds.cell(r, 4).value = action
        ds.cell(r, 4).font = FONT_MUTED
        for c in range(1, 5):
            ds.cell(r, c).border = THIN_ALL

    # ── Column widths ────────────────────────────────────────────────────
    for col, width in {1: 20, 2: 16, 3: 16, 4: 38, 5: 16, 6: 16, 7: 16, 8: 16}.items():
        ds.column_dimensions[get_column_letter(col)].width = width

    # ── Footer ───────────────────────────────────────────────────────────
    footer_row = 30
    ds.merge_cells(f"A{footer_row}:H{footer_row}")
    ds[f"A{footer_row}"].value = "Definian  ·  Oracle Cloud FBDI Integration Analysis  ·  Confidential"
    ds[f"A{footer_row}"].font = Font(name="Calibri", size=9, color=MID_GRAY, italic=True)
    ds[f"A{footer_row}"].alignment = Alignment(horizontal="center")

    ds.sheet_view.showGridLines = False
    return ds


def format_mapping_workbook():
    """Format the main mapping workbook."""
    print("Formatting fbdi_applaud_mapping.xlsx...")
    wb = openpyxl.load_workbook(MAPPING_PATH)
    ws = wb["FBDI Mapping"]
    ws.sheet_properties.tabColor = DEEP_GREEN

    # Rename headers for presentation
    display_headers = {
        1: "FBDI Template",
        2: "FBDI Tab",
        3: "Applaud Table",
        4: "Prefix",
        5: "Status",
        6: "Module",
        7: "Notes",
        8: "Match Type",
        9: "Confidence",
    }
    for col, name in display_headers.items():
        ws.cell(1, col).value = name

    # Style header
    style_header_row(ws, 9)

    # Column widths
    widths = {1: 44, 2: 40, 3: 40, 4: 10, 5: 16, 6: 18, 7: 55, 8: 14, 9: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Style data rows
    for r in range(2, ws.max_row + 1):
        in_scope = ws.cell(r, 5).value
        match_type = ws.cell(r, 8).value
        confidence = ws.cell(r, 9).value

        # Base styling with alternating bands
        for c in range(1, 10):
            style_data_cell(ws.cell(r, c), r)

        # Override with status-based styling
        apply_mapping_status_styling(ws, r, match_type, confidence, in_scope)

        ws.row_dimensions[r].height = 22

    # Freeze and filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    # Hide gridlines for cleaner look
    ws.sheet_view.showGridLines = False

    # Add dashboard
    add_dashboard_sheet(wb, ws)

    # Save
    try:
        wb.save(MAPPING_PATH)
        print(f"  OK: Saved {MAPPING_PATH.name}")
    except PermissionError:
        alt = MAPPING_PATH.with_stem(MAPPING_PATH.stem + "_formatted")
        wb.save(alt)
        print(f"  ! Locked — saved as {alt.name}")


def format_coverage_workbook():
    """Convert coverage CSV to a polished XLSX."""
    print("Creating applaud_table_coverage.xlsx...")
    df = pd.read_csv(COVERAGE_CSV)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applaud Coverage"
    ws.sheet_properties.tabColor = STEEL_BLUE

    # ── Title row ────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"].value = "Applaud Table Coverage Analysis"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=GOLD)
    ws["A1"].fill = FILL_TITLE_BG
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 6):
        ws.cell(1, c).fill = FILL_TITLE_BG
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    mapped_count = (df["status"] == "MAPPED").sum()
    total_count = len(df)
    ws["A2"].value = f"{mapped_count} of {total_count} Applaud tables mapped to FBDI templates  ·  {round(mapped_count/total_count*100, 1)}% coverage"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].fill = FILL_TITLE_BG
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, 6):
        ws.cell(2, c).fill = FILL_TITLE_BG
    ws.row_dimensions[2].height = 24

    # Gold accent
    for c in range(1, 6):
        ws.cell(3, c).border = ACCENT_BOTTOM
    ws.row_dimensions[3].height = 6

    # ── Headers (row 4) ──────────────────────────────────────────────────
    header_row = 4
    headers = ["#", "Applaud Table", "Status", "Prefix", "FBDI Template Mappings"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(header_row, i)
        cell.value = h
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = Border(bottom=Side("medium", color=GOLD))
    ws.row_dimensions[header_row].height = 32

    # ── Data rows ────────────────────────────────────────────────────────
    # Sort: MAPPED first, then UNMAPPED
    df_sorted = df.sort_values(
        ["status", "applaud_table"], ascending=[True, True]
    ).reset_index(drop=True)

    # Find boundary between MAPPED and UNMAPPED
    mapped_end = (df_sorted["status"] == "MAPPED").sum()

    for idx, (_, row) in enumerate(df_sorted.iterrows()):
        r = header_row + 1 + idx
        is_mapped = row["status"] == "MAPPED"

        # Row number
        ws.cell(r, 1).value = idx + 1
        ws.cell(r, 1).font = FONT_MUTED
        ws.cell(r, 1).alignment = ALIGN_CENTER

        # Table name
        ws.cell(r, 2).value = row["applaud_table"]
        ws.cell(r, 2).font = FONT_DATA_BOLD if is_mapped else FONT_DATA

        # Status
        ws.cell(r, 3).value = row["status"]
        if is_mapped:
            ws.cell(r, 3).font = FONT_MAPPED
            ws.cell(r, 3).fill = FILL_MAPPED
        else:
            ws.cell(r, 3).font = FONT_UNMAPPED
            ws.cell(r, 3).fill = FILL_UNMAPPED
        ws.cell(r, 3).alignment = ALIGN_CENTER

        # Prefix
        prefix = row["prefix"]
        ws.cell(r, 4).value = prefix if pd.notna(prefix) else ""
        ws.cell(r, 4).alignment = ALIGN_CENTER
        ws.cell(r, 4).font = FONT_DATA_BOLD

        # Mappings
        mappings = row["fbdi_mappings"]
        ws.cell(r, 5).value = str(mappings) if pd.notna(mappings) else ""
        ws.cell(r, 5).font = FONT_DATA if is_mapped else FONT_MUTED

        # Alternating bands
        band_fill = FILL_BAND_ODD if idx % 2 == 0 else FILL_BAND_EVEN
        for c in range(1, 6):
            cell = ws.cell(r, c)
            if c != 3:  # Don't override status fill
                cell.fill = band_fill
            cell.border = THIN_ALL

        # Section divider between MAPPED and UNMAPPED
        if idx == mapped_end:
            for c in range(1, 6):
                ws.cell(r, c).border = Border(
                    top=Side("medium", color=NAVY),
                    left=Side("hair", color=LIGHT_GRAY),
                    right=Side("hair", color=LIGHT_GRAY),
                    bottom=Side("hair", color=LIGHT_GRAY),
                )

        ws.row_dimensions[r].height = 20

    # Column widths
    for col, w in {1: 6, 2: 36, 3: 12, 4: 10, 5: 80}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze and filter
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:E{header_row + len(df)}"

    ws.sheet_view.showGridLines = False

    wb.save(COVERAGE_XLSX)
    print(f"  OK: Saved {COVERAGE_XLSX.name}")


if __name__ == "__main__":
    format_mapping_workbook()
    format_coverage_workbook()
    print("\nDone. Both workbooks formatted.")
