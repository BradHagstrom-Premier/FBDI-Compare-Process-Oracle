"""
build_mapping.py — One-shot utility to build the FBDI-to-Applaud mapping spreadsheet.

Scans baselines/25d/ and baselines/26a/, enumerates tabs from every xlsm file,
merges 9 known Applaud mappings, and writes fbdi_applaud_mapping.xlsx at repo root.

Run as module:  python -m fbdi.build_mapping
"""

from __future__ import annotations

from pathlib import Path
from typing import NamedTuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .config import MAX_FILE_SIZE_BYTES, SKIP_TABS

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).parent.parent
BASELINES = {
    "25d": REPO_ROOT / "baselines" / "25d",
    "26a": REPO_ROOT / "baselines" / "26a",
}
OUTPUT_PATH = REPO_ROOT / "fbdi_applaud_mapping.xlsx"

# ---------------------------------------------------------------------------
# Known Applaud mappings  (fbdi_file_stem, fbdi_tab) → (applaud_table, prefix, module)
# ---------------------------------------------------------------------------

KNOWN_MAPPINGS: dict[tuple[str, str], tuple[str, str, str]] = {
    ("AutoInvoiceImportTemplate",            "RA_INTERFACE_LINES_ALL"):        ("T_RA_INTERFACE_LINES_ALL",     "TA4", "Financials"),
    ("FixedAssetMassAdditionsImportTemplate", "FA_MC_MASS_RATES"):             ("T_FA_MC_MASS_RATES",           "T67", "Financials"),
    ("ItemStructureImportTemplate",          "EGP_COMPONENTS_INTERFACE"):      ("T_EGP_COMPONENTS_INTERFACE",   "T91", "Supply Chain & Manufacturing"),
    ("ProcessWorkDefinitionTemplate",        "Work Definition Headers"):       ("T_WIS_WORK_DEFINITIONS_INT",   "TD6", "Supply Chain & Manufacturing"),
    ("ReceivingReceiptImportTemplate",       "RCV_HEADERS_INTERFACE"):         ("T_RCV_HEADERS_INTERFACE",      "TH7", "Procurement"),
    ("ReceivingReceiptImportTemplate",       "RCV_TRANSACTIONS_INTERFACE"):    ("T_RCV_TRANSACTIONS_INTERFACE", "TH8", "Procurement"),
    ("SourceSalesOrderImportTemplate",       "DOO_ORDER_HEADERS_ALL_INT"):     ("T_DOO_ORDER_HEADERS_ALL",      "TC4", "Supply Chain & Manufacturing"),
    ("SourceSalesOrderImportTemplate",       "DOO_ORDER_LINES_ALL_INT"):       ("T_DOO_ORDER_LINES_ALL",        "TC5", "Supply Chain & Manufacturing"),
    ("WorkOrderMaterialTransactionTemplate", "Material Transaction Lots"):     ("T_INV_TRANSACTION_LOTS_INT",   "T48", "Supply Chain & Manufacturing"),
}

# Extra notes to attach to specific (file, tab) rows
EXTRA_NOTES: dict[tuple[str, str], str] = {
    ("ReceivingReceiptImportTemplate", "RCV_TRANSACTIONS_INTERFACE"):
        "Field name truncation rule: Oracle names >30 chars truncated to 30 chars",
}

# Notes to attach to FILE_ERROR / FILE_TOO_LARGE rows by file stem
PROBLEM_NOTES: dict[str, str] = {
    "MntMaintenanceProgramImport":
        "File appeared in 26A comparison report (had changes) but was skipped by comparison engine due to size",
}

# ---------------------------------------------------------------------------
# Tab enumeration
# ---------------------------------------------------------------------------

STATUS_OK = "OK"
STATUS_TOO_LARGE = "FILE_TOO_LARGE"
STATUS_ERROR = "FILE_ERROR"


def get_tabs(path: Path) -> tuple[list[str], str]:
    """Return (tab_names, status).

    status is one of STATUS_OK, STATUS_TOO_LARGE, STATUS_ERROR.
    SKIP_TABS entries are filtered out before returning.
    """
    if path.stat().st_size > MAX_FILE_SIZE_BYTES:
        return [], STATUS_TOO_LARGE
    try:
        wb = load_workbook(path, read_only=True)
        try:
            tabs = [s for s in wb.sheetnames if s not in SKIP_TABS]
        finally:
            wb.close()
        return tabs, STATUS_OK
    except Exception:
        return [], STATUS_ERROR


# ---------------------------------------------------------------------------
# Scan both releases → build union of (file_stem, tab) pairs
# ---------------------------------------------------------------------------

def scan_baselines() -> dict[str, dict[str, str]]:
    """Return mapping of file_stem → {tab: release_note, ...}.

    release_note is "" (both releases), "25D only", or "26A only".
    For FILE_ERROR / FILE_TOO_LARGE rows the tab key is "" and the value
    is the status string so callers can distinguish them.
    """
    stems_25d: dict[str, tuple[list[str], str]] = {}
    stems_26a: dict[str, tuple[list[str], str]] = {}

    for stem, store in [("25d", stems_25d), ("26a", stems_26a)]:
        folder = BASELINES[stem]
        for fpath in sorted(folder.glob("*.xlsm")):
            tabs, status = get_tabs(fpath)
            store[fpath.stem] = (tabs, status)

    all_stems = sorted(set(stems_25d) | set(stems_26a))
    result: dict[str, dict[str, str]] = {}

    for file_stem in all_stems:
        tabs_25d, status_25d = stems_25d.get(file_stem, ([], STATUS_OK))
        tabs_26a, status_26a = stems_26a.get(file_stem, ([], STATUS_OK))

        # Use worst status across releases for error/size flags
        if STATUS_ERROR in (status_25d, status_26a):
            effective_status = STATUS_ERROR
        elif STATUS_TOO_LARGE in (status_25d, status_26a):
            effective_status = STATUS_TOO_LARGE
        else:
            effective_status = STATUS_OK

        if effective_status in (STATUS_ERROR, STATUS_TOO_LARGE):
            result[file_stem] = {"": effective_status}
            continue

        # Build union of tabs with release presence notes
        set_25d = set(tabs_25d)
        set_26a = set(tabs_26a)
        tab_map: dict[str, str] = {t: "" for t in set_25d & set_26a}
        tab_map.update({t: "25D only" for t in set_25d - set_26a})
        tab_map.update({t: "26A only" for t in set_26a - set_25d})
        result[file_stem] = tab_map

    return result


# ---------------------------------------------------------------------------
# Build row data
# ---------------------------------------------------------------------------

class Row(NamedTuple):
    fbdi_file: str
    fbdi_tab: str
    applaud_table: str
    prefix: str
    in_scope: str
    module: str
    notes: str


def build_rows(scan: dict[str, dict[str, str]]) -> list[Row]:
    """Convert scan results into output rows sorted per spec."""
    problem_rows: list[Row] = []
    normal_rows: list[Row] = []

    for file_stem in sorted(scan):
        tab_map = scan[file_stem]

        for tab in sorted(tab_map):
            release_note = tab_map[tab]

            # Error / oversized
            if release_note in (STATUS_ERROR, STATUS_TOO_LARGE):
                notes = PROBLEM_NOTES.get(file_stem, "")
                problem_rows.append(Row(file_stem, "", "", "", release_note, "", notes))
                continue

            # Normal tab row
            key = (file_stem, tab)
            mapping = KNOWN_MAPPINGS.get(key)

            if mapping:
                applaud_table, prefix, module = mapping
                in_scope = "YES"
            else:
                applaud_table, prefix, module = "", "", ""
                in_scope = "TBD"

            notes_parts: list[str] = []
            if release_note:
                notes_parts.append(release_note)
            extra = EXTRA_NOTES.get(key, "")
            if extra:
                notes_parts.append(extra)

            normal_rows.append(Row(file_stem, tab, applaud_table, prefix, in_scope, module, "; ".join(notes_parts)))

    return problem_rows + normal_rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADERS = ["fbdi_file", "fbdi_tab", "applaud_table", "prefix", "in_scope", "module", "notes"]
COL_WIDTHS = [48, 40, 38, 10, 16, 28, 60]

FILLS = {
    "YES":            PatternFill("solid", fgColor="E2EFDA"),
    "TBD":            PatternFill("solid", fgColor="FFF2CC"),
    "NO":             PatternFill("solid", fgColor="FCE4D6"),
    "FILE_ERROR":     PatternFill("solid", fgColor="F4B942"),
    "FILE_TOO_LARGE": PatternFill("solid", fgColor="F4B942"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
DATA_FONT_BOLD = Font(name="Calibri", size=11, bold=True)


def write_xlsx(rows: list[Row], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FBDI Mapping"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        fill = FILLS.get(row_data.in_scope)
        is_problem = row_data.in_scope in (STATUS_ERROR, STATUS_TOO_LARGE)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            cell.font = DATA_FONT_BOLD if is_problem else DATA_FONT
            if col_idx == 7:  # notes column — wrap text
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Scanning baselines...")
    scan = scan_baselines()
    print(f"  Found {len(scan)} unique file stems")

    rows = build_rows(scan)
    print(f"  Built {len(rows)} output rows")

    problem_count = sum(1 for r in rows if r.in_scope in (STATUS_ERROR, STATUS_TOO_LARGE))
    yes_count = sum(1 for r in rows if r.in_scope == "YES")
    print(f"  Problem rows (FILE_ERROR/FILE_TOO_LARGE): {problem_count}")
    print(f"  YES rows (known mappings): {yes_count}")

    write_xlsx(rows, OUTPUT_PATH)
    print(f"\nWrote: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
