"""Oracle <-> Applaud field-name correspondence layer.

Pure-Python, no MCP/live I/O (same discipline as applaud_appmap.py). Proposes
candidate field correspondences for HITL confirmation, persists confirmed pairs
in a committed workbook, and resolves them into a {applaud_bare: oracle_key}
alias the audit applies before set-intersection.

Verified live against ORACLE_MASTER; see
docs/superpowers/AUDIT_RESULTS_field-correspondence.md for the data behind the
truncation/abbreviation/type rules.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

from fbdi.align import AlignedField
from fbdi.applaud_snapshot import DataColumn
from fbdi.audit_applaud import expected_shape, actual_shape

logger = logging.getLogger(__name__)

# Applaud's application-level name cap (NOT the DataDictionary.Name TEXT(60)
# schema). Longest observed bare is 27 with a 3-char prefix (audit §2.1).
APPLAUD_NAME_CAP = 30

# Longest suffix Applaud appends-then-truncates (NUMBER=6). Bounds how much a
# normalized name may exceed its counterpart on the "appended suffix" path so a
# genuinely short coincidental prefix is not read as a truncation hit (audit §2.2).
MAX_SUFFIX_SLACK = 6

# Data-grounded abbreviation seed (audit §8.3). Abbrev -> full expansion; expanded
# on both sides before comparison. Seed ONLY from post-exact residual divergences;
# do NOT add Oracle's own spellings (they already match in the exact pre-pass).
ABBREVIATIONS: dict[str, str] = {
    "BU": "BUSINESSUNIT",
    "BUS": "BUSINESS",
    "DISC": "DISCOUNT",
    "NUM": "NUMBER",
    "DESCR": "DESCRIPTION",
    "DESC": "DESCRIPTION",
    "AMT": "AMOUNT",
    "INV": "INVOICE",
    "COMP": "COMPONENT",
    "REFER": "REFERENCE",
}

_BOOL_SUFFIX_RE = re.compile(r"_(FLAG|FLG|F)$")
_TRAILING_DIGITS_RE = re.compile(r"(\d+)$")


def _split_trailing_digits(s: str) -> tuple[str, str]:
    """('TIMESTAMP10') -> ('TIMESTAMP', '10'); ('VENDOR_NAME') -> ('VENDOR_NAME', '')."""
    m = _TRAILING_DIGITS_RE.search(s)
    if not m:
        return s, ""
    return s[: m.start()], m.group(1)


def expand_abbreviations(name: str) -> str:
    """Token-wise abbreviation expansion on the underscore-delimited form.
    Unknown tokens pass through; already-expanded input is stable (idempotent)."""
    return "_".join(ABBREVIATIONS.get(tok, tok) for tok in name.split("_"))


def normalize_name(name: str) -> str:
    """Canonical comparison form: upper, strip '*', strip a full boolean suffix,
    expand abbreviations, then FULL underscore squash (audit §2.3 — collapse
    position carries no information). Returns a contiguous A-Z0-9 string."""
    s = (name or "").strip().upper().strip("*")
    s = _BOOL_SUFFIX_RE.sub("", s)
    s = expand_abbreviations(s)
    return s.replace("_", "")


def truncation_window(prefix: str | None) -> int:
    """Per-table truncation window = 30 - len(prefix) (audit §2.1)."""
    return APPLAUD_NAME_CAP - len(prefix or "")


# Score weights (audit §2.4 keeps position weak — DO NOT raise the 0.15).
_W_NAME, _W_TYPE, _W_POSITION = 0.6, 0.25, 0.15

# Confidence bands on the single weighted score, highest first. EXACT is handled
# in the pre-pass and never persisted, so it is not a band here.
TIER_BANDS: list[tuple[str, float]] = [("HIGH", 0.85), ("PROBABLE", 0.55), ("WEAK", 0.0)]


@dataclass
class FieldCorrespondence:
    applaud_table: str
    oracle_key: str
    applaud_bare: str
    applaud_ddid: str
    confidence: str               # HIGH | PROBABLE | WEAK  (or "confirmed"/"rejected" origin)
    origin: str = "derived"       # derived | confirmed | rejected
    score: float = 0.0
    signals: str = ""
    notes: str = ""


def names_correspond(oracle_norm: str, applaud_norm: str, applaud_bare_len: int,
                     window: int) -> bool:
    """True if the two normalized names plausibly denote the same field.

    Implements audit §1.1 (digit-run preservation), §2.1 (derived window) and
    §2.2 (truncation-aware prefix-of-other). `applaud_bare_len` is the RAW stored
    bare length (truncation happened on the stored name)."""
    # Digit-run preservation (audit §1.1): equal trailing digits, stem-match the rest.
    o_stem, o_dig = _split_trailing_digits(oracle_norm)
    a_stem, a_dig = _split_trailing_digits(applaud_norm)
    if o_dig or a_dig:
        if o_dig != a_dig:
            return False
        oracle_norm, applaud_norm = o_stem, a_stem

    if oracle_norm == applaud_norm:
        return True

    # Applaud right-truncated from a longer Oracle logical name.
    if oracle_norm.startswith(applaud_norm):
        delta = len(oracle_norm) - len(applaud_norm)
        # Valid if only a short suffix was lost, OR Applaud was cut at the cap
        # (a hard truncation legitimately drops many trailing chars).
        return delta <= MAX_SUFFIX_SLACK or applaud_bare_len >= window - 1

    # Applaud appended a suffix (NAME/FLAG/...) then possibly truncated it.
    if applaud_norm.startswith(oracle_norm):
        return len(applaud_norm) - len(oracle_norm) <= MAX_SUFFIX_SLACK

    return False


def _name_score(oracle_key: str, applaud_bare: str, applaud_bare_len: int,
                window: int) -> float:
    """1.0 if normalized-equal (a non-truncation divergence, e.g. abbreviation or
    underscore-only), 0.8 if a truncation/suffix match, 0.0 if no correspondence."""
    o, a = normalize_name(oracle_key), normalize_name(applaud_bare)
    if o == a:
        return 1.0
    if names_correspond(o, a, applaud_bare_len, window):
        return 0.8
    return 0.0


def _type_class_conflict(of: AlignedField, col: DataColumn) -> bool:
    """Char-vs-numeric clash only (audit §1.2) — never date-vs-char."""
    return {expected_shape(of)[0], actual_shape(col)[0]} == {"char", "numeric"}


def _type_score(of: AlignedField, col: DataColumn) -> float:
    exp_cls, act_cls = expected_shape(of)[0], actual_shape(col)[0]
    if exp_cls and act_cls and exp_cls == act_cls:
        return 1.0
    return 0.5


def _tier(score: float) -> str:
    for name, floor in TIER_BANDS:
        if score >= floor:
            return name
    return "WEAK"


def score_candidate(oracle_key: str, of: AlignedField, col: DataColumn, window: int,
                    position_score: float) -> tuple[float, str]:
    """Weighted score + a human-readable signals string.

    The name score is computed from `oracle_key` (= oracle_match_key(of)) — NEVER
    re-derived from of.technical, which is empty on the label-derived-key path and
    would mis-score name=0.00 (audit §2.1). Caller has already confirmed a non-zero
    name correspondence and a passing type veto."""
    ns = _name_score(oracle_key, col.bare, len(col.bare), window)
    ts = _type_score(of, col)
    score = _W_NAME * ns + _W_TYPE * ts + _W_POSITION * position_score
    signals = f"name={ns:.2f} type={ts:.2f} pos={position_score:.2f}"
    return round(score, 4), signals


def _candidate_excluded(col: DataColumn, prefix: str | None) -> bool:
    """Audit §1.3/§1.4: @-audit fields and non-prefix working columns (X_PHANTOM)
    never enter the candidate pool. Defensive — build_table already drops them."""
    if col.bare.lstrip().startswith("@") or col.ddid.lstrip().startswith("@"):
        return True
    if prefix and not col.ddid.upper().startswith(prefix.upper()):
        return True
    return False


def derive_table_correspondences(
    applaud_table: str, prefix: str | None,
    oracle_by_key: dict[str, AlignedField],
    applaud_columns: list[DataColumn],
    decided: set[tuple[str, str]],
) -> list[FieldCorrespondence]:
    """Propose one-to-one correspondences for the residual after the exact pre-pass.

    `oracle_by_key` maps oracle_match_key -> AlignedField. `decided` is the set of
    (table, oracle_key) pairs already in the committed map (confirmed/rejected) —
    never re-proposed."""
    window = truncation_window(prefix)
    cols = [c for c in applaud_columns if not _candidate_excluded(c, prefix)]

    # 1. Exact pre-pass: keys present verbatim on both sides need no map entry.
    # oracle_match_key always returns UPPER_SNAKE_CASE (audit LOW #3), but normalize
    # the key set once so the pre-pass is correct even on mixed-case input.
    oracle_keys_upper = {k.upper() for k in oracle_by_key}
    applaud_bares = {c.bare.upper() for c in cols}
    residual_oracle = {k: of for k, of in oracle_by_key.items()
                       if k.upper() not in applaud_bares
                       and (applaud_table, k) not in decided}
    residual_cols = [c for c in cols if c.bare.upper() not in oracle_keys_upper]

    # Position is a WEAK tiebreak only (audit §2.4): a gentle order-agreement gradient
    # over Oracle position vs Applaud ROW order. Row-sort the residual columns once so
    # a_idx genuinely means row order, not raw list order (audit §2.4 — the prior code
    # built a row-sorted list but then scored over the unsorted list).
    residual_cols_by_row = sorted(residual_cols, key=lambda c: c.row)
    n_oracle = len(residual_oracle)
    n_applaud = len(residual_cols_by_row)

    # 2. Build candidate pairs (name-match + passing type veto), scored.
    candidates: list[tuple[float, str, str, DataColumn, str]] = []  # (score, tier, okey, col, signals)
    for o_idx, (okey, of) in enumerate(residual_oracle.items()):
        for a_idx, col in enumerate(residual_cols_by_row):
            if _name_score(okey, col.bare, len(col.bare), window) == 0.0:
                continue
            if _type_class_conflict(of, col):
                continue   # char-vs-numeric veto (audit §1.2)
            pos = _position_score(o_idx, a_idx, n_oracle, n_applaud)
            score, signals = score_candidate(okey, of, col, window, pos)
            candidates.append((score, _tier(score), okey, col, signals))

    # 3. Greedy one-to-one bijection: best score first, both sides must be free.
    candidates.sort(key=lambda t: t[0], reverse=True)
    used_oracle: set[str] = set()
    used_bare: set[str] = set()
    out: list[FieldCorrespondence] = []
    for score, tier, okey, col, signals in candidates:
        if okey in used_oracle or col.bare.upper() in used_bare:
            continue
        used_oracle.add(okey)
        used_bare.add(col.bare.upper())
        out.append(FieldCorrespondence(
            applaud_table=applaud_table, oracle_key=okey, applaud_bare=col.bare,
            applaud_ddid=col.ddid, confidence=tier, origin="derived",
            score=score, signals=signals))
    return out


def _position_score(o_idx: int, a_idx: int, n_oracle: int, n_applaud: int) -> float:
    """Gentle order-agreement gradient in [0,1]; a weak tiebreak only (audit §2.4)."""
    span = max(n_oracle, n_applaud, 1)
    return max(0.0, 1.0 - abs(o_idx - a_idx) / span)


def derive_correspondences(
    tables: dict[str, tuple[str | None, dict[str, AlignedField], list[DataColumn]]],
    decided: set[tuple[str, str]],
) -> list[FieldCorrespondence]:
    """Derive across many tables. `tables` maps applaud_table ->
    (prefix, oracle_by_key, applaud_columns). Output sorted table -> tier -> score."""
    out: list[FieldCorrespondence] = []
    for table in sorted(tables):
        prefix, oracle_by_key, cols = tables[table]
        out.extend(derive_table_correspondences(table, prefix, oracle_by_key, cols, decided))
    tier_rank = {name: i for i, (name, _) in enumerate(TIER_BANDS)}
    out.sort(key=lambda fc: (fc.applaud_table, tier_rank.get(fc.confidence, 9), -fc.score))
    return out


# ---------------------------------------------------------------------------
# Fieldmap workbook I/O
# ---------------------------------------------------------------------------

_FIELDMAP_HEADERS = ["Applaud Table", "Oracle Key", "Applaud Bare", "Applaud DDID",
                     "Confidence", "Origin", "Notes"]


def write_fieldmap_workbook(rows: list[FieldCorrespondence], path: Path) -> None:
    """Write the committed Oracle<->Applaud field map (sheet 'Field Map')."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Map"
    ws.append(_FIELDMAP_HEADERS)
    for r in sorted(rows, key=lambda fc: (fc.applaud_table, fc.oracle_key)):
        ws.append([r.applaud_table, r.oracle_key, r.applaud_bare, r.applaud_ddid,
                   r.confidence, r.origin, r.notes])
    ws.freeze_panes = "A2"
    wb.save(path)


def _validate_headers(ws, expected: list[str], path: Path) -> None:
    """Fail loud if the worksheet's header row does not match `expected`.

    Both loaders parse human-editable workbooks by fixed column position, so a
    reordered/renamed/deleted column would otherwise be read as silently wrong data.
    Compares only the first len(expected) header cells (stripped)."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    actual = [(str(c).strip() if c is not None else "") for c in header_row][:len(expected)]
    if actual != expected:
        raise ValueError(
            f"{path}: unexpected header row {actual!r}; expected {expected!r}. "
            "Do not reorder, rename, or delete columns in the workbook.")


def load_fieldmap_workbook(path: Path) -> dict[str, list[FieldCorrespondence]]:
    """Load the committed field map into {applaud_table: [FieldCorrespondence, ...]}.

    Precedence invariant (audit §1.1): the committed map holds only confirmed/rejected
    rows. Any stray `derived` row (hand-edited, or written by a future flow) is dropped
    with a WARNING so it can never silently block a future decision.

    Duplicate (table, oracle_key) rows: the last row wins (matching dict-collapse
    behaviour downstream) and a WARNING is emitted so hand-edits that create
    duplicates are never silent.

    Fails loud (ValueError) if the header row does not match `_FIELDMAP_HEADERS`."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Field Map"] if "Field Map" in wb.sheetnames else wb.active
    _validate_headers(ws, _FIELDMAP_HEADERS, path)
    # Use an ordered dict keyed by (table, oracle_key) so duplicate detection and
    # last-wins collapse happen in a single pass.
    seen: dict[tuple[str, str], FieldCorrespondence] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        table, okey, bare, ddid, conf, origin, notes = (list(row) + [None] * 7)[:7]
        if not table or not okey:
            continue
        origin_s = str(origin).strip().lower() if origin else "derived"
        if origin_s not in ("confirmed", "rejected"):
            logger.warning("Dropping non-decision (origin=%r) row from committed field map: "
                           "%s / %s — only confirmed/rejected persist.", origin_s, table, okey)
            continue
        key = (str(table), str(okey))
        if key in seen:
            logger.warning("Duplicate (table, oracle_key) in committed field map: "
                           "%s / %s — keeping last row.", str(table), str(okey))
        seen[key] = FieldCorrespondence(
            applaud_table=str(table), oracle_key=str(okey),
            applaud_bare=(str(bare) if bare else ""), applaud_ddid=(str(ddid) if ddid else ""),
            confidence=(str(conf) if conf else ""), origin=origin_s,
            notes=(str(notes) if notes else ""))
    wb.close()
    # Rebuild the {table: [rows]} structure preserving insertion order within each table.
    out: dict[str, list[FieldCorrespondence]] = {}
    for fc in seen.values():
        out.setdefault(fc.applaud_table, []).append(fc)
    return out


# ---------------------------------------------------------------------------
# Merge functions (opposite precedence — audit §1.1)
# ---------------------------------------------------------------------------

def merge_fieldmap(
    derived: list[FieldCorrespondence],
    committed: dict[str, list[FieldCorrespondence]],
) -> dict[str, list[FieldCorrespondence]]:
    """Confirmed/rejected rows win; derived rows fill only undecided (table, oracle_key).
    Clones merge_appmap semantics (applaud_appmap.py:165). Idempotent across re-derives."""
    out: dict[str, dict[str, FieldCorrespondence]] = {}
    for table, rows in committed.items():
        out[table] = {r.oracle_key: r for r in rows}
    for r in derived:
        bucket = out.setdefault(r.applaud_table, {})
        if r.oracle_key not in bucket:
            bucket[r.oracle_key] = r
    return {table: [bucket[k] for k in sorted(bucket)] for table, bucket in sorted(out.items())}


def merge_decisions(
    decisions: list[FieldCorrespondence],
    committed: dict[str, list[FieldCorrespondence]],
) -> dict[str, list[FieldCorrespondence]]:
    """Confirm-time merge (audit §1.1): incoming human DECISIONS WIN, untouched committed
    rows carry forward. The inverse precedence of merge_fieldmap — a reviewer can revise a
    prior confirmation/rejection through the tooling instead of hand-editing the xlsx."""
    out: dict[str, dict[str, FieldCorrespondence]] = {}
    for table, rows in committed.items():
        out[table] = {r.oracle_key: r for r in rows}
    for r in decisions:
        out.setdefault(r.applaud_table, {})[r.oracle_key] = r   # incoming overrides
    return {table: [bucket[k] for k in sorted(bucket)] for table, bucket in sorted(out.items())}


# ---------------------------------------------------------------------------
# Review workbook I/O + apply_review_decisions
# ---------------------------------------------------------------------------

_REVIEW_HEADERS = ["Applaud Table", "Oracle Key", "Oracle Type", "Candidate Applaud Bare",
                   "Applaud DDID", "Applaud Type", "Confidence", "Score", "Signals",
                   "Conflicts/Alternatives", "Confirm?", "Corrected Bare"]


@dataclass
class ReviewRow:
    applaud_table: str
    oracle_key: str
    oracle_type: str
    candidate_bare: str
    applaud_ddid: str
    applaud_type: str
    confidence: str
    score: float
    signals: str
    alternatives: str
    confirm: str = ""          # reviewer input: 'Y' | 'N' | ''
    corrected_bare: str = ""   # reviewer input: substitute bare


class InvalidCorrectedBareError(ValueError):
    """Raised when a reviewer-entered Corrected Bare is not an actual bare in the
    table (audit §4.1) — fail loud rather than commit an alias that maps to nothing."""


def write_review_workbook(rows: list[ReviewRow], path: Path,
                          exact_counts: dict[str, tuple[int, int]] | None = None) -> None:
    """Write the disposable HITL review workbook. `exact_counts` maps table ->
    (exact_matched, total) so the reviewer sees denominator context (audit §6)."""
    exact_counts = exact_counts or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.append(_REVIEW_HEADERS)
    current_table = None
    for r in rows:
        if r.applaud_table != current_table:
            current_table = r.applaud_table
            matched, total = exact_counts.get(current_table, (0, 0))
            ws.append([f"--- {current_table}: {matched} of {total} matched exactly; "
                       f"deciding the residual {max(total - matched, 0)} ---"])
        ws.append([r.applaud_table, r.oracle_key, r.oracle_type, r.candidate_bare,
                   r.applaud_ddid, r.applaud_type, r.confidence, r.score, r.signals,
                   r.alternatives, r.confirm, r.corrected_bare])
    ws.freeze_panes = "A2"
    wb.save(path)


def load_review_workbook(path: Path) -> list[ReviewRow]:
    """Load reviewer decisions. Header-separator rows (a single '--- ...' cell) are skipped.

    Fails loud (ValueError) if the header row does not match `_REVIEW_HEADERS` — a
    reordered/renamed column would otherwise be parsed by position into wrong fields."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Review"] if "Review" in wb.sheetnames else wb.active
    _validate_headers(ws, _REVIEW_HEADERS, path)
    out: list[ReviewRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = (list(row) + [None] * 12)[:12]
        table, okey = cells[0], cells[1]
        if not table or not okey:
            continue   # blank row or a "--- table header ---" separator
        out.append(ReviewRow(
            applaud_table=str(table), oracle_key=str(okey),
            oracle_type=str(cells[2] or ""), candidate_bare=str(cells[3] or ""),
            applaud_ddid=str(cells[4] or ""), applaud_type=str(cells[5] or ""),
            confidence=str(cells[6] or ""), score=float(cells[7] or 0.0),
            signals=str(cells[8] or ""), alternatives=str(cells[9] or ""),
            confirm=str(cells[10] or "").strip(), corrected_bare=str(cells[11] or "").strip()))
    wb.close()
    return out


def apply_review_decisions(
    rows: list[ReviewRow],
    valid_bares_by_table: dict[str, set[str]],
) -> list[FieldCorrespondence]:
    """Turn reviewer input into FieldCorrespondence rows.

    Corrected Bare (if present) wins -> confirmed with the substitute, VALIDATED
    against the table's bare set (audit §4.1, fail loud). The stored bare is the
    *canonical* casing from the table's valid set (not the reviewer's literal typing)
    to keep the committed map self-consistent. The DDID of the *rejected candidate*
    is NOT carried forward — it would be misleading in a human-auditable artifact;
    applaud_ddid is stored as "" (consistent with the rejected branch).

    Else Confirm? 'Y' -> confirmed; 'N' -> rejected; blank -> skipped (undecided).
    A non-blank Confirm? value that is neither 'Y' nor 'N' (case-insensitive) is
    warned loudly (audit §4.1 fail-loud philosophy) and the row is skipped."""
    out: list[FieldCorrespondence] = []
    for r in rows:
        raw_valid = valid_bares_by_table.get(r.applaud_table, set())
        # upper-to-canonical map for case-insensitive lookup while preserving stored casing
        canonical_map: dict[str, str] = {b.upper(): b for b in raw_valid}
        if r.corrected_bare:
            upper_corrected = r.corrected_bare.upper()
            if upper_corrected not in canonical_map:
                raise InvalidCorrectedBareError(
                    f"{r.applaud_table}: Corrected Bare {r.corrected_bare!r} for Oracle "
                    f"key {r.oracle_key!r} is not a column in that table. Fix the typo "
                    "or clear the cell; refusing to commit an alias that maps to nothing.")
            # Use canonical casing from the table's valid set (not as-typed by reviewer).
            # DDID of the rejected candidate must not be stored — it is misleading for the
            # corrected target bare; store "" consistent with the rejected branch.
            out.append(FieldCorrespondence(
                applaud_table=r.applaud_table, oracle_key=r.oracle_key,
                applaud_bare=canonical_map[upper_corrected], applaud_ddid="",
                confidence="HIGH", origin="confirmed", notes="reviewer-corrected"))
        elif r.confirm.upper() == "Y":
            upper_candidate = r.candidate_bare.upper()
            if upper_candidate not in canonical_map:
                raise InvalidCorrectedBareError(
                    f"{r.applaud_table}: Confirm?='Y' candidate bare {r.candidate_bare!r} for "
                    f"Oracle key {r.oracle_key!r} is not a column in that table. The workbook "
                    "candidate cell may have been edited. Fix or use Corrected Bare instead; "
                    "refusing to commit an alias that maps to nothing.")
            notes = f"confirmed at {r.confidence}" + (f"; {r.signals}" if r.signals else "")
            out.append(FieldCorrespondence(
                applaud_table=r.applaud_table, oracle_key=r.oracle_key,
                applaud_bare=r.candidate_bare, applaud_ddid=r.applaud_ddid,
                confidence=r.confidence, origin="confirmed", score=r.score,
                signals=r.signals, notes=notes))
        elif r.confirm.upper() == "N":
            out.append(FieldCorrespondence(
                applaud_table=r.applaud_table, oracle_key=r.oracle_key,
                applaud_bare="", applaud_ddid="", confidence=r.confidence,
                origin="rejected", notes="reviewer-rejected"))
        elif r.confirm:
            # Non-blank value that is not Y/N — warn loudly (audit §4.1 fail-loud philosophy)
            # and skip the row so an accidental "YES" / "TRUE" / "X" never vanishes silently.
            logger.warning(
                "Unrecognized Confirm? value %r for %s / %s — expected 'Y', 'N', or blank. "
                "Row skipped; correct the workbook cell to proceed.",
                r.confirm, r.applaud_table, r.oracle_key)
        # blank Confirm? + no Corrected Bare -> undecided, skip silently
    return out


# ---------------------------------------------------------------------------
# Alias resolver
# ---------------------------------------------------------------------------

def assemble_derivation_inputs(
    snapshot, catalog: dict[tuple[str, str], list[AlignedField]],
    mapping: dict[tuple[str, str], dict],
) -> dict[str, tuple[str | None, dict[str, AlignedField], list[DataColumn]]]:
    """Group the audit's (template, tab)->table chain into per-table derivation inputs:
    {applaud_table: (prefix, {oracle_match_key: AlignedField}, [DataColumn, ...])}.
    Mirrors run_audit's loop so derivation sees exactly the audit's column set.

    When several (template, tab) rows map to ONE Applaud table (the known multi-mapping
    rows from the first-pass mapping audit), MERGE their Oracle keys rather than
    overwrite (audit §2.2) — otherwise only the last tab's divergent fields would ever
    get correspondence candidates, so the derivation would be lossier than the audit it
    serves. Tables with no catalog fields or no snapshot entry are skipped with an INFO
    log (audit LOW #6) so an empty review workbook is explainable."""
    from fbdi.audit_applaud import oracle_match_key
    out: dict[str, tuple[str | None, dict[str, AlignedField], list[DataColumn]]] = {}
    for (template, tab), info in mapping.items():
        table_name = info.get("applaud_table")
        if not table_name:
            continue
        oracle_fields = catalog.get((template, tab), [])
        table = snapshot.tables.get(table_name)
        if not oracle_fields or table is None:
            logger.info("correspondence: skipping (%s, %s) -> %s — %s.",
                        template, tab, table_name,
                        "no catalog fields" if not oracle_fields else "no snapshot table")
            continue
        oracle_by_key = {oracle_match_key(f): f for f in oracle_fields if oracle_match_key(f)}
        if table_name in out:
            prev_prefix, prev_keys, prev_cols = out[table_name]
            if prev_prefix != table.prefix:
                logger.warning("correspondence: prefix disagreement for %s (%r vs %r); "
                               "keeping first.", table_name, prev_prefix, table.prefix)
            prev_keys.update(oracle_by_key)             # merge this tab's keys (audit §2.2)
            out[table_name] = (prev_prefix, prev_keys, prev_cols)
        else:
            out[table_name] = (table.prefix, oracle_by_key, list(table.columns))
    return out


def build_alias(fieldmap_for_table: list[FieldCorrespondence],
                accept_confidence: str = "confirmed") -> dict[str, str]:
    """Resolve a table's field map into {applaud_bare_upper: oracle_key_upper}.

    'confirmed' (default): only origin=confirmed rows. A tier name ('HIGH' /
    'PROBABLE' / 'WEAK') additionally admits origin=derived rows at or above that
    tier (pre-review pass). origin=rejected is never aliased.

    If two admitted rows map the same applaud_bare to *different* oracle keys (only
    possible via a stale map after a rename — Pass-1 audit LOW #5), the first wins and
    a WARNING is emitted, so the collision is never a silent last-row-wins overwrite."""
    tier_rank = {name: i for i, (name, _) in enumerate(TIER_BANDS)}  # HIGH=0 best
    gate = (accept_confidence or "confirmed").strip()
    alias: dict[str, str] = {}
    for fc in fieldmap_for_table:
        if fc.origin == "rejected" or not fc.applaud_bare:
            continue
        if fc.origin == "confirmed":
            admit = True
        elif gate == "confirmed":
            admit = False
        else:
            admit = tier_rank.get(fc.confidence, 99) <= tier_rank.get(gate, -1)
        if not admit:
            continue
        bare_u, okey_u = fc.applaud_bare.upper(), fc.oracle_key.upper()
        existing = alias.get(bare_u)
        if existing is not None and existing != okey_u:
            logger.warning("build_alias: conflicting alias for bare %r — keeping %r, "
                           "ignoring %r (stale map after a rename?).",
                           bare_u, existing, okey_u)
            continue
        alias[bare_u] = okey_u
    return alias
