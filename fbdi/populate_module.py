"""Surgical Module-column updater for FBDI_to_ApplaudTables_Mapping.xlsx.

Reads file_modules.json from a NEW and OLD release, looks up each row's
FBDI Template (col A) against the merged dict (NEW wins), writes column
F (Module) only. All other cells, formatting, formulas, merged cells,
validations, and freeze-panes are preserved by openpyxl's full-mode load.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


SHEET_NAME = "FBDI Mapping"
TEMPLATE_COL = 1  # A
MODULE_COL = 6    # F


def _stem(name) -> str:
    """Normalize an FBDI Template identifier: strip .xlsm suffix and whitespace."""
    if name is None:
        return ""
    s = str(name).strip()
    if s.lower().endswith(".xlsm"):
        s = s[:-5]
    return s


def _load_modules_json(path: Path) -> dict[str, str]:
    """Load a file_modules.json. Returns {} if path is missing."""
    if not path.is_file():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def populate_module_column(
    mapping_path: Path,
    new_modules: dict[str, str],
    old_modules: dict[str, str],
) -> dict[str, int]:
    """Update the Module column (F) in place. Returns counts dict.

    Lookup order: NEW release wins; OLD fills only when NEW lacks the file.
    Files in neither release leave the cell blank.

    Returns: {'populated': N, 'blank': M, 'overwritten': K}
      - populated: rows that ended with a non-blank Module value
      - blank: rows with non-blank FBDI Template that found no match
      - overwritten: rows whose pre-existing Module value was changed
    """
    # Merge: new_modules takes precedence — Python dict merge semantics
    # mean the right-hand operand wins for duplicate keys.
    merged = {_stem(k): v for k, v in old_modules.items()}
    merged.update({_stem(k): v for k, v in new_modules.items()})

    wb = load_workbook(mapping_path)  # full mode preserves everything
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {mapping_path}")

    ws = wb[SHEET_NAME]
    populated = 0
    blank = 0
    overwritten = 0

    # Iterate data rows (skip header at row 1)
    for row_idx in range(2, ws.max_row + 1):
        template_cell = ws.cell(row=row_idx, column=TEMPLATE_COL)
        template = _stem(template_cell.value)
        if not template:
            continue  # blank rows don't count

        module_cell = ws.cell(row=row_idx, column=MODULE_COL)
        previous = module_cell.value
        new_value = merged.get(template)

        if new_value:
            if previous and previous != new_value:
                overwritten += 1
            elif previous == new_value:
                # Idempotent re-write: count as overwritten so callers can
                # detect re-runs.
                overwritten += 1
            module_cell.value = new_value
            populated += 1
        else:
            blank += 1

    wb.save(mapping_path)
    wb.close()

    return {"populated": populated, "blank": blank, "overwritten": overwritten}
