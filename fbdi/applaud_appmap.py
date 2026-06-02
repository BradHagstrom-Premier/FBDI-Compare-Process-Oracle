"""Applaud application-map bridge: prefix derivation + table<->IF/EF derivation.

Pure-Python. Fed raw Application / get_application results (and DatabaseDetail
DDIDs for the prefix fallback). No MCP I/O.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook

_log = logging.getLogger(__name__)

# Matches a trailing "(T32)" / "(O33)" style prefix tag in a table description.
_PAREN_PREFIX_RE = re.compile(r"\(([A-Z0-9]+)\)\s*$")

# Applaud TableId code: a letter followed by two alphanumerics, prepended to every
# DDID with no separator (e.g. T32COUNTRY, O33BANK_NAME, TA1..., TL2...). Used only
# as a fallback when the description parenthetical is absent. A longest-common-prefix
# approach is wrong here: two field names sharing a leading letter (BANK_NAME /
# BRANCH_NUMBER -> "B") would extend the prefix past the 3-char TableId code.
_TABLEID_RE = re.compile(r"^[A-Z][A-Z0-9]{2}")


def derive_prefix(description: str, column_ddids: list[str]) -> tuple[str | None, bool]:
    """Return (prefix, used_fallback).

    Parenthetical first (authoritative — present on all in-scope T_* tables).
    Otherwise derive the 3-char TableId code from the first business DDID, logged
    as a fallback. `@`-audit fields are excluded from the fallback input.
    """
    m = _PAREN_PREFIX_RE.search((description or "").strip())
    if m:
        return m.group(1), False
    business = [d.upper() for d in column_ddids if not d.lstrip().startswith("@")]
    for ddid in business:
        tm = _TABLEID_RE.match(ddid)
        if tm:
            prefix = tm.group(0)
            _log.warning(
                "Prefix fallback for %r: no description parenthetical; derived %r "
                "from the TableId-code pattern on %r.", description, prefix, ddid,
            )
            return prefix, True
    _log.warning(
        "Prefix fallback for %r: no parenthetical and no TableId-pattern DDID; "
        "prefix is None.", description,
    )
    return None, True


# ---------------------------------------------------------------------------
# App-map derivation (Task 4)
# ---------------------------------------------------------------------------

@dataclass
class AppMapRow:
    target_table: str
    import_files: list[str] = field(default_factory=list)
    export_files: list[str] = field(default_factory=list)
    source_applications: list[str] = field(default_factory=list)
    origin: str = "derived"          # "derived" | "confirmed"


def _steps_of_type(app: dict, func_type: str) -> list[str]:
    steps = sorted(app.get("steps", []), key=lambda s: s.get("order", 0))
    return [s["func_name"] for s in steps if s.get("func_type") == func_type]


def is_validation_file(name: str) -> bool:
    """`*_VAL` files are Applaud validation exports, not the FBDI-fields export.
    Excluded from the derived app-map by default (they carry a different field set
    and would generate coverage/ordering noise). A consultant can add one back
    manually in the confirmed workbook if a specific _VAL EF should be audited."""
    return name.strip().upper().endswith("_VAL")


def derive_appmap(applications: dict, target_tables: set[str],
                  exclude_validation: bool = True) -> list[AppMapRow]:
    """One AppMapRow per target table. Apps are matched by DBID; IF/EF file names
    come from the apps' get_application steps in execution order. By default,
    `*_VAL` validation exports are excluded (see is_validation_file)."""
    rows: list[AppMapRow] = []
    for table in sorted(target_tables):
        imports: list[str] = []
        exports: list[str] = []
        sources: list[str] = []
        for app_name in sorted(applications):
            app = applications[app_name]
            if app.get("dbid") != table:
                continue
            ifs = _steps_of_type(app, "IF")
            efs = _steps_of_type(app, "EF")
            if exclude_validation:
                ifs = [f for f in ifs if not is_validation_file(f)]
                efs = [f for f in efs if not is_validation_file(f)]
            if ifs or efs:
                sources.append(app_name)
            for f in ifs:
                if f not in imports:
                    imports.append(f)
            for f in efs:
                if f not in exports:
                    exports.append(f)
        rows.append(AppMapRow(target_table=table, import_files=imports,
                              export_files=exports, source_applications=sources,
                              origin="derived"))
    return rows


# ---------------------------------------------------------------------------
# App-map workbook I/O + merge (Task 5)
# ---------------------------------------------------------------------------

_APPMAP_HEADERS = ["Target Table", "Import Files", "Export Files",
                   "Source Applications", "Origin"]


def _join(items: list[str]) -> str:
    return "; ".join(items)


def _split(cell) -> list[str]:
    if cell is None:
        return []
    return [p.strip() for p in str(cell).split(";") if p.strip()]


def write_appmap_workbook(rows: list[AppMapRow], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "App Map"
    ws.append(_APPMAP_HEADERS)
    for r in rows:
        ws.append([r.target_table, _join(r.import_files), _join(r.export_files),
                   _join(r.source_applications), r.origin])
    ws.freeze_panes = "A2"
    wb.save(path)


def load_appmap_workbook(path: Path) -> dict[str, AppMapRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["App Map"] if "App Map" in wb.sheetnames else wb.active
    out: dict[str, AppMapRow] = {}
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        table, imports, exports, sources, origin = (list(row) + [None] * 5)[:5]
        if not table:
            continue
        out[str(table)] = AppMapRow(
            target_table=str(table), import_files=_split(imports),
            export_files=_split(exports), source_applications=_split(sources),
            origin=(str(origin) if origin else "derived"),
        )
    wb.close()
    return out


def merge_appmap(derived: list[AppMapRow],
                 confirmed: dict[str, AppMapRow]) -> list[AppMapRow]:
    """Confirmed rows win; derived rows fill only tables not already confirmed."""
    out: dict[str, AppMapRow] = dict(confirmed)
    for r in derived:
        if r.target_table not in out:
            out[r.target_table] = r
    return [out[k] for k in sorted(out)]
