"""Applaud compliance audit engine (Step B).

Pure-Python over an ApplaudSnapshot + Oracle catalog + FBDI mapping + confirmed
app-map. Runs the dimension checks and writes an Excel findings workbook. Every
finding is an addressable delta: (object_type, object_name, field, attribute,
current -> expected), so a future write phase can replay accepted findings.

Matching notes (verified live against ORACLE_MASTER):
  - Scope is T_* target tables; within that family IF/EF/table share one TableId
    prefix, so intra-Applaud matching (Dim 5) is exact-DDID.
  - The Oracle side is normalized via oracle_match_key(): the FBDI catalog leaves
    `technical` = None on thin tabs (e.g. Bank Account, label-only), so the key
    falls back to _label_to_technical(label) to match the Applaud bare name.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from fbdi.align import AlignedField, Change, align_tabs
from fbdi.applaud_snapshot import ApplaudSnapshot, DataColumn, FileField
from fbdi.applaud_appmap import AppMapRow
from fbdi.applaud_type import applaud_type_for
from fbdi.audit import _style_header_row, _label_to_technical
from fbdi.type_parser import ParsedType


# ---------------------------------------------------------------------------
# Mapping filter (--tables scope support)
# ---------------------------------------------------------------------------

class UnknownTableError(ValueError):
    """Raised when --tables names a target table absent from the FBDI mapping —
    so a typo fails loud instead of silently narrowing audit scope."""


def filter_mapping_to_tables(
    mapping: dict[tuple[str, str], dict],
    table_names: list[str],
) -> dict[tuple[str, str], dict]:
    """Restrict the FBDI->table mapping to rows whose Applaud target table is in
    `table_names` (case-insensitive). Fail loud via UnknownTableError if any
    requested name is absent from the mapping. Returns a new dict; input untouched."""
    requested = {t.strip().upper() for t in table_names if t and t.strip()}
    present = {(info.get("applaud_table") or "").upper() for info in mapping.values()}
    unknown = sorted(requested - present)
    if unknown:
        raise UnknownTableError(
            "Unknown table(s) not in mapping: " + ", ".join(unknown))
    return {key: info for key, info in mapping.items()
            if (info.get("applaud_table") or "").upper() in requested}


# ---------------------------------------------------------------------------
# Finding model (Task 6)
# ---------------------------------------------------------------------------

@dataclass
class Finding:
    finding_id: str
    dimension: str
    severity: str                 # HIGH | MED | INFO
    fbdi_template: str
    fbdi_tab: str
    oracle_field: str
    oracle_type: str
    applaud_object_type: str      # DATA_ELEMENT | IMPORT | EXPORT | TABLE
    applaud_object_name: str
    applaud_field: str
    attribute: str                # SIZE | SCALE | TYPE_CLASS | PRESENCE | ORDER
    current_value: str
    expected_value: str
    message: str
    status: str = ""              # Phase-2: ACCEPTED | DEFERRED | ACTIONED
    notes: str = ""


def make_finding_id(*, dimension: str, applaud_object_type: str,
                    applaud_object_name: str, applaud_field: str,
                    attribute: str) -> str:
    """Stable 12-char id for a finding, keyed on (dimension, object, field, attribute)
    so the same delta reconciles across re-runs (Phase-2 triage continuity)."""
    key = "|".join([dimension, applaud_object_type, applaud_object_name,
                    applaud_field, attribute])
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]


# ---------------------------------------------------------------------------
# Dim 1 — data element sizing (Task 7) + the shared Oracle match key (§3.2)
# ---------------------------------------------------------------------------

Shape = tuple[str, int | None, int | None]   # (class, size, scale)


def oracle_match_key(of: AlignedField) -> str:
    """Normalized Oracle identity used by every matching dimension (§3.2).

    The FBDI catalog leaves `technical` = None on thin tabs (e.g. the canonical
    RapidImplementationForCashManagement / 'Bank Account' tab, which exposes only
    labels like 'Bank Name'). Prefer technical when present, else the label; either
    way normalize via audit._label_to_technical, because some catalog tabs store a
    display header in `technical` (e.g. 'Supplier Name*', with spaces and a trailing
    '*'). Normalizing both paths makes 'Supplier Name*' -> 'SUPPLIER_NAME' match the
    Applaud bare name, while a clean technical name ('ITEM_NUMBER') passes through
    unchanged (idempotent). Returns UPPER_SNAKE_CASE (or '' if neither is present)."""
    raw = of.technical if of.technical else (of.label or "")
    return _label_to_technical(raw).upper()


def _shape_from_applaud_str(s: str) -> Shape:
    """Parse 'char 100' / 'numeric 18,4' / 'date' into (class, size, scale)."""
    parts = s.split()
    cls = parts[0] if parts else ""
    size = scale = None
    if len(parts) == 2:
        nums = parts[1].split(",")
        size = int(nums[0]) if nums[0].isdigit() else None
        if len(nums) == 2 and nums[1].isdigit():
            scale = int(nums[1])
    return (cls, size, scale)


def expected_shape(of: AlignedField) -> Shape:
    """Oracle field's expected Applaud shape (class, size, scale) via applaud_type_for."""
    pt = ParsedType(data_type=(of.data_type or ""), length=of.length,
                    scale=of.scale, parse_warning=False)
    return _shape_from_applaud_str(applaud_type_for(pt))


def actual_shape(col: DataColumn) -> Shape:
    """Applaud column's actual shape (class, size, scale) from its DataDictionary type.
    X and U (Unicode text, audit §1.2) -> char; N -> numeric; else the lowercased code
    (e.g. D -> 'd' for DATE columns)."""
    dt = (col.data_type or "").strip().upper()
    if dt in ("X", "U"):
        return ("char", col.size, None)
    if dt == "N":
        return ("numeric", col.size, col.dec_places)
    return (dt.lower(), col.size, col.dec_places)


def check_sizing(fbdi_template: str, fbdi_tab: str, table_name: str,
                 oracle_by_bare: dict[str, AlignedField],
                 columns: list[DataColumn]) -> list[Finding]:
    """Dim 1: flag undersized / precision-loss / type-class mismatches between each Oracle
    field and its Applaud target-table column. Oversize is not a defect."""
    col_by_bare = {c.bare.upper(): c for c in columns}
    findings: list[Finding] = []
    for bare, of in oracle_by_bare.items():
        col = col_by_bare.get(bare.upper())
        if col is None:
            continue   # presence is Dim 4's job, not Dim 1
        exp_cls, exp_size, exp_scale = expected_shape(of)
        act_cls, act_size, act_scale = actual_shape(col)
        oracle_type = applaud_type_for(ParsedType(of.data_type or "", of.length,
                                                  of.scale, False))
        common = dict(fbdi_template=fbdi_template, fbdi_tab=fbdi_tab,
                      oracle_field=bare, oracle_type=oracle_type,
                      applaud_object_type="TABLE", applaud_object_name=table_name,
                      applaud_field=col.ddid)

        def mk(attribute, severity, current, expected, message):
            return Finding(finding_id=make_finding_id(
                dimension="1-SIZING", applaud_object_type="TABLE",
                applaud_object_name=table_name, applaud_field=col.ddid,
                attribute=attribute), dimension="1-SIZING", severity=severity,
                attribute=attribute, current_value=current, expected_value=expected,
                message=message, **common)

        if exp_cls != act_cls and exp_cls not in ("", act_cls):
            findings.append(mk("TYPE_CLASS", "HIGH",
                f"{act_cls} {act_size}", f"{exp_cls} {exp_size}",
                f"Type-class mismatch: Applaud {act_cls} vs Oracle {exp_cls}"))
            continue
        if exp_cls == "char" and exp_size and act_size is not None and act_size < exp_size:
            findings.append(mk("SIZE", "HIGH", f"char {act_size}", f"char {exp_size}",
                f"Undersized: Applaud char {act_size} < Oracle char {exp_size}"))
        elif exp_cls == "numeric":
            if exp_size and act_size is not None and act_size < exp_size:
                findings.append(mk("SIZE", "HIGH",
                    f"numeric {act_size},{act_scale or 0}",
                    f"numeric {exp_size},{exp_scale or 0}",
                    f"Precision loss: Applaud {act_size} < Oracle {exp_size} digits"))
            elif (exp_scale or 0) > (act_scale or 0):
                findings.append(mk("SCALE", "HIGH",
                    f"numeric {act_size},{act_scale or 0}",
                    f"numeric {exp_size},{exp_scale or 0}",
                    f"Scale loss: Applaud scale {act_scale or 0} < Oracle {exp_scale}"))
    return findings


# ---------------------------------------------------------------------------
# Dim 2 / 3 — IF/EF coverage & ordering (Task 8)
# ---------------------------------------------------------------------------

def _lcs_sequence(a: list[str], b: list[str]) -> list[str]:
    """Longest common subsequence of two string lists (the in-order overlap)."""
    m, n = len(a), len(b)
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m - 1, -1, -1):
        for j in range(n - 1, -1, -1):
            dp[i][j] = (dp[i + 1][j + 1] + 1) if a[i] == b[j] else max(dp[i + 1][j], dp[i][j + 1])
    out: list[str] = []
    i = j = 0
    while i < m and j < n:
        if a[i] == b[j]:
            out.append(a[i]); i += 1; j += 1
        elif dp[i + 1][j] >= dp[i][j + 1]:
            i += 1
        else:
            j += 1
    return out


def _presence_finding(fbdi_template, fbdi_tab, object_type, object_name, dimension,
                      *, oracle_field, applaud_field, severity, message) -> Finding:
    id_field = applaud_field if applaud_field != "(missing)" else oracle_field
    return Finding(
        finding_id=make_finding_id(dimension=dimension, applaud_object_type=object_type,
            applaud_object_name=object_name, applaud_field=id_field, attribute="PRESENCE"),
        dimension=dimension, severity=severity, fbdi_template=fbdi_template,
        fbdi_tab=fbdi_tab, oracle_field=oracle_field, oracle_type="",
        applaud_object_type=object_type, applaud_object_name=object_name,
        applaud_field=applaud_field, attribute="PRESENCE",
        current_value="absent" if applaud_field == "(missing)" else "present",
        expected_value="present" if applaud_field == "(missing)" else "(advisory)",
        message=message)


def check_file_coverage(fbdi_template: str, fbdi_tab: str, object_name: str,
                        object_type: str, dimension: str,
                        oracle_fields: list[AlignedField],
                        file_fields: list[FileField]) -> list[Finding]:
    """Dim 2 (IF) / Dim 3 (EF): coverage (set difference) + ordering (LCS over
    fields common to both). Identity is the bare DDID (Applaud side) vs the
    normalized Oracle key; never ColumnHeader (empty on real EFs)."""
    oracle_order = [oracle_match_key(f) for f in oracle_fields if oracle_match_key(f)]
    oracle_set = set(oracle_order)
    file_sorted = sorted(file_fields, key=lambda x: x.row)
    file_order = [f.bare.upper() for f in file_sorted]
    file_set = set(file_order)
    findings: list[Finding] = []

    # PRESENCE — Oracle field missing from the file (HIGH)
    for f in oracle_fields:
        name = oracle_match_key(f)
        if name and name not in file_set:
            findings.append(_presence_finding(
                fbdi_template, fbdi_tab, object_type, object_name, dimension,
                oracle_field=name, applaud_field="(missing)", severity="HIGH",
                message=f"Missing field: Oracle {name} not in {object_name}"))

    # PRESENCE — extra file field with no Oracle counterpart (INFO)
    for ff in file_sorted:
        if ff.bare.upper() not in oracle_set:
            findings.append(_presence_finding(
                fbdi_template, fbdi_tab, object_type, object_name, dimension,
                oracle_field="", applaud_field=ff.ddid, severity="INFO",
                message=f"Extra field: {object_name} has {ff.ddid} with no Oracle counterpart"))

    # ORDER — among fields common to both, is relative order preserved? (MED)
    common_in_oracle = [n for n in oracle_order if n in file_set]
    common_in_file = [n for n in file_order if n in oracle_set]
    if common_in_oracle != common_in_file:
        keep = set(_lcs_sequence(common_in_oracle, common_in_file))
        ddid_by_bare = {f.bare.upper(): f.ddid for f in file_sorted}
        for idx, n in enumerate(common_in_file):
            if n in keep:
                continue
            field_id = ddid_by_bare.get(n, n)
            findings.append(Finding(
                finding_id=make_finding_id(dimension=dimension, applaud_object_type=object_type,
                    applaud_object_name=object_name, applaud_field=field_id, attribute="ORDER"),
                dimension=dimension, severity="MED", fbdi_template=fbdi_template,
                fbdi_tab=fbdi_tab, oracle_field=n, oracle_type="",
                applaud_object_type=object_type, applaud_object_name=object_name,
                applaud_field=field_id, attribute="ORDER",
                current_value=f"file pos {idx + 1}", expected_value="Oracle relative order",
                message=f"Ordering violation: {n} is out of Oracle field order in {object_name}"))
    return findings


# ---------------------------------------------------------------------------
# Dim 4 — target-table field coverage (Task 10)
# ---------------------------------------------------------------------------

def check_table_coverage(fbdi_template: str, fbdi_tab: str, table_name: str,
                         oracle_fields: list[AlignedField],
                         columns: list[DataColumn]) -> list[Finding]:
    """Dim 4: flag each mapped Oracle field that has no column in the target table
    (matched on bare name / ODBCName)."""
    present = set()
    for c in columns:
        present.add(c.bare.upper())
        if c.odbc_name:
            present.add(c.odbc_name.upper())
    findings: list[Finding] = []
    for of in oracle_fields:
        tech = oracle_match_key(of)
        if not tech or tech in present:
            continue
        findings.append(Finding(
            finding_id=make_finding_id(dimension="4-TABLE", applaud_object_type="TABLE",
                applaud_object_name=table_name, applaud_field=tech, attribute="PRESENCE"),
            dimension="4-TABLE", severity="HIGH", fbdi_template=fbdi_template,
            fbdi_tab=fbdi_tab, oracle_field=tech,
            oracle_type="", applaud_object_type="TABLE", applaud_object_name=table_name,
            applaud_field="(missing)", attribute="PRESENCE",
            current_value="absent", expected_value="present",
            message=f"Missing column: Oracle {tech} has no column in {table_name}"))
    return findings


# ---------------------------------------------------------------------------
# Dim 5 — data element <-> target-table consistency (Task 11)
# ---------------------------------------------------------------------------

def check_orphans(fbdi_template: str, fbdi_tab: str, table_name: str,
                  object_name: str, object_type: str,
                  table_columns: list[DataColumn],
                  file_fields: list[FileField]) -> list[Finding]:
    """Dim 5: flag IF/EF data elements (exact DDID) absent from the target table —
    orphans that load into nothing."""
    table_ddids = {c.ddid.upper() for c in table_columns}
    findings: list[Finding] = []
    for f in file_fields:
        if f.ddid.upper() in table_ddids:
            continue
        findings.append(Finding(
            finding_id=make_finding_id(dimension="5-ORPHAN", applaud_object_type=object_type,
                applaud_object_name=object_name, applaud_field=f.ddid, attribute="PRESENCE"),
            dimension="5-ORPHAN", severity="MED", fbdi_template=fbdi_template,
            fbdi_tab=fbdi_tab, oracle_field="", oracle_type="",
            applaud_object_type=object_type, applaud_object_name=object_name,
            applaud_field=f.ddid, attribute="PRESENCE",
            current_value=f"in {object_name}", expected_value=f"column in {table_name}",
            message=f"Orphaned data element: {f.ddid} used in {object_name} but absent "
                    f"from table {table_name}"))
    return findings


# ---------------------------------------------------------------------------
# Dim 6b — release-delta cross-check (Task 12)
# ---------------------------------------------------------------------------

def build_release_changes(
    old_catalog: dict[tuple[str, str], list[AlignedField]],
    new_catalog: dict[tuple[str, str], list[AlignedField]],
) -> dict[tuple[str, str], list[Change]]:
    """Per-tab align(old, new) so Dim 6b can see Oracle's added/removed fields.
    Both sides are catalog rows (same numbering space) — the correct use of align_tabs."""
    out: dict[tuple[str, str], list[Change]] = {}
    for key, new_rows in new_catalog.items():
        old_rows = old_catalog.get(key)
        if old_rows:
            out[key] = align_tabs(old_rows, new_rows)
    return out


def check_release_delta(fbdi_template: str, fbdi_tab: str, table_name: str,
                        changes: list[Change], applaud_bares: set[str],
                        old_release: str, new_release: str) -> list[Finding]:
    """Dim 6b: flag Oracle fields added in the new release but absent from Applaud (behind),
    and fields removed by Oracle that still linger in Applaud (stale)."""
    present = {b.upper() for b in applaud_bares}
    findings: list[Finding] = []
    for ch in changes:
        if ch.change_type == "ADDED" and ch.new_field is not None:
            name = oracle_match_key(ch.new_field)
            if name and name not in present:
                findings.append(_release_finding(fbdi_template, fbdi_tab, table_name,
                    name, "HIGH", "ADDED",
                    f"Behind release: Oracle added {name} in {new_release}; "
                    f"absent from Applaud {table_name}"))
        elif ch.change_type == "REMOVED" and ch.old_field is not None:
            name = oracle_match_key(ch.old_field)
            if name and name in present:
                findings.append(_release_finding(fbdi_template, fbdi_tab, table_name,
                    name, "MED", "REMOVED",
                    f"Stale field: Oracle removed {name} in {new_release}; "
                    f"still present in Applaud {table_name}"))
    return findings


def _release_finding(fbdi_template, fbdi_tab, table_name, name, severity,
                     change_type, message) -> Finding:
    return Finding(
        finding_id=make_finding_id(dimension="6b-RELEASE", applaud_object_type="TABLE",
            applaud_object_name=table_name, applaud_field=name, attribute=change_type),
        dimension="6b-RELEASE", severity=severity, fbdi_template=fbdi_template,
        fbdi_tab=fbdi_tab, oracle_field=name, oracle_type="",
        applaud_object_type="TABLE", applaud_object_name=table_name,
        applaud_field=name, attribute=change_type,
        current_value=("absent" if change_type == "ADDED" else "present"),
        expected_value=("present" if change_type == "ADDED" else "removed"),
        message=message)


# ---------------------------------------------------------------------------
# Dim 6c — unmapped-but-present + coverage gaps (Task 13)
# ---------------------------------------------------------------------------

def check_unmapped(snapshot_tables: set[str], mapped_tables: set[str]) -> list[Finding]:
    """Dim 6c: INFO-flag T_* tables in the snapshot that have no FBDI mapping row."""
    findings: list[Finding] = []
    for t in sorted(snapshot_tables):
        if not t.upper().startswith("T_") or t in mapped_tables:
            continue
        findings.append(Finding(
            finding_id=make_finding_id(dimension="6c-UNMAPPED", applaud_object_type="TABLE",
                applaud_object_name=t, applaud_field="", attribute="PRESENCE"),
            dimension="6c-UNMAPPED", severity="INFO", fbdi_template="", fbdi_tab="",
            oracle_field="", oracle_type="", applaud_object_type="TABLE",
            applaud_object_name=t, applaud_field="", attribute="PRESENCE",
            current_value="present", expected_value="(no FBDI mapping)",
            message=f"Unmapped Applaud table: {t} has no FBDI mapping row"))
    return findings


def coverage_gaps(mapped_tables: set[str],
                  appmap: dict[str, tuple[list[str], list[str]]]) -> list[tuple[str, str]]:
    """Mapped tables whose confirmed app-map resolved no IF/EF — what the audit could not
    check, so silence is never mistaken for a pass (Coverage sheet)."""
    gaps: list[tuple[str, str]] = []
    for t in sorted(mapped_tables):
        ifs, efs = appmap.get(t, ([], []))
        if not ifs and not efs:
            gaps.append((t, "no IF/EF resolved in app-map"))
    return gaps


# ---------------------------------------------------------------------------
# Excel findings writer (Task 14)
# ---------------------------------------------------------------------------

_SEVERITY_FILLS = {
    "HIGH": PatternFill("solid", fgColor="FCE4D6"),
    "MED":  PatternFill("solid", fgColor="FFF2CC"),
    "INFO": PatternFill("solid", fgColor="E2EFDA"),
}

_FINDINGS_HEADERS = [
    "Finding ID", "Dimension", "Severity", "FBDI Template", "FBDI Tab",
    "Oracle Field", "Oracle Type", "Applaud Object Type", "Applaud Object",
    "Applaud Field", "Attribute", "Current", "Expected", "Message", "Status", "Notes",
]


def _finding_row(f: Finding) -> list:
    return [f.finding_id, f.dimension, f.severity, f.fbdi_template, f.fbdi_tab,
            f.oracle_field, f.oracle_type, f.applaud_object_type, f.applaud_object_name,
            f.applaud_field, f.attribute, f.current_value, f.expected_value,
            f.message, f.status, f.notes]


def _write_findings_sheet(ws, findings: list[Finding]) -> None:
    ws.append(_FINDINGS_HEADERS)
    _style_header_row(ws, len(_FINDINGS_HEADERS))
    for f in findings:
        ws.append(_finding_row(f))
        fill = _SEVERITY_FILLS.get(f.severity)
        if fill:
            for col in range(1, len(_FINDINGS_HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    ws.freeze_panes = "A2"


def write_findings_workbook(findings: list[Finding], coverage: list[tuple[str, str]],
                            meta: dict, path: Path) -> None:
    """Write the 4-sheet findings workbook (Summary / Findings / High Priority / Coverage),
    severity-sorted and colored, with empty Status/Notes columns for Phase-2 triage."""
    sev_order = {"HIGH": 0, "MED": 1, "INFO": 2}
    findings = sorted(findings, key=lambda f: (sev_order.get(f.severity, 9), f.dimension))

    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Applaud Compliance Report"])
    ws_sum.append(["System", meta.get("system", "")])
    ws_sum.append(["Release", meta.get("release", "")])
    ws_sum.append(["Snapshot extracted", meta.get("extracted_at", "")])
    ws_sum.append([])
    ws_sum.append(["Dimension", "HIGH", "MED", "INFO"])
    dims = sorted({f.dimension for f in findings})
    for d in dims:
        ws_sum.append([d,
            sum(1 for f in findings if f.dimension == d and f.severity == "HIGH"),
            sum(1 for f in findings if f.dimension == d and f.severity == "MED"),
            sum(1 for f in findings if f.dimension == d and f.severity == "INFO")])

    _write_findings_sheet(wb.create_sheet("Findings"), findings)
    _write_findings_sheet(wb.create_sheet("High Priority"),
                          [f for f in findings if f.severity == "HIGH"])

    ws_cov = wb.create_sheet("Coverage")
    ws_cov.append(["Table", "Coverage Note"])
    _style_header_row(ws_cov, 2)
    for table, note in coverage:
        ws_cov.append([table, note])
    ws_cov.freeze_panes = "A2"

    wb.save(path)


# ---------------------------------------------------------------------------
# Orchestration (Task 15)
# ---------------------------------------------------------------------------

def run_audit(snapshot: ApplaudSnapshot,
              catalog: dict[tuple[str, str], list[AlignedField]],
              mapping: dict[tuple[str, str], dict],
              appmap: dict[str, AppMapRow],
              release: str,
              release_changes: dict[tuple[str, str], list[Change]],
              out_path: Path,
              old_release: str | None = None) -> list[Finding]:
    """Run all dimension checks over the mapped (template, tab)->table chain and write the
    findings workbook. Returns the flat findings list. `old_release` (optional) enables Dim 6b."""
    findings: list[Finding] = []
    mapped_tables: set[str] = set()
    appmap_pairs: dict[str, tuple[list[str], list[str]]] = {}

    for (template, tab), info in mapping.items():
        table_name = info.get("applaud_table")
        if not table_name:
            continue
        oracle_fields = catalog.get((template, tab), [])
        if not oracle_fields:
            continue
        table = snapshot.tables.get(table_name)
        row = appmap.get(table_name)
        mapped_tables.add(table_name)
        ifs = row.import_files if row else []
        efs = row.export_files if row else []
        appmap_pairs[table_name] = (ifs, efs)

        oracle_by_bare = {oracle_match_key(f): f
                          for f in oracle_fields if oracle_match_key(f)}

        if table is not None:
            findings += check_sizing(template, tab, table_name, oracle_by_bare, table.columns)
            findings += check_table_coverage(template, tab, table_name, oracle_fields, table.columns)

        for if_name in ifs:
            if_fields = snapshot.imports.get(if_name, [])
            findings += check_file_coverage(template, tab, if_name, "IMPORT", "2-IF",
                                            oracle_fields, if_fields)
            if table is not None:
                findings += check_orphans(template, tab, table_name, if_name, "IMPORT",
                                          table.columns, if_fields)
        for ef_name in efs:
            ef_fields = snapshot.exports.get(ef_name, [])
            findings += check_file_coverage(template, tab, ef_name, "EXPORT", "3-EF",
                                            oracle_fields, ef_fields)
            if table is not None:
                findings += check_orphans(template, tab, table_name, ef_name, "EXPORT",
                                          table.columns, ef_fields)

        changes = release_changes.get((template, tab))
        if changes and table is not None:
            # Mirror check_table_coverage's match set (bare + ODBCName) so a field that
            # matches via ODBCName isn't falsely flagged ADDED/REMOVED by Dim 6b.
            applaud_bares = {c.bare.upper() for c in table.columns}
            applaud_bares |= {c.odbc_name.upper() for c in table.columns if c.odbc_name}
            findings += check_release_delta(template, tab, table_name, changes,
                                            applaud_bares, old_release=(old_release or "(prior)"),
                                            new_release=release)

    findings += check_unmapped(set(snapshot.tables), mapped_tables)
    coverage = coverage_gaps(mapped_tables, appmap_pairs)

    write_findings_workbook(findings, coverage,
        {"system": snapshot.system, "release": release,
         "extracted_at": snapshot.extracted_at}, out_path)
    return findings
