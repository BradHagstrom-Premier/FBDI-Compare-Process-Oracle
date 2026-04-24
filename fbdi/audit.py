"""
audit.py — FBDI ↔ Applaud mapping audit engine.

Consumes applaud_snapshot.json + FBDI_Master_Catalog.xlsx (26B) +
fbdi_applaud_mapping.xlsx and produces Claude_fbdi_applaud_mapping.xlsx
(3 sheets) + Claude_fbdi_applaud_mapping_audit.md.

Run: python -m fbdi.audit
"""
from __future__ import annotations

import json
import logging
import re
import warnings
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).parent.parent
SNAPSHOT_PATH = REPO_ROOT / "baselines" / "applaud" / "applaud_snapshot.json"
CATALOG_PATH = REPO_ROOT / "FBDI_Master_Catalog.xlsx"
PRIOR_MAPPING_PATH = REPO_ROOT / "fbdi_applaud_mapping.xlsx"
OUTPUT_MAPPING_PATH = REPO_ROOT / "Claude_fbdi_applaud_mapping.xlsx"
OUTPUT_AUDIT_PATH = REPO_ROOT / "Claude_fbdi_applaud_mapping_audit.md"
CATALOG_RELEASE = "26B"
SNAPSHOT_MAX_AGE_DAYS = 30

_STRIP_SUFFIXES = ("_ALL", "_INT", "_INTERFACE")

_LABEL_NORMALIZE_RE = re.compile(r"[^A-Z0-9]+")


def _label_to_technical(label: str) -> str:
    """Normalize an FBDI column label to a technical-ish UPPER_SNAKE_CASE token.

    Used as a fallback when the catalog row has no column_technical (thin tabs
    where Oracle's source template exposes only human-readable labels). Strips
    the required marker, any non-alphanumerics, and collapses to underscores.
    """
    if not label:
        return ""
    s = label.strip().lstrip("*").strip().upper()
    return _LABEL_NORMALIZE_RE.sub("_", s).strip("_")
_log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class SnapshotField:
    name: str
    bare_name: str
    is_legacy_tracking: bool
    data_type: str
    length: int


@dataclass
class SnapshotKeySeq:
    seq: str
    keys: list[str]


@dataclass
class SnapshotTable:
    name: str
    prefix: str | None
    description: str
    type: str
    key_sequences: list[SnapshotKeySeq]
    fields: list[SnapshotField]

    def business_fields(self) -> list[SnapshotField]:
        return [f for f in self.fields if not f.is_legacy_tracking]

    def key_bare_names(self) -> set[str]:
        bare: set[str] = set()
        for seq in self.key_sequences:
            for k in seq.keys:
                # Keys are stored as full prefixed names; strip prefix
                if self.prefix and k.upper().startswith(self.prefix.upper()):
                    bare.add(k[len(self.prefix):].upper())
                else:
                    bare.add(k.upper())
        return bare


@dataclass
class ApplaudSnapshot:
    mdb_path: str
    extracted_at: str
    extractor_version: str
    tables: list[SnapshotTable]
    missing_tables: list[dict]

    def table_by_name(self) -> dict[str, SnapshotTable]:
        return {t.name: t for t in self.tables}

    def missing_set(self) -> set[str]:
        return {m["name"] for m in self.missing_tables}


@dataclass
class Candidate:
    fbdi_file: str
    fbdi_tab: str
    name_alignment: str           # EXACT | PARTIAL | NONE
    key_coverage: float
    column_overlap: float
    prefix_conformance: bool
    applaud_key_fields_matched: list[str]
    applaud_fields_matched: list[str]
    applaud_fields_missing: list[str]


@dataclass
class EvidenceBundle:
    candidates_evaluated: list[Candidate] = field(default_factory=list)
    rejected_alternatives: list[Candidate] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


@dataclass
class PriorRow:
    applaud_table: str
    prior_status: str
    prefix: str
    mapping_text: str
    module: str
    notes: str


@dataclass
class AuditRow:
    applaud_table: str
    prefix: str
    verdict: str                   # YES | UNMAPPED | NEEDS_REVIEW | FILE_TOO_LARGE | FILE_ERROR
    fbdi_mapping: str
    confidence: str                # H | M | L | ""
    rationale: str
    prior_verdict: str
    changed: bool
    needs_deep_rationale: bool
    evidence: EvidenceBundle


# Type aliases
CatalogIndex = dict[tuple[str, str], set[str]]   # {(file_name, tab_name): set[column_technical]}
CandidateIndex = dict[str, list[Candidate]]       # {applaud_table_name: sorted candidates}


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_snapshot(path: Path = SNAPSHOT_PATH) -> ApplaudSnapshot:
    if not path.exists():
        raise FileNotFoundError(f"Snapshot missing — run Step A first: {path}")
    data = json.loads(path.read_text(encoding="utf-8"))
    tables = []
    for t in data["tables"]:
        fields = [
            SnapshotField(
                name=f["name"],
                bare_name=f["bare_name"],
                is_legacy_tracking=f["is_legacy_tracking"],
                data_type=f["data_type"],
                length=f["length"],
            )
            for f in t["fields"]
        ]
        key_seqs = [
            SnapshotKeySeq(seq=k["seq"], keys=k["keys"])
            for k in t["key_sequences"]
        ]
        tables.append(SnapshotTable(
            name=t["name"],
            prefix=t.get("prefix"),
            description=t.get("description", ""),
            type=t.get("type", ""),
            key_sequences=key_seqs,
            fields=fields,
        ))
    return ApplaudSnapshot(
        mdb_path=data["mdb_path"],
        extracted_at=data["extracted_at"],
        extractor_version=data["extractor_version"],
        tables=tables,
        missing_tables=data.get("missing_tables", []),
    )


def load_catalog(
    path: Path = CATALOG_PATH, release: str = CATALOG_RELEASE
) -> CatalogIndex:
    if not path.exists():
        raise FileNotFoundError(f"Catalog missing: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if release not in wb.sheetnames:
            raise ValueError(f"No '{release}' tab in catalog. Available: {wb.sheetnames}")
        ws = wb[release]
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter)
        headers = [str(h).strip().lower() if h else "" for h in raw_headers]
        try:
            file_col = headers.index("file_name")
            tab_col = headers.index("tab_name")
            tech_col = headers.index("column_technical")
            label_col = headers.index("column_label")
        except ValueError as exc:
            raise ValueError(f"Catalog missing expected header: {exc}. Got: {headers}")
        index: CatalogIndex = {}
        for row in rows_iter:
            fname = str(row[file_col]).strip() if row[file_col] else ""
            tab = str(row[tab_col]).strip() if row[tab_col] else ""
            tech = str(row[tech_col]).strip() if row[tech_col] else ""
            label = str(row[label_col]).strip() if row[label_col] else ""
            if fname and tab:
                key = (fname, tab)
                index.setdefault(key, set())
                if tech:
                    index[key].add(tech.upper())
                elif label:
                    # Thin-tab fallback: Oracle source has no technical name,
                    # only a human-readable label. Normalize the label so it
                    # can participate in column-overlap / key-coverage signals.
                    normalized = _label_to_technical(label)
                    if normalized:
                        index[key].add(normalized)
        return index
    finally:
        wb.close()


def load_prior_mapping(path: Path = PRIOR_MAPPING_PATH) -> dict[str, PriorRow]:
    if not path.exists():
        raise FileNotFoundError(f"Prior mapping missing: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Find Applaud Tables sheet by name (case-insensitive), fall back to index 1
        ws = None
        for name in wb.sheetnames:
            if "applaud" in name.lower():
                ws = wb[name]
                break
        if ws is None:
            if len(wb.sheetnames) >= 2:
                ws = wb.worksheets[1]
            else:
                raise ValueError(
                    f"No 'Applaud Tables' sheet found. Sheets: {wb.sheetnames}"
                )
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter)
        headers = [
            str(h).strip().lower().replace(" ", "_") if h else ""
            for h in raw_headers
        ]

        def _col(name: str) -> int:
            return headers.index(name) if name in headers else -1

        col_table = _col("applaud_table")
        col_status = _col("status")
        col_prefix = _col("prefix")
        col_mapping = _col("fbdi_template_mappings")
        col_module = _col("module")
        col_notes = _col("notes")

        if col_table == -1:
            raise ValueError(
                f"Sheet2 missing 'applaud_table' column. Headers found: {headers}"
            )

        result: dict[str, PriorRow] = {}
        for row in rows_iter:
            def _val(idx: int) -> str:
                if idx == -1 or idx >= len(row):
                    return ""
                v = row[idx]
                return str(v).strip() if v is not None else ""

            table_name = _val(col_table)
            if not table_name or table_name.startswith("#"):
                continue
            result[table_name] = PriorRow(
                applaud_table=table_name,
                prior_status=_val(col_status),
                prefix=_val(col_prefix),
                mapping_text=_val(col_mapping),
                module=_val(col_module),
                notes=_val(col_notes),
            )
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Prefix + bare_name utilities
# ---------------------------------------------------------------------------

_PREFIX_RE = re.compile(r'\(([A-Z0-9]+)\)\s*$')


def extract_prefix(description: str) -> str | None:
    m = _PREFIX_RE.search(description.strip())
    return m.group(1) if m else None


def derive_bare_name(field_name: str, prefix: str) -> tuple[str, bool]:
    """Return (bare_name, is_legacy_tracking)."""
    name = field_name
    is_legacy = False
    if name.startswith("@"):
        is_legacy = True
        name = name[1:]  # strip @
    upper_prefix = prefix.upper()
    if name.upper().startswith(upper_prefix):
        return name[len(prefix):], is_legacy
    return name, is_legacy


# ---------------------------------------------------------------------------
# Signal computation
# ---------------------------------------------------------------------------

def compute_name_alignment(applaud_table: str, fbdi_tab: str) -> str:
    """Compare Applaud table name (strip ``T_``) against FBDI tab name.

    Oracle ships two tab-naming conventions: technical (``RA_INTERFACE_LINES_ALL``)
    and human-readable (``Award Budget Periods``). Both get normalized to
    UPPER_SNAKE_CASE before comparison so name alignment works either way.

    Matching is tiered:
      * EXACT — names are identical after strip-T_ and label normalization.
      * PARTIAL — names match after stripping common suffixes (_ALL / _INT /
        _INTERFACE) from either side.
      * PARTIAL (loose) — names match after additionally stripping ``_T``
        (Oracle temp-table convention), collapsing all separators, and
        normalizing trailing-S plurals. Guards against spurious hits by
        requiring the collapsed form to be ≥ 4 characters.
      * NONE — no match.
    """
    stripped = _label_to_technical(applaud_table.removeprefix("T_").removeprefix("t_"))
    tab_upper = _label_to_technical(fbdi_tab)

    if stripped == tab_upper:
        return "EXACT"

    def _base(s: str) -> str:
        for suffix in _STRIP_SUFFIXES:
            if s.endswith(suffix):
                return s[: -len(suffix)]
        return s

    s_base = _base(stripped)
    t_base = _base(tab_upper)
    if s_base == t_base:
        return "PARTIAL"
    if s_base == tab_upper or stripped == t_base:
        return "PARTIAL"

    # Loose PARTIAL: separators collapsed + singular/plural + _T suffix
    def _loose(s: str) -> str:
        s = s.removesuffix("_T")
        s = s.replace("_", "")
        if s.endswith("S"):
            s = s[:-1]
        return s

    ls = _loose(s_base)
    lt = _loose(t_base)
    if ls == lt and len(ls) >= 4:
        return "PARTIAL"

    return "NONE"


def compute_key_coverage(
    applaud_key_bare_names: set[str], fbdi_columns: set[str]
) -> float:
    if not applaud_key_bare_names:
        return 0.0
    fbdi_upper = {c.upper() for c in fbdi_columns}
    matched = sum(1 for k in applaud_key_bare_names if k.upper() in fbdi_upper)
    return matched / len(applaud_key_bare_names)


def compute_column_overlap(
    applaud_fields: list[SnapshotField], fbdi_columns: set[str]
) -> float:
    biz_fields = [f for f in applaud_fields if not f.is_legacy_tracking]
    if not biz_fields:
        return 0.0
    fbdi_upper = {c.upper() for c in fbdi_columns}
    matched = sum(1 for f in biz_fields if f.bare_name.upper() in fbdi_upper)
    return matched / len(biz_fields)


def check_prefix_conformance(
    applaud_table: str, prefix: str, fbdi_tab: str
) -> bool:
    """True when Applaud table name minus T_ exactly equals the FBDI tab name."""
    return applaud_table.upper().removeprefix("T_") == fbdi_tab.upper()


# ---------------------------------------------------------------------------
# Pass 1 — candidate index
# ---------------------------------------------------------------------------

_PASS1_MIN_NAME_ALIGNMENT = {"EXACT", "PARTIAL"}
_PASS1_MIN_KEY_COVERAGE = 0.5
_PASS1_MIN_COLUMN_OVERLAP = 0.3


def _sort_key(c: Candidate) -> tuple:
    align_order = {"EXACT": 0, "PARTIAL": 1, "NONE": 2}
    return (align_order[c.name_alignment], -c.key_coverage, -c.column_overlap)


def build_candidate_index(
    snapshot: ApplaudSnapshot, catalog: CatalogIndex
) -> CandidateIndex:
    index: CandidateIndex = {}
    table_by_name = snapshot.table_by_name()

    for applaud_table_name, snap_table in table_by_name.items():
        candidates: list[Candidate] = []
        key_bare = snap_table.key_bare_names()

        for (fbdi_file, fbdi_tab), fbdi_cols in catalog.items():
            name_align = compute_name_alignment(applaud_table_name, fbdi_tab)
            key_cov = compute_key_coverage(key_bare, fbdi_cols)
            col_ovlp = compute_column_overlap(snap_table.fields, fbdi_cols)
            prefix_ok = check_prefix_conformance(
                applaud_table_name, snap_table.prefix or "", fbdi_tab
            )

            # Pass-1 threshold: keep if any signal clears its floor
            if (
                name_align in _PASS1_MIN_NAME_ALIGNMENT
                or key_cov >= _PASS1_MIN_KEY_COVERAGE
                or col_ovlp >= _PASS1_MIN_COLUMN_OVERLAP
            ):
                fbdi_upper = {c.upper() for c in fbdi_cols}
                biz_fields = snap_table.business_fields()
                matched = [f.bare_name for f in biz_fields if f.bare_name.upper() in fbdi_upper]
                missing = [f.bare_name for f in biz_fields if f.bare_name.upper() not in fbdi_upper]
                key_matched = [k for k in key_bare if k.upper() in fbdi_upper]

                candidates.append(Candidate(
                    fbdi_file=fbdi_file,
                    fbdi_tab=fbdi_tab,
                    name_alignment=name_align,
                    key_coverage=key_cov,
                    column_overlap=col_ovlp,
                    prefix_conformance=prefix_ok,
                    applaud_key_fields_matched=key_matched,
                    applaud_fields_matched=matched,
                    applaud_fields_missing=missing,
                ))

        candidates.sort(key=_sort_key)
        index[applaud_table_name] = candidates

    return index


# ---------------------------------------------------------------------------
# Prior-mapping text parser
# ---------------------------------------------------------------------------

def parse_prior_mapping(mapping_text: str) -> list[tuple[str, str]]:
    """Parse "Template / Tab[; Template / Tab]" → [(file, tab), ...]."""
    result: list[tuple[str, str]] = []
    if not mapping_text or not mapping_text.strip():
        return result
    for segment in mapping_text.split(";"):
        segment = segment.strip()
        if not segment:
            continue
        parts = segment.split(" / ", maxsplit=1)
        if len(parts) != 2 or not parts[0].strip() or not parts[1].strip():
            _log.warning("Malformed prior mapping segment (skipping): %r", segment)
            continue
        result.append((parts[0].strip(), parts[1].strip()))
    return result


# ---------------------------------------------------------------------------
# Confidence tier evaluator
# ---------------------------------------------------------------------------

def evaluate_confidence(candidate: Candidate) -> str:
    """Return H, M, or L. First matching rule wins.

    Signals are name alignment, key coverage, and column overlap. The rubric
    evolved from the original spec (§6.2) in two ways to avoid counter-
    intuitive verdicts against real 26B data:

    1. EXACT name alignment alone → at least M. The original rubric had
       EXACT + weak data fall to L while PARTIAL → M regardless. That
       inverted signal strength and flagged canonical mappings (e.g.
       ``T_RA_INTERFACE_LINES_ALL`` → ``RA_INTERFACE_LINES_ALL``) as
       NEEDS_REVIEW on thin tabs (label-only, no technical columns).
    2. 100% key coverage alone → at least M. Keys are the most
       semantically-loaded columns (e.g. ``BANK_NAME``, ``ACCOUNT_NUMBER``),
       so full key match is strong evidence of a real mapping even when
       names diverge (e.g. ``T_BANKS_BRANCHES`` → ``Bank Account``).

    High column overlap *alone* is intentionally NOT promoted because
    Oracle's generic DFF columns (``ATTRIBUTE1``–``20``, ``ATTRIBUTE_DATE*``)
    inflate overlap across unrelated tabs.
    """
    if candidate.name_alignment == "EXACT" and (
        candidate.key_coverage == 1.0 or candidate.column_overlap >= 0.7
    ):
        return "H"
    # 100% key coverage promotes to M only when 2+ keys matched — single
    # generic keys (SEQUENCE_NUMBER, ACCOUNT_NUMBER) aren't discriminative
    # enough and would match many unrelated interface tabs.
    full_key_match_discriminative = (
        candidate.key_coverage == 1.0
        and len(candidate.applaud_key_fields_matched) >= 2
    )
    if (
        candidate.name_alignment in ("EXACT", "PARTIAL")
        or full_key_match_discriminative
        or (0 < candidate.key_coverage < 1.0 and candidate.column_overlap >= 0.4)
    ):
        return "M"
    return "L"


# ---------------------------------------------------------------------------
# Pass 2 — adjudication
# ---------------------------------------------------------------------------

_CARRYTHROUGH_VERDICTS = {"FILE_TOO_LARGE", "FILE_ERROR"}


def _find_candidate(
    candidates: list[Candidate], fbdi_file: str, fbdi_tab: str
) -> Candidate | None:
    for c in candidates:
        if c.fbdi_file == fbdi_file and c.fbdi_tab == fbdi_tab:
            return c
    return None


def adjudicate_table(
    applaud_table: str,
    snap_table: SnapshotTable | None,
    candidates: list[Candidate],
    prior: PriorRow,
) -> AuditRow:
    evidence = EvidenceBundle(candidates_evaluated=list(candidates))
    prefix = snap_table.prefix or prior.prefix if snap_table else prior.prefix

    # ── PREFLIGHT ────────────────────────────────────────────────────────────
    if snap_table is None and prior.prior_status not in _CARRYTHROUGH_VERDICTS:
        return AuditRow(
            applaud_table=applaud_table, prefix=prefix,
            verdict="UNMAPPED", fbdi_mapping="",
            confidence="H", rationale="Applaud table not present in MDB snapshot",
            prior_verdict=prior.prior_status, changed=False,
            needs_deep_rationale=False, evidence=evidence,
        )

    if prior.prior_status in _CARRYTHROUGH_VERDICTS:
        return AuditRow(
            applaud_table=applaud_table, prefix=prefix,
            verdict=prior.prior_status, fbdi_mapping=prior.mapping_text,
            confidence="", rationale="Sized out / unreadable in 26B — unchanged from prior",
            prior_verdict=prior.prior_status, changed=False,
            needs_deep_rationale=False, evidence=evidence,
        )

    # ── PRIOR MAPPING PARSE ──────────────────────────────────────────────────
    prior_claims = parse_prior_mapping(prior.mapping_text)
    best_candidate = candidates[0] if candidates else None

    verdict: str
    fbdi_mapping: str
    confidence: str
    rationale: str

    # ── UNMAPPED / blank ─────────────────────────────────────────────────────
    if prior.prior_status in ("UNMAPPED", "") or (
        prior.prior_status == "YES" and not prior.mapping_text.strip()
    ):
        if best_candidate:
            conf = evaluate_confidence(best_candidate)
            if conf == "H":
                verdict = "YES"
                fbdi_mapping = f"{best_candidate.fbdi_file} / {best_candidate.fbdi_tab}"
                confidence = "H"
                rationale = (
                    f"Promoted from UNMAPPED — EXACT name match, "
                    f"key={best_candidate.key_coverage:.0%}, overlap={best_candidate.column_overlap:.0%}"
                )
            elif conf == "M":
                verdict = "NEEDS_REVIEW"
                fbdi_mapping = f"{best_candidate.fbdi_file} / {best_candidate.fbdi_tab}"
                confidence = "M"
                rationale = "Potential new mapping — Medium confidence; verify with Brad"
            else:
                verdict = "UNMAPPED"
                fbdi_mapping = ""
                confidence = "H"
                rationale = "No FBDI tab in 26B catalog scores above threshold"
        else:
            verdict = "UNMAPPED"
            fbdi_mapping = ""
            confidence = "H"
            rationale = "No FBDI tab in 26B catalog scores above threshold"

    # ── SINGLE prior claim ───────────────────────────────────────────────────
    elif len(prior_claims) == 1:
        file, tab = prior_claims[0]
        matched_c = _find_candidate(candidates, file, tab)
        if matched_c:
            conf = evaluate_confidence(matched_c)
            if conf in ("H", "M"):
                verdict = "YES"
                fbdi_mapping = f"{file} / {tab}"
                confidence = conf
                rationale = (
                    f"name={matched_c.name_alignment}, "
                    f"key={matched_c.key_coverage:.0%}, "
                    f"overlap={matched_c.column_overlap:.0%}"
                )
            else:
                verdict = "NEEDS_REVIEW"
                fbdi_mapping = f"{file} / {tab}"
                confidence = "L"
                rationale = "Prior claim scores Low against 26B catalog — verify"
        else:
            verdict = "NEEDS_REVIEW"
            fbdi_mapping = f"{file} / {tab}"
            confidence = "L" if best_candidate else "H"
            rationale = (
                "Prior references file/tab not found in 26B catalog or below all thresholds"
            )

    # ── MULTI prior claims ───────────────────────────────────────────────────
    else:
        high_or_med: list[tuple[str, str, Candidate, str]] = []
        low_or_absent: list[tuple[str, str]] = []
        for file, tab in prior_claims:
            c = _find_candidate(candidates, file, tab)
            if c:
                conf = evaluate_confidence(c)
                if conf in ("H", "M"):
                    high_or_med.append((file, tab, c, conf))
                else:
                    low_or_absent.append((file, tab))
                    evidence.rejected_alternatives.append(c)
            else:
                low_or_absent.append((file, tab))

        if len(high_or_med) == len(prior_claims):
            # All claims score High or Medium → keep multi
            verdict = "YES"
            fbdi_mapping = "; ".join(f"{f} / {t}" for f, t, _, _ in high_or_med)
            confidence = "H" if all(conf == "H" for _, _, _, conf in high_or_med) else "M"
            rationale = f"Multi-mapping retained — {len(high_or_med)} legs verified"
        elif len(high_or_med) == 1:
            # One good leg — collapse to single
            file, tab, c, conf = high_or_med[0]
            verdict = "YES"
            fbdi_mapping = f"{file} / {tab}"
            confidence = conf
            rationale = (
                f"Collapsed from multi — 1/{len(prior_claims)} legs scored {conf}; "
                f"rest below threshold"
            )
        else:
            verdict = "NEEDS_REVIEW"
            fbdi_mapping = "; ".join(f"{f} / {t}" for f, t in prior_claims)
            confidence = "M" if high_or_med else "L"
            rationale = "Multi-mapping contested — see audit.md for per-leg evidence"

    # ── PREFIX AUDIT (all verdicts) ──────────────────────────────────────────
    if verdict == "YES" and fbdi_mapping:
        # Check prefix_conformance on the first/primary chosen candidate
        first_claim = parse_prior_mapping(fbdi_mapping)
        if first_claim:
            chosen_c = _find_candidate(candidates, first_claim[0][0], first_claim[0][1])
            if chosen_c and not chosen_c.prefix_conformance:
                evidence.notes.append(
                    f"Prefix mismatch — expected T_<tab> convention, "
                    f"got prefix={prefix} for tab={first_claim[0][1]}"
                )

    changed = verdict != prior.prior_status or fbdi_mapping != prior.mapping_text.strip()
    needs_deep = (
        verdict == "NEEDS_REVIEW"
        or changed
        or confidence == "L"
        or bool(evidence.notes)
    )

    return AuditRow(
        applaud_table=applaud_table, prefix=prefix,
        verdict=verdict, fbdi_mapping=fbdi_mapping,
        confidence=confidence, rationale=rationale,
        prior_verdict=prior.prior_status, changed=changed,
        needs_deep_rationale=needs_deep, evidence=evidence,
    )


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_DATA_FONT = Font(name="Calibri", size=11)

_VERDICT_FILLS = {
    "YES":            PatternFill("solid", fgColor="E2EFDA"),
    "UNMAPPED":       PatternFill("solid", fgColor="FCE4D6"),
    "NEEDS_REVIEW":   PatternFill("solid", fgColor="FFF2CC"),
    "FILE_TOO_LARGE": PatternFill("solid", fgColor="F4B942"),
    "FILE_ERROR":     PatternFill("solid", fgColor="F4B942"),
}

_S1_HEADERS = [
    "FBDI Template", "FBDI Tab", "Applaud Table", "Prefix",
    "Status", "Module", "Notes", "Match Type", "Confidence",
]
_S2_HEADERS = [
    "#", "Applaud Table", "Status", "Prefix", "FBDI Template Mappings",
    "Confidence", "Rationale", "Changed From Prior", "Prior Status",
]


def _style_header_row(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_sheet1(ws, audit_rows: list[AuditRow], catalog: CatalogIndex) -> None:
    """FBDI Mapping — one row per (file, tab) in the 26B catalog."""
    tab_to_row: dict[tuple[str, str], AuditRow] = {}
    for ar in audit_rows:
        for file, tab in parse_prior_mapping(ar.fbdi_mapping):
            tab_to_row[(file, tab)] = ar

    ws.append(_S1_HEADERS)
    _style_header_row(ws, len(_S1_HEADERS))

    for (fbdi_file, fbdi_tab) in sorted(catalog):
        ar = tab_to_row.get((fbdi_file, fbdi_tab))
        if ar:
            match_type = "EXACT" if ar.confidence == "H" else "PARTIAL" if ar.confidence == "M" else "PRIOR-CARRYOVER"
            row = [fbdi_file, fbdi_tab, ar.applaud_table, ar.prefix,
                   ar.verdict, "", "", match_type, ar.confidence]
        else:
            row = [fbdi_file, fbdi_tab, "", "", "UNMAPPED", "", "", "", ""]
        ws.append(row)
        fill = _VERDICT_FILLS.get(row[4])
        if fill:
            for col in range(1, len(_S1_HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    ws.freeze_panes = "A2"


def _write_sheet2(ws, audit_rows: list[AuditRow]) -> None:
    """Applaud Tables — one row per Applaud table."""
    ws.append(_S2_HEADERS)
    _style_header_row(ws, len(_S2_HEADERS))

    for i, ar in enumerate(audit_rows, start=1):
        changed_mark = "✓" if ar.changed else ""
        row = [i, ar.applaud_table, ar.verdict, ar.prefix, ar.fbdi_mapping,
               ar.confidence, ar.rationale, changed_mark, ar.prior_verdict]
        ws.append(row)
        fill = _VERDICT_FILLS.get(ar.verdict)
        if fill:
            for col in range(1, len(_S2_HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    ws.freeze_panes = "A2"


def _write_sheet3(ws, audit_rows: list[AuditRow]) -> None:
    """Needs Review — filtered subset, sorted by priority."""
    ws.append(_S2_HEADERS)
    _style_header_row(ws, len(_S2_HEADERS))

    deep_rows = [ar for ar in audit_rows if ar.needs_deep_rationale]

    def _sort(ar: AuditRow) -> tuple:
        return (ar.verdict != "NEEDS_REVIEW", not ar.changed, ar.confidence != "L")

    deep_rows.sort(key=_sort)

    for i, ar in enumerate(deep_rows, start=1):
        changed_mark = "✓" if ar.changed else ""
        rationale = ar.rationale + " → see audit.md"
        row = [i, ar.applaud_table, ar.verdict, ar.prefix, ar.fbdi_mapping,
               ar.confidence, rationale, changed_mark, ar.prior_verdict]
        ws.append(row)
        fill = _VERDICT_FILLS.get(ar.verdict)
        if fill:
            for col in range(1, len(_S2_HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    ws.freeze_panes = "A2"


def write_output_xlsx(
    audit_rows: list[AuditRow],
    catalog: CatalogIndex,
    output_path: Path = OUTPUT_MAPPING_PATH,
) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "FBDI Mapping"
    _write_sheet1(ws1, audit_rows, catalog)

    ws2 = wb.create_sheet("Applaud Tables")
    _write_sheet2(ws2, audit_rows)

    ws3 = wb.create_sheet("Needs Review")
    _write_sheet3(ws3, audit_rows)

    wb.save(output_path)
    print(f"Wrote: {output_path}")


def write_audit_md(
    audit_rows: list[AuditRow],
    snapshot_meta: dict,
    output_path: Path = OUTPUT_AUDIT_PATH,
) -> None:
    deep_rows = [ar for ar in audit_rows if ar.needs_deep_rationale]
    needs_review = [ar for ar in deep_rows if ar.verdict == "NEEDS_REVIEW"]
    changed = [ar for ar in deep_rows if ar.changed and ar.verdict != "NEEDS_REVIEW"]

    total = len(audit_rows)
    yes_count = sum(1 for ar in audit_rows if ar.verdict == "YES")
    unmapped_count = sum(1 for ar in audit_rows if ar.verdict == "UNMAPPED")
    nr_count = len(needs_review)
    changed_count = sum(1 for ar in audit_rows if ar.changed)

    lines: list[str] = [
        "# FBDI ↔ Applaud Mapping Audit — 26B",
        "",
        f"**Generated:** {datetime.now(timezone.utc).isoformat()}",
        f"**Snapshot:** baselines/applaud/applaud_snapshot.json @ {snapshot_meta.get('extracted_at', 'unknown')}",
        "**Catalog:** FBDI_Master_Catalog.xlsx 26B tab",
        "**Prior mapping:** fbdi_applaud_mapping.xlsx",
        "",
        "## Summary",
        "",
        f"Of {total} Applaud tables audited: "
        f"{yes_count} YES, {unmapped_count} UNMAPPED, {nr_count} NEEDS_REVIEW. "
        f"{changed_count} rows changed from prior.",
        "",
    ]

    if needs_review:
        lines += [f"## Needs Review ({len(needs_review)} rows)", ""]
        for ar in needs_review:
            lines += _md_section(ar)

    if changed:
        lines += ["## Changed From Prior", ""]
        for ar in changed:
            lines += _md_section(ar)

    prefix_mismatches = [ar for ar in audit_rows if ar.evidence.notes]
    if prefix_mismatches:
        lines += ["## Prefix Mismatches", ""]
        lines += ["| Applaud Table | Prefix | Notes |", "|---|---|---|"]
        for ar in prefix_mismatches:
            for note in ar.evidence.notes:
                lines.append(f"| {ar.applaud_table} | {ar.prefix} | {note} |")
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote: {output_path}")


def _md_section(ar: AuditRow) -> list[str]:
    lines = [
        f"### {ar.applaud_table} (prefix: {ar.prefix}) — {ar.verdict}",
        f"- **Prior:** {ar.prior_verdict} → `{ar.fbdi_mapping or '(none)'}`",
        f"- **Decision:** {ar.rationale}",
    ]
    if ar.evidence.candidates_evaluated:
        lines.append("- **Candidates evaluated:**")
        for c in ar.evidence.candidates_evaluated[:5]:
            conf = evaluate_confidence(c)
            lines.append(
                f"  - `{c.fbdi_file} / {c.fbdi_tab}` — "
                f"name={c.name_alignment}, "
                f"keys={c.key_coverage:.0%}, "
                f"cols={c.column_overlap:.0%} → {conf}"
            )
    for note in ar.evidence.notes:
        lines.append(f"- **Note:** {note}")
    lines.append("")
    return lines


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def run_audit(
    snapshot_path: Path = SNAPSHOT_PATH,
    catalog_path: Path = CATALOG_PATH,
    prior_mapping_path: Path = PRIOR_MAPPING_PATH,
    output_xlsx_path: Path = OUTPUT_MAPPING_PATH,
    output_md_path: Path = OUTPUT_AUDIT_PATH,
) -> list[AuditRow]:
    snap = load_snapshot(snapshot_path)
    try:
        extracted = datetime.fromisoformat(snap.extracted_at.replace("Z", "+00:00"))
        age_days = (datetime.now(timezone.utc) - extracted).days
        if age_days > SNAPSHOT_MAX_AGE_DAYS:
            warnings.warn(
                f"Snapshot is {age_days} days old (>{SNAPSHOT_MAX_AGE_DAYS}). "
                "Re-run Step A if the MDB has changed.",
                stacklevel=2,
            )
    except Exception:
        pass

    catalog = load_catalog(catalog_path)
    prior_mapping = load_prior_mapping(prior_mapping_path)

    print(f"Snapshot: {len(snap.tables)} tables, {len(snap.missing_set())} missing")
    print(f"Catalog: {len(catalog)} (file, tab) pairs")
    print(f"Prior mapping: {len(prior_mapping)} Applaud tables")

    candidate_index = build_candidate_index(snap, catalog)

    table_by_name = snap.table_by_name()
    missing_set = snap.missing_set()
    audit_rows: list[AuditRow] = []

    for table_name, prior_row in prior_mapping.items():
        snap_table = table_by_name.get(table_name)
        if table_name in missing_set:
            snap_table = None
        candidates = candidate_index.get(table_name, [])
        row = adjudicate_table(table_name, snap_table, candidates, prior_row)
        audit_rows.append(row)

    verdicts: dict[str, int] = {}
    for ar in audit_rows:
        verdicts[ar.verdict] = verdicts.get(ar.verdict, 0) + 1
    changed_count = sum(1 for ar in audit_rows if ar.changed)
    print(f"\nResults: {verdicts}")
    print(f"Changed from prior: {changed_count}")
    print(f"Needs deep rationale: {sum(1 for ar in audit_rows if ar.needs_deep_rationale)}")

    write_output_xlsx(audit_rows, catalog, output_xlsx_path)
    write_audit_md(audit_rows, {"extracted_at": snap.extracted_at}, output_md_path)

    return audit_rows


if __name__ == "__main__":
    run_audit()
