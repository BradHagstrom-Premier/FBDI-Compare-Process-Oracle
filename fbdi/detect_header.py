"""Dynamic header row detection for Oracle FBDI worksheets.

Many Oracle FBDI templates have a multi-row header structure:
  Row 1: Human-readable labels ("Name", "*Calendar Code", "Batch Code")
  Row 2: Descriptions
  Row 3: Data types
  Row 4: Required/Optional indicators
  Row 5+: Technical column names (UPPER_SNAKE_CASE like "INTERFACE_BATCH_CODE")

The VBA macro hardcodes which row to use per template. This module detects
the correct row dynamically using a two-tier approach:
  Tier 1: Look for a row dominated by UPPER_SNAKE_CASE names (technical headers)
  Tier 2: Fall back to the best header-like row (mixed case, asterisk-prefixed)
"""

import logging
import re

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from fbdi.config import MIN_CELLS

logger = logging.getLogger(__name__)

# Oracle FBDI technical column names: UPPER_SNAKE_CASE
UPPER_SNAKE_PATTERN = re.compile(r"^[A-Z][A-Z0-9_]+$")

# Maximum length for header-like strings (instruction text is longer)
HEADER_LIKE_MAX_LEN = 50

# Tier 1 threshold: fraction of non-empty cells that must be UPPER_SNAKE_CASE
TIER1_PATTERN_THRESHOLD = 0.5

# Tier 1 fill floor: Tier-1 candidates must cover at least this fraction of the
# sheet's column span.  Data rows are sparse (fill ~0.05) even when their few
# non-empty cells happen to be UPPER_SNAKE_CASE; real header rows fill the sheet.
TIER1_MIN_FILL = 0.15

# Tier 2 threshold: minimum combined score for header-like rows
TIER2_SCORE_THRESHOLD = 0.35


def _is_header_like(value: str) -> bool:
    """Check if a string looks like a column header (vs instruction/description text)."""
    s = value.strip().lstrip("*").strip()
    if not s:
        return False
    if UPPER_SNAKE_PATTERN.match(s):
        return True
    if len(s) > HEADER_LIKE_MAX_LEN:
        return False
    # Sentence indicators
    if s.endswith(".") or s.count(",") > 1:
        return False
    lower = s.lower()
    if any(phrase in lower for phrase in [
        "please ", "must be", "should be", "refer to", "note:",
        "indicates the", "contains one of", "user-defined",
        "code that identifies", "associated with",
    ]):
        return False
    return True


def _scan_rows(ws: Worksheet, max_scan: int) -> list[dict]:
    """Scan rows and compute metrics for each.

    Uses iter_rows() (streaming) rather than ws.cell(row, col) lookups.
    In read-only mode, ws.cell() is O(n) per call (re-scans streamed rows),
    so per-cell access on a 500-wide sheet is O(n*m) overall. iter_rows() is
    O(n) for the whole scan. On wide sheets (e.g. 679 columns) this turns
    a 74-second detection into <1 second.
    """
    # Cap at 500 columns — real headers never exceed this, and some sheets
    # report max_column=16384 (Excel max) due to phantom formatting.
    max_col = min(ws.max_column or 1, 500)
    row_data = []

    for row_idx, row_cells in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, max_col=max_col),
        start=1,
    ):
        cells = []
        actual_max_col = 0
        for col_idx, cell in enumerate(row_cells, start=1):
            if isinstance(cell, MergedCell):
                cells.append(None)
            else:
                cells.append(cell.value)
                if cell.value is not None and str(cell.value).strip() != "":
                    actual_max_col = col_idx

        non_empty = [v for v in cells if v is not None and str(v).strip() != ""]
        if len(non_empty) < MIN_CELLS:
            continue

        str_cells = [v for v in non_empty if isinstance(v, str)]
        upper_snake = [v for v in str_cells if UPPER_SNAKE_PATTERN.match(str(v).strip())]
        header_like = [v for v in str_cells if _is_header_like(str(v))]
        short_strs = [v for v in str_cells if len(str(v).strip()) < HEADER_LIKE_MAX_LEN]

        n = len(non_empty)
        ns = len(str_cells)

        row_data.append({
            "row": row_idx,
            "non_empty": n,
            "upper_snake_ratio": len(upper_snake) / n,
            "header_like_ratio": len(header_like) / n,
            # fill_ratio is recomputed below using the sheet-wide max column span.
            # Storing actual_max_col here temporarily; per-row extent is wrong for
            # narrow rows (e.g. a 2-cell legend row in cols 1-2 gets 2/2=1.0 the
            # same as a 45-col header row, which causes the legend row to win).
            "_actual_max_col": actual_max_col,
            "str_ratio": ns / n,
            "brevity_ratio": len(short_strs) / ns if ns > 0 else 0.0,
        })

    # fill_ratio must reflect what fraction of the *template's* column span each
    # row covers, not each row's own span.  Use the sheet-wide maximum so that a
    # 2-cell legend row in a 45-column template scores ~0.04, not 1.0.
    sheet_max_col = max((r["_actual_max_col"] for r in row_data), default=1)
    for r in row_data:
        r["fill_ratio"] = r["non_empty"] / sheet_max_col if sheet_max_col > 0 else 0.0
        del r["_actual_max_col"]

    return row_data


def detect_header_row(ws: Worksheet, max_scan: int = 20) -> int | None:
    """Detect the header row in an FBDI worksheet.

    Two-tier approach:
      Tier 1: Find the row with the highest UPPER_SNAKE_CASE concentration.
              If >50% of non-empty cells match, use it. This catches technical
              header rows at positions 3, 5, 8, 10, 11, 16.
      Tier 2: If no strong UPPER_SNAKE_CASE row, score rows on header-like
              characteristics (short labels, high fill ratio, all strings).
              This catches templates with mixed-case headers at row 1 or 4.

    Returns 1-indexed row number, or None if no confident match found.
    """
    rows = _scan_rows(ws, max_scan)
    if not rows:
        logger.warning("No candidate rows found in '%s'", ws.title)
        return None

    # Tier 1: Look for UPPER_SNAKE_CASE dominated row.
    # Require fill_ratio >= TIER1_MIN_FILL to exclude sparse data rows that
    # happen to have >= 50% UPPER_SNAKE values across their few non-empty cells
    # (e.g. a data row with BMRX / RECLASS / POST / ACK out of 12 cells in a
    # 260-column sheet scores snake=0.50, fill=0.05 and would otherwise beat
    # the real 260-column mixed header at snake=0.49).
    tier1_candidates = [
        r for r in rows
        if r["upper_snake_ratio"] >= TIER1_PATTERN_THRESHOLD
        and r["fill_ratio"] >= TIER1_MIN_FILL
    ]
    if tier1_candidates:
        # Among tier1 candidates, prefer highest upper_snake_ratio * fill_ratio
        best = max(
            tier1_candidates,
            key=lambda r: r["upper_snake_ratio"] * 0.6 + r["fill_ratio"] * 0.4,
        )
        logger.debug(
            "Tier 1 match: row %d (upper_snake=%.2f, fill=%.2f)",
            best["row"], best["upper_snake_ratio"], best["fill_ratio"],
        )
        return best["row"]

    # Tier 2: Score on general header-like characteristics.
    # Tiebreak by non_empty count: a real header row spans many columns, while
    # narrow annotation rows (e.g. "* Required") may score identically but have
    # far fewer cells.
    best_row = None
    best_score = 0.0
    best_non_empty = 0
    for r in rows:
        score = (
            0.40 * r["header_like_ratio"]
            + 0.35 * r["fill_ratio"]
            + 0.15 * r["str_ratio"]
            + 0.10 * r["brevity_ratio"]
        )
        logger.debug(
            "Tier 2 row %d: score=%.3f (header_like=%.2f, fill=%.2f, str=%.2f, brevity=%.2f)",
            r["row"], score, r["header_like_ratio"], r["fill_ratio"],
            r["str_ratio"], r["brevity_ratio"],
        )
        if (score, r["non_empty"]) > (best_score, best_non_empty):
            best_score = score
            best_non_empty = r["non_empty"]
            best_row = r["row"]

    if best_row is not None and best_score > TIER2_SCORE_THRESHOLD:
        logger.debug("Tier 2 match: row %d (score=%.3f)", best_row, best_score)
        return best_row

    logger.warning(
        "No confident header row found in '%s' (best score=%.3f at row %s)",
        ws.title, best_score, best_row,
    )
    return None
