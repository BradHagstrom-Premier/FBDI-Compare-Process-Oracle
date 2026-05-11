"""FBDI Compliance Report generator.

Reads the FBDI Master Catalog (per-release sheets) and the FBDI-to-Applaud
mapping, runs alignment per (file, tab), filters to the in-scope universe
(MAPPED only; pending-base routed to a separate section), and emits an
HTML and PDF report from one Jinja2 template.

This module exposes:
- build_report_context(...) — pure view-model construction (testable in isolation)
- generate_report(...)      — top-level: load -> build -> render -> write (TBD)

The view-model dataclasses (ReportContext, FileSection, ChangeRow,
PendingBaseEntry) are the contract between this module and the template.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from fbdi.align import AlignedField, Change, align_tabs
from fbdi.applaud_type import applaud_type_for
from fbdi.catalog_normalize import normalize_label
from fbdi.type_parser import parse_data_type


APPLAUD_NAME_LIMIT = 30


@dataclass
class ChangeRow:
    """One row in a per-file change-type table (view-model)."""
    change_type: str
    applaud_field_name: str
    name_length: int
    name_exceeds_30: bool
    old_position: int | None
    new_position: int | None
    label: str
    oracle_type_str: str           # e.g. "VARCHAR2(30)" — empty when not applicable
    applaud_type_str: str          # e.g. "char 30"
    required: bool | None
    axes: tuple[str, ...]
    sub_kinds: tuple[str, ...]
    # For RENAMED / MODIFIED / MULTI — old vs new values to display side-by-side
    old_label: str | None = None
    new_label: str | None = None
    old_oracle_type_str: str | None = None
    new_oracle_type_str: str | None = None
    old_required: bool | None = None
    new_required: bool | None = None


@dataclass
class FileSection:
    """One per-file section in the main body."""
    file: str
    tab: str
    applaud_table: str
    prefix: str
    module: str
    status: str                    # "MAPPED" | "NEEDS_REVIEW"
    in_base_note: str | None       # e.g. the "Multiple mapping is possible..." string when present
    changes_by_type: dict[str, list[ChangeRow]] = field(default_factory=dict)
    shift_summary: str | None = None  # e.g. "20 fields shifted from positions 19-39 to 20-40"


@dataclass
class PendingBaseEntry:
    """One entry in the pending base-system tables list."""
    file: str
    tab: str
    applaud_table: str
    prefix: str
    module: str
    change_count: int


@dataclass
class ReportContext:
    """Top-level view-model passed to the Jinja2 template."""
    old_release: str
    new_release: str
    generated_date: str
    file_sections: list[FileSection]
    pending_base: list[PendingBaseEntry]


# Public: scope filtering and view-model construction ---------------------------

def build_report_context(
    catalog_old: dict[tuple[str, str], list[AlignedField]],
    catalog_new: dict[tuple[str, str], list[AlignedField]],
    mapping: dict[tuple[str, str], dict],
    old_release: str,
    new_release: str,
    generated_date: str | None = None,
) -> ReportContext:
    """Build the report context from grouped catalog data + mapping lookup.

    catalog_old / catalog_new keys are (file, tab) tuples. Values are
    AlignedField rows already in catalog form. mapping keys match the
    catalog keys; values are dicts with 'applaud_table', 'prefix',
    'module', 'in_base'.

    Scope rules:
    - keys absent from mapping are silently excluded (UNMAPPED universe).
    - mapping rows whose 'in_base' contains "Needs to be created in base
      system" are routed to the pending_base list rather than file_sections.
    - keys with no detected changes are dropped (no empty sections).
    """
    from datetime import date as _date
    if generated_date is None:
        generated_date = _date.today().isoformat()

    file_sections: list[FileSection] = []
    pending_base: list[PendingBaseEntry] = []
    all_keys = set(catalog_old.keys()) | set(catalog_new.keys())

    for key in sorted(all_keys):
        if key not in mapping:
            continue  # UNMAPPED — silently exclude
        m = mapping[key]
        file_name, tab = key

        old_rows = catalog_old.get(key, [])
        new_rows = catalog_new.get(key, [])
        changes = align_tabs(old_rows, new_rows)
        if not changes:
            continue

        in_base = m.get("in_base") or ""
        if "Needs to be created in base system" in in_base:
            pending_base.append(PendingBaseEntry(
                file=file_name, tab=tab,
                applaud_table=m["applaud_table"],
                prefix=m["prefix"],
                module=m["module"],
                change_count=len(changes),
            ))
            continue

        in_base_note = in_base if in_base else None

        section = FileSection(
            file=file_name, tab=tab,
            applaud_table=m["applaud_table"],
            prefix=m["prefix"],
            module=m["module"],
            status=m.get("status", "MAPPED"),
            in_base_note=in_base_note,
        )
        section.changes_by_type = _bucket_changes(changes, prefix=m["prefix"])
        section.shift_summary = _build_shift_summary(section.changes_by_type.get("SHIFTED", []))
        file_sections.append(section)

    # Sort by (module, file, tab) for stable ordering — also drives
    # the template's groupby('module') so groups appear in this order.
    file_sections.sort(key=lambda s: (s.module or "", s.file, s.tab))
    pending_base.sort(key=lambda p: (p.module or "", p.file, p.tab))

    return ReportContext(
        old_release=old_release,
        new_release=new_release,
        generated_date=generated_date,
        file_sections=file_sections,
        pending_base=pending_base,
    )


def _applaud_field_name(prefix: str, technical: str | None, label: str | None) -> str:
    """Construct the Applaud field name: prefix + technical (or normalized label).

    Technical UPPER_SNAKE_CASE names are used verbatim when present (already
    canonical). Otherwise the user-facing label is normalized — punctuation
    stripped, whitespace collapsed — to keep the suffix Applaud-compatible.
    """
    if technical:
        suffix = technical
    else:
        suffix = normalize_label(label or "")
    return f"{prefix}{suffix}"


def _oracle_type_str(f: AlignedField | None) -> str:
    """Return the Oracle-style type string for a field.

    Prefers data_type_raw (preserves CHAR unit, e.g. VARCHAR2(30 CHAR)) when
    present; falls back to reconstructing from parsed parts.
    """
    if f is None or not f.data_type:
        return ""
    if f.data_type_raw:
        return f.data_type_raw
    if f.length is not None and f.scale is not None:
        return f"{f.data_type}({f.length},{f.scale})"
    if f.length is not None:
        return f"{f.data_type}({f.length})"
    return f.data_type


def _applaud_type_str_for(f: AlignedField | None) -> str:
    """Translate an AlignedField's type into the Applaud-side type string."""
    raw = _oracle_type_str(f)
    if not raw:
        return ""
    return applaud_type_for(parse_data_type(raw))


def _bucket_changes(changes: list[Change], prefix: str) -> dict[str, list[ChangeRow]]:
    """Group classified changes into per-type buckets of ChangeRow view-models.

    The "primary" field for naming/typing is the new field when present
    (ADDED, MODIFIED, RENAMED, SHIFTED, MULTI) and the old field for
    REMOVED. Old/new pairs are also stamped onto the row so the template
    can render side-by-side comparisons for RENAMED / MODIFIED / MULTI.
    """
    buckets: dict[str, list[ChangeRow]] = defaultdict(list)
    for c in changes:
        primary = c.new_field if c.new_field is not None else c.old_field
        # primary is non-None for every classified change (align_tabs guarantee).
        applaud_name = _applaud_field_name(prefix, primary.technical, primary.label)
        oracle_type = _oracle_type_str(primary)
        applaud_type = _applaud_type_str_for(primary)

        row = ChangeRow(
            change_type=c.change_type,
            applaud_field_name=applaud_name,
            name_length=len(applaud_name),
            name_exceeds_30=len(applaud_name) > APPLAUD_NAME_LIMIT,
            old_position=c.old_position,
            new_position=c.new_position,
            label=primary.label or "",
            oracle_type_str=oracle_type,
            applaud_type_str=applaud_type,
            required=primary.required,
            axes=c.axes,
            sub_kinds=c.sub_kinds,
            old_label=c.old_field.label if c.old_field else None,
            new_label=c.new_field.label if c.new_field else None,
            old_oracle_type_str=_oracle_type_str(c.old_field) if c.old_field else None,
            new_oracle_type_str=_oracle_type_str(c.new_field) if c.new_field else None,
            old_required=c.old_field.required if c.old_field else None,
            new_required=c.new_field.required if c.new_field else None,
        )
        buckets[c.change_type].append(row)
    return dict(buckets)


def _build_shift_summary(shifted_rows: list[ChangeRow]) -> str | None:
    """Build the inline shift-summary sentence used in the SHIFTED block."""
    if not shifted_rows:
        return None
    old_positions = sorted(r.old_position for r in shifted_rows)
    new_positions = sorted(r.new_position for r in shifted_rows)
    n = len(shifted_rows)
    return (
        f"{n} field{'s' if n != 1 else ''} shifted from positions "
        f"{old_positions[0]}-{old_positions[-1]} to {new_positions[0]}-{new_positions[-1]}."
    )


# Public: on-disk loaders -------------------------------------------------------

def load_catalog_release(catalog_path: Path, release: str) -> dict[tuple[str, str], list[AlignedField]]:
    """Read one release sheet from the master catalog and group by (file, tab).

    Catalog schema (verified against FBDI_Master_Catalog.xlsx):
    release | file_name | tab_name | position | column_label |
    column_technical | data_type | length | scale | data_type_raw | required
    """
    wb = load_workbook(catalog_path, read_only=True, data_only=True)
    if release not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Release sheet '{release}' not found in {catalog_path}")
    ws = wb[release]

    grouped: dict[tuple[str, str], list[AlignedField]] = defaultdict(list)
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        # Schema: release, file_name, tab_name, position, column_label,
        # column_technical, data_type, length, scale, data_type_raw, required
        _rel, file_name, tab_name, position, label, technical, data_type, length, scale, data_type_raw, required = row
        if file_name is None or tab_name is None:
            continue
        grouped[(file_name, tab_name)].append(AlignedField(
            position=int(position),
            label=label,
            technical=(technical or None),
            data_type=(data_type or None),
            length=(int(length) if length is not None and length != "" else None),
            scale=(int(scale) if scale is not None and scale != "" else None),
            required=_parse_required(required),
            data_type_raw=(str(data_type_raw).strip() if data_type_raw is not None and data_type_raw != "" else None),
        ))

    wb.close()
    # Sort each group's rows by position to be safe
    for k in grouped:
        grouped[k].sort(key=lambda f: f.position)
    return dict(grouped)


def _parse_required(v) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().upper()
    if s == "TRUE":
        return True
    if s == "FALSE":
        return False
    return None


def load_mapping(mapping_path: Path) -> dict[tuple[str, str], dict]:
    """Read FBDI_to_ApplaudTables_Mapping.xlsx and return MAPPED-status rows.

    UNMAPPED rows are filtered out at load time (they're noise per the spec).
    NEEDS_REVIEW rows are kept so the report can flag them visually.

    Mapping schema (verified):
    FBDI Template | FBDI Tab | Applaud Table | Prefix | Status | Module |
    In Base System?
    """
    wb = load_workbook(mapping_path, read_only=True, data_only=True)
    ws = wb["FBDI Mapping"]
    out: dict[tuple[str, str], dict] = {}
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        # Schema: FBDI Template, FBDI Tab, Applaud Table, Prefix, Status,
        # Module, In Base System?
        template, tab, applaud_table, prefix, status, module, in_base = row[:7]
        if template is None or tab is None:
            continue
        if status not in ("MAPPED", "NEEDS_REVIEW"):
            continue
        out[(str(template), str(tab))] = {
            "applaud_table": applaud_table,
            "prefix": prefix,
            "module": module,
            "status": status,
            "in_base": in_base,
        }
    wb.close()
    return out


# Public: top-level entry point ------------------------------------------------

# Probed in order; first match wins. MSYS2's mingw64 ships current Pango (1.50+),
# so it goes first — required for weasyprint >= 53. The standalone GtkD installer
# is a fallback but caps at Pango 1.43, which can't run weasyprint >= 53.
_GTK_WINDOWS_BIN_CANDIDATES = (
    r"C:\msys64\mingw64\bin",
    r"C:\Program Files\GTK3-Runtime Win64\bin",
    r"C:\Program Files\Gtk-Runtime\bin",
    r"C:\Program Files (x86)\GTK3-Runtime Win64\bin",
)


def _register_windows_gtk_dlls() -> None:
    """On Windows, make a known GTK install dir loadable by weasyprint.

    Python 3.8+ ignores PATH for direct cffi.dlopen calls, so we register the
    GTK bin dir via os.add_dll_directory. cairocffi (used by weasyprint < 53)
    falls back to ctypes.util.find_library which honors PATH, so we prepend
    there too. No-op on non-Windows or when GTK is not at a known path.
    """
    import os
    import sys

    if sys.platform != "win32":
        return
    for candidate in _GTK_WINDOWS_BIN_CANDIDATES:
        if Path(candidate, "libgobject-2.0-0.dll").is_file():
            os.add_dll_directory(candidate)
            if candidate not in os.environ.get("PATH", ""):
                os.environ["PATH"] = candidate + os.pathsep + os.environ.get("PATH", "")
            return


def generate_report(
    catalog_path: Path,
    mapping_path: Path,
    old_release: str,
    new_release: str,
    out_dir: Path,
) -> tuple[Path, Path]:
    """Load -> build -> render -> write HTML and PDF.

    Returns (html_path, pdf_path).
    """
    _register_windows_gtk_dlls()
    import jinja2
    import weasyprint

    catalog_old = load_catalog_release(catalog_path, old_release)
    catalog_new = load_catalog_release(catalog_path, new_release)
    mapping = load_mapping(mapping_path)

    ctx = build_report_context(
        catalog_old=catalog_old,
        catalog_new=catalog_new,
        mapping=mapping,
        old_release=old_release,
        new_release=new_release,
    )

    template_dir = Path(__file__).parent / "templates"
    env = jinja2.Environment(
        loader=jinja2.FileSystemLoader(template_dir),
        autoescape=jinja2.select_autoescape(["html", "j2"]),
    )
    tpl = env.get_template("report.html.j2")

    out_dir.mkdir(parents=True, exist_ok=True)
    base = f"FBDI_Compliance_Report_{old_release}_{new_release}"
    html_path = out_dir / f"{base}.html"
    pdf_path = out_dir / f"{base}.pdf"

    html_path.write_text(tpl.render(ctx=ctx, print_mode=False), encoding="utf-8")

    pdf_html = tpl.render(ctx=ctx, print_mode=True)
    # base_url lets weasyprint resolve the bundled fonts in templates/fonts/
    weasyprint.HTML(string=pdf_html, base_url=str(template_dir)).write_pdf(str(pdf_path))

    return html_path, pdf_path
