"""Header detection diagnostic module for Oracle FBDI worksheets.

Runs header detection on every tab in every file and produces a diagnostic
.xlsx report showing detection outcomes, scores, and failure reasons.
"""

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from fbdi.config import MAX_FILE_SIZE_BYTES, SKIP_TABS
from fbdi.detect_header import TIER2_SCORE_THRESHOLD, _scan_rows, detect_header_row

logger = logging.getLogger(__name__)

DIAGNOSTIC_HEADERS = [
    "FBDI File",
    "FBDI Tab",
    "Detection Result",
    "Detected Row",
    "Best Score",
    "Failure Reason",
    "Notes",
]


@dataclass
class DiagnosticRow:
    fbdi_file: str
    fbdi_tab: str
    detection_result: str  # DETECTED / NO_HEADER / SKIPPED_TAB / FILE_ERROR / FILE_TOO_LARGE
    detected_row: int | None
    best_score: float | None
    failure_reason: str
    notes: str


def _best_score_for_ws(ws) -> float:
    """Compute the best Tier 2 score seen for a worksheet without modifying detect_header_row."""
    rows = _scan_rows(ws, max_scan=20)
    if not rows:
        return 0.0
    best = 0.0
    for r in rows:
        score = (
            0.40 * r["header_like_ratio"]
            + 0.35 * r["fill_ratio"]
            + 0.15 * r["str_ratio"]
            + 0.10 * r["brevity_ratio"]
        )
        if score > best:
            best = score
    return best


def diagnose_file(file_path: Path) -> list[DiagnosticRow]:
    """Run header detection diagnostics on all tabs in a single FBDI file.

    Returns one DiagnosticRow per tab (or one row for the file if it cannot be opened).
    """
    file_path = Path(file_path)
    file_stem = file_path.stem
    rows: list[DiagnosticRow] = []

    # Check file size first
    if file_path.stat().st_size > MAX_FILE_SIZE_BYTES:
        size_mb = file_path.stat().st_size / (1024 * 1024)
        logger.warning("DIAGNOSTIC: Skipping %s — %.1fMB exceeds limit", file_stem, size_mb)
        rows.append(DiagnosticRow(
            fbdi_file=file_stem,
            fbdi_tab="",
            detection_result="FILE_TOO_LARGE",
            detected_row=None,
            best_score=None,
            failure_reason=f"File size {size_mb:.1f}MB exceeds {MAX_FILE_SIZE_BYTES // (1024 * 1024)}MB limit",
            notes="Manual review required",
        ))
        return rows

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("DIAGNOSTIC: Cannot load %s: %s", file_path.name, e)
        rows.append(DiagnosticRow(
            fbdi_file=file_stem,
            fbdi_tab="",
            detection_result="FILE_ERROR",
            detected_row=None,
            best_score=None,
            failure_reason=str(e),
            notes="",
        ))
        return rows

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_TABS:
            rows.append(DiagnosticRow(
                fbdi_file=file_stem,
                fbdi_tab=sheet_name,
                detection_result="SKIPPED_TAB",
                detected_row=None,
                best_score=None,
                failure_reason="",
                notes=f"Tab skipped per SKIP_TABS config",
            ))
            continue

        ws = wb[sheet_name]
        detected_row = detect_header_row(ws)

        if detected_row is not None:
            rows.append(DiagnosticRow(
                fbdi_file=file_stem,
                fbdi_tab=sheet_name,
                detection_result="DETECTED",
                detected_row=detected_row,
                best_score=None,
                failure_reason="",
                notes="",
            ))
        else:
            best_score = _best_score_for_ws(ws)
            scan_rows = _scan_rows(ws, max_scan=20)
            if not scan_rows:
                reason = "No rows with MIN_CELLS >= 3 found in first 20 rows"
            else:
                reason = f"Best score {best_score:.3f} below {TIER2_SCORE_THRESHOLD} threshold"
            rows.append(DiagnosticRow(
                fbdi_file=file_stem,
                fbdi_tab=sheet_name,
                detection_result="NO_HEADER",
                detected_row=None,
                best_score=best_score,
                failure_reason=reason,
                notes="",
            ))

    wb.close()
    return rows


def write_diagnostic_report(rows: list[DiagnosticRow], output_path: Path) -> Path:
    """Write diagnostic rows to an .xlsx report file."""
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostic"

    bold = Font(name="Calibri", size=11, bold=True)
    normal = Font(name="Calibri", size=11)

    for col_idx, header in enumerate(DIAGNOSTIC_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_num, r in enumerate(rows, start=2):
        ws.cell(row=row_num, column=1, value=r.fbdi_file).font = normal
        ws.cell(row=row_num, column=2, value=r.fbdi_tab).font = normal
        ws.cell(row=row_num, column=3, value=r.detection_result).font = normal
        ws.cell(row=row_num, column=4, value=r.detected_row).font = normal
        ws.cell(row=row_num, column=5, value=r.best_score).font = normal
        ws.cell(row=row_num, column=6, value=r.failure_reason).font = normal
        ws.cell(row=row_num, column=7, value=r.notes).font = normal

    ws.auto_filter.ref = f"A1:G{max(len(rows) + 1, 1)}"

    wb.save(output_path)
    wb.close()
    logger.info("Diagnostic report written: %s (%d rows)", output_path, len(rows))
    return output_path
