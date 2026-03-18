"""Core FBDI comparison engine."""

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from fbdi.config import REPORT_HEADERS, SKIP_TABS
from fbdi.detect_header import detect_header_row
from fbdi.utils import col_index_to_letter, match_fbdi_files

logger = logging.getLogger(__name__)


@dataclass
class ComparisonRow:
    fbdi_file: str
    fbdi_tab: str
    column_letter: str
    column_number: int
    old_field_name: str | None
    new_field_name: str | None
    difference: str


def _read_header_values(ws: Worksheet, header_row: int) -> list[str | None]:
    """Read all header values from the detected header row."""
    max_col = 1
    # Find the last populated column in the header row
    for col_idx in range(1, (ws.max_column or 1) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        if cell.value is not None and str(cell.value).strip() != "":
            max_col = col_idx

    values = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if isinstance(cell, MergedCell):
            values.append(None)
        elif cell.value is not None:
            values.append(str(cell.value).strip())
        else:
            values.append(None)
    return values


def compare_fbdi_pair(old_path: Path, new_path: Path) -> list[ComparisonRow]:
    """Compare one old/new FBDI template pair across all non-skipped tabs."""
    file_stem = old_path.stem
    rows: list[ComparisonRow] = []

    old_wb = load_workbook(old_path, read_only=True, data_only=True)
    new_wb = load_workbook(new_path, read_only=True, data_only=True)

    new_sheet_names = {name.lower(): name for name in new_wb.sheetnames}

    # Iterate tabs from old file, look up same tab name in new file (VBA behavior)
    for old_sheet_name in old_wb.sheetnames:
        if old_sheet_name in SKIP_TABS:
            continue

        # Find matching tab in new file (case-insensitive)
        new_sheet_actual = new_sheet_names.get(old_sheet_name.lower())
        if new_sheet_actual is None:
            logger.warning(
                "%s: tab '%s' exists in old but not in new — skipping",
                file_stem, old_sheet_name,
            )
            continue

        old_ws = old_wb[old_sheet_name]
        new_ws = new_wb[new_sheet_actual]

        # Detect header row independently per file per tab
        old_header_row = detect_header_row(old_ws)
        new_header_row = detect_header_row(new_ws)

        if old_header_row is None:
            logger.warning(
                "%s: could not detect header row in old file tab '%s' — skipping",
                file_stem, old_sheet_name,
            )
            continue
        if new_header_row is None:
            logger.warning(
                "%s: could not detect header row in new file tab '%s' — skipping",
                file_stem, old_sheet_name,
            )
            continue

        old_headers = _read_header_values(old_ws, old_header_row)
        new_headers = _read_header_values(new_ws, new_header_row)

        # Align by position
        max_cols = max(len(old_headers), len(new_headers))
        for col_idx in range(max_cols):
            old_val = old_headers[col_idx] if col_idx < len(old_headers) else None
            new_val = new_headers[col_idx] if col_idx < len(new_headers) else None
            col_num = col_idx + 1

            rows.append(ComparisonRow(
                fbdi_file=file_stem,
                fbdi_tab=old_sheet_name,
                column_letter=col_index_to_letter(col_num),
                column_number=col_num,
                old_field_name=old_val,
                new_field_name=new_val,
                difference="YES" if old_val != new_val else "NO",
            ))

    # Log tabs only in new file
    old_sheet_lower = {name.lower() for name in old_wb.sheetnames}
    for new_name in new_wb.sheetnames:
        if new_name in SKIP_TABS:
            continue
        if new_name.lower() not in old_sheet_lower:
            logger.info(
                "%s: tab '%s' exists in new but not in old",
                file_stem, new_name,
            )

    old_wb.close()
    new_wb.close()
    return rows


def compare_all(
    old_dir: Path,
    new_dir: Path,
    output_path: Path,
    changes_only: bool = True,
) -> Path:
    """Compare all matched FBDI pairs and write Comparison_Report.xlsx."""
    old_dir = Path(old_dir)
    new_dir = Path(new_dir)
    output_path = Path(output_path)

    matched, old_only, new_only = match_fbdi_files(old_dir, new_dir)

    for f in old_only:
        logger.warning("Old-only file (no match in new): %s", f.name)
    for f in new_only:
        logger.warning("New-only file (no match in old): %s", f.name)

    all_rows: list[ComparisonRow] = []
    for old_path, new_path in matched:
        logger.info("Comparing: %s", old_path.stem)
        pair_rows = compare_fbdi_pair(old_path, new_path)
        all_rows.extend(pair_rows)

    # Write output
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    font = Font(name="Calibri", size=11, bold=True)
    for col_idx, header in enumerate(REPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font

    # Data rows
    data_font = Font(name="Calibri", size=11)
    row_num = 2
    change_count = 0
    for r in all_rows:
        if r.difference == "YES":
            change_count += 1
        if changes_only and r.difference != "YES":
            continue
        ws.cell(row=row_num, column=1, value=r.fbdi_file).font = data_font
        ws.cell(row=row_num, column=2, value=r.fbdi_tab).font = data_font
        ws.cell(row=row_num, column=3, value=r.column_letter).font = data_font
        ws.cell(row=row_num, column=4, value=r.column_number).font = data_font
        ws.cell(row=row_num, column=5, value=r.old_field_name or "").font = data_font
        ws.cell(row=row_num, column=6, value=r.new_field_name or "").font = data_font
        ws.cell(row=row_num, column=7, value=r.difference).font = data_font
        row_num += 1

    # Autofilter
    ws.auto_filter.ref = f"A1:G{max(row_num - 1, 1)}"

    # Autofit column widths (approximate)
    for col_idx in range(1, 8):
        max_len = len(REPORT_HEADERS[col_idx - 1])
        for row in range(2, row_num):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_index_to_letter(col_idx)].width = max_len + 2

    wb.save(output_path)
    wb.close()

    logger.info(
        "Comparison complete: %d file pairs, %d changes found, output: %s",
        len(matched), change_count, output_path,
    )

    return output_path
