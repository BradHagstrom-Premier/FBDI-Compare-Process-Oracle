"""CLI entry point for the FBDI comparison engine."""

import argparse
import logging
import sys
from pathlib import Path

from fbdi.compare import compare_all
from fbdi.utils import match_fbdi_files


def _resolve_dir(path: Path) -> Path:
    """Resolve a release label to its baselines originals directory.

    If path is already a directory, return it unchanged.
    Otherwise, try baselines/<path>/originals/ as a convenience shorthand.
    Falls through to the original path if no match (caller handles the error).
    """
    if path.is_dir():
        return path
    candidate = Path("baselines") / str(path) / "originals"
    if candidate.is_dir():
        return candidate
    return path


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        prog="fbdi",
        description="Oracle FBDI template comparison engine",
    )
    subparsers = parser.add_subparsers(dest="command")

    compare_parser = subparsers.add_parser(
        "compare",
        help="Compare FBDI templates between two release versions",
    )
    compare_parser.add_argument(
        "--old", required=True, type=Path,
        help="Path to directory containing old FBDI templates",
    )
    compare_parser.add_argument(
        "--new", required=True, type=Path,
        help="Path to directory containing new FBDI templates",
    )
    compare_parser.add_argument(
        "--output", type=Path, default=Path("Comparison_Report.xlsx"),
        help="Output file path (default: Comparison_Report.xlsx)",
    )
    compare_parser.add_argument(
        "--all-rows", action="store_true",
        help="Include unchanged rows in output (default: changes only)",
    )
    compare_parser.add_argument(
        "--verbose", action="store_true",
        help="Set logging to DEBUG (shows header detection scores)",
    )

    diagnose_parser = subparsers.add_parser(
        "diagnose",
        help="Diagnose header detection outcomes for FBDI templates",
    )
    diagnose_parser.add_argument(
        "--release", type=str, default=None,
        help="Release label (e.g. 26a) — looks in baselines/<release>/",
    )
    diagnose_parser.add_argument(
        "--old", type=Path, default=None,
        help="Path to old release directory",
    )
    diagnose_parser.add_argument(
        "--new", type=Path, default=None,
        help="Path to new release directory",
    )
    diagnose_parser.add_argument(
        "--output", type=Path, default=None,
        help="Output file path (default: Diagnostic_Report_<label>.xlsx)",
    )
    diagnose_parser.add_argument(
        "--verbose", action="store_true",
        help="Set logging to DEBUG",
    )

    catalog_parser = subparsers.add_parser(
        "catalog",
        help="Generate or update the FBDI master catalog for a release",
    )
    catalog_parser.add_argument(
        "--release", required=True, type=str,
        help="Release label (e.g. 26B) — looks in baselines/<release>/originals",
    )
    catalog_parser.add_argument(
        "--baselines-dir", type=Path, default=None,
        help="Explicit path to release originals dir (overrides --release resolution)",
    )
    catalog_parser.add_argument(
        "--master", type=Path, default=Path("FBDI_Master_Catalog.xlsx"),
        help="Output master workbook path (default: FBDI_Master_Catalog.xlsx)",
    )
    catalog_parser.add_argument(
        "--timeout", type=int, default=120,
        help="Per-file subprocess timeout in seconds (default: 120)",
    )
    catalog_parser.add_argument(
        "--verbose", action="store_true",
        help="Set logging to DEBUG",
    )

    populate_parser = subparsers.add_parser(
        "populate-module",
        help="Populate the Module column in FBDI_to_ApplaudTables_Mapping.xlsx",
    )
    populate_parser.add_argument(
        "--new", required=True, type=str,
        help="Newer release label (e.g. 26B) — reads baselines/<new>/file_modules.json",
    )
    populate_parser.add_argument(
        "--old", required=True, type=str,
        help="Older release label (e.g. 26A) — reads baselines/<old>/file_modules.json as fallback",
    )
    populate_parser.add_argument(
        "--mapping", type=Path,
        default=Path("FBDI_to_ApplaudTables_Mapping.xlsx"),
        help="Path to the mapping spreadsheet (default: ./FBDI_to_ApplaudTables_Mapping.xlsx)",
    )

    report_parser = subparsers.add_parser(
        "report",
        help="Generate the FBDI Compliance Report (HTML + PDF) from the catalog + mapping",
    )
    report_parser.add_argument(
        "--old", required=True, type=str,
        help="Older release label (e.g. 26A)",
    )
    report_parser.add_argument(
        "--new", required=True, type=str,
        help="Newer release label (e.g. 26B)",
    )
    report_parser.add_argument(
        "--out-dir", type=Path, default=Path("."),
        help="Output directory (default: ./)",
    )
    report_parser.add_argument(
        "--catalog", type=Path, default=Path("FBDI_Master_Catalog.xlsx"),
        help="Path to the master catalog (default: ./FBDI_Master_Catalog.xlsx)",
    )
    report_parser.add_argument(
        "--mapping", type=Path,
        default=Path("FBDI_to_ApplaudTables_Mapping.xlsx"),
        help="Path to the mapping spreadsheet (default: ./FBDI_to_ApplaudTables_Mapping.xlsx)",
    )

    audit_applaud_parser = subparsers.add_parser(
        "audit-applaud",
        help="Audit an Applaud system against the Oracle FBDI release it targets",
    )
    audit_applaud_parser.add_argument("--release", required=True, help="Release tag, e.g. 26B")
    audit_applaud_parser.add_argument(
        "--old-release", default=None,
        help="Prior release tag for Dim 6b (e.g. 26A); aligns the catalog's old sheet "
             "against --release. Omit to skip 6b.",
    )
    audit_applaud_parser.add_argument("--system", default="ORACLE_MASTER",
                                      help="Applaud system alias (default: ORACLE_MASTER)")
    audit_applaud_parser.add_argument("--catalog", type=Path,
                                      default=Path("FBDI_Master_Catalog.xlsx"))
    audit_applaud_parser.add_argument("--mapping", type=Path,
                                      default=Path("FBDI_to_ApplaudTables_Mapping.xlsx"))
    audit_applaud_parser.add_argument("--appmap", type=Path,
                                      default=Path("FBDI_to_Applaud_AppMap.xlsx"))
    audit_applaud_parser.add_argument("--output", type=Path, default=None)
    audit_applaud_parser.add_argument("--fieldmap", type=Path,
                                      default=Path("FBDI_to_Applaud_FieldMap.xlsx"),
                                      help="Committed Oracle<->Applaud field map (loaded if present)")
    audit_applaud_parser.add_argument("--accept-confidence", default="confirmed",
                                      choices=["confirmed", "HIGH", "PROBABLE", "WEAK"],
                                      help="Minimum origin/tier to alias (default: confirmed). The "
                                           "committed field map holds only confirmed/rejected rows, so "
                                           "HIGH/PROBABLE/WEAK currently behave the same as 'confirmed' "
                                           "(tier gates apply only to derived rows, which the committed "
                                           "map never stores).")
    audit_applaud_parser.add_argument(
        "--tables", default=None,
        help="Comma-separated Applaud target tables to scope the audit to "
             "(e.g. T_BANKS_BRANCHES,T_AP_INVOICE_INT). Omit to audit the full mapping. "
             "An unknown table name fails loud.")

    corr_derive_parser = subparsers.add_parser(
        "correspondence-derive",
        help="Propose Oracle<->Applaud field correspondences into a review workbook")
    corr_derive_parser.add_argument("--release", required=True, help="Release tag, e.g. 26B")
    corr_derive_parser.add_argument("--system", default="ORACLE_MASTER")
    corr_derive_parser.add_argument("--catalog", type=Path,
                                    default=Path("FBDI_Master_Catalog.xlsx"))
    corr_derive_parser.add_argument("--mapping", type=Path,
                                    default=Path("FBDI_to_ApplaudTables_Mapping.xlsx"))
    corr_derive_parser.add_argument("--map", dest="fieldmap", type=Path,
                                    default=Path("FBDI_to_Applaud_FieldMap.xlsx"),
                                    help="Committed field map; already-decided pairs are skipped")
    corr_derive_parser.add_argument("--tables", default=None,
                                    help="Comma-separated Applaud target tables to scope to")
    corr_derive_parser.add_argument("--output", type=Path, default=None)

    corr_confirm_parser = subparsers.add_parser(
        "correspondence-confirm",
        help="Merge reviewer decisions from a review workbook into the committed field map")
    corr_confirm_parser.add_argument("--review", type=Path, required=True)
    corr_confirm_parser.add_argument("--system", default="ORACLE_MASTER")
    corr_confirm_parser.add_argument("--map", dest="fieldmap", type=Path,
                                     default=Path("FBDI_to_Applaud_FieldMap.xlsx"))

    args = parser.parse_args(argv)

    if args.command is None:
        parser.print_help()
        sys.exit(1)

    if args.command == "compare":
        _run_compare(args)
    elif args.command == "diagnose":
        _run_diagnose(args)
    elif args.command == "catalog":
        _run_catalog(args)
    elif args.command == "populate-module":
        _run_populate_module(args)
    elif args.command == "report":
        _run_report(args)
    elif args.command == "audit-applaud":
        _run_audit_applaud(args)
    elif args.command == "correspondence-derive":
        _run_correspondence_derive(args)
    elif args.command == "correspondence-confirm":
        _run_correspondence_confirm(args)


def _run_compare(args: argparse.Namespace) -> None:
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(name)s: %(message)s",
    )

    old_dir = _resolve_dir(args.old)
    new_dir = _resolve_dir(args.new)

    if not old_dir.is_dir():
        print(f"Error: old directory not found: {old_dir}")
        sys.exit(1)
    if not new_dir.is_dir():
        print(f"Error: new directory not found: {new_dir}")
        sys.exit(1)

    # Print summary of matched files before comparison
    matched, old_only, new_only = match_fbdi_files(old_dir, new_dir)

    print(f"Matched file pairs: {len(matched)}")
    if old_only:
        print(f"Old-only files ({len(old_only)}):")
        for f in old_only:
            print(f"  - {f.name}")
    if new_only:
        print(f"New-only files ({len(new_only)}):")
        for f in new_only:
            print(f"  - {f.name}")

    print(f"\nComparing {len(matched)} file pairs...")

    output_path, timed_out = compare_all(
        old_dir,
        new_dir,
        args.output,
        changes_only=not args.all_rows,
    )

    # Count changes in output
    from openpyxl import load_workbook
    wb = load_workbook(output_path, read_only=True)
    ws = wb.active
    change_count = max((ws.max_row or 1) - 1, 0)
    wb.close()

    print(f"\nChanges found: {change_count}")
    print(f"Output written to: {output_path}")

    if timed_out:
        print(f"\n{'=' * 60}")
        print(f"WARNING: {len(timed_out)} file(s) timed out and were excluded from this report.")
        print("These files require manual review:")
        for name in timed_out:
            print(f"  - {name}")
        print(f"{'=' * 60}")


def _run_diagnose(args: argparse.Namespace) -> None:
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(name)s: %(message)s",
    )

    from fbdi.diagnose import diagnose_file, write_diagnostic_report

    # Resolve directories
    dirs: list[Path] = []
    label_parts: list[str] = []

    if args.release:
        release_dir = Path("baselines") / args.release
        if not release_dir.is_dir():
            print(f"Error: release directory not found: {release_dir}")
            sys.exit(1)
        dirs.append(release_dir)
        label_parts.append(args.release.upper())
    elif args.old or args.new:
        if not args.old or not args.new:
            print("Error: --old and --new must be used together")
            sys.exit(1)
        for d, flag in [(args.old, "--old"), (args.new, "--new")]:
            if not d.is_dir():
                print(f"Error: directory not found ({flag}): {d}")
                sys.exit(1)
        dirs.extend([args.old, args.new])
        label_parts.extend([args.old.name.upper(), args.new.name.upper()])
    else:
        print("Error: provide --release or --old/--new")
        sys.exit(1)

    # Determine output path
    label = "_".join(label_parts)
    output_path = args.output or Path(f"Diagnostic_Report_{label}.xlsx")

    # Scan files
    all_rows = []
    for directory in dirs:
        xlsm_files = sorted(directory.glob("*.xlsm"))
        print(f"Scanning {len(xlsm_files)} files in {directory} ...")
        for file_path in xlsm_files:
            rows = diagnose_file(file_path)
            all_rows.extend(rows)

    detected = sum(1 for r in all_rows if r.detection_result == "DETECTED")
    no_header = sum(1 for r in all_rows if r.detection_result == "NO_HEADER")
    skipped_tab = sum(1 for r in all_rows if r.detection_result == "SKIPPED_TAB")
    file_too_large = sum(1 for r in all_rows if r.detection_result == "FILE_TOO_LARGE")
    file_error = sum(1 for r in all_rows if r.detection_result == "FILE_ERROR")

    write_diagnostic_report(all_rows, output_path)

    print(f"\nDiagnostic complete: {len(all_rows)} tab entries")
    print(f"  DETECTED:       {detected}")
    print(f"  NO_HEADER:      {no_header}")
    print(f"  SKIPPED_TAB:    {skipped_tab}")
    print(f"  FILE_TOO_LARGE: {file_too_large}")
    print(f"  FILE_ERROR:     {file_error}")
    print(f"Output written to: {output_path}")


def _run_catalog(args: argparse.Namespace) -> None:
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(name)s: %(message)s",
    )

    from fbdi.catalog import generate_catalog

    # Resolve baselines dir
    if args.baselines_dir:
        baselines_dir = args.baselines_dir
    else:
        candidate = Path("baselines") / args.release / "originals"
        baselines_dir = candidate

    if not baselines_dir.is_dir():
        print(f"Error: baselines directory not found: {baselines_dir}")
        sys.exit(1)

    xlsm_count = len(list(baselines_dir.glob("*.xlsm")))
    if xlsm_count == 0:
        print(f"Error: no .xlsm files found in {baselines_dir}")
        sys.exit(1)

    print(f"Cataloging release {args.release.upper()} from {baselines_dir}")
    print(f"  {xlsm_count} .xlsm files")
    print(f"  Output: {args.master}")
    print()

    generate_catalog(
        release=args.release.upper(),
        baselines_dir=baselines_dir,
        master_path=args.master,
        timeout=args.timeout,
    )

    # Summary from the written workbook
    from openpyxl import load_workbook as _lw
    wb = _lw(args.master, read_only=True)
    release_tabs = [sn for sn in wb.sheetnames if sn not in {"Issues", "Drift"}]
    issue_count = max(0, (wb["Issues"].max_row or 1) - 1)
    drift_count = max(0, (wb["Drift"].max_row or 1) - 1)
    wb.close()

    print(f"\nCatalog updated: {args.master}")
    print(f"  Release tabs: {', '.join(release_tabs)}")
    print(f"  Issues: {issue_count}")
    print(f"  Drift rows: {drift_count}")


def _run_populate_module(args: argparse.Namespace) -> None:
    """Surgically populate the Module column in the mapping spreadsheet.

    Exit codes:
      0 — success (JSON summary printed) or mapping file absent (logged).
      2 — required file_modules.json missing for --new or --old release.
          Note: argparse also uses 2 for usage errors. Stage 6.5 of the
          fbdi-compare-release skill treats both as "halt and surface to user".
      3 — mapping spreadsheet is open in Excel (PermissionError).
    """
    import json

    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(name)s: %(message)s",
    )

    from fbdi.populate_module import populate_module_column

    new_path = Path("baselines") / args.new.lower() / "file_modules.json"
    old_path = Path("baselines") / args.old.lower() / "file_modules.json"

    if not new_path.is_file():
        print(f"Error: {new_path} not found. Run downloader for {args.new} first.")
        sys.exit(2)
    if not old_path.is_file():
        print(f"Error: {old_path} not found. Run downloader for {args.old} first.")
        sys.exit(2)

    if not args.mapping.is_file():
        print(json.dumps({
            "mapping": str(args.mapping),
            "new_release": args.new.upper(),
            "old_release": args.old.upper(),
            "status": "skipped",
            "reason": "mapping file not present",
        }, indent=2))
        return  # exit 0; the orchestrator skill pre-checks for this case

    with open(new_path, "r", encoding="utf-8") as f:
        new_modules = json.load(f)
    with open(old_path, "r", encoding="utf-8") as f:
        old_modules = json.load(f)

    try:
        result = populate_module_column(args.mapping, new_modules, old_modules)
    except PermissionError:
        print(f"Error: {args.mapping} is open in Excel — close it and re-run.")
        sys.exit(3)

    print(json.dumps({
        "mapping": str(args.mapping),
        "new_release": args.new.upper(),
        "old_release": args.old.upper(),
        **result,
    }, indent=2))


def _run_report(args: argparse.Namespace) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(name)s: %(message)s",
    )

    if not args.catalog.is_file():
        print(f"Error: catalog file not found: {args.catalog}")
        sys.exit(1)
    if not args.mapping.is_file():
        print(f"Error: mapping file not found: {args.mapping}")
        sys.exit(1)

    from fbdi.report import generate_report

    html_path, pdf_path = generate_report(
        catalog_path=args.catalog,
        mapping_path=args.mapping,
        old_release=args.old.upper(),
        new_release=args.new.upper(),
        out_dir=args.out_dir,
    )

    print(f"HTML: {html_path}")
    print(f"PDF : {pdf_path}")


def _run_audit_applaud(args: argparse.Namespace) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(name)s: %(message)s",
    )

    from fbdi.applaud_snapshot import ApplaudSnapshot
    from fbdi.applaud_appmap import load_appmap_workbook
    from fbdi.audit_applaud import run_audit, build_release_changes
    from fbdi.report import load_catalog_release, load_mapping
    from fbdi.config import applaud_snapshot_path

    if not args.catalog.is_file():
        print(f"Error: catalog file not found: {args.catalog}")
        sys.exit(1)
    if not args.mapping.is_file():
        print(f"Error: mapping file not found: {args.mapping}")
        sys.exit(1)
    snap_path = applaud_snapshot_path(args.system)
    if not snap_path.exists():
        print(f"Error: snapshot not found: {snap_path}. Run Step A (agent-driven extraction) first.")
        sys.exit(1)

    # Catalog sheet names are uppercase (26A/26B) and matched exactly by load_catalog_release.
    release = args.release.upper()
    old_release = args.old_release.upper() if args.old_release else None

    snapshot = ApplaudSnapshot.load(snap_path)
    try:
        catalog = load_catalog_release(args.catalog, release)
    except ValueError as exc:
        print(f"Error: {exc}")
        sys.exit(1)
    mapping = load_mapping(args.mapping)
    if args.tables:
        from fbdi.audit_applaud import filter_mapping_to_tables, UnknownTableError
        names = [t.strip() for t in args.tables.split(",") if t.strip()]
        if not names:
            print("Error: --tables must list at least one table name "
                  "(got empty/whitespace-only input).")
            sys.exit(1)
        try:
            mapping = filter_mapping_to_tables(mapping, names)
        except UnknownTableError as exc:
            print(f"Error: {exc}")
            sys.exit(1)
        print(f"Scoped audit to {len({i['applaud_table'] for i in mapping.values()})} "
              f"table(s) via --tables.")
    appmap = load_appmap_workbook(args.appmap) if args.appmap.exists() else {}

    from fbdi.correspondence import load_fieldmap_workbook
    fieldmap = load_fieldmap_workbook(args.fieldmap) if args.fieldmap.exists() else None

    release_changes = {}
    if old_release:
        try:
            old_catalog = load_catalog_release(args.catalog, old_release)
        except ValueError as exc:
            print(f"Error: {exc}")
            sys.exit(1)
        release_changes = build_release_changes(old_catalog, catalog)

    out = args.output or Path(f"Applaud_Compliance_Report_{release}_{args.system}.xlsx")

    findings = run_audit(snapshot, catalog, mapping, appmap, release=release,
                         release_changes=release_changes, out_path=out, old_release=old_release,
                         fieldmap=fieldmap, accept_confidence=args.accept_confidence)
    print(f"Findings: {len(findings)}  (HIGH={sum(1 for f in findings if f.severity=='HIGH')})")
    print(f"Output written to: {out}")


def _run_correspondence_derive(args: argparse.Namespace) -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(name)s: %(message)s")
    from fbdi.applaud_snapshot import ApplaudSnapshot
    from fbdi.report import load_catalog_release, load_mapping
    from fbdi.config import applaud_snapshot_path
    from fbdi.correspondence import (
        assemble_derivation_inputs, derive_correspondences, load_fieldmap_workbook,
        write_review_workbook, ReviewRow,
    )
    from fbdi.audit_applaud import expected_shape, actual_shape

    snap_path = applaud_snapshot_path(args.system)
    if not snap_path.exists():
        print(f"Error: snapshot not found: {snap_path}. Run Step A extraction first.")
        sys.exit(1)
    snapshot = ApplaudSnapshot.load(snap_path)
    release = args.release.upper()
    try:
        catalog = load_catalog_release(args.catalog, release)
    except ValueError as exc:
        print(f"Error: {exc}")
        sys.exit(1)
    mapping = load_mapping(args.mapping)
    if args.tables:
        from fbdi.audit_applaud import filter_mapping_to_tables, UnknownTableError
        names = [t.strip() for t in args.tables.split(",") if t.strip()]
        try:
            mapping = filter_mapping_to_tables(mapping, names)
        except UnknownTableError as exc:
            print(f"Error: {exc}")
            sys.exit(1)

    committed = load_fieldmap_workbook(args.fieldmap) if args.fieldmap.exists() else {}
    decided = {(t, fc.oracle_key) for t, rows in committed.items() for fc in rows}

    inputs = assemble_derivation_inputs(snapshot, catalog, mapping)
    derived = derive_correspondences(inputs, decided)

    # exact_counts for the reviewer's denominator context (audit §6).
    exact_counts: dict[str, tuple[int, int]] = {}
    for table, (_prefix, oracle_by_key, cols) in inputs.items():
        bares = {c.bare.upper() for c in cols}
        exact = sum(1 for k in oracle_by_key if k.upper() in bares)
        exact_counts[table] = (exact, len(oracle_by_key))

    col_by_table_bare = {(t, c.bare): c for t, (_, _, cols) in inputs.items() for c in cols}
    of_by_table_key = {(t, k): of for t, (_, obk, _) in inputs.items()
                       for k, of in obk.items()}

    rows: list[ReviewRow] = []
    for fc in derived:
        of = of_by_table_key.get((fc.applaud_table, fc.oracle_key))
        col = col_by_table_bare.get((fc.applaud_table, fc.applaud_bare))
        rows.append(ReviewRow(
            applaud_table=fc.applaud_table, oracle_key=fc.oracle_key,
            oracle_type=" ".join(str(x) for x in expected_shape(of) if x not in (None, "")) if of else "",
            candidate_bare=fc.applaud_bare, applaud_ddid=fc.applaud_ddid,
            applaud_type=" ".join(str(x) for x in actual_shape(col) if x not in (None, "")) if col else "",
            confidence=fc.confidence, score=fc.score, signals=fc.signals, alternatives=""))

    out = args.output or Path(f"Applaud_FieldMap_Review_{release}_{args.system}.xlsx")
    write_review_workbook(rows, out, exact_counts=exact_counts)
    print(f"Derived {len(rows)} candidate correspondence(s) across "
          f"{len(inputs)} table(s). Review workbook: {out}")


def _run_correspondence_confirm(args: argparse.Namespace) -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(name)s: %(message)s")
    from fbdi.applaud_snapshot import ApplaudSnapshot
    from fbdi.config import applaud_snapshot_path
    from fbdi.correspondence import (
        load_review_workbook, apply_review_decisions, InvalidCorrectedBareError,
        load_fieldmap_workbook, merge_decisions, write_fieldmap_workbook,
    )
    if not args.review.is_file():
        print(f"Error: review workbook not found: {args.review}")
        sys.exit(1)
    snap_path = applaud_snapshot_path(args.system)
    if not snap_path.exists():
        print(f"Error: snapshot not found: {snap_path}.")
        sys.exit(1)
    snapshot = ApplaudSnapshot.load(snap_path)
    valid_bares = {name: {c.bare for c in t.columns}
                   for name, t in snapshot.tables.items()}

    review_rows = load_review_workbook(args.review)
    try:
        decisions = apply_review_decisions(review_rows, valid_bares)
    except InvalidCorrectedBareError as exc:
        print(f"Error: {exc}")
        sys.exit(1)

    committed = load_fieldmap_workbook(args.fieldmap) if args.fieldmap.exists() else {}
    merged = merge_decisions(decisions, committed)   # confirm-time: new decisions win (audit §1.1)
    flat = [fc for rows in merged.values() for fc in rows]
    write_fieldmap_workbook(flat, args.fieldmap)
    n_conf = sum(1 for fc in flat if fc.origin == "confirmed")
    n_rej = sum(1 for fc in flat if fc.origin == "rejected")
    print(f"Merged {len(decisions)} decision(s). Field map now: {n_conf} confirmed, "
          f"{n_rej} rejected -> {args.fieldmap}")
