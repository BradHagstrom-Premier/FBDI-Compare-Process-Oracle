"""Smart clearing for Oracle FBDI template workbooks.

Clears sample data from FBDI templates while preserving all header rows
and metadata. Uses detect_header_row() to dynamically find the header
boundary per sheet, then clears everything below it.

Used by tools/download_and_clear.py to produce blank copies for client
Oracle projects. NOT used by the comparison engine (which reads originals).
"""

import logging

from openpyxl import load_workbook
from openpyxl.styles.fonts import Font

from fbdi.detect_header import detect_header_row

logger = logging.getLogger(__name__)


def clear_workbook(src_path: str, dst_path: str) -> list[str]:
    """Clear sample data from an FBDI workbook, preserving all headers.

    For each data sheet (skipping sheet 1 / instructions):
      1. Run detect_header_row() to find the actual header row
      2. Clear all cell values from header_row + 1 onward
      3. Strip legacy_drawing references (prevents openpyxl save errors)

    If header detection fails for a sheet, that sheet is left untouched.

    Args:
        src_path: Path to the original FBDI .xlsm file.
        dst_path: Path to write the cleared copy.

    Returns:
        List of sheet names where header detection failed (skipped).
    """
    # Oracle FBDI files use font family values >14 which openpyxl rejects.
    Font.family.max = 255

    wb = load_workbook(src_path, keep_vba=True)
    skipped_sheets: list[str] = []

    for sheet in wb.worksheets[1:]:
        header_row = detect_header_row(sheet)
        if header_row is not None:
            for row in sheet.iter_rows(min_row=header_row + 1):
                for cell in row:
                    try:
                        cell.value = None
                    except AttributeError:
                        pass  # MergedCell — read-only, skip
        else:
            logger.warning(
                "Header detection failed for sheet '%s' — skipping (data preserved)",
                sheet.title,
            )
            skipped_sheets.append(sheet.title)

        # Strip malformed VML XML references on ALL sheets to prevent save errors.
        sheet.legacy_drawing = None

    wb.save(dst_path)
    wb.close()
    return skipped_sheets
