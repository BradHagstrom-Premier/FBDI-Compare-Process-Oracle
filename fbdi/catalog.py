"""FBDI Master Catalog — per-release snapshot generator.

Generates FBDI_Master_Catalog.xlsx with:
  - One tab per Oracle release (e.g., 26A, 26B): flat (file, tab, position,
    label, technical, type, length, scale, required) snapshot.
  - Issues tab: consolidated coverage gaps across all releases.
  - Drift tab: position-aligned diff between the two most-recent releases.

Uses subprocess-per-file isolation (mirroring compare.py) with a 120s
timeout to handle openpyxl resource accumulation. Re-running for an
existing release regenerates only that release's tab plus Issues/Drift.
"""

import logging
import multiprocessing
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from fbdi._subprocess_util import run_worker
from fbdi.catalog_normalize import normalize_label
from fbdi.config import CATALOG_TIMEOUT, SKIP_TABS
from fbdi.detect_header import UPPER_SNAKE_PATTERN, detect_header_row
from fbdi.type_parser import parse_data_type

logger = logging.getLogger(__name__)

_MAX_COL = 500

# Column-A keyword -> metadata role. Case-insensitive match after BOM strip.
# Note: header row (tier 1) has col A starting with "Column name of the Table";
# that's already handled by detect_header_row, we don't re-discover it here.
_COL_A_ROLE_KEYWORDS = {
    "name": "label",
    "data type": "type",
    "required or optional": "required",
}


@dataclass
class CatalogRow:
    """One row per (release, file, tab, column position)."""
    release: str
    file_name: str
    tab_name: str
    position: int               # 1-based column index
    column_label: str           # normalized user-friendly label
    column_technical: str       # UPPER_SNAKE_CASE; blank for thin tabs
    data_type: str              # uppercase; blank for thin tabs or parse failures
    length: int | None          # None when absent; blank in output
    scale: int | None           # None when absent; blank in output
    data_type_raw: str          # original string; blank for thin tabs
    required: bool | None       # True/False; None when unknown


@dataclass
class IssueRow:
    """One row per coverage gap or error condition."""
    release: str
    file: str
    tab: str                    # empty for file-level issues
    issue_type: str             # FILE_ERROR | TIMEOUT | SUBPROCESS_FAILED | NO_HEADER | TYPE_PARSE_WARNING
    detail: str


@dataclass
class DriftRow:
    """One row per classified change between two releases for one (file, tab).

    Schema is alignment-driven, not naive per-position. position columns
    are split into old/new because SHIFTED rows have a different position
    on each side.
    """
    file: str
    tab: str
    change_type: str            # ADDED | REMOVED | MODIFIED | RENAMED | SHIFTED | MULTI
    old_position: int | None
    new_position: int | None
    col_label_old: str
    col_label_new: str
    col_technical_old: str
    col_technical_new: str
    data_type_old: str
    data_type_new: str
    length_old: str
    length_new: str
    required_old: str
    required_new: str
    sub_kinds: str              # comma-joined ("type", "length", "required") for MODIFIED/MULTI; empty otherwise


def _read_row_values(ws: Worksheet, row_idx: int) -> list[str | None]:
    """Read one row into a list of trimmed string values (or None), trailing blanks trimmed."""
    row_cells = next(
        iter(ws.iter_rows(
            min_row=row_idx, max_row=row_idx,
            max_col=min(ws.max_column or 1, _MAX_COL),
        )),
        None,
    )
    if row_cells is None:
        return []
    raw = []
    for cell in row_cells:
        if isinstance(cell, MergedCell):
            raw.append(None)
        elif cell.value is not None and str(cell.value).strip() != "":
            raw.append(str(cell.value).strip())
        else:
            raw.append(None)
    # Trim trailing None
    last = 0
    for i, v in enumerate(raw, start=1):
        if v is not None:
            last = i
    return raw[:last]


def _is_tier1_header(values: list[str | None]) -> bool:
    """True if the row is dominated by UPPER_SNAKE_CASE technical names.

    Oracle Tier-1 (technical-names) rows never have asterisk-prefixed cells.
    Asterisks are a label-row convention marking required fields.  If any
    cell starts with '*', the row is a label/thin-tab row, not tier-1.
    """
    non_empty = [v for v in values if v]
    if not non_empty:
        return False
    if any(isinstance(v, str) and v.strip().startswith("*") for v in non_empty):
        return False
    snake = sum(
        1 for v in non_empty
        if isinstance(v, str) and UPPER_SNAKE_PATTERN.match(v.strip())
    )
    return (snake / len(non_empty)) >= 0.5


def extract_tab_rows(
    ws: Worksheet,
    file_stem: str,
    release: str,
) -> tuple[list[CatalogRow], list[IssueRow]]:
    """Extract catalog rows + any issues for one worksheet.

    Returns (rows, issues). On NO_HEADER, returns ([], [IssueRow]).
    Dispatches to thin-tab or rich-tab extraction based on the detected
    header row's content.
    """
    header_row = detect_header_row(ws)
    if header_row is None:
        return [], [IssueRow(
            release=release,
            file=file_stem,
            tab=ws.title,
            issue_type="NO_HEADER",
            detail=f"no confident header row in '{ws.title}'",
        )]

    header_values = _read_row_values(ws, header_row)
    if _is_tier1_header(header_values):
        return _extract_rich(ws, file_stem, release, header_row, header_values)
    return _extract_thin(ws, file_stem, release, header_row, header_values)


def _extract_thin(
    ws: Worksheet,
    file_stem: str,
    release: str,
    header_row: int,
    header_values: list[str | None],
) -> tuple[list[CatalogRow], list[IssueRow]]:
    """Thin-tab extraction: header row is a list of user-friendly labels
    (possibly asterisk-prefixed for required). No type/length/technical info."""
    rows: list[CatalogRow] = []
    for idx, raw in enumerate(header_values, start=1):
        if raw is None:
            continue
        raw_str = str(raw)
        required = raw_str.lstrip().startswith("*")
        rows.append(CatalogRow(
            release=release,
            file_name=file_stem,
            tab_name=ws.title,
            position=idx,
            column_label=normalize_label(raw_str),
            column_technical="",
            data_type="",
            length=None,
            scale=None,
            data_type_raw="",
            required=required,
        ))
    return rows, []


def _find_metadata_rows(
    ws: Worksheet, header_row: int
) -> dict[str, int]:
    """Scan col A of rows 1..header_row-1. Return {role: row_idx} for matched rows.

    Match is case-insensitive with BOM prefix stripped. Unmatched rows are
    silently ignored (e.g., 'Description', 'Reserved for Future Use').
    """
    found: dict[str, int] = {}
    for r in range(1, header_row):
        cell = ws.cell(row=r, column=1).value
        if cell is None:
            continue
        key = str(cell).lstrip("\ufeff").strip().lower()
        role = _COL_A_ROLE_KEYWORDS.get(key)
        if role and role not in found:
            found[role] = r
    return found


def _extract_rich(
    ws: Worksheet,
    file_stem: str,
    release: str,
    header_row: int,
    header_values: list[str | None],
) -> tuple[list[CatalogRow], list[IssueRow]]:
    """Rich-tab extraction. header_values is the tier-1 (technical) row.

    Uses col-A keyword matching to locate the label/type/required rows
    above header_row. Missing role rows leave the corresponding field
    blank; other fields still populate. Unparseable data types emit
    TYPE_PARSE_WARNING issues but the row still emits (raw preserved).
    """
    roles = _find_metadata_rows(ws, header_row)

    label_values = _read_row_values(ws, roles["label"]) if "label" in roles else []
    type_values = _read_row_values(ws, roles["type"]) if "type" in roles else []
    required_values = _read_row_values(ws, roles["required"]) if "required" in roles else []

    rows: list[CatalogRow] = []
    issues: list[IssueRow] = []

    def _val_at(values: list[str | None], sheet_col: int) -> str:
        """Return the cell value at sheet column `sheet_col` (1-based) or ''."""
        idx = sheet_col - 1
        if 0 <= idx < len(values):
            v = values[idx]
            return str(v) if v is not None else ""
        return ""

    # Most Oracle FBDI rich tabs have a row-label sentinel in col A of the
    # technical-names row (e.g. "Column name of the Table"), so data columns
    # start at col B.  A few tabs (e.g. CST_I_INCOMING_TXN_COSTS) have no
    # sentinel and begin data at col A.  Detect by checking whether col A
    # itself is UPPER_SNAKE_CASE.
    col_a = header_values[0] if header_values else None
    col_a_is_data = bool(
        col_a and UPPER_SNAKE_PATTERN.match(str(col_a).strip())
    )
    start_col = 1 if col_a_is_data else 2

    for sheet_col in range(start_col, len(header_values) + 1):
        tech_raw = header_values[sheet_col - 1]
        if not tech_raw:
            continue
        technical = str(tech_raw).strip()
        label_raw = _val_at(label_values, sheet_col)
        type_raw = _val_at(type_values, sheet_col)
        req_raw = _val_at(required_values, sheet_col)

        parsed = parse_data_type(type_raw)
        if parsed.parse_warning:
            issues.append(IssueRow(
                release=release,
                file=file_stem,
                tab=ws.title,
                issue_type="TYPE_PARSE_WARNING",
                detail=type_raw,
            ))

        rows.append(CatalogRow(
            release=release,
            file_name=file_stem,
            tab_name=ws.title,
            position=sheet_col - start_col + 1,  # 1-based data-column index
            column_label=normalize_label(label_raw),
            column_technical=technical,
            data_type=parsed.data_type,
            length=parsed.length,
            scale=parsed.scale,
            data_type_raw=type_raw,
            required=_parse_required_flag(req_raw),
        ))

    return rows, issues


def _parse_required_flag(raw: str) -> bool | None:
    """Parse 'Required'/'Optional' (case-insensitive) to bool. Unknown -> None."""
    if not raw:
        return None
    v = raw.strip().lower()
    if v.startswith("required"):
        return True
    if v.startswith("optional"):
        return False
    return None


def extract_file(
    path: Path, release: str
) -> tuple[list[CatalogRow], list[IssueRow]]:
    """Open one .xlsm and extract catalog rows for every non-skipped tab.

    Returns (rows, issues). FILE_ERROR on load failure produces one
    IssueRow with tab="". Each data tab that extract_tab_rows flags with
    issues contributes its issues to the combined list.
    """
    file_stem = path.stem
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [], [IssueRow(
            release=release,
            file=file_stem,
            tab="",
            issue_type="FILE_ERROR",
            detail=f"{type(e).__name__}: {str(e)[:200]}",
        )]

    all_rows: list[CatalogRow] = []
    all_issues: list[IssueRow] = []
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_TABS:
                continue
            ws = wb[sheet_name]
            rows, issues = extract_tab_rows(ws, file_stem=file_stem, release=release)
            all_rows.extend(rows)
            all_issues.extend(issues)
    finally:
        wb.close()
    return all_rows, all_issues


def _rows_to_tuples(rows: list[CatalogRow]) -> list[tuple]:
    return [(
        r.release, r.file_name, r.tab_name, r.position,
        r.column_label, r.column_technical,
        r.data_type, r.length, r.scale, r.data_type_raw,
        r.required,
    ) for r in rows]


def _issues_to_tuples(issues: list[IssueRow]) -> list[tuple]:
    return [(i.release, i.file, i.tab, i.issue_type, i.detail) for i in issues]


def _tuples_to_rows(tuples: list[tuple]) -> list[CatalogRow]:
    return [CatalogRow(*t) for t in tuples]


def _tuples_to_issues(tuples: list[tuple]) -> list[IssueRow]:
    return [IssueRow(*t) for t in tuples]


def _catalog_worker(path_str: str, release: str, queue: multiprocessing.Queue) -> None:
    """Subprocess entry point. Mirrors _compare_worker for resource isolation."""
    # Some templates hit openpyxl's Font.family.max=14 cap; 255 matches compare.
    from openpyxl.styles.fonts import Font as WorkerFont
    WorkerFont.family.max = 255

    try:
        rows, issues = extract_file(Path(path_str), release=release)
        queue.put((_rows_to_tuples(rows), _issues_to_tuples(issues)))
    except Exception as e:
        queue.put(f"ERROR: {type(e).__name__}: {e}")


def _run_file_in_subprocess(
    path: Path, release: str, timeout: int = CATALOG_TIMEOUT
) -> tuple[list[CatalogRow], list[IssueRow]]:
    """Run extract_file in a fresh subprocess with timeout.

    Translates WorkerOutcome into catalog-specific IssueRows on failure
    paths (TIMEOUT, SUBPROCESS_FAILED). Happy path unpacks the
    (row_tuples, issue_tuples) payload.
    """
    outcome = run_worker(_catalog_worker, args=(str(path), release), timeout=timeout)
    file_stem = path.stem

    if outcome.status == "timeout":
        return [], [IssueRow(
            release=release, file=file_stem, tab="",
            issue_type="TIMEOUT", detail=f"exceeded {timeout}s",
        )]

    if outcome.status == "crashed":
        return [], [IssueRow(
            release=release, file=file_stem, tab="",
            issue_type="SUBPROCESS_FAILED",
            detail=f"exit code {outcome.exitcode}",
        )]

    result = outcome.payload
    if isinstance(result, str) and result.startswith("ERROR:"):
        return [], [IssueRow(
            release=release, file=file_stem, tab="",
            issue_type="SUBPROCESS_FAILED",
            detail=result,
        )]

    row_tuples, issue_tuples = result
    return _tuples_to_rows(row_tuples), _tuples_to_issues(issue_tuples)


def _compute_drift(
    old_rows: list[CatalogRow],
    new_rows: list[CatalogRow],
    release_old: str,
    release_new: str,
) -> list[DriftRow]:
    """Alignment-driven diff between two release row sets.

    Groups rows by (file, tab); for each pair, calls fbdi.align.align_tabs
    to derive the classified Change list; emits one DriftRow per Change.
    Replaces the prior naive per-position diff that misclassified shift
    cascades as RENAMED/MULTI.
    """
    from fbdi.align import AlignedField, align_tabs

    def _to_aligned(r: CatalogRow) -> AlignedField:
        return AlignedField(
            position=r.position,
            label=r.column_label,
            technical=r.column_technical or None,
            data_type=r.data_type or None,
            length=r.length,
            required=r.required,
        )

    def _group(rows: list[CatalogRow]) -> dict[tuple[str, str], list[CatalogRow]]:
        out: dict[tuple[str, str], list[CatalogRow]] = {}
        for r in rows:
            out.setdefault((r.file_name, r.tab_name), []).append(r)
        for v in out.values():
            v.sort(key=lambda r: r.position)
        return out

    old_grouped = _group(old_rows)
    new_grouped = _group(new_rows)
    all_keys = sorted(set(old_grouped.keys()) | set(new_grouped.keys()))

    drift: list[DriftRow] = []
    for file, tab in all_keys:
        old_aligned = [_to_aligned(r) for r in old_grouped.get((file, tab), [])]
        new_aligned = [_to_aligned(r) for r in new_grouped.get((file, tab), [])]
        for change in align_tabs(old_aligned, new_aligned):
            drift.append(_drift_row_from_change(file, tab, change))
    return drift


def _drift_row_from_change(file: str, tab: str, change) -> DriftRow:
    """Build a DriftRow from a fbdi.align.Change."""
    old = change.old_field
    new = change.new_field
    return DriftRow(
        file=file,
        tab=tab,
        change_type=change.change_type,
        old_position=change.old_position,
        new_position=change.new_position,
        col_label_old=(old.label or "") if old else "",
        col_label_new=(new.label or "") if new else "",
        col_technical_old=(old.technical or "") if old else "",
        col_technical_new=(new.technical or "") if new else "",
        data_type_old=_fmt_type_align(old),
        data_type_new=_fmt_type_align(new),
        length_old=_fmt_length_align(old),
        length_new=_fmt_length_align(new),
        required_old=_fmt_required_align(old),
        required_new=_fmt_required_align(new),
        sub_kinds=",".join(change.sub_kinds),
    )


def _fmt_type_align(f) -> str:
    return (f.data_type or "") if f else ""


def _fmt_length_align(f) -> str:
    if f is None or f.length is None:
        return ""
    return str(f.length)


def _fmt_required_align(f) -> str:
    if f is None or f.required is None:
        return ""
    return "TRUE" if f.required else "FALSE"


_RELEASE_TAB_HEADERS = [
    "release", "file_name", "tab_name", "position",
    "column_label", "column_technical",
    "data_type", "length", "scale", "data_type_raw",
    "required",
]

_ISSUES_TAB_HEADERS = ["release", "file", "tab", "issue_type", "detail"]


def _drift_tab_headers(release_old: str | None, release_new: str | None) -> list[str]:
    """Build Drift tab headers with release names substituted."""
    old = release_old or "OLD"
    new = release_new or "NEW"
    return [
        "file", "tab", "change_type",
        f"position_{old}", f"position_{new}",
        f"col_label_{old}", f"col_label_{new}",
        f"col_technical_{old}", f"col_technical_{new}",
        f"data_type_{old}", f"data_type_{new}",
        f"length_{old}", f"length_{new}",
        f"required_{old}", f"required_{new}",
        "sub_kinds",
    ]


def _write_master_workbook(
    output_path: Path,
    rows_by_release: dict[str, list[CatalogRow]],
    issues: list[IssueRow],
    drift: list[DriftRow],
    release_old: str | None,
    release_new: str | None,
) -> None:
    """Write the master workbook. Release tabs, Issues, Drift.

    The caller is responsible for providing *all* release data (merged
    from any existing workbook + fresh run). This function writes
    atomically via .tmp + rename.
    """
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    bold = Font(name="Calibri", size=11, bold=True)
    plain = Font(name="Calibri", size=11)

    # Release tabs, in lexicographic order (26A < 26B < 26C < 27A)
    for release in sorted(rows_by_release.keys()):
        rows = rows_by_release[release]
        ws = wb.create_sheet(title=release)
        for col_idx, h in enumerate(_RELEASE_TAB_HEADERS, start=1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = bold
        for row_idx, r in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=r.release).font = plain
            ws.cell(row=row_idx, column=2, value=r.file_name).font = plain
            ws.cell(row=row_idx, column=3, value=r.tab_name).font = plain
            ws.cell(row=row_idx, column=4, value=r.position).font = plain
            ws.cell(row=row_idx, column=5, value=r.column_label).font = plain
            ws.cell(row=row_idx, column=6, value=r.column_technical).font = plain
            ws.cell(row=row_idx, column=7, value=r.data_type).font = plain
            ws.cell(row=row_idx, column=8, value=r.length).font = plain
            ws.cell(row=row_idx, column=9, value=r.scale).font = plain
            ws.cell(row=row_idx, column=10, value=r.data_type_raw).font = plain
            ws.cell(
                row=row_idx, column=11,
                value="" if r.required is None else ("TRUE" if r.required else "FALSE"),
            ).font = plain
        ws.auto_filter.ref = f"A1:K{max(len(rows) + 1, 1)}"

    # Issues tab
    ws = wb.create_sheet(title="Issues")
    for col_idx, h in enumerate(_ISSUES_TAB_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = bold
    for row_idx, i in enumerate(issues, start=2):
        ws.cell(row=row_idx, column=1, value=i.release).font = plain
        ws.cell(row=row_idx, column=2, value=i.file).font = plain
        ws.cell(row=row_idx, column=3, value=i.tab).font = plain
        ws.cell(row=row_idx, column=4, value=i.issue_type).font = plain
        ws.cell(row=row_idx, column=5, value=i.detail).font = plain
    ws.auto_filter.ref = f"A1:E{max(len(issues) + 1, 1)}"

    # Drift tab
    ws = wb.create_sheet(title="Drift")
    drift_headers = _drift_tab_headers(release_old, release_new)
    for col_idx, h in enumerate(drift_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = bold
    for row_idx, d in enumerate(drift, start=2):
        values = [
            d.file, d.tab, d.change_type,
            d.old_position if d.old_position is not None else "",
            d.new_position if d.new_position is not None else "",
            d.col_label_old, d.col_label_new,
            d.col_technical_old, d.col_technical_new,
            d.data_type_old, d.data_type_new,
            d.length_old, d.length_new,
            d.required_old, d.required_new,
            d.sub_kinds,
        ]
        for col_idx, v in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v).font = plain
    ws.auto_filter.ref = f"A1:P{max(len(drift) + 1, 1)}"

    # Atomic save
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    wb.save(tmp_path)
    wb.close()
    tmp_path.replace(output_path)


def _load_existing_release_rows(
    master_path: Path, current_release: str
) -> dict[str, list[CatalogRow]]:
    """Read rows for every release tab in master_path except `current_release`."""
    if not master_path.exists():
        return {}
    try:
        wb = load_workbook(master_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(
            "Could not load existing master at %s: %s — starting fresh",
            master_path, e,
        )
        return {}
    result: dict[str, list[CatalogRow]] = {}
    try:
        for sn in wb.sheetnames:
            if sn in {"Issues", "Drift"} or sn == current_release:
                continue
            rows = _read_release_tab_rows(wb[sn])
            if rows:
                result[sn] = rows
    finally:
        wb.close()
    return result


def _read_release_tab_rows(ws) -> list[CatalogRow]:
    """Reconstruct CatalogRows from an existing release tab in the master workbook."""
    rows: list[CatalogRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        padded = list(row) + [None] * (11 - len(row))
        (release_v, file_v, tab_v, pos_v, label_v, tech_v,
         dtype_v, length_v, scale_v, raw_v, required_v) = padded[:11]
        required: bool | None
        if required_v in (None, ""):
            required = None
        elif str(required_v).upper() == "TRUE":
            required = True
        else:
            required = False
        rows.append(CatalogRow(
            release=str(release_v) if release_v else "",
            file_name=str(file_v) if file_v else "",
            tab_name=str(tab_v) if tab_v else "",
            position=int(pos_v) if pos_v is not None else 0,
            column_label=str(label_v) if label_v else "",
            column_technical=str(tech_v) if tech_v else "",
            data_type=str(dtype_v) if dtype_v else "",
            length=int(length_v) if isinstance(length_v, (int, float)) else None,
            scale=int(scale_v) if isinstance(scale_v, (int, float)) else None,
            data_type_raw=str(raw_v) if raw_v else "",
            required=required,
        ))
    return rows


def _load_existing_issues_excluding(
    master_path: Path, current_release: str
) -> list[IssueRow]:
    """Read Issues tab, excluding rows for current_release."""
    if not master_path.exists():
        return []
    try:
        wb = load_workbook(master_path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        if "Issues" not in wb.sheetnames:
            return []
        ws = wb["Issues"]
        out: list[IssueRow] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            padded = list(row) + [None] * (5 - len(row))
            release_v, file_v, tab_v, itype_v, detail_v = padded[:5]
            if release_v == current_release:
                continue
            out.append(IssueRow(
                release=str(release_v) if release_v else "",
                file=str(file_v) if file_v else "",
                tab=str(tab_v) if tab_v else "",
                issue_type=str(itype_v) if itype_v else "",
                detail=str(detail_v) if detail_v else "",
            ))
        return out
    finally:
        wb.close()


def generate_catalog(
    release: str,
    baselines_dir: Path,
    master_path: Path,
    timeout: int = CATALOG_TIMEOUT,
) -> None:
    """Generate / update the master catalog for one release."""
    baselines_dir = Path(baselines_dir)
    master_path = Path(master_path)
    new_rows: list[CatalogRow] = []
    new_issues: list[IssueRow] = []
    xlsm_files = sorted(baselines_dir.glob("*.xlsm"))
    for i, path in enumerate(xlsm_files, 1):
        logger.info("[%d/%d] Cataloging: %s", i, len(xlsm_files), path.stem)
        rows, issues = _run_file_in_subprocess(path, release=release, timeout=timeout)
        new_rows.extend(rows)
        new_issues.extend(issues)
    rows_by_release = _load_existing_release_rows(master_path, current_release=release)
    rows_by_release[release] = new_rows
    preserved_issues = _load_existing_issues_excluding(master_path, current_release=release)
    all_issues = preserved_issues + new_issues
    sorted_releases = sorted(rows_by_release.keys())
    if len(sorted_releases) >= 2:
        release_old = sorted_releases[-2]
        release_new = sorted_releases[-1]
        drift = _compute_drift(
            rows_by_release[release_old],
            rows_by_release[release_new],
            release_old=release_old,
            release_new=release_new,
        )
    else:
        release_old = None
        release_new = sorted_releases[0] if sorted_releases else None
        drift = []
    _write_master_workbook(
        master_path,
        rows_by_release=rows_by_release,
        issues=all_issues,
        drift=drift,
        release_old=release_old,
        release_new=release_new,
    )
    logger.info(
        "Catalog written: %s (%d releases, %d rows, %d issues, %d drift)",
        master_path, len(rows_by_release),
        sum(len(v) for v in rows_by_release.values()),
        len(all_issues), len(drift),
    )
