"""Tests for fbdi.populate_module — surgical column-F updater for the mapping xlsx."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from fbdi.populate_module import populate_module_column


# Mirrors the actual columns of FBDI_to_ApplaudTables_Mapping.xlsx
# 'FBDI Mapping' sheet: A FBDI Template, B FBDI Tab, C Applaud Table,
# D Prefix, E Status, F Module, G In Base System?
HEADER_ROW = ["FBDI Template", "FBDI Tab", "Applaud Table",
              "Prefix", "Status", "Module", "In Base System?"]


def _make_mapping_workbook(path: Path, rows: list[list]):
    """Build a synthetic mapping xlsx with the production sheet structure."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FBDI Mapping"
    for col_idx, val in enumerate(HEADER_ROW, start=1):
        ws.cell(row=1, column=col_idx, value=val)
    for r_idx, row_vals in enumerate(rows, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # Add the second sheet that exists in production so the updater
    # has to find the right one by name.
    wb.create_sheet("Applaud Tables Reference")
    wb.save(path)
    wb.close()


def _read_module_col(path: Path) -> list:
    wb = load_workbook(path, read_only=True)
    ws = wb["FBDI Mapping"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        out.append(row[5])  # column F = index 5
    wb.close()
    return out


class TestPopulateModuleColumn:
    def test_populated_from_new_release(self, tmp_path):
        """Happy path: every row's FBDI Template appears in NEW; all populated."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL", "T_RA", "TA4", "", "", ""],
            ["ItemImportTemplate", "EGP_ITEMS_INTERFACE", "T_EGP", "T91", "", "", ""],
        ])
        new_modules = {
            "AutoInvoiceImportTemplate.xlsm": "Financials",
            "ItemImportTemplate.xlsm": "Supply Chain & Manufacturing",
        }
        result = populate_module_column(mapping, new_modules, old_modules={})

        assert result == {"populated": 2, "blank": 0, "overwritten": 0}
        assert _read_module_col(mapping) == [
            "Financials", "Supply Chain & Manufacturing",
        ]

    def test_falls_back_to_old_when_missing_from_new(self, tmp_path):
        """File only in OLD release: OLD module is used."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["LegacyTemplate", "LEGACY_TAB", "", "", "", "", ""],
        ])
        result = populate_module_column(
            mapping,
            new_modules={},
            old_modules={"LegacyTemplate.xlsm": "Procurement"},
        )
        assert result == {"populated": 1, "blank": 0, "overwritten": 0}
        assert _read_module_col(mapping) == ["Procurement"]

    def test_new_wins_when_present_in_both(self, tmp_path):
        """When file is in BOTH releases, NEW takes precedence."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["DualTemplate", "DUAL_TAB", "", "", "", "", ""],
        ])
        result = populate_module_column(
            mapping,
            new_modules={"DualTemplate.xlsm": "Financials"},
            old_modules={"DualTemplate.xlsm": "Procurement"},
        )
        assert result == {"populated": 1, "blank": 0, "overwritten": 0}
        assert _read_module_col(mapping) == ["Financials"]

    def test_blank_when_in_neither(self, tmp_path):
        """File in neither release: Module stays blank."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["GhostTemplate", "GHOST_TAB", "", "", "", "", ""],
        ])
        result = populate_module_column(mapping, new_modules={}, old_modules={})
        assert result == {"populated": 0, "blank": 1, "overwritten": 0}
        assert _read_module_col(mapping) == [None]

    def test_other_columns_preserved(self, tmp_path):
        """Manually-edited columns (A, B, C, D, E, G) survive the update."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL",
             "T_RA_INTERFACE_LINES_ALL", "TA4", "Mapped", "", "Yes"],
        ])
        populate_module_column(
            mapping,
            new_modules={"AutoInvoiceImportTemplate.xlsm": "Financials"},
            old_modules={},
        )
        wb = load_workbook(mapping, read_only=True)
        ws = wb["FBDI Mapping"]
        row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        wb.close()
        assert row == (
            "AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL",
            "T_RA_INTERFACE_LINES_ALL", "TA4", "Mapped", "Financials", "Yes",
        )

    def test_idempotency(self, tmp_path):
        """Running twice produces identical output and bumps overwritten count."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["AutoInvoiceImportTemplate", "RA_TAB", "", "", "", "", ""],
        ])
        new = {"AutoInvoiceImportTemplate.xlsm": "Financials"}
        first = populate_module_column(mapping, new_modules=new, old_modules={})
        second = populate_module_column(mapping, new_modules=new, old_modules={})
        assert first == {"populated": 1, "blank": 0, "overwritten": 0}
        assert second == {"populated": 1, "blank": 0, "overwritten": 1}
        assert _read_module_col(mapping) == ["Financials"]

    def test_xlsm_suffix_normalized(self, tmp_path):
        """JSON keys may have .xlsm; spreadsheet col A may not. Both should match."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["AutoInvoiceImportTemplate", "RA_TAB", "", "", "", "", ""],   # no suffix
            ["ItemImportTemplate.xlsm", "EGP_TAB", "", "", "", "", ""],     # with suffix
        ])
        new = {
            "AutoInvoiceImportTemplate.xlsm": "Financials",
            "ItemImportTemplate.xlsm": "Supply Chain & Manufacturing",
        }
        result = populate_module_column(mapping, new_modules=new, old_modules={})
        assert result == {"populated": 2, "blank": 0, "overwritten": 0}
        assert _read_module_col(mapping) == [
            "Financials", "Supply Chain & Manufacturing",
        ]

    def test_blank_template_row_skipped(self, tmp_path):
        """Rows with empty FBDI Template (col A) are skipped, not counted."""
        mapping = tmp_path / "mapping.xlsx"
        _make_mapping_workbook(mapping, [
            ["AutoInvoiceImportTemplate", "RA_TAB", "", "", "", "", ""],
            [None, None, None, None, None, None, None],  # blank row in middle
            ["ItemImportTemplate", "EGP_TAB", "", "", "", "", ""],
        ])
        new = {
            "AutoInvoiceImportTemplate.xlsm": "Financials",
            "ItemImportTemplate.xlsm": "Supply Chain & Manufacturing",
        }
        result = populate_module_column(mapping, new_modules=new, old_modules={})
        # 2 rows have a non-blank template; both populated. Blank row not counted.
        assert result == {"populated": 2, "blank": 0, "overwritten": 0}
