from openpyxl import load_workbook

from fbdi.align import AlignedField, Change
from fbdi.applaud_snapshot import (
    DataColumn, FileField, SnapshotTable, ApplaudSnapshot, build_table,
)
from fbdi.applaud_appmap import AppMapRow
from fbdi.audit_applaud import (
    Finding, make_finding_id, oracle_match_key, expected_shape, actual_shape,
    check_sizing, check_file_coverage, check_table_coverage, check_orphans,
    build_release_changes, check_release_delta, check_unmapped, coverage_gaps,
    write_findings_workbook, run_audit,
)


# --- Task 6: Finding model + finding_id ------------------------------------

def test_finding_id_is_stable_and_attribute_sensitive():
    base = dict(dimension="1-SIZING", applaud_object_type="TABLE",
                applaud_object_name="T_BANKS_BRANCHES", applaud_field="T32BANK_NAME")
    id_size = make_finding_id(attribute="SIZE", **base)
    id_size_again = make_finding_id(attribute="SIZE", **base)
    id_scale = make_finding_id(attribute="SCALE", **base)
    assert id_size == id_size_again
    assert id_size != id_scale
    assert len(id_size) == 12


def test_finding_defaults_status_and_notes_blank():
    f = Finding(finding_id="abc", dimension="1-SIZING", severity="HIGH",
                fbdi_template="t", fbdi_tab="tab", oracle_field="BANK_NAME",
                oracle_type="VARCHAR2(100)", applaud_object_type="TABLE",
                applaud_object_name="T_BANKS_BRANCHES", applaud_field="T32BANK_NAME",
                attribute="SIZE", current_value="char 30", expected_value="char 100",
                message="Undersized")
    assert f.status == "" and f.notes == ""


# --- Task 7: Dim 1 sizing + oracle_match_key -------------------------------

def test_oracle_match_key_normalizes_label_when_technical_missing():
    thin = AlignedField(2, "Bank Name", None, None, None, None, None)
    assert oracle_match_key(thin) == "BANK_NAME"
    rich = AlignedField(5, "Bank Name", "BANK_NAME", "VARCHAR2", 60, None, True)
    assert oracle_match_key(rich) == "BANK_NAME"


def test_shapes_char_and_numeric():
    of = AlignedField(position=1, label="Bank Name", technical="BANK_NAME",
                      data_type="VARCHAR2", length=100, scale=None, required=True)
    assert expected_shape(of) == ("char", 100, None)
    col = DataColumn(ddid="T32BANK_NAME", bare="BANK_NAME", data_type="X",
                     size=30, dec_places=None, odbc_name="BANK_NAME", row=1)
    assert actual_shape(col) == ("char", 30, None)


def test_actual_shape_reflects_datadictionary_not_blank_databasedetail():
    raw_cols = [{"Row": 1, "DDID": "T32BANK_NAME", "DataType": "", "Size": 0,
                 "DecPlaces": 0, "ODBCName": ""}]
    dd = {"T32BANK_NAME": {"DataType": "X", "Size": 100, "DecPlaces": 0}}
    table = build_table("T_BANKS_BRANCHES", "T32", False, "T_BANKS_BRANCHES (T32)",
                        [["T32COUNTRY"]], raw_cols, dd_by_ddid=dd)
    assert actual_shape(table.columns[0]) == ("char", 100, None)


def test_check_sizing_flags_undersized():
    of = AlignedField(1, "Bank Name", "BANK_NAME", "VARCHAR2", 100, None, True)
    col = DataColumn("T32BANK_NAME", "BANK_NAME", "X", 30, None, "BANK_NAME", 1)
    findings = check_sizing("Tmpl", "Bank Account", "T_BANKS_BRANCHES",
                            {"BANK_NAME": of}, [col])
    assert len(findings) == 1
    f = findings[0]
    assert f.attribute == "SIZE" and f.severity == "HIGH"
    assert f.current_value == "char 30" and f.expected_value == "char 100"


def test_check_sizing_flags_type_class_mismatch():
    of = AlignedField(1, "Amount", "AMOUNT", "NUMBER", 18, 4, False)
    col = DataColumn("T32AMOUNT", "AMOUNT", "X", 50, None, "AMOUNT", 1)
    findings = check_sizing("Tmpl", "Tab", "T_X", {"AMOUNT": of}, [col])
    assert findings[0].attribute == "TYPE_CLASS" and findings[0].severity == "HIGH"


def test_check_sizing_oversize_is_info_not_high():
    of = AlignedField(1, "Code", "CODE", "VARCHAR2", 10, None, False)
    col = DataColumn("T32CODE", "CODE", "X", 50, None, "CODE", 1)
    findings = check_sizing("Tmpl", "Tab", "T_X", {"CODE": of}, [col])
    assert findings == [] or all(f.severity == "INFO" for f in findings)


def test_check_sizing_date_stored_as_char_is_type_class_finding():
    of = AlignedField(1, "Effective Date", "EFFECTIVE_DATE", "DATE", None, None, False)
    col = DataColumn("T32EFFECTIVE_DATE", "EFFECTIVE_DATE", "X", 30, None, "", 1)
    findings = check_sizing("Tmpl", "Tab", "T_X", {"EFFECTIVE_DATE": of}, [col])
    assert len(findings) == 1 and findings[0].attribute == "TYPE_CLASS"
    assert findings[0].severity == "HIGH"


# --- Task 8/9: Dim 2/3 coverage + ordering ---------------------------------

def _of(pos, tech):
    return AlignedField(pos, tech, tech, "VARCHAR2", 50, None, False)


def test_check_if_flags_missing_extra_and_order():
    oracle = [_of(1, "COUNTRY"), _of(2, "BANK_NAME"), _of(3, "BANK_CODE")]
    if_fields = [
        FileField(1, "T32BANK_NAME", "BANK_NAME", "X(100)", "C", None),  # out of order
        FileField(2, "T32COUNTRY", "COUNTRY", "X(60)", "C", None),
        FileField(3, "T32EXTRA", "EXTRA", "X(10)", "C", None),           # extra
        # BANK_CODE missing
    ]
    findings = check_file_coverage("Tmpl", "Bank Account", "I_T_BANKS_BRANCHES",
                                   "IMPORT", "2-IF", oracle, if_fields)
    assert any(f.attribute == "PRESENCE" and f.oracle_field == "BANK_CODE"
               and f.severity == "HIGH" for f in findings)
    assert any(f.attribute == "PRESENCE" and "EXTRA" in f.applaud_field
               and f.severity == "INFO" for f in findings)
    assert any(f.attribute == "ORDER" and f.severity == "MED" for f in findings)


def test_check_ef_uses_bare_ddid_not_empty_column_header():
    oracle = [_of(1, "COUNTRY"), _of(2, "BANK_NAME")]
    ef_fields = [
        FileField(1, "T32COUNTRY", "COUNTRY", "X(60)", None, ""),     # ColumnHeader empty
        FileField(2, "T32BANK_NAME", "BANK_NAME", "X(100)", None, ""),
    ]
    findings = check_file_coverage("Tmpl", "Bank Account", "T_BANKS_BRANCHES",
                                   "EXPORT", "3-EF", oracle, ef_fields)
    assert findings == []


# --- Task 10: Dim 4 target-table coverage ----------------------------------

def test_check_table_coverage_flags_missing_column():
    oracle = [_of(1, "COUNTRY"), _of(2, "BANK_NAME")]
    cols = [DataColumn("T32COUNTRY", "COUNTRY", "X", 60, None, "COUNTRY", 1)]
    findings = check_table_coverage("Tmpl", "Bank Account", "T_BANKS_BRANCHES",
                                    oracle, cols)
    assert len(findings) == 1
    assert findings[0].oracle_field == "BANK_NAME"
    assert findings[0].attribute == "PRESENCE" and findings[0].severity == "HIGH"


def test_check_table_coverage_matches_on_odbcname():
    oracle = [_of(1, "BANK_NAME")]
    cols = [DataColumn("T32BNK", "BNK", "X", 60, None, "BANK_NAME", 1)]  # bare differs, ODBC matches
    findings = check_table_coverage("Tmpl", "Tab", "T_X", oracle, cols)
    assert findings == []


# --- Task 11: Dim 5 orphans ------------------------------------------------

def test_check_orphans_flags_if_field_absent_from_table():
    table_cols = [DataColumn("T32COUNTRY", "COUNTRY", "X", 60, None, "COUNTRY", 1)]
    if_fields = [
        FileField(1, "T32COUNTRY", "COUNTRY", "X(60)", "C", None),
        FileField(2, "T32GHOST", "GHOST", "X(10)", "C", None),   # not a table column
    ]
    findings = check_orphans("Tmpl", "Bank Account", "T_BANKS_BRANCHES",
                             "I_T_BANKS_BRANCHES", "IMPORT", table_cols, if_fields)
    assert len(findings) == 1
    assert findings[0].applaud_field == "T32GHOST"
    assert findings[0].attribute == "PRESENCE" and findings[0].severity == "MED"


def test_check_orphans_silent_when_all_match():
    table_cols = [DataColumn("T32COUNTRY", "COUNTRY", "X", 60, None, "COUNTRY", 1)]
    if_fields = [FileField(1, "T32COUNTRY", "COUNTRY", "X(60)", "C", None)]
    assert check_orphans("T", "tab", "T_X", "I_X", "IMPORT", table_cols, if_fields) == []


# --- Task 12: Dim 6b release-delta -----------------------------------------

def test_release_delta_flags_added_missing_and_removed_lingering():
    changes = [
        Change("ADDED", None, 3, None, _of(3, "NEW_FIELD")),
        Change("REMOVED", 2, None, _of(2, "OLD_FIELD"), None),
    ]
    applaud_bares = {"COUNTRY", "OLD_FIELD"}
    findings = check_release_delta("Tmpl", "Bank Account", "T_BANKS_BRANCHES",
                                   changes, applaud_bares, old_release="26A", new_release="26B")
    by_field = {f.oracle_field: f for f in findings}
    assert by_field["NEW_FIELD"].severity == "HIGH"
    assert "added" in by_field["NEW_FIELD"].message.lower()
    assert by_field["OLD_FIELD"].severity == "MED"
    assert "removed" in by_field["OLD_FIELD"].message.lower()


def test_release_delta_silent_when_applaud_in_sync():
    changes = [Change("ADDED", None, 3, None, _of(3, "NEW_FIELD"))]
    findings = check_release_delta("T", "tab", "T_X", changes,
                                   {"NEW_FIELD"}, "26A", "26B")
    assert findings == []


def test_build_release_changes_aligns_per_tab():
    old = {("Tmpl", "Bank Account"): [_of(1, "COUNTRY"), _of(2, "OLD_FIELD")]}
    new = {("Tmpl", "Bank Account"): [_of(1, "COUNTRY"), _of(2, "NEW_FIELD")]}
    changes = build_release_changes(old, new)
    kinds = {c.change_type for c in changes[("Tmpl", "Bank Account")]}
    assert "ADDED" in kinds and "REMOVED" in kinds


# --- Task 13: Dim 6c unmapped + coverage gaps ------------------------------

def test_check_unmapped_flags_snapshot_table_without_mapping():
    snapshot_tables = {"T_BANKS_BRANCHES", "T_ORPHAN_TABLE"}
    mapped_tables = {"T_BANKS_BRANCHES"}
    findings = check_unmapped(snapshot_tables, mapped_tables)
    assert len(findings) == 1
    assert findings[0].applaud_object_name == "T_ORPHAN_TABLE"
    assert findings[0].severity == "INFO" and findings[0].dimension == "6c-UNMAPPED"


def test_coverage_gaps_lists_mapped_tables_with_no_if_ef():
    gaps = coverage_gaps(
        mapped_tables={"T_A", "T_B"},
        appmap={"T_A": (["I_T_A"], ["E_T_A"]), "T_B": ([], [])},
    )
    assert gaps == [("T_B", "no IF/EF resolved in app-map")]


# --- Task 14: Excel findings writer ----------------------------------------

def _sample_finding(sev="HIGH"):
    return Finding(finding_id="abc123", dimension="1-SIZING", severity=sev,
                   fbdi_template="Tmpl", fbdi_tab="Bank Account", oracle_field="BANK_NAME",
                   oracle_type="VARCHAR2(100)", applaud_object_type="TABLE",
                   applaud_object_name="T_BANKS_BRANCHES", applaud_field="T32BANK_NAME",
                   attribute="SIZE", current_value="char 30", expected_value="char 100",
                   message="Undersized")


def test_write_findings_workbook_has_four_sheets_and_status_columns(tmp_path):
    path = tmp_path / "report.xlsx"
    write_findings_workbook(
        findings=[_sample_finding("HIGH"), _sample_finding("INFO")],
        coverage=[("T_B", "no IF/EF resolved in app-map")],
        meta={"system": "ORACLE_MASTER", "release": "26B",
              "extracted_at": "2026-06-02T00:00:00+00:00"},
        path=path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "Findings", "High Priority", "Coverage"]
    findings_headers = [c.value for c in wb["Findings"][1]]
    assert "Status" in findings_headers and "Notes" in findings_headers
    assert wb["High Priority"].max_row == 2   # header + 1 HIGH row


# --- Task 15: run_audit orchestration --------------------------------------

def _build_snapshot():
    return ApplaudSnapshot(
        system="ORACLE_MASTER", mdb_path="X", extracted_at="2026-06-02T00:00:00+00:00",
        extractor_version="1",
        tables={"T_BANKS_BRANCHES": SnapshotTable(
            name="T_BANKS_BRANCHES", prefix="T32", prefix_fallback=False,
            description="T_BANKS_BRANCHES (T32)", key_seqs=[["T32COUNTRY"]],
            columns=[DataColumn("T32COUNTRY", "COUNTRY", "X", 60, None, "COUNTRY", 1),
                     DataColumn("T32BANK_NAME", "BANK_NAME", "X", 30, None, "BANK_NAME", 2)])},
        imports={"I_T_BANKS_BRANCHES": [
            FileField(1, "T32COUNTRY", "COUNTRY", "X(60)", "C", None),
            FileField(2, "T32BANK_NAME", "BANK_NAME", "X(30)", "C", None)]},
        exports={"T_BANKS_BRANCHES": [
            FileField(1, "T32COUNTRY", "COUNTRY", "X(60)", None, ""),
            FileField(2, "T32BANK_NAME", "BANK_NAME", "X(30)", None, "")]},
        applications={})


def test_run_audit_produces_workbook_and_sizing_finding(tmp_path):
    snap = _build_snapshot()
    catalog = {("Tmpl", "Bank Account"): [
        AlignedField(1, "Country", "COUNTRY", "VARCHAR2", 60, None, False),
        AlignedField(2, "Bank Name", "BANK_NAME", "VARCHAR2", 100, None, True)]}
    mapping = {("Tmpl", "Bank Account"): {"applaud_table": "T_BANKS_BRANCHES",
               "prefix": "T32", "module": "Fin", "status": "MAPPED", "in_base": ""}}
    appmap = {"T_BANKS_BRANCHES": AppMapRow("T_BANKS_BRANCHES",
              ["I_T_BANKS_BRANCHES"], ["T_BANKS_BRANCHES"],
              ["I_T_BANKS_BRANCHES", "X_T_BANKS_BRANCHES"], "confirmed")}
    out = tmp_path / "report.xlsx"
    findings = run_audit(snap, catalog, mapping, appmap, release="26B",
                         release_changes={}, out_path=out)
    assert out.exists()
    assert any(f.dimension == "1-SIZING" and f.oracle_field == "BANK_NAME"
               and f.severity == "HIGH" for f in findings)


def test_run_audit_thin_tab_label_only_no_spurious_presence(tmp_path):
    """§3.2 integration check: the canonical Bank Account tab has technical=None
    (labels only). With label->technical normalization wired, the IF's known-good
    fields must produce ZERO spurious 2-IF PRESENCE findings."""
    bares = ["COUNTRY", "BANK_NAME", "BANK_CODE", "ALTERNATE_BANK_NAME"]
    labels = ["Country", "Bank Name", "Bank Code", "Alternate Bank Name"]
    snap = ApplaudSnapshot(
        system="ORACLE_MASTER", mdb_path="X", extracted_at="2026-06-02T00:00:00+00:00",
        extractor_version="1",
        tables={"T_BANKS_BRANCHES": SnapshotTable(
            name="T_BANKS_BRANCHES", prefix="T32", prefix_fallback=False,
            description="T_BANKS_BRANCHES (T32)", key_seqs=[["T32COUNTRY"]],
            columns=[DataColumn(f"T32{b}", b, "X", 100, None, "", i + 1)
                     for i, b in enumerate(bares)])},
        imports={"I_T_BANKS_BRANCHES": [
            FileField(i + 1, f"T32{b}", b, "X(100)", "C", None)
            for i, b in enumerate(bares)]},
        exports={}, applications={})
    catalog = {("RapidImplementationForCashManagement", "Bank Account"): [
        AlignedField(i + 1, lbl, None, None, None, None, None)   # technical=None
        for i, lbl in enumerate(labels)]}
    mapping = {("RapidImplementationForCashManagement", "Bank Account"): {
        "applaud_table": "T_BANKS_BRANCHES", "prefix": "T32", "module": "Fin",
        "status": "MAPPED", "in_base": ""}}
    appmap = {"T_BANKS_BRANCHES": AppMapRow("T_BANKS_BRANCHES",
              ["I_T_BANKS_BRANCHES"], [], ["I_T_BANKS_BRANCHES"], "confirmed")}
    findings = run_audit(snap, catalog, mapping, appmap, release="26B",
                         release_changes={}, out_path=tmp_path / "r.xlsx")
    if_presence = [f for f in findings
                   if f.dimension == "2-IF" and f.attribute == "PRESENCE"]
    assert if_presence == []
