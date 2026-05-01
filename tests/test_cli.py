"""Tests for fbdi.cli — CLI helpers."""

import pytest
from pathlib import Path
from fbdi.cli import _resolve_dir


class TestResolveDir:
    def test_existing_directory_passes_through(self, tmp_path):
        """A path that is already a directory is returned unchanged."""
        assert _resolve_dir(tmp_path) == tmp_path

    def test_release_label_resolves_to_originals(self, tmp_path, monkeypatch):
        """A non-directory path like '26A' resolves to baselines/26A/originals/."""
        baselines = tmp_path / "baselines" / "26A" / "originals"
        baselines.mkdir(parents=True)
        monkeypatch.chdir(tmp_path)
        result = _resolve_dir(Path("26A"))
        assert result == Path("baselines") / "26A" / "originals"
        assert result.is_dir()

    def test_nonexistent_path_passes_through(self, tmp_path, monkeypatch):
        """A path that doesn't exist and has no baselines match passes through."""
        monkeypatch.chdir(tmp_path)
        result = _resolve_dir(Path("nonexistent"))
        # Should return the original path for downstream error handling
        assert result == Path("nonexistent")

    def test_explicit_originals_path_passes_through(self, tmp_path):
        """An explicit path to originals/ is returned unchanged."""
        originals = tmp_path / "baselines" / "26A" / "originals"
        originals.mkdir(parents=True)
        assert _resolve_dir(originals) == originals


class TestCatalogCLI:
    def test_catalog_cli_requires_release(self, tmp_path, capsys):
        from fbdi.cli import main
        with pytest.raises(SystemExit):
            main(["catalog"])

    def test_catalog_cli_missing_baselines_errors(self, tmp_path, capsys):
        from fbdi.cli import main
        with pytest.raises(SystemExit):
            main([
                "catalog", "--release", "99Z",
                "--baselines-dir", str(tmp_path / "does-not-exist"),
                "--master", str(tmp_path / "M.xlsx"),
            ])
        captured = capsys.readouterr()
        assert "not found" in captured.out.lower()

    def test_catalog_cli_end_to_end(self, tmp_path):
        """Build a tiny release dir + run catalog CLI + verify file written."""
        from openpyxl import Workbook
        from fbdi.cli import main

        baselines = tmp_path / "baselines" / "TESTZ" / "originals"
        baselines.mkdir(parents=True)
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("MY_TAB")
        # Thin tab — need MIN_CELLS=2 for header detection
        ws.cell(row=4, column=1, value="*Only Field")
        ws.cell(row=4, column=2, value="Second Field")
        wb.save(baselines / "Tpl.xlsm")

        master = tmp_path / "Catalog.xlsx"
        main([
            "catalog", "--release", "TESTZ",
            "--baselines-dir", str(baselines),
            "--master", str(master),
            "--timeout", "30",
        ])
        assert master.exists()


def test_populate_module_subcommand_invocation(tmp_path, monkeypatch, capsys):
    """`python -m fbdi populate-module` invokes populate_module_column with the
    right args and prints the summary."""
    import fbdi.cli as cli_mod
    from openpyxl import Workbook

    # Build minimal artifacts in tmp_path
    (tmp_path / "baselines" / "26a").mkdir(parents=True)
    (tmp_path / "baselines" / "26b").mkdir(parents=True)
    (tmp_path / "baselines" / "26a" / "file_modules.json").write_text(
        '{"AutoInvoiceImportTemplate.xlsm": "Financials"}'
    )
    (tmp_path / "baselines" / "26b" / "file_modules.json").write_text(
        '{"AutoInvoiceImportTemplate.xlsm": "Financials"}'
    )

    mapping_path = tmp_path / "mapping.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "FBDI Mapping"
    headers = ["FBDI Template", "FBDI Tab", "Applaud Table", "Prefix",
               "Status", "Module", "In Base System?"]
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    ws.cell(row=2, column=1, value="AutoInvoiceImportTemplate")
    ws.cell(row=2, column=2, value="RA_TAB")
    wb.save(mapping_path)
    wb.close()

    monkeypatch.chdir(tmp_path)
    cli_mod.main(["populate-module", "--new", "26b", "--old", "26a",
                  "--mapping", str(mapping_path)])

    # Verify the workbook actually got the Module value written
    from openpyxl import load_workbook
    wb = load_workbook(mapping_path, read_only=True)
    ws = wb["FBDI Mapping"]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    wb.close()
    assert row2[5] == "Financials"  # column F

    # Verify the JSON summary on stdout includes the populated count
    out = capsys.readouterr().out
    assert '"populated": 1' in out
    assert '"new_release": "26B"' in out


def test_populate_module_missing_json_exits_2(tmp_path, monkeypatch):
    """Missing file_modules.json for either release exits 2."""
    import fbdi.cli as cli_mod

    # Only create the OLD baseline — NEW's file_modules.json is missing
    (tmp_path / "baselines" / "26a").mkdir(parents=True)
    (tmp_path / "baselines" / "26a" / "file_modules.json").write_text("{}")

    monkeypatch.chdir(tmp_path)
    with pytest.raises(SystemExit) as excinfo:
        cli_mod.main(["populate-module", "--new", "26b", "--old", "26a",
                      "--mapping", "ignored.xlsx"])
    assert excinfo.value.code == 2


class TestReportSubcommand:
    def test_report_subcommand_parses_old_and_new(self, monkeypatch, tmp_path):
        from fbdi import cli

        called = {}

        def fake_generate(catalog_path, mapping_path, old_release, new_release, out_dir):
            called.update(dict(
                catalog_path=catalog_path, mapping_path=mapping_path,
                old_release=old_release, new_release=new_release, out_dir=out_dir,
            ))
            return tmp_path / "x.html", tmp_path / "x.pdf"

        (tmp_path / "cat.xlsx").write_bytes(b"stub")
        (tmp_path / "map.xlsx").write_bytes(b"stub")
        monkeypatch.setattr("fbdi.report.generate_report", fake_generate)
        cli.main([
            "report", "--old", "26A", "--new", "26B",
            "--out-dir", str(tmp_path),
            "--catalog", str(tmp_path / "cat.xlsx"),
            "--mapping", str(tmp_path / "map.xlsx"),
        ])
        assert called["old_release"] == "26A"
        assert called["new_release"] == "26B"
        assert called["out_dir"] == tmp_path
