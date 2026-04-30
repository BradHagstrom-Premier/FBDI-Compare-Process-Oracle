"""Tests for the verify_rerun.py macro-signal validator."""

import importlib.util
import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook


SKILL_SCRIPT = Path(__file__).resolve().parent.parent / ".claude" / "skills" / \
    "fbdi-compare-release" / "scripts" / "verify_rerun.py"


def _load_module():
    spec = importlib.util.spec_from_file_location("verify_rerun", SKILL_SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    sys.modules["verify_rerun"] = mod
    spec.loader.exec_module(mod)
    return mod


def _make_catalog(path: Path, release_rows: dict[str, int]):
    """Build a synthetic FBDI_Master_Catalog with N rows on each release sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    for release, n_rows in release_rows.items():
        ws = wb.create_sheet(release)
        ws.cell(row=1, column=1, value="file")
        ws.cell(row=1, column=2, value="tab")
        for i in range(n_rows):
            ws.cell(row=i + 2, column=1, value=f"file{i}.xlsm")
            ws.cell(row=i + 2, column=2, value=f"tab{i}")
    wb.save(path)
    wb.close()


def _make_compare_report(path: Path, n_changes: int):
    wb = Workbook()
    ws = wb.active
    ws.title = "Changes"
    ws.cell(row=1, column=1, value="file")
    for i in range(n_changes):
        ws.cell(row=i + 2, column=1, value=f"row{i}")
    wb.save(path)
    wb.close()


def _make_mapping(path: Path, total_rows: int, populated_rows: int):
    wb = Workbook()
    ws = wb.active
    ws.title = "FBDI Mapping"
    headers = ["FBDI Template", "FBDI Tab", "Applaud Table", "Prefix",
               "Status", "Module", "In Base System?"]
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    for i in range(total_rows):
        ws.cell(row=i + 2, column=1, value=f"Template{i}")
        if i < populated_rows:
            ws.cell(row=i + 2, column=6, value="Financials")
    wb.save(path)
    wb.close()


class TestVerifyRerun:
    def test_all_green(self, tmp_path):
        mod = _load_module()
        # post = pre (no delta)
        _make_catalog(tmp_path / "post.xlsx", {"26A": 12000, "26B": 12000})
        _make_catalog(tmp_path / "pre.xlsx",  {"26A": 12000, "26B": 12000})
        _make_compare_report(tmp_path / "report.xlsx", 706)
        _make_mapping(tmp_path / "mapping.xlsx", 100, 96)

        result = mod.run_checks(
            new_catalog=tmp_path / "post.xlsx",
            baseline_catalog=tmp_path / "pre.xlsx",
            compare_report=tmp_path / "report.xlsx",
            mapping=tmp_path / "mapping.xlsx",
            release="26B",
        )
        assert result["regressions"] == []
        assert result["catalog_delta_pct"] == pytest.approx(0.0, abs=0.01)
        assert result["module_pct_populated"] == pytest.approx(96.0, abs=0.01)

    def test_catalog_row_delta_exceeds_threshold(self, tmp_path):
        mod = _load_module()
        _make_catalog(tmp_path / "post.xlsx", {"26B": 11000})  # -8.3% vs 12000
        _make_catalog(tmp_path / "pre.xlsx",  {"26B": 12000})
        _make_compare_report(tmp_path / "report.xlsx", 706)
        _make_mapping(tmp_path / "mapping.xlsx", 100, 96)

        result = mod.run_checks(
            new_catalog=tmp_path / "post.xlsx",
            baseline_catalog=tmp_path / "pre.xlsx",
            compare_report=tmp_path / "report.xlsx",
            mapping=tmp_path / "mapping.xlsx",
            release="26B",
        )
        assert any("catalog" in r.lower() for r in result["regressions"])

    def test_compare_changes_delta_exceeds_threshold(self, tmp_path):
        mod = _load_module()
        _make_catalog(tmp_path / "post.xlsx", {"26B": 12000})
        _make_catalog(tmp_path / "pre.xlsx",  {"26B": 12000})
        _make_compare_report(tmp_path / "report.xlsx", 900)  # 706 ± 50 → fail
        _make_mapping(tmp_path / "mapping.xlsx", 100, 96)

        result = mod.run_checks(
            new_catalog=tmp_path / "post.xlsx",
            baseline_catalog=tmp_path / "pre.xlsx",
            compare_report=tmp_path / "report.xlsx",
            mapping=tmp_path / "mapping.xlsx",
            release="26B",
            expected_compare_changes=706,
        )
        assert any("compare" in r.lower() for r in result["regressions"])

    def test_module_pct_below_threshold(self, tmp_path):
        mod = _load_module()
        _make_catalog(tmp_path / "post.xlsx", {"26B": 12000})
        _make_catalog(tmp_path / "pre.xlsx",  {"26B": 12000})
        _make_compare_report(tmp_path / "report.xlsx", 706)
        _make_mapping(tmp_path / "mapping.xlsx", 100, 80)  # 80% < 95%

        result = mod.run_checks(
            new_catalog=tmp_path / "post.xlsx",
            baseline_catalog=tmp_path / "pre.xlsx",
            compare_report=tmp_path / "report.xlsx",
            mapping=tmp_path / "mapping.xlsx",
            release="26B",
        )
        assert any("module" in r.lower() for r in result["regressions"])

    def test_baseline_catalog_missing_skips_delta(self, tmp_path):
        """If pre-rerun catalog isn't available (first run), skip the delta check."""
        mod = _load_module()
        _make_catalog(tmp_path / "post.xlsx", {"26B": 12000})
        _make_compare_report(tmp_path / "report.xlsx", 706)
        _make_mapping(tmp_path / "mapping.xlsx", 100, 96)

        result = mod.run_checks(
            new_catalog=tmp_path / "post.xlsx",
            baseline_catalog=tmp_path / "missing.xlsx",
            compare_report=tmp_path / "report.xlsx",
            mapping=tmp_path / "mapping.xlsx",
            release="26B",
        )
        # No regression from the missing baseline; the result records that the
        # check was skipped instead.
        assert result["catalog_delta_pct"] is None
        assert "regressions" in result
