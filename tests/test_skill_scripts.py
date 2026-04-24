"""Tests for the fbdi-compare-release skill's bundled scripts."""

import json
import subprocess
import sys
from pathlib import Path


SKILL_ROOT = Path(__file__).resolve().parent.parent / ".claude" / "skills" / "fbdi-compare-release"
# Make `from scripts import <module>` resolve the skill's bundled scripts.
sys.path.insert(0, str(SKILL_ROOT))


def test_skill_folder_exists():
    assert SKILL_ROOT.is_dir(), f"expected skill folder at {SKILL_ROOT}"


def test_skill_md_has_frontmatter():
    skill_md = SKILL_ROOT / "SKILL.md"
    assert skill_md.is_file()
    text = skill_md.read_text(encoding="utf-8")
    assert text.startswith("---\n")
    assert "\nname: fbdi-compare-release\n" in text
    assert "\ndescription:" in text


def test_scripts_dir_is_python_package():
    scripts_dir = SKILL_ROOT / "scripts"
    assert scripts_dir.is_dir()
    assert (scripts_dir / "__init__.py").is_file()


from scripts import check_env  # noqa: E402 — importable when cwd is repo root


def _run_check_env(tmp_path):
    """Invoke check_env.py as a subprocess with a cwd of tmp_path and return (exit_code, stdout_json)."""
    cmd = [sys.executable, str(SKILL_ROOT / "scripts" / "check_env.py")]
    proc = subprocess.run(cmd, cwd=tmp_path, capture_output=True, text=True)
    return proc.returncode, json.loads(proc.stdout)


def test_check_env_exposes_main():
    assert hasattr(check_env, "main")


def test_check_env_python_version_check_passes_on_314():
    result = check_env.check_python_version(current=(3, 14, 3))
    assert result["ok"] is True
    assert "3.14" in result["detail"]


def test_check_env_python_version_check_fails_on_old():
    result = check_env.check_python_version(current=(3, 11, 0))
    assert result["ok"] is False
    assert "3.14" in result["detail"]


def test_check_env_deps_check_detects_missing():
    result = check_env.check_deps(required=["definitely_not_a_real_package_xyz"])
    assert result["ok"] is False
    assert "definitely_not_a_real_package_xyz" in result["detail"]


def test_check_env_deps_check_passes_on_stdlib():
    # json is stdlib — always importable
    result = check_env.check_deps(required=["json"])
    assert result["ok"] is True


def test_check_env_baselines_dir_creates_if_missing(tmp_path):
    result = check_env.check_baselines_dir(root=tmp_path)
    assert result["ok"] is True
    assert (tmp_path / "baselines").is_dir()


def test_check_env_produces_structured_json(tmp_path):
    """check_env.py always emits a parseable payload with the documented
    shape. Exit code is not asserted here because a dev machine may legitimately
    be missing Chrome (→ exit 1) or deps (→ exit 2); those paths have their
    own unit tests via the helper functions."""
    (tmp_path / "baselines").mkdir()
    (tmp_path / "baseline_files.txt").write_text("stub\n")
    _, payload = _run_check_env(tmp_path)
    assert "checks" in payload
    assert "missing_deps" in payload
    assert "fatal" in payload


def test_check_env_json_output_parseable(tmp_path):
    exit_code, payload = _run_check_env(tmp_path)
    assert isinstance(payload, dict)
    assert isinstance(payload["checks"], list)


from scripts import verify_download  # noqa: E402


INVENTORY_FIXTURE = """\
FBDI Baseline File Inventory
Generated: 2026-04-23
============================

26A has 3 files. 26B has 4 files.

============================
26A ORIGINALS (3 files)
============================
AccountCombinationsImportTemplate.xlsm
BudgetImportTemplate.xlsm
RapidImplementationForCashManagement.xlsm

============================
26B ORIGINALS (4 files)
============================
AccountCombinationsImportTemplate.xlsm
BudgetImportTemplate.xlsm
ItemImportReferenceOrgTemplate.xlsm
RapidImplementationForCashManagement.xlsm

============================
DIFFERENCES
============================
Only in 26B: ItemImportReferenceOrgTemplate.xlsm
"""


def test_parse_inventory_extracts_both_sections():
    inventory = verify_download.parse_inventory(INVENTORY_FIXTURE)
    assert set(inventory.keys()) == {"26A", "26B"}
    assert inventory["26A"] == [
        "AccountCombinationsImportTemplate.xlsm",
        "BudgetImportTemplate.xlsm",
        "RapidImplementationForCashManagement.xlsm",
    ]
    assert len(inventory["26B"]) == 4


def test_parse_inventory_ignores_differences_footer():
    inventory = verify_download.parse_inventory(INVENTORY_FIXTURE)
    # The "DIFFERENCES" block is after the last ORIGINALS header;
    # its content must NOT leak into 26B.
    assert "Only in 26B: ItemImportReferenceOrgTemplate.xlsm" not in inventory["26B"]


def test_parse_inventory_empty_text():
    assert verify_download.parse_inventory("") == {}


def test_diff_clean_case():
    inventory = {"26A": ["A.xlsm", "B.xlsm"]}
    result = verify_download.diff_against_inventory(
        release="26A",
        downloaded_names=["A.xlsm", "B.xlsm"],
        inventory=inventory,
        manual_files=[],
    )
    assert result["missing"] == []
    assert result["extras"] == []


def test_diff_detects_missing_and_extras():
    inventory = {"26A": ["A.xlsm", "B.xlsm", "C.xlsm"]}
    result = verify_download.diff_against_inventory(
        release="26A",
        downloaded_names=["A.xlsm", "B.xlsm", "D.xlsm"],
        inventory=inventory,
        manual_files=[],
    )
    assert result["missing"] == ["C.xlsm"]
    assert result["extras"] == ["D.xlsm"]


def test_diff_excludes_manual_files_from_missing():
    inventory = {"26A": ["A.xlsm", "RapidImplementationForCashManagement.xlsm"]}
    result = verify_download.diff_against_inventory(
        release="26A",
        downloaded_names=["A.xlsm"],
        inventory=inventory,
        manual_files=["RapidImplementationForCashManagement.xlsm"],
    )
    assert result["missing"] == []  # manual file excluded


def test_diff_is_locale_agnostic():
    """Guard against non-`LC_ALL=C` environments where Mac default sort
    misorders mixed-case filenames. Set operations are locale-independent —
    we verify the diff is identical regardless of filename case ordering."""
    inventory = {"26A": ["AccountCombinationsImportTemplate.xlsm", "zxCustomTemplate.xlsm"]}
    # Downloaded in a different case-order
    result = verify_download.diff_against_inventory(
        release="26A",
        downloaded_names=["zxCustomTemplate.xlsm", "AccountCombinationsImportTemplate.xlsm"],
        inventory=inventory,
        manual_files=[],
    )
    assert result["missing"] == []
    assert result["extras"] == []


def test_group_missing_by_module_basic():
    groups = verify_download.group_missing_by_module([
        "POBlanketPurchaseAgreementImportTemplate.xlsm",
        "SupplierImportTemplate.xlsm",
        "FixedAssetMassAdditionsImportTemplate.xlsm",
        "ScpItemCostImportTemplate.xlsm",
        "ImportAwards.xlsm",
        "WeirdUnknownFileXYZ.xlsm",
    ])
    assert "POBlanketPurchaseAgreementImportTemplate.xlsm" in groups["procurement"]
    assert "SupplierImportTemplate.xlsm" in groups["procurement"]
    assert "FixedAssetMassAdditionsImportTemplate.xlsm" in groups["financials"]
    assert "ScpItemCostImportTemplate.xlsm" in groups["supply-chain-and-manufacturing"]
    assert "ImportAwards.xlsm" in groups["project-management"]
    assert "WeirdUnknownFileXYZ.xlsm" in groups["other"]


def test_group_missing_by_module_empty():
    assert verify_download.group_missing_by_module([]) == {}


def test_group_missing_by_module_longest_prefix_wins():
    """Regression guard: project-management's 'Import' prefix must not
    swallow financials-specific 'ImportStandaloneFiscal*' filenames. The
    longer prefix should win across module boundaries."""
    groups = verify_download.group_missing_by_module([
        "ImportStandaloneFiscalDocumentTemplate.xlsm",
        "ImportAwards.xlsm",
    ])
    assert "ImportStandaloneFiscalDocumentTemplate.xlsm" in groups["financials"]
    assert "ImportAwards.xlsm" in groups["project-management"]


def test_group_missing_by_module_covers_real_inventory():
    """Regression guard: every non-manual file in the committed baseline_files.txt
    must classify to a known module (not 'other'). If a future release adds a
    file with an unrecognized prefix, this test fails and MODULE_PREFIXES needs
    updating."""
    inv_path = Path(__file__).resolve().parent.parent / "baseline_files.txt"
    if not inv_path.is_file():
        import pytest
        pytest.skip("baseline_files.txt not present (e.g., CI without committed inventory)")
    inventory = verify_download.parse_inventory(inv_path.read_text(encoding="utf-8"))
    all_files = set()
    for files in inventory.values():
        all_files.update(files)
    # Filter out known manual-only files
    candidates = sorted(all_files - set(verify_download.MANUAL_FILES))
    groups = verify_download.group_missing_by_module(candidates)
    other = groups.get("other", [])
    assert other == [], (
        f"{len(other)} file(s) fell into 'other' — add matching prefixes to "
        f"MODULE_PREFIXES in verify_download.py:\n  " + "\n  ".join(other)
    )


def test_compute_first_run_delta_within_threshold():
    inventory = {"26A": ["a.xlsm"] * 212, "26B": ["a.xlsm"] * 213}
    result = verify_download.compute_first_run_delta(
        downloaded_count=215,
        inventory=inventory,
    )
    assert result["prior_release"] == "26B"
    assert result["prior_count"] == 213
    assert abs(result["delta_pct"] - (2 / 213)) < 1e-6
    assert result["over_threshold"] is False


def test_compute_first_run_delta_over_threshold():
    inventory = {"26B": ["a.xlsm"] * 213}
    result = verify_download.compute_first_run_delta(
        downloaded_count=107,  # -49.8% — matches the 2026-04-23 module-silent-failure case
        inventory=inventory,
    )
    assert result["prior_release"] == "26B"
    assert result["over_threshold"] is True


def test_compute_first_run_delta_no_prior():
    # Empty inventory = no prior to compare against — non-fatal
    result = verify_download.compute_first_run_delta(downloaded_count=100, inventory={})
    assert result["prior_release"] is None
    assert result["over_threshold"] is False


def test_most_recent_release_sorts_ascii():
    inventory = {"25D": [], "26A": [], "26B": []}
    assert verify_download.most_recent_release(inventory) == "26B"


def test_most_recent_release_empty():
    assert verify_download.most_recent_release({}) is None


def test_commit_inventory_replaces_existing_section():
    inventory_text = INVENTORY_FIXTURE
    new_26b = ["A.xlsm", "B.xlsm"]  # shrunk from 4 to 2
    result = verify_download.commit_inventory(
        inventory_text, release="26B", filenames=new_26b,
    )
    parsed = verify_download.parse_inventory(result)
    assert parsed["26B"] == ["A.xlsm", "B.xlsm"]
    assert parsed["26A"] == [
        "AccountCombinationsImportTemplate.xlsm",
        "BudgetImportTemplate.xlsm",
        "RapidImplementationForCashManagement.xlsm",
    ]
    assert "26B ORIGINALS (2 files)" in result


def test_commit_inventory_appends_new_section():
    inventory_text = INVENTORY_FIXTURE
    filenames_26c = ["NewFileA.xlsm", "NewFileB.xlsm", "NewFileC.xlsm"]
    result = verify_download.commit_inventory(
        inventory_text, release="26C", filenames=filenames_26c,
    )
    parsed = verify_download.parse_inventory(result)
    assert parsed["26C"] == sorted(filenames_26c)
    assert "26C ORIGINALS (3 files)" in result
    # 26B section must still be present and unchanged
    assert parsed["26B"] == sorted([
        "AccountCombinationsImportTemplate.xlsm",
        "BudgetImportTemplate.xlsm",
        "ItemImportReferenceOrgTemplate.xlsm",
        "RapidImplementationForCashManagement.xlsm",
    ])


def test_commit_inventory_sorts_filenames_ascii():
    result = verify_download.commit_inventory(
        INVENTORY_FIXTURE, release="26C",
        filenames=["Zebra.xlsm", "AAA.xlsm", "Middle.xlsm"],
    )
    idx_aaa = result.index("AAA.xlsm")
    idx_middle = result.index("Middle.xlsm")
    idx_zebra = result.index("Zebra.xlsm")
    assert idx_aaa < idx_middle < idx_zebra


def test_commit_inventory_cli_writes_file_in_place(tmp_path):
    inv_path = tmp_path / "baseline_files.txt"
    inv_path.write_text(INVENTORY_FIXTURE, encoding="utf-8")

    originals = tmp_path / "baselines" / "26C" / "originals"
    originals.mkdir(parents=True)
    for name in ("A.xlsm", "B.xlsm"):
        (originals / name).touch()

    exit_code = verify_download.main([
        "--release", "26C",
        "--inventory", str(inv_path),
        "--originals", str(originals),
        "--commit-inventory",
    ])
    assert exit_code == 0
    parsed = verify_download.parse_inventory(inv_path.read_text(encoding="utf-8"))
    assert parsed["26C"] == ["A.xlsm", "B.xlsm"]


def test_commit_inventory_roundtrips_real_baseline_files_txt():
    """Regression guard: committing each release's own filenames back into
    the real baseline_files.txt must produce an identical-parse result
    (set of releases → same sorted file lists). Any formatting drift that
    breaks the writer's compatibility with the real inventory file format
    will fail this test."""
    inv_path = Path(__file__).resolve().parent.parent / "baseline_files.txt"
    if not inv_path.is_file():
        import pytest
        pytest.skip("baseline_files.txt not present")
    original = inv_path.read_text(encoding="utf-8")
    parsed_before = verify_download.parse_inventory(original)
    # Commit 26B's own files back — should be idempotent on parse semantics.
    rewritten = verify_download.commit_inventory(
        original, release="26B", filenames=parsed_before["26B"],
    )
    parsed_after = verify_download.parse_inventory(rewritten)
    assert parsed_after == parsed_before


def test_commit_inventory_handles_missing_trailing_newline():
    """Guard against input that doesn't end in \\n. The section-match regex
    requires lines to be newline-terminated; the function must normalize."""
    # Drop the trailing newline from the fixture
    text = INVENTORY_FIXTURE.rstrip("\n")
    assert not text.endswith("\n")
    result = verify_download.commit_inventory(
        text, release="26B", filenames=["A.xlsm", "B.xlsm"],
    )
    parsed = verify_download.parse_inventory(result)
    assert parsed["26B"] == ["A.xlsm", "B.xlsm"]
    # 26A must not have been mangled
    assert parsed["26A"] == [
        "AccountCombinationsImportTemplate.xlsm",
        "BudgetImportTemplate.xlsm",
        "RapidImplementationForCashManagement.xlsm",
    ]


def test_format_section_singular_for_one_file():
    """A release with exactly 1 file emits 'file', not 'files'."""
    text = verify_download._format_section("26C", ["Only.xlsm"])
    assert "26C ORIGINALS (1 file)" in text
    assert "1 files)" not in text


from openpyxl import Workbook
from scripts import summarize_report  # noqa: E402


def _make_comparison_report(path, rows):
    """rows = list of (fbdi_file, fbdi_tab, col_letter, col_num, old, new, diff)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["FBDI File", "FBDI Tab", "Column Letter", "Column Number",
               "Old FBDI Field Name", "New FBDI Field Name", "Difference?"])
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    wb.close()


def test_summarize_counts_changes(tmp_path):
    path = tmp_path / "cmp.xlsx"
    _make_comparison_report(path, [
        ("FileA", "Tab1", "A", 1, "old1", "new1", "YES"),
        ("FileA", "Tab1", "B", 2, "old2", "new2", "YES"),
        ("FileB", "Tab1", "A", 1, "old3", "new3", "YES"),
    ])
    result = summarize_report.summarize(path)
    assert result["total_changes"] == 3
    assert result["files_with_changes"] == 2


def test_summarize_top_files_ordered(tmp_path):
    path = tmp_path / "cmp.xlsx"
    rows = (
        [("FileB", "T", "A", 1, "o", "n", "YES")] * 10
        + [("FileA", "T", "A", 1, "o", "n", "YES")] * 5
        + [("FileC", "T", "A", 1, "o", "n", "YES")] * 3
    )
    _make_comparison_report(path, rows)
    result = summarize_report.summarize(path)
    assert [t["file"] for t in result["top_files"]][:3] == ["FileB", "FileA", "FileC"]
    assert result["top_files"][0]["changes"] == 10


def test_summarize_top_files_capped_at_5(tmp_path):
    path = tmp_path / "cmp.xlsx"
    rows = [(f"File{i}", "T", "A", 1, "o", "n", "YES") for i in range(10)]
    _make_comparison_report(path, rows)
    result = summarize_report.summarize(path)
    assert len(result["top_files"]) <= 5


def test_summarize_empty_report(tmp_path):
    path = tmp_path / "cmp.xlsx"
    _make_comparison_report(path, [])
    result = summarize_report.summarize(path)
    assert result["total_changes"] == 0
    assert result["files_with_changes"] == 0
    assert result["top_files"] == []


def test_summarize_cli_passthrough_timeouts(tmp_path):
    path = tmp_path / "cmp.xlsx"
    _make_comparison_report(path, [])
    exit_code = summarize_report.main([
        "--report", str(path),
        "--catalog", "dummy.xlsx",
        "--timeouts", "Foo.xlsm,Bar.xlsm",
    ])
    assert exit_code == 0


def test_summarize_against_ground_truth():
    """Spec §8 eval #2 reference: 26A→26B run produced 706 changes in 19 files."""
    report = Path("Comparison_Report_26A_26B.xlsx")
    if not report.is_file():
        import pytest
        pytest.skip("ground-truth report not present")
    result = summarize_report.summarize(report)
    assert result["total_changes"] == 706
    assert result["files_with_changes"] == 19


from scripts import verify_run  # noqa: E402


def _make_catalog_with_issues(path, issues_by_release):
    """issues_by_release: {release: [(file, tab, issue_type, detail), ...]}"""
    wb = Workbook()
    # Remove default + add per-release tabs (any content, we only read Issues)
    wb.remove(wb.active)
    for release in issues_by_release:
        wb.create_sheet(release)
    issues_ws = wb.create_sheet("Issues")
    issues_ws.append(["release", "file", "tab", "issue_type", "detail"])
    for release, rows in issues_by_release.items():
        for row in rows:
            issues_ws.append([release, *row])
    wb.create_sheet("Drift")
    wb.save(path)
    wb.close()


def test_verify_run_catalog_check_no_regression(tmp_path):
    catalog = tmp_path / "cat.xlsx"
    _make_catalog_with_issues(catalog, {
        "26A": [("F", "T", "TYPE_PARSE_WARNING", "x")] * 4,
        "26B": [("F", "T", "TYPE_PARSE_WARNING", "x")] * 5,
    })
    result = verify_run.check_catalog_issues(catalog, release="26B")
    assert result["release_issue_count"] == 5
    assert result["prior_issue_count"] == 4
    assert result["regression"] is False


def test_verify_run_catalog_check_regression_2x(tmp_path):
    catalog = tmp_path / "cat.xlsx"
    _make_catalog_with_issues(catalog, {
        "26A": [("F", "T", "TYPE_PARSE_WARNING", "x")] * 4,
        "26B": [("F", "T", "TYPE_PARSE_WARNING", "x")] * 10,
    })
    result = verify_run.check_catalog_issues(catalog, release="26B")
    assert result["regression"] is True


def test_verify_run_catalog_check_regression_absolute(tmp_path):
    catalog = tmp_path / "cat.xlsx"
    _make_catalog_with_issues(catalog, {
        "26A": [("F", "T", "TYPE_PARSE_WARNING", "x")] * 100,
        "26B": [("F", "T", "TYPE_PARSE_WARNING", "x")] * 160,
    })
    result = verify_run.check_catalog_issues(catalog, release="26B")
    # delta = 60, >50 → regression even though <2x
    assert result["regression"] is True


def test_verify_run_catalog_check_no_prior(tmp_path):
    catalog = tmp_path / "cat.xlsx"
    _make_catalog_with_issues(catalog, {
        "26B": [("F", "T", "TYPE_PARSE_WARNING", "x")] * 5,
    })
    result = verify_run.check_catalog_issues(catalog, release="26B")
    assert result["prior_release"] is None
    assert result["regression"] is False
