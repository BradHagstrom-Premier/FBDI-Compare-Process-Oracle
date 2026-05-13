"""Tests for fbdi.catalog — master catalog generation."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from openpyxl.comments import Comment

from fbdi.catalog import (
    CatalogRow,
    IssueRow,
    DriftRow,
    extract_tab_rows,
    _write_master_workbook,
)


def _make_thin_tab(ws, labels: list[str], header_row: int = 4):
    """Build a thin-tab workbook: just a title/legend and a label row."""
    ws.cell(row=2, column=1, value="Some Import")
    ws.cell(row=3, column=1, value="* Required")
    for col_idx, label in enumerate(labels, start=1):
        ws.cell(row=header_row, column=col_idx, value=label)


class TestExtractTabRowsThin:
    def test_thin_tab_labels_only(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "XCC_BUDGET_INTERFACE"
        _make_thin_tab(ws, [
            "*Source Budget Type",
            "*Source Budget Name",
            "Line Number",
            "Amount",
        ])
        rows, issues = extract_tab_rows(
            ws, file_stem="BudgetImportTemplate", release="26B"
        )

        assert issues == []
        assert len(rows) == 4
        # Position 1: required (had asterisk), normalized label
        assert rows[0].position == 1
        assert rows[0].column_label == "Source Budget Type"
        assert rows[0].column_technical == ""
        assert rows[0].data_type == ""
        assert rows[0].length is None
        assert rows[0].scale is None
        assert rows[0].data_type_raw == ""
        assert rows[0].required is True
        # Position 2: required
        assert rows[1].column_label == "Source Budget Name"
        assert rows[1].required is True
        # Position 3: not required (no asterisk)
        assert rows[2].column_label == "Line Number"
        assert rows[2].required is False
        # Position 4: not required
        assert rows[3].column_label == "Amount"
        assert rows[3].required is False

    def test_thin_tab_sets_release_and_file_and_tab(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "MY_TAB"
        _make_thin_tab(ws, ["*Field A", "Field B"])
        rows, _ = extract_tab_rows(ws, file_stem="MyTemplate", release="26A")
        assert rows[0].release == "26A"
        assert rows[0].file_name == "MyTemplate"
        assert rows[0].tab_name == "MY_TAB"

    def test_thin_tab_no_header_emits_issue(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "EmptyTab"
        # Only a title — no detectable header row
        ws.cell(row=1, column=1, value="Just a title")

        rows, issues = extract_tab_rows(
            ws, file_stem="Tpl", release="26A"
        )

        assert rows == []
        assert len(issues) == 1
        assert issues[0].issue_type == "NO_HEADER"
        assert issues[0].tab == "EmptyTab"
        assert issues[0].release == "26A"
        assert issues[0].file == "Tpl"


def _make_rich_tab(
    ws,
    labels: list[str],
    descriptions: list[str] | None = None,
    data_types: list[str] | None = None,
    required_flags: list[str] | None = None,
    technicals: list[str] | None = None,
    header_row: int = 5,
    table_name: str = "RCS_ATTACHMENTS_INT",
):
    """Build a rich-tab workbook with metadata rows + technical header."""
    def _row(row_idx, label_a, values):
        ws.cell(row=row_idx, column=1, value=label_a)
        for col_idx, v in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=v)
    # Header row: "Column name of the Table X" in col A, then tech names col B..
    _row(header_row, f"Column name of the Table {table_name}", technicals or labels)
    # Name row above, then Description, Data Type, Required
    if header_row >= 2:
        _row(header_row - 1, "Required or Optional", required_flags or ["Optional"] * len(labels))
    if header_row >= 3:
        _row(header_row - 2, "Data Type", data_types or [""] * len(labels))
    if header_row >= 4:
        _row(header_row - 3, "Description", descriptions or [""] * len(labels))
    if header_row >= 5:
        _row(header_row - 4, "Name", labels)


class TestExtractTabRowsRich:
    def test_rich_tab_all_metadata_rows(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attachment Details"
        _make_rich_tab(
            ws,
            labels=["Attachment Type", "Attachment Name", "Document ID"],
            data_types=["VARCHAR2(5 CHAR)", "VARCHAR2(2048 CHAR)", "NUMBER(18)"],
            required_flags=["Required", "Required", "Optional"],
            technicals=["ATTACHMENT_TYPE", "ATTACHMENT_NAME", "DOCUMENT_ID"],
            table_name="RCS_ATTACHMENTS_INT",
            header_row=5,
        )

        rows, issues = extract_tab_rows(ws, file_stem="AttachmentsImportTemplate", release="26B")

        assert issues == []
        assert len(rows) == 3
        r0 = rows[0]
        assert r0.position == 1
        assert r0.column_label == "Attachment Type"
        assert r0.column_technical == "ATTACHMENT_TYPE"
        assert r0.data_type == "VARCHAR2"
        assert r0.length == 5
        assert r0.scale is None
        assert r0.data_type_raw == "VARCHAR2(5 CHAR)"
        assert r0.required is True
        assert rows[1].length == 2048
        assert rows[2].data_type == "NUMBER"
        assert rows[2].length == 18
        assert rows[2].required is False

    def test_rich_tab_with_bom_on_required_row(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "TabWithBOM"
        _make_rich_tab(
            ws, labels=["Col A"],
            data_types=["VARCHAR2(80)"],
            technicals=["COL_A"],
        )
        # Overwrite the 'Required or Optional' col-A label with BOM-prefixed variant
        # In _make_rich_tab, required is at header_row - 1 = row 4
        ws.cell(row=4, column=1, value="\ufeffRequired or Optional")
        ws.cell(row=4, column=2, value="Required")

        rows, _ = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].required is True

    def test_rich_tab_case_insensitive_col_a_match(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "MixedCase"
        _make_rich_tab(
            ws, labels=["Col A"],
            data_types=["VARCHAR2(80)"],
            technicals=["COL_A"],
        )
        # Lowercase the 'Data Type' label
        ws.cell(row=3, column=1, value="DATA type")  # row 3 = Data Type row
        rows, _ = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].data_type == "VARCHAR2"
        assert rows[0].length == 80

    def test_rich_tab_missing_data_type_row(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "NoDataType"
        _make_rich_tab(
            ws, labels=["Col A"],
            data_types=["VARCHAR2(80)"],
            technicals=["COL_A"],
        )
        # Clear the Data Type row's column A label so it's not recognized
        ws.cell(row=3, column=1, value="Reserved for Future Use")

        rows, _ = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        # Type fields blank; other fields still populate
        assert rows[0].column_label == "Col A"
        assert rows[0].column_technical == "COL_A"
        assert rows[0].data_type == ""
        assert rows[0].length is None
        assert rows[0].data_type_raw == ""

    def test_rich_tab_unparseable_type_emits_warning_issue(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "WeirdType"
        _make_rich_tab(
            ws, labels=["Col A"],
            data_types=["???junk???"],
            technicals=["COL_A"],
        )
        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        # Row still emitted, raw preserved, parsed fields blank
        assert rows[0].data_type == ""
        assert rows[0].length is None
        assert rows[0].data_type_raw == "???junk???"
        # One warning issue for that raw string
        warnings = [i for i in issues if i.issue_type == "TYPE_PARSE_WARNING"]
        assert len(warnings) == 1
        assert warnings[0].detail == "???junk???"

    def test_rich_tab_asterisk_in_label_stripped(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "StarLabel"
        _make_rich_tab(
            ws, labels=["*Required Label"],
            data_types=["VARCHAR2(80)"],
            technicals=["REQ_LABEL"],
        )
        rows, _ = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        # Asterisk stripped by normalize_label; required comes from R4 row anyway
        assert rows[0].column_label == "Required Label"

    def test_rich_tab_data_starts_at_col_a(self, tmp_path):
        """Rich tab where col A of the Tier-1 row is a technical name, not a sentinel.

        Standard Oracle structure puts "Column name of the Table X" in col A.
        When col A is itself UPPER_SNAKE_CASE (no sentinel), _extract_rich must
        include it as data column 1.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "NO_SENTINEL_TAB"
        # Row 1-3: non-UPPER_SNAKE_CASE text so detection doesn't false-fire.
        ws.cell(row=1, column=1, value="Some Import")
        ws.cell(row=2, column=1, value="Version 26B")
        ws.cell(row=3, column=1, value="Fill in required fields")
        # Row 4: tier-1 header with no sentinel — data starts at col A
        ws.cell(row=4, column=1, value="FIELD_ALPHA")
        ws.cell(row=4, column=2, value="FIELD_BETA")
        ws.cell(row=4, column=3, value="FIELD_GAMMA")

        rows, issues = extract_tab_rows(
            ws, file_stem="SomeTemplate", release="26B"
        )

        assert issues == []
        assert len(rows) == 3
        assert rows[0].position == 1
        assert rows[0].column_technical == "FIELD_ALPHA"
        assert rows[1].position == 2
        assert rows[1].column_technical == "FIELD_BETA"
        assert rows[2].position == 3
        assert rows[2].column_technical == "FIELD_GAMMA"


class TestExtractTabRowsCommentFallback:
    """Cell comments on header cells can carry Oracle metadata.

    Oracle ships some FBDI templates with TECHNICAL_NAME / data type /
    description embedded in a cell comment rather than in dedicated
    metadata rows. Example (PayablesStandardInvoiceImportTemplate,
    "Landed Cost Enabled"):

        LCM_ENABLED_FLAG
        VARCHAR2(1 CHAR)
        Flag which indicates whether invoice line is enabled for...
    """

    def test_thin_tab_recovers_technical_and_type_from_comment(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PAY_INV_LINES_IFACE"
        _make_thin_tab(ws, ["Landed Cost Enabled", "Plain Column"])
        ws.cell(row=4, column=1).comment = Comment(
            "LCM_ENABLED_FLAG\nVARCHAR2(1 CHAR)\nFlag which indicates...",
            "oracle",
        )

        rows, issues = extract_tab_rows(
            ws, file_stem="PayablesStandardInvoiceImportTemplate", release="26B",
        )

        assert issues == []
        assert len(rows) == 2
        r0 = rows[0]
        assert r0.column_label == "Landed Cost Enabled"
        assert r0.column_technical == "LCM_ENABLED_FLAG"
        assert r0.data_type == "VARCHAR2"
        assert r0.length == 1
        assert r0.data_type_raw == "VARCHAR2(1 CHAR)"
        # Second column has no comment — fields stay blank
        assert rows[1].column_technical == ""
        assert rows[1].data_type == ""

    def test_rich_tab_comment_fills_missing_type(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "RICH_WITH_GAP"
        _make_rich_tab(
            ws,
            labels=["Col A"],
            data_types=[""],
            technicals=["COL_A"],
        )
        ws.cell(row=5, column=2).comment = Comment(
            "COL_A\nVARCHAR2(80)\nDescription",
            "oracle",
        )

        rows, _ = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        # Row-based provided technical; comment fills the type gap
        assert rows[0].column_technical == "COL_A"
        assert rows[0].data_type == "VARCHAR2"
        assert rows[0].length == 80
        assert rows[0].data_type_raw == "VARCHAR2(80)"

    def test_rich_tab_label_row_comment_fills_missing_type(self):
        """Oracle sometimes attaches the metadata comment to the
        user-facing label cell (above the tier-1 technical row) rather
        than the tier-1 cell itself — rich-tab fallback must probe both.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "RICH_LABEL_COMMENT"
        _make_rich_tab(
            ws,
            labels=["Col A"],
            data_types=[""],
            technicals=["COL_A"],
        )
        # Label row is at row 1 (header_row=5, header_row-4=1)
        ws.cell(row=1, column=2).comment = Comment(
            "COL_A\nVARCHAR2(120)\nVisible label tooltip",
            "oracle",
        )

        rows, _ = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].column_technical == "COL_A"
        assert rows[0].data_type == "VARCHAR2"
        assert rows[0].length == 120
        assert rows[0].data_type_raw == "VARCHAR2(120)"

    def test_description_only_comment_emits_no_warning(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "TabWithNotes"
        _make_thin_tab(ws, ["Field Name", "Other Column"])
        ws.cell(row=4, column=1).comment = Comment(
            "This is a free-form note.\nNo technical name or type here.",
            "author",
        )

        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].column_technical == ""
        assert rows[0].data_type == ""
        assert [i for i in issues if i.issue_type == "COMMENT_PARSE_WARNING"] == []

    def test_malformed_length_in_comment_falls_back_to_bare_type(self):
        """When Oracle ships a malformed length spec like 'VARCHAR2(BOGUS_LEN)',
        the type-prefix extractor recovers the bare type 'VARCHAR2' (length
        unknown) rather than discarding the whole line.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "BadLength"
        _make_thin_tab(ws, ["Field", "Other"])
        ws.cell(row=4, column=1).comment = Comment(
            "BAD_FIELD_NAME\nVARCHAR2(BOGUS_LEN)\nA description",
            "author",
        )

        rows, _ = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].column_technical == "BAD_FIELD_NAME"
        assert rows[0].data_type == "VARCHAR2"
        assert rows[0].length is None
        assert rows[0].data_type_raw == "VARCHAR2"

    def test_technical_only_comment_no_warning(self):
        """Single-line UPPER_SNAKE_CASE comment is acceptable — no warning."""
        wb = Workbook()
        ws = wb.active
        ws.title = "TechOnly"
        _make_thin_tab(ws, ["Field", "Other"])
        ws.cell(row=4, column=1).comment = Comment(
            "TECH_NAME_ONLY",
            "author",
        )

        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].column_technical == "TECH_NAME_ONLY"
        assert rows[0].data_type == ""
        assert [i for i in issues if i.issue_type == "COMMENT_PARSE_WARNING"] == []

    def test_prose_only_comment_extracts_mixed_case_technical_no_warning(self):
        """Real-world AutoInvoiceImportTemplate shape: technical name on
        line 1 (mixed-case, underscores), no type line, just prose
        description. Extract the technical; do NOT warn (it's the
        documented Oracle convention for prose-only metadata comments).
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "RA_INTERFACE_LINES_ALL"
        _make_thin_tab(ws, ["Some Field", "Other"])
        ws.cell(row=4, column=1).comment = Comment(
            "bill_customer_account_number\n\n"
            "Value used to uniquely identify the Bill-to customer account number\n"
            "of the transaction.\n\n"
            "To identify an existing customer account:\n"
            "1. Navigate to Billing Work Area.",
            "oracle",
        )

        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].column_technical == "bill_customer_account_number"
        assert rows[0].data_type == ""
        assert [i for i in issues if i.issue_type == "COMMENT_PARSE_WARNING"] == []

    def test_type_with_trailing_prose_on_same_line(self):
        """Real-world InventoryTransactionImportTemplate shape: type spec
        and description text are squished onto a single line, e.g.
        'VARCHAR2(300) This column is used to store...'. Extract just the
        type prefix and avoid firing a warning.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "INV_TRANSACTIONS_INTERFACE"
        _make_thin_tab(ws, ["Some Field", "Other"])
        ws.cell(row=4, column=1).comment = Comment(
            "EXTERNAL_SYS_TXN_REFERENCE\n\n"
            "VARCHAR2(300) This column is used to store the link to the "
            "transaction references passed by external WMS or 3PL systems.",
            "oracle",
        )

        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].column_technical == "EXTERNAL_SYS_TXN_REFERENCE"
        assert rows[0].data_type == "VARCHAR2"
        assert rows[0].length == 300
        assert rows[0].data_type_raw == "VARCHAR2(300)"
        assert [i for i in issues if i.issue_type == "COMMENT_PARSE_WARNING"] == []

    def test_prose_only_comment_with_single_word_lines_no_warning(self):
        """A prose comment whose second line happens to be a one-word token
        like 'Description' must NOT fire COMMENT_PARSE_WARNING — only lines
        that start with a known Oracle type word qualify as drift signals.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "ProseTab"
        _make_thin_tab(ws, ["Field A", "Field B"])
        ws.cell(row=4, column=1).comment = Comment(
            "my_field_name\nDescription\nof what this field does",
            "oracle",
        )

        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        assert rows[0].column_technical == "my_field_name"
        assert rows[0].data_type == ""
        assert [i for i in issues if i.issue_type == "COMMENT_PARSE_WARNING"] == []

    def test_non_oracle_type_word_does_not_false_anchor(self):
        """A non-metadata comment with one-word lines like 'Required' or
        'FlexField' must NOT be classified as Oracle type metadata.
        parse_data_type accepts bare alpha tokens permissively; comment
        mining needs the Oracle-type allowlist to reject them.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "FreeForm"
        _make_thin_tab(ws, ["Some Field", "Other"])
        ws.cell(row=4, column=1).comment = Comment(
            "Notes\nRequired field\nY/N",
            "author",
        )

        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        # No technical, no type — "Required" / "Notes" must not be picked
        assert rows[0].column_technical == ""
        assert rows[0].data_type == ""
        assert rows[0].data_type_raw == ""

    def test_mixed_case_technical_with_bare_type_in_comment(self):
        """Real-world AutoInvoiceImportTemplate shape: mixed-case technical
        name on line 1, bare type ('NUMBER') on the next non-blank line,
        description below. The type line anchors the technical extraction.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "RA_INTERFACE_DISTRIBUTIONS_ALL"
        _make_thin_tab(ws, ["Global Attribute Number1", "Other Field"])
        ws.cell(row=4, column=1).comment = Comment(
            "Global_Attribute_Number1\n\nNUMBER\n\nSegment of the Receivables Line "
            "Regional Information global descriptive flexfield...",
            "oracle",
        )

        rows, issues = extract_tab_rows(ws, file_stem="Tpl", release="26B")
        # Bare type "NUMBER" must not be picked as the technical name —
        # anchor strategy preserves the mixed-case name on the line above.
        assert rows[0].column_technical == "Global_Attribute_Number1"
        assert rows[0].data_type == "NUMBER"
        assert rows[0].data_type_raw == "NUMBER"
        assert [i for i in issues if i.issue_type == "COMMENT_PARSE_WARNING"] == []


class TestExtractTabRowsThinAsteriskColA:
    def test_asterisk_col_a_routes_to_thin_not_rich(self, tmp_path):
        """Regression: CST_I_INCOMING_TXN_COSTS-style tabs.

        Row 4 is '*TRANSACTION_COST_IDENTIFIER | COST_COMPONENT_CODE | COST'.
        _is_tier1_header previously returned True (2/3 UPPER_SNAKE) and routed
        to _extract_rich, which skipped col A — TRANSACTION_COST_IDENTIFIER was
        lost and the remaining columns were numbered 1, 2 instead of 2, 3.

        Fix: _is_tier1_header returns False when any cell has an asterisk prefix,
        routing to _extract_thin which includes all columns.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "CST_I_INCOMING_TXN_COSTS"
        ws.cell(row=2, column=1, value="Costs")
        ws.cell(row=3, column=1, value="* Required")
        ws.cell(row=4, column=1, value="*TRANSACTION_COST_IDENTIFIER")
        ws.cell(row=4, column=2, value="COST_COMPONENT_CODE")
        ws.cell(row=4, column=3, value="COST")

        rows, issues = extract_tab_rows(
            ws, file_stem="InventoryTransactionImportTemplate", release="26B"
        )

        assert issues == []
        assert len(rows) == 3
        # Thin-tab path: column_label populated, column_technical empty
        assert rows[0].position == 1
        assert rows[0].column_label == "TRANSACTION_COST_IDENTIFIER"  # asterisk stripped
        assert rows[0].column_technical == ""
        assert rows[0].required is True
        assert rows[1].position == 2
        assert rows[1].column_label == "COST_COMPONENT_CODE"
        assert rows[2].position == 3
        assert rows[2].column_label == "COST"


from fbdi.catalog import _compute_drift


def _row(**kwargs) -> CatalogRow:
    defaults = dict(
        release="26A", file_name="F", tab_name="T", position=1,
        column_label="L", column_technical="T", data_type="VARCHAR2",
        length=50, scale=None, data_type_raw="VARCHAR2(50)", required=False,
    )
    defaults.update(kwargs)
    return CatalogRow(**defaults)


class TestComputeDrift:
    def test_unchanged_rows_not_in_drift(self):
        old = [_row(release="26A")]
        new = [_row(release="26B")]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert drift == []

    def test_added_column(self):
        old = []
        new = [_row(release="26B")]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert len(drift) == 1
        assert drift[0].change_type == "ADDED"
        assert drift[0].col_label_old == ""
        assert drift[0].col_label_new == "L"

    def test_removed_column(self):
        old = [_row(release="26A")]
        new = []
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert len(drift) == 1
        assert drift[0].change_type == "REMOVED"
        assert drift[0].col_label_new == ""

    def test_renamed_label_only(self):
        old = [_row(release="26A", column_label="Old Name")]
        new = [_row(release="26B", column_label="New Name")]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert len(drift) == 1
        assert drift[0].change_type == "RENAMED"

    def test_technical_rename_yields_removed_and_added(self):
        # Under LCS-by-identity, a pure technical-name change is indistinguishable
        # from removal+addition: the identity keys ("tech", "OLD_NAME") and
        # ("tech", "NEW_NAME") never match, so neither row enters a matched pair.
        old = [_row(release="26A", column_label="", column_technical="OLD_NAME")]
        new = [_row(release="26B", column_label="", column_technical="NEW_NAME")]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert {d.change_type for d in drift} == {"ADDED", "REMOVED"}

    def test_type_changed_only(self):
        old = [_row(release="26A", data_type="VARCHAR2")]
        new = [_row(release="26B", data_type="NUMBER")]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert drift[0].change_type == "MODIFIED"
        assert drift[0].sub_kinds == "type"

    def test_length_changed_only(self):
        old = [_row(release="26A", length=50)]
        new = [_row(release="26B", length=100)]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert drift[0].change_type == "MODIFIED"
        assert drift[0].sub_kinds == "length"

    def test_scale_only_change_emits_modified(self):
        # Scale is an alignment metadata axis. NUMBER(18) → NUMBER(18,4) keeps
        # data_type and length identical but flips scale from None to 4 — a real
        # semantic shift (integer column gains decimal places). Must classify
        # as MODIFIED with sub_kinds="scale".
        old = [_row(release="26A", data_type="NUMBER", length=18, scale=None, data_type_raw="NUMBER(18)")]
        new = [_row(release="26B", data_type="NUMBER", length=18, scale=4, data_type_raw="NUMBER(18,4)")]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert len(drift) == 1
        assert drift[0].change_type == "MODIFIED"
        assert drift[0].sub_kinds == "scale"
        assert drift[0].scale_old == ""
        assert drift[0].scale_new == "4"

    def test_required_changed_only(self):
        old = [_row(release="26A", required=False)]
        new = [_row(release="26B", required=True)]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert drift[0].change_type == "MODIFIED"
        assert drift[0].sub_kinds == "required"

    def test_multi_metadata_kinds_emit_modified_with_joined_sub_kinds(self):
        # Multiple metadata sub-kinds change but they're all on one axis (metadata)
        # — that is MODIFIED with comma-joined sub_kinds, not MULTI.
        # MULTI requires 2+ distinct axes (label, metadata, position).
        old = [_row(release="26A", data_type="VARCHAR2", length=50, required=False)]
        new = [_row(release="26B", data_type="NUMBER", length=18, required=True)]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert drift[0].change_type == "MODIFIED"
        kinds = drift[0].sub_kinds.split(",")
        assert set(kinds) == {"type", "length", "required"}

    def test_multi_change_label_plus_metadata(self):
        # Two distinct axes change (label + metadata) → MULTI
        old = [_row(release="26A", column_label="Old Name", length=50)]
        new = [_row(release="26B", column_label="New Name", length=100)]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert drift[0].change_type == "MULTI"
        assert "length" in drift[0].sub_kinds

    def test_aligns_by_file_tab_position(self):
        old = [
            _row(release="26A", file_name="F1", tab_name="T1", position=1),
            _row(release="26A", file_name="F2", tab_name="T2", position=1),
        ]
        new = [
            _row(release="26B", file_name="F1", tab_name="T1", position=1, column_label="NEW"),
            _row(release="26B", file_name="F2", tab_name="T2", position=1),
        ]
        drift = _compute_drift(old, new, release_old="26A", release_new="26B")
        assert len(drift) == 1
        assert drift[0].file == "F1"


from fbdi.catalog import extract_file


class TestExtractFile:
    def test_extract_file_multiple_tabs(self, tmp_path):
        path = tmp_path / "MultiTab.xlsm"
        wb = Workbook()
        wb.remove(wb.active)

        # Thin tab
        ws1 = wb.create_sheet("THIN_TAB")
        _make_thin_tab(ws1, ["*Field A", "Field B"])

        # Rich tab
        ws2 = wb.create_sheet("RICH_TAB")
        _make_rich_tab(
            ws2,
            labels=["Col A", "Col B"],
            data_types=["VARCHAR2(50)", "NUMBER(10)"],
            required_flags=["Required", "Optional"],
            technicals=["COL_A", "COL_B"],
        )
        wb.save(path)

        rows, issues = extract_file(path, release="26B")
        tabs = {r.tab_name for r in rows}
        assert tabs == {"THIN_TAB", "RICH_TAB"}
        # Thin tab contributes 2 rows, rich tab contributes 2 rows
        assert len(rows) == 4
        assert issues == []

    def test_extract_file_skips_instruction_tabs(self, tmp_path):
        from fbdi.config import SKIP_TABS
        path = tmp_path / "WithInstructions.xlsm"
        wb = Workbook()
        wb.remove(wb.active)
        for name in list(SKIP_TABS)[:2]:
            ws = wb.create_sheet(name)
            ws.cell(row=1, column=1, value="Instruction content")
        data_ws = wb.create_sheet("DATA_TAB")
        _make_thin_tab(data_ws, ["*Field One", "Field Two"])
        wb.save(path)

        rows, issues = extract_file(path, release="26B")
        tabs = {r.tab_name for r in rows}
        assert tabs == {"DATA_TAB"}
        assert issues == []

    def test_extract_file_load_error_yields_issue(self, tmp_path):
        path = tmp_path / "Corrupt.xlsm"
        path.write_bytes(b"not a real xlsx file")

        rows, issues = extract_file(path, release="26B")
        assert rows == []
        assert len(issues) == 1
        assert issues[0].issue_type == "FILE_ERROR"
        assert issues[0].file == "Corrupt"
        assert issues[0].tab == ""
        assert issues[0].release == "26B"


class TestWriteMasterWorkbook:
    def test_writes_release_tab_with_correct_headers(self, tmp_path):
        out = tmp_path / "Master.xlsx"
        rows_by_release = {
            "26A": [_row(release="26A", file_name="F", tab_name="T", position=1)],
        }
        _write_master_workbook(
            out,
            rows_by_release=rows_by_release,
            issues=[],
            drift=[],
            release_old=None,
            release_new="26A",
        )
        assert out.exists()
        wb = load_workbook(out)
        assert "26A" in wb.sheetnames
        assert "Issues" in wb.sheetnames
        assert "Drift" in wb.sheetnames
        ws = wb["26A"]
        headers = [c.value for c in ws[1]]
        assert headers == [
            "release", "file_name", "tab_name", "position",
            "column_label", "column_technical",
            "data_type", "length", "scale", "data_type_raw",
            "required",
        ]
        # Data row
        row2 = [c.value for c in ws[2]]
        assert row2[0] == "26A"
        assert row2[1] == "F"

    def test_writes_issues_tab(self, tmp_path):
        out = tmp_path / "Master.xlsx"
        issues = [IssueRow("26B", "F", "T", "FILE_ERROR", "boom")]
        _write_master_workbook(
            out, rows_by_release={}, issues=issues, drift=[],
            release_old=None, release_new=None,
        )
        wb = load_workbook(out)
        ws = wb["Issues"]
        headers = [c.value for c in ws[1]]
        assert headers == ["release", "file", "tab", "issue_type", "detail"]
        assert [c.value for c in ws[2]] == ["26B", "F", "T", "FILE_ERROR", "boom"]

    def test_writes_drift_tab(self, tmp_path):
        out = tmp_path / "Master.xlsx"
        drift = [DriftRow(
            file="F", tab="T", change_type="MODIFIED",
            old_position=1, new_position=1,
            col_label_old="A", col_label_new="A",
            col_technical_old="A1", col_technical_new="A1",
            data_type_old="NUMBER", data_type_new="NUMBER",
            length_old="18", length_new="18",
            scale_old="", scale_new="4",
            required_old="FALSE", required_new="FALSE",
            sub_kinds="scale",
        )]
        _write_master_workbook(
            out, rows_by_release={}, issues=[], drift=drift,
            release_old="26A", release_new="26B",
        )
        wb = load_workbook(out)
        ws = wb["Drift"]
        headers = [c.value for c in ws[1]]
        assert "col_label_26A" in headers
        assert "col_label_26B" in headers
        assert "change_type" in headers
        assert "sub_kinds" in headers
        assert "position_26A" in headers
        assert "position_26B" in headers
        assert "scale_26A" in headers
        assert "scale_26B" in headers
        # Spot-check the data row matches the new column order
        row2 = [c.value for c in ws[2]]
        assert row2[0] == "F"
        assert row2[1] == "T"
        assert row2[2] == "MODIFIED"
        assert row2[3] == 1                # position_26A
        assert row2[4] == 1                # position_26B
        # scale columns sit between length and required
        scale_a = headers.index("scale_26A")
        scale_b = headers.index("scale_26B")
        assert row2[scale_a] in ("", None)   # openpyxl reads blank cells as None
        assert row2[scale_b] == "4"
        assert row2[-1] == "scale"

    def test_idempotent_content(self, tmp_path):
        out1 = tmp_path / "M1.xlsx"
        out2 = tmp_path / "M2.xlsx"
        rows_by_release = {"26A": [_row(release="26A")]}
        for out in (out1, out2):
            _write_master_workbook(
                out, rows_by_release=rows_by_release, issues=[], drift=[],
                release_old=None, release_new="26A",
            )
        wb1 = load_workbook(out1)
        wb2 = load_workbook(out2)
        assert wb1.sheetnames == wb2.sheetnames
        for sn in wb1.sheetnames:
            r1 = [[c.value for c in row] for row in wb1[sn].iter_rows()]
            r2 = [[c.value for c in row] for row in wb2[sn].iter_rows()]
            assert r1 == r2, f"Tab {sn} content differs"

    def test_preserves_existing_release_tabs(self, tmp_path):
        """Writing release X shouldn't wipe release Y if Y was already present in the file."""
        out = tmp_path / "Master.xlsx"
        # First run: writes 26A
        _write_master_workbook(
            out, rows_by_release={"26A": [_row(release="26A")]},
            issues=[], drift=[],
            release_old=None, release_new="26A",
        )
        # Second run: writes 26B but must preserve 26A
        # (caller has loaded 26A rows from existing workbook and passes both)
        _write_master_workbook(
            out, rows_by_release={
                "26A": [_row(release="26A")],
                "26B": [_row(release="26B")],
            },
            issues=[], drift=[],
            release_old="26A", release_new="26B",
        )
        wb = load_workbook(out)
        assert "26A" in wb.sheetnames
        assert "26B" in wb.sheetnames


from fbdi.catalog import generate_catalog


def _make_rich_xlsm(path: Path, tab_name: str, labels, types, techs, required):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=tab_name)
    _make_rich_tab(
        ws, labels=labels, data_types=types,
        required_flags=required, technicals=techs,
    )
    wb.save(path)


class TestGenerateCatalog:
    def test_end_to_end_single_release(self, tmp_path):
        release_dir = tmp_path / "baselines" / "TESTA" / "originals"
        release_dir.mkdir(parents=True)
        _make_rich_xlsm(
            release_dir / "Fake.xlsm",
            tab_name="MY_TAB",
            labels=["Col A", "Col B"],
            types=["VARCHAR2(50)", "NUMBER(18)"],
            techs=["COL_A", "COL_B"],
            required=["Required", "Optional"],
        )
        master = tmp_path / "Catalog.xlsx"
        generate_catalog(
            release="TESTA",
            baselines_dir=release_dir,
            master_path=master,
            timeout=60,
        )
        assert master.exists()
        wb = load_workbook(master)
        assert "TESTA" in wb.sheetnames
        assert "Issues" in wb.sheetnames
        assert "Drift" in wb.sheetnames
        ws = wb["TESTA"]
        data = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        assert len(data) == 2
        drift_ws = wb["Drift"]
        drift_rows = [
            r for r in drift_ws.iter_rows(min_row=2)
            if any(c.value is not None for c in r)
        ]
        assert drift_rows == []

    def test_end_to_end_two_releases_drift_classifications(self, tmp_path):
        testa_dir = tmp_path / "baselines" / "TESTA" / "originals"
        testa_dir.mkdir(parents=True)
        _make_rich_xlsm(
            testa_dir / "Fake.xlsm",
            tab_name="MY_TAB",
            labels=["Col A", "Col B", "Col C"],
            types=["VARCHAR2(50)", "NUMBER(18)", "DATE"],
            techs=["COL_A", "COL_B", "COL_C"],
            required=["Required", "Optional", "Optional"],
        )
        testb_dir = tmp_path / "baselines" / "TESTB" / "originals"
        testb_dir.mkdir(parents=True)
        _make_rich_xlsm(
            testb_dir / "Fake.xlsm",
            tab_name="MY_TAB",
            labels=["Col A", "Col B", "Col C", "Col D"],
            types=["VARCHAR2(50)", "NUMBER(32)", "DATE", "VARCHAR2(10)"],
            techs=["COL_A_RENAMED", "COL_B", "COL_C", "COL_D"],
            required=["Required", "Optional", "Required", "Optional"],
        )
        master = tmp_path / "Catalog.xlsx"
        generate_catalog(
            release="TESTA", baselines_dir=testa_dir,
            master_path=master, timeout=60,
        )
        generate_catalog(
            release="TESTB", baselines_dir=testb_dir,
            master_path=master, timeout=60,
        )
        wb = load_workbook(master)
        assert "TESTA" in wb.sheetnames
        assert "TESTB" in wb.sheetnames
        drift_ws = wb["Drift"]
        drift = [[c.value for c in row] for row in drift_ws.iter_rows(min_row=2)]
        # Schema: file, tab, change_type, position_old, position_new, ..., sub_kinds
        change_types = {r[2] for r in drift}
        sub_kinds_seen = {r[-1] for r in drift if r[-1]}
        # COL_A → COL_A_RENAMED: identity key changes → REMOVED + ADDED
        assert "REMOVED" in change_types
        # COL_D fresh in TESTB
        assert "ADDED" in change_types
        # COL_B length 18→32 and COL_C required Optional→Required → MODIFIED
        assert "MODIFIED" in change_types
        assert "length" in sub_kinds_seen
        assert "required" in sub_kinds_seen

    def test_end_to_end_file_error_in_issues(self, tmp_path):
        release_dir = tmp_path / "baselines" / "TESTA" / "originals"
        release_dir.mkdir(parents=True)
        (release_dir / "Broken.xlsm").write_bytes(b"not a real xlsx file")
        master = tmp_path / "Catalog.xlsx"
        generate_catalog(
            release="TESTA", baselines_dir=release_dir,
            master_path=master, timeout=60,
        )
        wb = load_workbook(master)
        ws = wb["Issues"]
        issue_rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        assert any(r[3] == "FILE_ERROR" and r[1] == "Broken" for r in issue_rows)

    def test_end_to_end_idempotent(self, tmp_path):
        release_dir = tmp_path / "baselines" / "TESTA" / "originals"
        release_dir.mkdir(parents=True)
        _make_rich_xlsm(
            release_dir / "Fake.xlsm",
            tab_name="MY_TAB",
            labels=["Col A"],
            types=["VARCHAR2(50)"],
            techs=["COL_A"],
            required=["Required"],
        )
        master = tmp_path / "Catalog.xlsx"
        generate_catalog(release="TESTA", baselines_dir=release_dir,
                         master_path=master, timeout=60)
        wb1 = load_workbook(master)
        snap1 = {sn: [[c.value for c in row] for row in wb1[sn].iter_rows()]
                 for sn in wb1.sheetnames}
        generate_catalog(release="TESTA", baselines_dir=release_dir,
                         master_path=master, timeout=60)
        wb2 = load_workbook(master)
        snap2 = {sn: [[c.value for c in row] for row in wb2[sn].iter_rows()]
                 for sn in wb2.sheetnames}
        assert snap1 == snap2


import time

from fbdi.catalog import _run_file_in_subprocess


class TestRunFileInSubprocessLargePayload:
    def test_large_file_does_not_deadlock(self, tmp_path):
        # Regression for the Windows pipe-buffer deadlock that caused
        # ChangeOrderImportTemplate and ItemImportTemplate to report
        # bogus TIMEOUTs in the 26A/26B catalog. Build a rich tab with
        # ~499 data columns (capped by _MAX_COL=500 in catalog.py). The
        # resulting pickled CatalogRow payload (~150 KB) comfortably
        # exceeds the ~64 KB pipe buffer, so this exercises the
        # drain-before-join path added to run_worker().
        # Build the sheet with 1500 writes so that if _MAX_COL is ever
        # raised, the test still passes (we only assert on the rows
        # that catalog actually returns, not on the raw sheet size).
        wb = Workbook()
        ws = wb.active
        ws.title = "BIG_TAB"
        n_sheet = 1500
        # header_row = 5; metadata rows above it
        ws.cell(row=2, column=1, value="Name")
        ws.cell(row=3, column=1, value="Data Type")
        ws.cell(row=4, column=1, value="Required or Optional")
        ws.cell(row=5, column=1, value="Column name of the Table BIG_TAB")
        for i in range(1, n_sheet + 1):
            ws.cell(row=2, column=i + 1, value=f"Label {i}")
            ws.cell(row=3, column=i + 1, value="VARCHAR2(80)")
            ws.cell(row=4, column=i + 1, value="Required" if i % 2 else "Optional")
            ws.cell(row=5, column=i + 1, value=f"COL_{i:04d}")
        path = tmp_path / "BigTemplate.xlsm"
        wb.save(path)

        t0 = time.perf_counter()
        rows, issues = _run_file_in_subprocess(path, release="26A", timeout=30)
        elapsed = time.perf_counter() - t0

        assert issues == []
        # catalog caps column scanning at _MAX_COL=500 → 499 data columns
        assert len(rows) >= 499, f"got only {len(rows)} rows"
        assert rows[0].column_technical == "COL_0001"
        assert rows[-1].column_technical == f"COL_{len(rows):04d}"
        assert elapsed < 20, f"returned in {elapsed:.1f}s (possible deadlock)"
