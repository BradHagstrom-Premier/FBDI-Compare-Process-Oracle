"""Tests for fbdi.detect_header — dynamic header row detection."""

import pytest
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from fbdi.detect_header import detect_header_row


def _make_ws_with_headers(header_row: int, headers: list[str], extra_rows: dict[int, list] | None = None):
    """Create a worksheet with Oracle-style headers at a specific row."""
    wb = Workbook()
    ws = wb.active
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    if extra_rows:
        for row_num, values in extra_rows.items():
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row_num, column=col_idx, value=val)
    return ws


SAMPLE_HEADERS = [
    "INVOICE_NUMBER", "PO_LINE_ID", "ATTRIBUTE_CATEGORY",
    "VENDOR_NAME", "AMOUNT", "CURRENCY_CODE", "DESCRIPTION",
    "LINE_NUMBER", "TAX_CODE", "UNIT_PRICE",
]


class TestDetectHeaderRow:
    def test_headers_at_row_1(self):
        ws = _make_ws_with_headers(1, SAMPLE_HEADERS)
        assert detect_header_row(ws) == 1

    def test_headers_at_row_3(self):
        ws = _make_ws_with_headers(3, SAMPLE_HEADERS, extra_rows={
            1: ["Instructions for importing data"],
            2: ["Please fill in all required fields below"],
        })
        assert detect_header_row(ws) == 3

    def test_headers_at_row_4(self):
        ws = _make_ws_with_headers(4, SAMPLE_HEADERS, extra_rows={
            1: ["Oracle FBDI Template"],
            2: ["Version 25D"],
            3: ["Fill in the columns below with your data"],
        })
        assert detect_header_row(ws) == 4

    def test_headers_at_row_5(self):
        ws = _make_ws_with_headers(5, SAMPLE_HEADERS, extra_rows={
            1: ["Work Definition Template"],
            2: [""],
            3: ["Instructions"],
            4: ["Required fields are marked with *"],
        })
        assert detect_header_row(ws) == 5

    def test_headers_at_row_8(self):
        ws = _make_ws_with_headers(8, SAMPLE_HEADERS, extra_rows={
            1: ["Item Import Template"],
            2: [""],
            3: ["Section 1"],
            4: ["Notes about this template"],
            5: [""],
            6: [""],
            7: ["Additional instructions here"],
        })
        assert detect_header_row(ws) == 8

    def test_headers_at_row_10(self):
        ws = _make_ws_with_headers(10, SAMPLE_HEADERS, extra_rows={
            1: ["Scp Template"],
            4: ["Some text", "More text"],
        })
        assert detect_header_row(ws) == 10

    def test_headers_at_row_16(self):
        ws = _make_ws_with_headers(16, SAMPLE_HEADERS, extra_rows={
            1: ["XLA Mappings Import Template"],
            4: ["Field descriptions below"],
        })
        assert detect_header_row(ws) == 16

    def test_empty_worksheet_returns_none(self):
        wb = Workbook()
        ws = wb.active
        assert detect_header_row(ws) is None

    def test_instruction_text_not_detected_as_headers(self):
        """Instruction rows with mixed case and natural language should not score high."""
        ws = _make_ws_with_headers(4, SAMPLE_HEADERS, extra_rows={
            1: ["This template is used for importing invoice data into Oracle"],
            2: ["Please ensure all required columns are populated before upload"],
            3: ["Fields marked with * are mandatory"],
        })
        assert detect_header_row(ws) == 4

    def test_highest_scoring_row_wins(self):
        """When multiple rows contain uppercase text, the header row should win."""
        wb = Workbook()
        ws = wb.active
        # Row 2: some uppercase words but not Oracle-style
        for col, val in enumerate(["STATUS", "Name", "Type", "Notes"], start=1):
            ws.cell(row=2, column=col, value=val)
        # Row 4: proper Oracle headers
        for col, val in enumerate(SAMPLE_HEADERS, start=1):
            ws.cell(row=4, column=col, value=val)
        assert detect_header_row(ws) == 4

    def test_merged_cells_above_header(self):
        """Merged cells (instruction rows) should not confuse detection."""
        wb = Workbook()
        ws = wb.active
        # Simulate merged instruction row by putting text in A1 only
        ws.cell(row=1, column=1, value="Instructions for this FBDI template - please read carefully")
        ws.merge_cells("A1:J1")
        # Headers at row 4
        for col, val in enumerate(SAMPLE_HEADERS, start=1):
            ws.cell(row=4, column=col, value=val)
        assert detect_header_row(ws) == 4

    def test_few_columns_still_detected(self):
        """Templates with only 3-4 columns should still be detected."""
        ws = _make_ws_with_headers(4, ["FIELD_A", "FIELD_B", "FIELD_C"])
        assert detect_header_row(ws) == 4

    def test_below_minimum_cells_returns_none(self):
        """Rows with fewer than 2 non-empty cells should be skipped."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="HEADER_A")
        # Only 1 cell — below MIN_CELLS=2 threshold
        assert detect_header_row(ws) is None

    def test_two_column_header_detected(self):
        """Headers with exactly 2 columns should now be detected (MIN_CELLS=2)."""
        ws = _make_ws_with_headers(4, ["FIELD_A", "FIELD_B"])
        assert detect_header_row(ws) == 4

    def test_legend_row_does_not_beat_wider_header(self):
        """A narrow legend annotation row must lose to a wider mixed-case header row.

        Oracle Upload-style templates (e.g. UploadCustomersTemplate) have a
        2-cell legend row ('* Required', '** Conditionally required') before the
        real header row.  The legend row used to win Tier 2 because its
        per-row fill_ratio was 2/2 = 1.0, identical to the 12-cell header.
        After the fix, fill_ratio is normalised to the sheet-wide column span,
        so the legend row scores 2/12 ≈ 0.17 while the header row scores 1.0.
        """
        wb = Workbook()
        ws = wb.active
        # Row 3: legend key (2 cells — columns 1 and 2 only)
        ws.cell(row=3, column=1, value="* Required")
        ws.cell(row=3, column=2, value="** Conditionally required")
        # Row 5: actual mixed-case header row spanning 12 columns
        mixed_headers = [
            "*Source System", "*Account Number", "Site Number", "*Person Number",
            "*First Name", "Middle Name", "*Last Name", "Primary Contact Indicator",
            "Department", "Job Title", "Responsibility Type", "Primary Responsibility Indicator",
        ]
        for col, h in enumerate(mixed_headers, start=1):
            ws.cell(row=5, column=col, value=h)
        assert detect_header_row(ws) == 5

    def test_sparse_data_row_does_not_beat_wide_mixed_header(self):
        """Regression: FA_ADJUSTMENTS_T — sparse data row must not win over real header.

        Row 4 (header): 22 columns, 10 UPPER_SNAKE + 12 human-readable labels
          → snake=0.45, fill=1.0  — Tier-1 threshold miss, falls to Tier-2
        Row 5 (data):   3 non-empty cells, 2 UPPER_SNAKE + 1 mixed
          → snake=0.67, fill=3/22=0.14 — OLD: Tier-1 winner (only candidate)
                                         NEW: excluded (fill < TIER1_MIN_FILL)

        Before the fix, row 5 was the only Tier-1 candidate and won, producing a
        catalog with sample data values (BMRX / OPS CORP / RECLASS) as column names.
        After the fix, row 5 is excluded from Tier-1 by the fill floor; row 4 wins
        Tier-2 on fill=1.0 and header_like=1.0.
        """
        wb = Workbook()
        ws = wb.active
        # Row 2: title (1 cell, not a candidate)
        ws.cell(row=2, column=1, value="Mass Financial Trans")
        # Row 3: legend (1 cell, not a candidate)
        ws.cell(row=3, column=1, value="* Required")
        # Row 4: real header — mixed human-readable labels and ATTRIBUTE* fields.
        # 12 human-readable (non-UPPER_SNAKE) + 10 UPPER_SNAKE = 10/22 ≈ 0.45 snake,
        # fill = 22/22 = 1.0
        human_labels = [
            "*Batch Name", "*Asset Book", "*Asset Number", "*Transaction Type",
            "Transaction Name", "Amortize", "Cost", "Date Placed",
            "Prorate Convention", "Depreciate", "Salvage Value", "Method",
        ]
        attr_fields = [f"ATTRIBUTE{i}" for i in range(1, 11)]
        for col, val in enumerate(human_labels + attr_fields, start=1):
            ws.cell(row=4, column=col, value=val)
        # Row 5: sparse data row — 3 non-empty, 2 are UPPER_SNAKE → snake=0.67,
        # fill = 3/22 = 0.136 < TIER1_MIN_FILL=0.15
        ws.cell(row=5, column=1, value="BMRX")       # UPPER_SNAKE
        ws.cell(row=5, column=2, value="OPS CORP")   # has space, not UPPER_SNAKE
        ws.cell(row=5, column=3, value="RECLASS")    # UPPER_SNAKE

        assert detect_header_row(ws) == 4

    def test_phantom_wide_columns_do_not_suppress_detection(self):
        """Headers should be detected even when max_column is phantom-wide."""
        wb = Workbook()
        ws = wb.active
        # Place headers at row 4 in columns 1-6 only
        for col, val in enumerate(SAMPLE_HEADERS[:6], start=1):
            ws.cell(row=4, column=col, value=val)
        # Simulate phantom wide column by writing a single value far out,
        # then clearing it — openpyxl still reports high max_column
        ws.cell(row=20, column=400, value="phantom")
        ws.cell(row=20, column=400, value=None)
        # Detection should still find row 4
        assert detect_header_row(ws) == 4
