"""Tests for fbdi.diagnose — header detection diagnostic module."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from fbdi.diagnose import DiagnosticRow, diagnose_file, write_diagnostic_report
from fbdi.config import MAX_FILE_SIZE_BYTES, SKIP_TABS


def _make_workbook(path: Path, sheets: dict) -> None:
    """Create a workbook with given sheet name -> list of (row, col, value) triples."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, cells in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row, col, value in cells:
            ws.cell(row=row, column=col, value=value)
    wb.save(path)


def _make_fbdi_workbook(path: Path, sheet_name: str, headers: list[str], header_row: int = 4) -> None:
    """Create a workbook with UPPER_SNAKE_CASE headers at header_row."""
    cells = [(header_row, col + 1, h) for col, h in enumerate(headers)]
    _make_workbook(path, {sheet_name: cells})


ORACLE_HEADERS = ["INVOICE_NUMBER", "PO_LINE_ID", "VENDOR_NAME", "AMOUNT", "CURRENCY_CODE"]


class TestDiagnoseFile:
    def test_clean_oracle_headers_detected(self, tmp_path):
        """A tab with UPPER_SNAKE_CASE headers at row 4 is DETECTED."""
        f = tmp_path / "Template.xlsm"
        _make_fbdi_workbook(f, "Data", ORACLE_HEADERS, header_row=4)

        rows = diagnose_file(f)
        data_rows = [r for r in rows if r.fbdi_tab == "Data"]
        assert len(data_rows) == 1
        assert data_rows[0].detection_result == "DETECTED"
        assert data_rows[0].detected_row == 4
        assert data_rows[0].fbdi_file == "Template"

    def test_no_header_tab_returns_no_header(self, tmp_path):
        """A tab with only 1 non-empty cell (below MIN_CELLS=2) returns NO_HEADER."""
        f = tmp_path / "Template.xlsm"
        cells = [(1, 1, "a")]
        _make_workbook(f, {"Data": cells})

        rows = diagnose_file(f)
        data_rows = [r for r in rows if r.fbdi_tab == "Data"]
        assert len(data_rows) == 1
        assert data_rows[0].detection_result == "NO_HEADER"
        assert data_rows[0].detected_row is None

    def test_skip_tab_returns_skipped_tab(self, tmp_path):
        """A tab in SKIP_TABS is recorded as SKIPPED_TAB without running detection."""
        skip_tab = next(iter(SKIP_TABS))
        f = tmp_path / "Template.xlsm"
        _make_fbdi_workbook(f, skip_tab, ORACLE_HEADERS)

        rows = diagnose_file(f)
        skip_rows = [r for r in rows if r.fbdi_tab == skip_tab]
        assert len(skip_rows) == 1
        assert skip_rows[0].detection_result == "SKIPPED_TAB"
        assert skip_rows[0].detected_row is None

    def test_oversized_file_returns_file_too_large(self, tmp_path):
        """A file exceeding MAX_FILE_SIZE_BYTES is recorded as FILE_TOO_LARGE."""
        f = tmp_path / "HugeFile.xlsm"
        f.write_bytes(b"0" * (MAX_FILE_SIZE_BYTES + 1))

        rows = diagnose_file(f)
        assert len(rows) == 1
        assert rows[0].detection_result == "FILE_TOO_LARGE"
        assert rows[0].fbdi_file == "HugeFile"
        assert rows[0].fbdi_tab == ""

    def test_corrupt_file_returns_file_error(self, tmp_path):
        """A corrupt/unreadable file produces a FILE_ERROR row."""
        f = tmp_path / "Corrupt.xlsm"
        f.write_bytes(b"not a zip file at all")

        rows = diagnose_file(f)
        assert len(rows) == 1
        assert rows[0].detection_result == "FILE_ERROR"
        assert rows[0].fbdi_file == "Corrupt"

    def test_no_header_row_has_best_score(self, tmp_path):
        """NO_HEADER rows have best_score populated (may be 0.0 if no candidates)."""
        f = tmp_path / "Template.xlsm"
        # 1 cell only — below MIN_CELLS=2, no candidates found by _scan_rows
        cells = [(1, 1, "a")]
        _make_workbook(f, {"Data": cells})

        rows = diagnose_file(f)
        data_rows = [r for r in rows if r.fbdi_tab == "Data"]
        assert data_rows[0].detection_result == "NO_HEADER"
        assert data_rows[0].best_score is not None  # 0.0 is valid


class TestWriteDiagnosticReport:
    def test_output_has_correct_headers(self, tmp_path):
        """Written .xlsx has the 7 required column headers."""
        rows = [
            DiagnosticRow(
                fbdi_file="Template",
                fbdi_tab="Data",
                detection_result="DETECTED",
                detected_row=4,
                best_score=0.85,
                failure_reason="",
                notes="",
            )
        ]
        output = tmp_path / "report.xlsx"
        write_diagnostic_report(rows, output)

        wb = load_workbook(output, read_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 8)]
        wb.close()

        assert headers[0] == "FBDI File"
        assert headers[1] == "FBDI Tab"
        assert headers[2] == "Detection Result"
        assert headers[3] == "Detected Row"
        assert headers[4] == "Best Score"
        assert headers[5] == "Failure Reason"
        assert headers[6] == "Notes"

    def test_output_has_data_rows(self, tmp_path):
        """Written .xlsx has at least one data row."""
        rows = [
            DiagnosticRow(
                fbdi_file="Template",
                fbdi_tab="Data",
                detection_result="DETECTED",
                detected_row=4,
                best_score=0.85,
                failure_reason="",
                notes="",
            )
        ]
        output = tmp_path / "report.xlsx"
        write_diagnostic_report(rows, output)

        wb = load_workbook(output, read_only=True)
        ws = wb.active
        assert ws.max_row >= 2
        wb.close()
