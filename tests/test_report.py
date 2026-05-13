"""Tests for fbdi.report — view-model construction and scope filtering."""

import pytest
from pathlib import Path

import jinja2

from fbdi.report import (
    FileSection,
    ReportContext,
    PendingBaseEntry,
    build_report_context,
)
from fbdi.align import AlignedField, Change


def _render_report(ctx, print_mode=False):
    """Render the report template with the given context. Returns HTML string."""
    template_dir = Path(__file__).parent.parent / "fbdi" / "templates"
    env = jinja2.Environment(
        loader=jinja2.FileSystemLoader(template_dir),
        autoescape=jinja2.select_autoescape(["html", "j2"]),
    )
    return env.get_template("report.html.j2").render(ctx=ctx, print_mode=print_mode)


# ---- helpers ----

def _aligned(position, label, technical, data_type=None, length=None, required=None, scale=None):
    return AlignedField(position=position, label=label, technical=technical,
                       data_type=data_type, length=length, scale=scale, required=required)


def _mapping(template, tab, applaud_table="T_X", prefix="TX1",
             module="Financials", in_base=None, status="MAPPED"):
    """Build one mapping dict entry."""
    return {
        (template, tab): {
            "applaud_table": applaud_table,
            "prefix": prefix,
            "module": module,
            "in_base": in_base,
            "status": status,
        }
    }


# ---- tests ----

class TestScopeFiltering:
    def test_unmapped_file_is_silently_excluded(self):
        catalog_old = {("UnmappedFile", "TabA"): [_aligned(1, "A", "A")]}
        catalog_new = {("UnmappedFile", "TabA"): [_aligned(1, "A", "A"),
                                                  _aligned(2, "B", "B")]}
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping={}, old_release="26A", new_release="26B",
        )
        assert ctx.file_sections == []
        assert ctx.pending_base == []

    def test_mapped_in_base_routes_to_main_body(self):
        catalog_old = {("MappedFile", "TabA"): [_aligned(1, "Old", "OLD_F")]}
        catalog_new = {("MappedFile", "TabA"): [_aligned(1, "New", "NEW_F")]}
        mapping = _mapping("MappedFile", "TabA", in_base=None)
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        assert len(ctx.file_sections) == 1
        assert ctx.pending_base == []

    def test_pending_base_routes_to_separate_section(self):
        catalog_old = {("PendingFile", "TabA"): [_aligned(1, "Old", "OLD_F")]}
        catalog_new = {("PendingFile", "TabA"): [_aligned(1, "Old", "OLD_F"),
                                                 _aligned(2, "New", "NEW_F")]}
        mapping = _mapping("PendingFile", "TabA",
                          in_base="Needs to be created in base system")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        assert ctx.file_sections == []
        assert len(ctx.pending_base) == 1
        assert ctx.pending_base[0].file == "PendingFile"
        assert ctx.pending_base[0].tab == "TabA"
        assert ctx.pending_base[0].change_count == 1


class TestApplaudFieldNameConstruction:
    def test_uses_technical_when_present(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Some Label", "FIELD_NAME",
                                              "VARCHAR2", 30, True)]}
        mapping = _mapping("F", "T", prefix="TX1")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        section = ctx.file_sections[0]
        added = section.changes_by_type["ADDED"]
        assert added[0].applaud_field_name == "TX1FIELD_NAME"

    def test_falls_back_to_underscore_joined_label_when_technical_is_none(self):
        """Applaud column names cannot contain spaces. When falling back to
        the user-facing label (no technical name available), whitespace is
        replaced with underscores so the suffix is Applaud-compatible.
        """
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Some Label!", None,
                                              None, None, True)]}
        mapping = _mapping("F", "T", prefix="TX1")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        section = ctx.file_sections[0]
        added = section.changes_by_type["ADDED"]
        # normalize_label strips '!' → "Some Label"; then space → underscore
        assert added[0].applaud_field_name == "TX1Some_Label"

    def test_landed_cost_enabled_underscore_joined(self):
        """Regression: real-world Oracle label 'Landed Cost Enabled' must
        produce an Applaud-safe field name when no technical name is set.
        """
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Landed Cost Enabled", None,
                                              None, None, True)]}
        mapping = _mapping("F", "T", prefix="TX1_")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        added = ctx.file_sections[0].changes_by_type["ADDED"]
        assert added[0].applaud_field_name == "TX1_Landed_Cost_Enabled"

    def test_thirty_char_warning_set_when_name_exceeds_limit(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "X",
                                              "COPY_LOTS_AND_SERIAL_NUMBERS_FROM_PARENT_TXN",
                                              "VARCHAR2", 1, True)]}
        mapping = _mapping("F", "T", prefix="TH8")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        added = ctx.file_sections[0].changes_by_type["ADDED"]
        assert added[0].name_exceeds_30
        assert added[0].name_length > 30


class TestLoaders:
    def test_load_catalog_release_groups_by_file_and_tab(self, tmp_path):
        from fbdi.report import load_catalog_release
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "26B"
        ws.append(["release", "file_name", "tab_name", "position",
                   "column_label", "column_technical",
                   "data_type", "length", "scale", "data_type_raw", "required"])
        ws.append(["26B", "F1", "T1", 1, "Lab", "TECH", "VARCHAR2", 30, None, "VARCHAR2(30)", "TRUE"])
        ws.append(["26B", "F1", "T1", 2, "Lab2", "TECH2", "NUMBER", 18, None, "NUMBER(18)", "FALSE"])
        ws.append(["26B", "F2", "T1", 1, "Lab", "TECH", None, None, None, "", "FALSE"])
        path = tmp_path / "cat.xlsx"
        wb.save(path)

        result = load_catalog_release(path, "26B")
        assert ("F1", "T1") in result
        assert len(result[("F1", "T1")]) == 2
        first = result[("F1", "T1")][0]
        assert first.position == 1
        assert first.technical == "TECH"
        assert first.required is True
        assert first.length == 30

    def test_load_mapping_filters_to_mapped_status(self, tmp_path):
        from fbdi.report import load_mapping
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "FBDI Mapping"
        ws.append(["FBDI Template", "FBDI Tab", "Applaud Table", "Prefix",
                   "Status", "Module", "In Base System?"])
        ws.append(["F1", "T1", "T_X", "TX1", "MAPPED", "Financials", None])
        ws.append(["F2", "T1", "T_Y", "TY1", "UNMAPPED", "SCM", None])
        ws.append(["F3", "T1", "T_Z", "TZ1", "MAPPED", "SCM", "Needs to be created in base system"])
        path = tmp_path / "mapping.xlsx"
        wb.save(path)

        result = load_mapping(path)
        assert ("F1", "T1") in result
        assert ("F3", "T1") in result   # MAPPED + pending-base still included
        assert ("F2", "T1") not in result  # UNMAPPED excluded
        assert result[("F1", "T1")]["module"] == "Financials"
        assert result[("F3", "T1")]["in_base"] == "Needs to be created in base system"


# ---- Phase 1 bug-fix tests ----

class TestNeedsReviewStatus:
    """NEEDS_REVIEW status from the mapping must propagate to FileSection."""

    def test_needs_review_status_propagates_to_file_section(self):
        catalog_old = {("F", "T"): [_aligned(1, "Old", "OLD_F")]}
        catalog_new = {("F", "T"): [_aligned(1, "New", "NEW_F")]}
        mapping = _mapping("F", "T", status="NEEDS_REVIEW")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        assert ctx.file_sections[0].status == "NEEDS_REVIEW"

    def test_mapped_status_propagates_to_file_section(self):
        catalog_old = {("F", "T"): [_aligned(1, "Old", "OLD_F")]}
        catalog_new = {("F", "T"): [_aligned(1, "New", "NEW_F")]}
        mapping = _mapping("F", "T", status="MAPPED")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        assert ctx.file_sections[0].status == "MAPPED"

    def test_needs_review_renders_warning_badge_in_html(self):
        catalog_old = {("F", "T"): [_aligned(1, "Old", "OLD_F")]}
        catalog_new = {("F", "T"): [_aligned(1, "New", "NEW_F")]}
        mapping = _mapping("F", "T", status="NEEDS_REVIEW")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        assert '<span class="needs-review-badge">' in html

    def test_mapped_does_not_render_warning_badge(self):
        catalog_old = {("F", "T"): [_aligned(1, "Old", "OLD_F")]}
        catalog_new = {("F", "T"): [_aligned(1, "New", "NEW_F")]}
        mapping = _mapping("F", "T", status="MAPPED")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        # The badge element must not appear; the CSS class exists in <style> but no <span> is emitted
        assert '<span class="needs-review-badge">' not in html

    def test_load_mapping_includes_needs_review_rows(self, tmp_path):
        from fbdi.report import load_mapping
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "FBDI Mapping"
        ws.append(["FBDI Template", "FBDI Tab", "Applaud Table", "Prefix",
                   "Status", "Module", "In Base System?"])
        ws.append(["F1", "T1", "T_X", "TX1", "NEEDS_REVIEW", "Financials", None])
        path = tmp_path / "mapping.xlsx"
        wb.save(path)

        result = load_mapping(path)
        assert ("F1", "T1") in result
        assert result[("F1", "T1")]["status"] == "NEEDS_REVIEW"


class TestRequiredNoneRendering:
    """Required=None must render as — not empty/FALSE in both ADDED and MODIFIED tables."""

    def test_required_none_renders_dash_in_added_row(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=None)]}
        mapping = _mapping("F", "T")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        # Only one ADDED field; its Required cell must render a placeholder (not empty).
        # Current broken behavior: <td class="center"></td>  (empty)
        # Fixed behavior:          <td class="center">—</td>
        # ">—<" matches content between tags, not inline text like "— 26A" in the title.
        assert ">—<" in html

    def test_required_true_still_renders_true_in_added_row(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=True)]}
        mapping = _mapping("F", "T")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        assert "TRUE" in html

    def test_required_false_still_renders_false_in_added_row(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=False)]}
        mapping = _mapping("F", "T")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        assert "FALSE" in html

    def test_old_required_none_renders_dash_not_false_in_modified_row(self):
        # Field: old required=None, new required=True → classified MODIFIED
        # The old_required cell must not show FALSE (None is not False)
        catalog_old = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=None)]}
        catalog_new = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=True)]}
        mapping = _mapping("F", "T")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        # The required change cell must NOT show "FALSE" as the old value
        # It must show the dash placeholder for None
        assert "&#x2014;" in html or "—" in html or "&mdash;" in html
        # Verify it doesn't incorrectly show FALSE → TRUE (which would be wrong)
        assert "FALSE" not in html


class TestRemovedTableApplaudType:
    """REMOVED rows must include the Applaud Type column — data is in the view-model."""

    def test_removed_row_renders_applaud_type_in_html(self):
        catalog_old = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=True)]}
        catalog_new = {("F", "T"): []}
        mapping = _mapping("F", "T")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        # Verify the view-model has applaud_type_str populated for REMOVED rows
        removed = ctx.file_sections[0].changes_by_type["REMOVED"]
        assert removed[0].applaud_type_str != ""

        # Verify the template renders the Applaud Type column in the REMOVED table
        html = _render_report(ctx)
        # The REMOVED table header must include an Applaud Type column
        assert "Applaud Type" in html

    def test_removed_row_applaud_type_value_appears_in_html(self):
        catalog_old = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=True)]}
        catalog_new = {("F", "T"): []}
        mapping = _mapping("F", "T")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        removed = ctx.file_sections[0].changes_by_type["REMOVED"]
        expected_type = removed[0].applaud_type_str  # e.g. "char 30"
        html = _render_report(ctx)
        assert expected_type in html


# ---- Phase 2 improvement tests ----

class TestOracleTypeStrCharUnit:
    """_oracle_type_str must preserve the CHAR length unit from data_type_raw."""

    def test_char_unit_preserved_when_data_type_raw_present(self):
        from fbdi.report import _oracle_type_str
        field = AlignedField(
            position=1, label="L", technical="T",
            data_type="VARCHAR2", length=30, scale=None, required=None,
            data_type_raw="VARCHAR2(30 CHAR)",
        )
        assert _oracle_type_str(field) == "VARCHAR2(30 CHAR)"

    def test_falls_back_to_reconstructed_when_data_type_raw_absent(self):
        from fbdi.report import _oracle_type_str
        field = AlignedField(
            position=1, label="L", technical="T",
            data_type="VARCHAR2", length=30, scale=None, required=None,
        )
        assert _oracle_type_str(field) == "VARCHAR2(30)"

    def test_falls_back_to_reconstructed_when_data_type_raw_empty(self):
        from fbdi.report import _oracle_type_str
        field = AlignedField(
            position=1, label="L", technical="T",
            data_type="VARCHAR2", length=30, scale=None, required=None,
            data_type_raw="",
        )
        assert _oracle_type_str(field) == "VARCHAR2(30)"


class TestSummaryTotalsRow:
    """Summary table must include a totals row summing all per-section counts."""

    def test_totals_row_sums_adds_across_sections(self):
        # Section 1: 2 ADDED fields
        cat_old_1 = {("F1", "T1"): []}
        cat_new_1 = {("F1", "T1"): [_aligned(1, "A", "A_F"), _aligned(2, "B", "B_F")]}
        # Section 2: 1 ADDED field
        cat_old_2 = {("F2", "T2"): []}
        cat_new_2 = {("F2", "T2"): [_aligned(1, "C", "C_F")]}
        catalog_old = {**cat_old_1, **cat_old_2}
        catalog_new = {**cat_new_1, **cat_new_2}
        mapping = {
            ("F1", "T1"): {"applaud_table": "T_X", "prefix": "TX1", "module": "SCM", "in_base": None, "status": "MAPPED"},
            ("F2", "T2"): {"applaud_table": "T_Y", "prefix": "TY1", "module": "SCM", "in_base": None, "status": "MAPPED"},
        }
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        # Total adds = 3; must appear in a tfoot totals row
        assert "<tfoot>" in html

    def test_totals_row_correct_add_count(self):
        catalog_old = {("F1", "T1"): [], ("F2", "T2"): []}
        catalog_new = {
            ("F1", "T1"): [_aligned(1, "A", "A_F"), _aligned(2, "B", "B_F")],
            ("F2", "T2"): [_aligned(1, "C", "C_F")],
        }
        mapping = {
            ("F1", "T1"): {"applaud_table": "T_X", "prefix": "TX1", "module": "SCM", "in_base": None, "status": "MAPPED"},
            ("F2", "T2"): {"applaud_table": "T_Y", "prefix": "TY1", "module": "SCM", "in_base": None, "status": "MAPPED"},
        }
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        # 2+1=3 total adds; the number 3 must appear inside the tfoot
        tfoot_start = html.find("<tfoot>")
        tfoot_end = html.find("</tfoot>")
        assert tfoot_start != -1
        tfoot_html = html[tfoot_start:tfoot_end]
        assert ">3<" in tfoot_html


class TestCopyAndGuidance:
    """MULTI must be renamed; SHIFTED, MULTI, and Required must have guidance notes."""

    def _multi_ctx(self):
        # MULTI: label + metadata both change (2 axes → MULTI)
        catalog_old = {("F", "T"): [_aligned(1, "OldLabel", "FIELD", "VARCHAR2", 30)]}
        catalog_new = {("F", "T"): [_aligned(1, "NewLabel", "FIELD", "NUMBER", 18)]}
        return build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=_mapping("F", "T"), old_release="26A", new_release="26B",
        )

    def _shifted_ctx(self):
        # SHIFTED: only position changes
        catalog_old = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30)]}
        catalog_new = {("F", "T"): [_aligned(2, "Label", "FIELD", "VARCHAR2", 30)]}
        return build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=_mapping("F", "T"), old_release="26A", new_release="26B",
        )

    def _added_ctx(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30, required=True)]}
        return build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=_mapping("F", "T"), old_release="26A", new_release="26B",
        )

    def test_multi_renders_combined_changes_heading(self):
        html = _render_report(self._multi_ctx())
        assert "Combined changes" in html

    def test_multi_does_not_render_multi_axis_heading(self):
        html = _render_report(self._multi_ctx())
        assert "Multi-axis changes" not in html

    def test_shifted_block_has_guidance_note(self):
        html = _render_report(self._shifted_ctx())
        # A guidance note explaining shifts are informational must be present
        assert "informational" in html.lower() or "no db" in html.lower() or "position shift" in html.lower()

    def test_multi_block_has_guidance_note(self):
        html = _render_report(self._multi_ctx())
        # A guidance note for MULTI must be present (explaining to review actions individually)
        assert "individually" in html.lower() or "combined" in html.lower()

    def test_added_block_has_required_column_footnote(self):
        html = _render_report(self._added_ctx())
        # A footnote element (class required-note) must appear below the ADDED table.
        # The class does not currently exist in the template; this test fails until added.
        assert 'class="required-note"' in html


class TestAccessibility:
    """Action checkboxes must have aria-label; change table th must have scope=col."""

    def test_added_checkboxes_have_aria_label_with_field_name(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Label", "MYFIELD", "VARCHAR2", 30)]}
        mapping = _mapping("F", "T", prefix="TX1")
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=mapping, old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        # The checkbox for this ADDED row must include aria-label containing the field name
        assert 'aria-label=' in html
        assert "TX1MYFIELD" in html  # field name must appear in aria-label context

    def test_change_table_th_have_scope_col(self):
        catalog_old = {("F", "T"): []}
        catalog_new = {("F", "T"): [_aligned(1, "Label", "FIELD", "VARCHAR2", 30)]}
        ctx = build_report_context(
            catalog_old=catalog_old, catalog_new=catalog_new,
            mapping=_mapping("F", "T"), old_release="26A", new_release="26B",
        )
        html = _render_report(ctx)
        assert 'scope="col"' in html
