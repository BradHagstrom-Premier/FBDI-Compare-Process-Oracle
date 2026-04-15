"""Tests for fbdi.clear — smart FBDI template clearing."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from fbdi.clear import clear_workbook


UPPER_SNAKE_HEADERS = [
    "TRANSACTION_TYPE", "BATCH_ID", "ORGANIZATION_CODE",
    "CHANGE_NOTICE", "CHANGE_NAME", "DESCRIPTION",
]

HUMAN_HEADERS = [
    "*Transaction Type", "*Batch ID", "Organization Code",
    "Change Number", "Change Name", "Description",
]

SAMPLE_DATA_ROW = ["Create Order", "10017", "V1-org", "eco test 1", "Test Change", "A description"]


def _make_fbdi_workbook(
    path: Path,
    header_row: int,
    headers: list[str],
    data_rows: list[list] | None = None,
    extra_rows: dict[int, list] | None = None,
    include_instructions_sheet: bool = True,
):
    """Create a test FBDI workbook with headers and sample data."""
    wb = Workbook()
    if include_instructions_sheet:
        ws_instr = wb.active
        ws_instr.title = "Instructions and CSV Generation"
        ws_instr.cell(row=1, column=1, value="Instructions")
        ws_instr.cell(row=5, column=1, value="This row should NOT be cleared")
        ws_instr.cell(row=6, column=1, value="Neither should this one")
    else:
        wb.remove(wb.active)

    ws = wb.create_sheet(title="DATA_INTERFACE")
    if extra_rows:
        for row_num, values in extra_rows.items():
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row_num, column=col_idx, value=val)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    if data_rows:
        for offset, row_data in enumerate(data_rows, start=1):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=header_row + offset, column=col_idx, value=val)

    wb.save(path)
    wb.close()


class TestClearWorkbook:
    def test_headers_at_row_4_preserved(self, tmp_path):
        """Standard pattern: headers at row 4, data at row 5+."""
        src = tmp_path / "Template.xlsx"
        dst = tmp_path / "Template_cleared.xlsx"
        _make_fbdi_workbook(
            src,
            header_row=4,
            headers=HUMAN_HEADERS,
            data_rows=[SAMPLE_DATA_ROW, SAMPLE_DATA_ROW],
            extra_rows={2: ["Invoice Headers"], 3: ["* Required"]},
        )

        skipped = clear_workbook(str(src), str(dst))

        wb = load_workbook(dst, read_only=True)
        ws = wb["DATA_INTERFACE"]
        # Headers at row 4 preserved
        assert ws.cell(row=4, column=1).value == "*Transaction Type"
        assert ws.cell(row=4, column=2).value == "*Batch ID"
        # Metadata above headers preserved
        assert ws.cell(row=2, column=1).value == "Invoice Headers"
        assert ws.cell(row=3, column=1).value == "* Required"
        # Data rows cleared
        assert ws.cell(row=5, column=1).value is None
        assert ws.cell(row=6, column=1).value is None
        assert skipped == []
        wb.close()

    def test_headers_at_row_5_preserved(self, tmp_path):
        """Pattern 1: technical UPPER_SNAKE headers at row 5."""
        src = tmp_path / "Template.xlsx"
        dst = tmp_path / "Template_cleared.xlsx"
        _make_fbdi_workbook(
            src,
            header_row=5,
            headers=UPPER_SNAKE_HEADERS,
            data_rows=[SAMPLE_DATA_ROW, SAMPLE_DATA_ROW],
            extra_rows={
                1: ["Name", "Transaction Type", "Batch ID", "Change Type", "Change Number", "Description"],
                2: ["Description", "Indicates the...", "Indicates the...", "Value that...", "Value that...", "Description of..."],
                3: ["Data Type", "VARCHAR2(10)", "NUMBER(18)", "VARCHAR2(80)", "VARCHAR2(50)", "VARCHAR2(2000)"],
                4: ["Required", "Required", "Required", "Required", "Optional", "Optional"],
            },
        )

        skipped = clear_workbook(str(src), str(dst))

        wb = load_workbook(dst, read_only=True)
        ws = wb["DATA_INTERFACE"]
        # Technical headers at row 5 preserved
        assert ws.cell(row=5, column=1).value == "TRANSACTION_TYPE"
        assert ws.cell(row=5, column=2).value == "BATCH_ID"
        # Metadata rows 1-4 preserved
        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=4, column=1).value == "Required"
        # Data rows cleared
        assert ws.cell(row=6, column=1).value is None
        assert ws.cell(row=7, column=1).value is None
        assert skipped == []
        wb.close()

    def test_headers_at_row_8_preserved(self, tmp_path):
        """Pattern 2: deeper headers at row 8 (e.g., ChangeOrderImportTemplate)."""
        src = tmp_path / "Template.xlsx"
        dst = tmp_path / "Template_cleared.xlsx"
        _make_fbdi_workbook(
            src,
            header_row=8,
            headers=UPPER_SNAKE_HEADERS,
            data_rows=[SAMPLE_DATA_ROW],
            extra_rows={
                4: HUMAN_HEADERS,
                5: ["Description"] * 6,
                6: ["VARCHAR2(10)"] * 6,
                7: ["Reserved for Future Use"],
            },
        )

        skipped = clear_workbook(str(src), str(dst))

        wb = load_workbook(dst, read_only=True)
        ws = wb["DATA_INTERFACE"]
        # Technical headers at row 8 preserved
        assert ws.cell(row=8, column=1).value == "TRANSACTION_TYPE"
        # All metadata rows 1-7 preserved
        assert ws.cell(row=4, column=1).value == "*Transaction Type"
        assert ws.cell(row=7, column=1).value == "Reserved for Future Use"
        # Data row 9 cleared
        assert ws.cell(row=9, column=1).value is None
        assert skipped == []
        wb.close()

    def test_instructions_sheet_untouched(self, tmp_path):
        """Sheet 1 (Instructions) should never be modified."""
        src = tmp_path / "Template.xlsx"
        dst = tmp_path / "Template_cleared.xlsx"
        _make_fbdi_workbook(
            src,
            header_row=4,
            headers=UPPER_SNAKE_HEADERS,
            data_rows=[SAMPLE_DATA_ROW],
        )

        clear_workbook(str(src), str(dst))

        wb = load_workbook(dst, read_only=True)
        ws_instr = wb["Instructions and CSV Generation"]
        assert ws_instr.cell(row=1, column=1).value == "Instructions"
        assert ws_instr.cell(row=5, column=1).value == "This row should NOT be cleared"
        assert ws_instr.cell(row=6, column=1).value == "Neither should this one"
        wb.close()

    def test_detection_fails_sheet_skipped_not_destroyed(self, tmp_path):
        """If detect_header_row returns None, leave the sheet untouched."""
        src = tmp_path / "Template.xlsx"
        dst = tmp_path / "Template_cleared.xlsx"

        wb = Workbook()
        ws_instr = wb.active
        ws_instr.title = "Instructions and CSV Generation"
        ws_data = wb.create_sheet(title="WeirdSheet")
        # Only 1 cell — below MIN_CELLS=2, detection will return None
        ws_data.cell(row=3, column=1, value="SOLO_HEADER")
        ws_data.cell(row=4, column=1, value="solo_data")
        wb.save(src)
        wb.close()

        skipped = clear_workbook(str(src), str(dst))

        assert "WeirdSheet" in skipped

        wb = load_workbook(dst, read_only=True)
        ws = wb["WeirdSheet"]
        # Data should be untouched since detection failed
        assert ws.cell(row=3, column=1).value == "SOLO_HEADER"
        assert ws.cell(row=4, column=1).value == "solo_data"
        wb.close()

    def test_multiple_data_sheets_cleared_independently(self, tmp_path):
        """Each sheet gets its own header detection and clearing."""
        src = tmp_path / "Template.xlsx"
        dst = tmp_path / "Template_cleared.xlsx"

        wb = Workbook()
        ws_instr = wb.active
        ws_instr.title = "Instructions and CSV Generation"

        # Sheet 2: headers at row 4
        ws1 = wb.create_sheet(title="SHEET_A")
        for col, val in enumerate(UPPER_SNAKE_HEADERS, start=1):
            ws1.cell(row=4, column=col, value=val)
        for col, val in enumerate(SAMPLE_DATA_ROW, start=1):
            ws1.cell(row=5, column=col, value=val)

        # Sheet 3: headers at row 8
        ws2 = wb.create_sheet(title="SHEET_B")
        for col, val in enumerate(HUMAN_HEADERS, start=1):
            ws2.cell(row=4, column=col, value=val)
        for col, val in enumerate(["Description"] * 6, start=1):
            ws2.cell(row=5, column=col, value=val)
        for col, val in enumerate(["VARCHAR2(10)"] * 6, start=1):
            ws2.cell(row=6, column=col, value=val)
        ws2.cell(row=7, column=1, value="Reserved")
        for col, val in enumerate(UPPER_SNAKE_HEADERS, start=1):
            ws2.cell(row=8, column=col, value=val)
        for col, val in enumerate(SAMPLE_DATA_ROW, start=1):
            ws2.cell(row=9, column=col, value=val)

        wb.save(src)
        wb.close()

        clear_workbook(str(src), str(dst))

        wb = load_workbook(dst, read_only=True)
        # Sheet A: headers at row 4, data at row 5 cleared
        ws_a = wb["SHEET_A"]
        assert ws_a.cell(row=4, column=1).value == "TRANSACTION_TYPE"
        assert ws_a.cell(row=5, column=1).value is None

        # Sheet B: headers at row 8, data at row 9 cleared
        ws_b = wb["SHEET_B"]
        assert ws_b.cell(row=4, column=1).value == "*Transaction Type"
        assert ws_b.cell(row=8, column=1).value == "TRANSACTION_TYPE"
        assert ws_b.cell(row=9, column=1).value is None
        wb.close()
