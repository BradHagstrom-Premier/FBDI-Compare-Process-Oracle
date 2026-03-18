"""Acceptance test: compare Python comparison output against VBA output.

The VBA-generated Comparison_Report_25D_26A.xlsx has invalid XML in its
stylesheet (common with macro-generated files), so we parse it directly
via zipfile + ElementTree rather than openpyxl.
"""

import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from fbdi.compare import compare_all


NS = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def load_vba_report(path: str) -> list[tuple]:
    """Load the VBA comparison report via zipfile extraction.
    Returns list of (fbdi_file, fbdi_tab, col_letter, col_number, old_name, new_name, difference).
    """
    with zipfile.ZipFile(path) as z:
        with z.open("xl/sharedStrings.xml") as f:
            ss_tree = ET.parse(f)
        shared_strings = []
        for si in ss_tree.getroot().findall(".//s:si", NS):
            texts = si.findall(".//s:t", NS)
            shared_strings.append("".join(t.text or "" for t in texts))

        with z.open("xl/worksheets/sheet1.xml") as f:
            tree = ET.parse(f)

    rows_el = tree.getroot().findall(".//s:sheetData/s:row", NS)
    data_rows = []

    for row_el in rows_el[1:]:  # skip header
        cells = {}
        for c in row_el.findall("s:c", NS):
            ref = c.get("r")
            col_letter = "".join(ch for ch in ref if ch.isalpha())
            cell_type = c.get("t", "")
            v_el = c.find("s:v", NS)
            if v_el is not None and v_el.text is not None:
                if cell_type == "s":
                    val = shared_strings[int(v_el.text)]
                else:
                    val = v_el.text
            else:
                val = ""
            cells[col_letter] = val

        data_rows.append((
            cells.get("A", ""),
            cells.get("B", ""),
            cells.get("C", ""),
            int(float(cells["D"])) if cells.get("D") else 0,
            cells.get("E", ""),
            cells.get("F", ""),
            cells.get("G", ""),
        ))

    return data_rows


def load_python_report(path: Path) -> list[tuple]:
    """Load the Python comparison report via openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        data_rows.append((
            str(row[0]),
            str(row[1]),
            str(row[2]),
            int(row[3]) if row[3] else 0,
            str(row[4]) if row[4] else "",
            str(row[5]) if row[5] else "",
            str(row[6]),
        ))
    wb.close()
    return data_rows


def compare_reports(vba_rows: list[tuple], py_rows: list[tuple]) -> dict:
    """Compare VBA and Python reports row by row.
    Uses (fbdi_file, fbdi_tab, col_number) as the join key.
    """
    def make_key(row):
        return (row[0], row[1], row[3])  # file, tab, col_number

    vba_dict = {}
    for r in vba_rows:
        key = make_key(r)
        vba_dict[key] = r

    py_dict = {}
    for r in py_rows:
        key = make_key(r)
        py_dict[key] = r

    vba_keys = set(vba_dict.keys())
    py_keys = set(py_dict.keys())

    exact_matches = 0
    value_mismatches = []
    for key in vba_keys & py_keys:
        vr = vba_dict[key]
        pr = py_dict[key]
        # Compare old_field, new_field, difference
        if vr[4] == pr[4] and vr[5] == pr[5] and vr[6] == pr[6]:
            exact_matches += 1
        else:
            value_mismatches.append((key, vr, pr))

    return {
        "vba_total": len(vba_rows),
        "py_total": len(py_rows),
        "exact_matches": exact_matches,
        "vba_only": sorted(vba_keys - py_keys),
        "py_only": sorted(py_keys - py_keys),  # Will fix below
        "value_mismatches": value_mismatches,
    }


def main():
    project_root = Path(__file__).resolve().parent.parent

    vba_path = project_root / "Comparison_Report_25D_26A.xlsx"
    old_dir = project_root / "25D"
    new_dir = project_root / "26A"
    py_output = project_root / "test_output.xlsx"

    if not vba_path.exists():
        print(f"ERROR: VBA report not found: {vba_path}")
        sys.exit(1)
    if not old_dir.exists() or not new_dir.exists():
        print(f"ERROR: FBDI directories not found")
        sys.exit(1)

    # Load VBA report
    print("Loading VBA report...")
    vba_rows = load_vba_report(str(vba_path))
    print(f"  VBA rows: {len(vba_rows)}")

    # Run Python comparison
    print("Running Python comparison...")
    compare_all(old_dir, new_dir, py_output, changes_only=True)

    # Load Python report
    print("Loading Python report...")
    py_rows = load_python_report(py_output)
    print(f"  Python rows: {len(py_rows)}")

    # Compare
    print("\nComparing reports...")
    results = compare_reports(vba_rows, py_rows)

    # Build Python-only set correctly
    vba_keys = {(r[0], r[1], r[3]) for r in vba_rows}
    py_keys = {(r[0], r[1], r[3]) for r in py_rows}
    py_only = sorted(py_keys - vba_keys)

    shared = len(vba_keys & py_keys)
    exact = results["exact_matches"]
    vba_only = results["vba_only"]
    value_mismatches = results["value_mismatches"]

    print(f"\n{'='*60}")
    print(f"VALIDATION REPORT")
    print(f"{'='*60}")
    print(f"VBA rows:           {len(vba_rows)}")
    print(f"Python rows:        {len(py_rows)}")
    print(f"Shared keys:        {shared}")
    print(f"Exact matches:      {exact} ({100*exact/shared:.1f}% of shared)" if shared else "")
    print(f"VBA-only rows:      {len(vba_only)}")
    print(f"Python-only rows:   {len(py_only)}")
    print(f"Value mismatches:   {len(value_mismatches)}")

    if shared:
        match_pct = 100 * exact / len(vba_rows)
        print(f"\nOverall match rate: {match_pct:.1f}% (exact matches / VBA total)")

    if vba_only:
        print(f"\nVBA-only entries (first 20):")
        # Group by file
        by_file = {}
        for key in vba_only[:50]:
            by_file.setdefault(key[0], []).append(key)
        for f, keys in sorted(by_file.items()):
            tabs = {k[1] for k in keys}
            print(f"  {f}: {len(keys)} rows across tabs: {', '.join(sorted(tabs)[:3])}...")

    if py_only:
        print(f"\nPython-only entries (first 20):")
        by_file = {}
        for key in py_only[:50]:
            by_file.setdefault(key[0], []).append(key)
        for f, keys in sorted(by_file.items()):
            tabs = {k[1] for k in keys}
            print(f"  {f}: {len(keys)} rows across tabs: {', '.join(sorted(tabs)[:3])}...")

    if value_mismatches:
        print(f"\nValue mismatches (first 10):")
        for key, vr, pr in value_mismatches[:10]:
            print(f"  {key[0]} | {key[1]} | Col {key[2]}")
            print(f"    VBA: old={vr[4]!r} new={vr[5]!r}")
            print(f"    Py:  old={pr[4]!r} new={pr[5]!r}")

    # Pass/fail
    print(f"\n{'='*60}")
    if shared and exact / len(vba_rows) >= 0.95:
        print("RESULT: PASS (>= 95% match rate)")
    elif shared and exact / len(vba_rows) >= 0.80:
        print("RESULT: ACCEPTABLE (>= 80% match rate, discrepancies documented)")
    else:
        print("RESULT: NEEDS INVESTIGATION")
    print(f"{'='*60}")


if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.WARNING)
    main()
