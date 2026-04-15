"""Tests for fbdi.compare — core comparison engine."""

import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from fbdi.compare import ComparisonRow, compare_fbdi_pair, compare_all
from fbdi.config import SKIP_TABS


HEADERS_OLD = ["INVOICE_NUMBER", "PO_LINE_ID", "VENDOR_NAME", "AMOUNT", "CURRENCY_CODE"]
HEADERS_NEW = ["INVOICE_NUMBER", "PO_LINE_ID", "VENDOR_NAME", "AMOUNT", "CURRENCY_CODE"]


def _create_fbdi_workbook(path: Path, sheets: dict[str, list[str]], header_row: int = 4):
    """Create a test FBDI workbook with headers at the specified row."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    for sheet_name, headers in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=header)
    wb.save(path)


class TestCompareFbdiPair:
    def test_identical_headers_all_no(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        _create_fbdi_workbook(old_path, {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_path, {"Data": HEADERS_NEW})

        rows = compare_fbdi_pair(old_path, new_path)
        assert len(rows) == 5
        assert all(r.difference == "NO" for r in rows)

    def test_changed_header_marked_yes(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        modified = HEADERS_NEW.copy()
        modified[2] = "SUPPLIER_NAME"  # Changed from VENDOR_NAME
        _create_fbdi_workbook(old_path, {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_path, {"Data": modified})

        rows = compare_fbdi_pair(old_path, new_path)
        changed = [r for r in rows if r.difference == "YES"]
        assert len(changed) == 1
        assert changed[0].old_field_name == "VENDOR_NAME"
        assert changed[0].new_field_name == "SUPPLIER_NAME"
        assert changed[0].column_letter == "C"
        assert changed[0].column_number == 3

    def test_new_columns_detected(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        extended = HEADERS_NEW + ["NEW_FIELD"]
        _create_fbdi_workbook(old_path, {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_path, {"Data": extended})

        rows = compare_fbdi_pair(old_path, new_path)
        additions = [r for r in rows if r.old_field_name is None]
        assert len(additions) == 1
        assert additions[0].new_field_name == "NEW_FIELD"
        assert additions[0].difference == "YES"

    def test_removed_columns_detected(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        shorter = HEADERS_NEW[:3]
        _create_fbdi_workbook(old_path, {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_path, {"Data": shorter})

        rows = compare_fbdi_pair(old_path, new_path)
        removals = [r for r in rows if r.new_field_name is None]
        assert len(removals) == 2
        assert all(r.difference == "YES" for r in removals)

    def test_skip_tabs_excluded(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        sheets = {
            "Instructions and CSV Generation": ["COL_A"],
            "Data": HEADERS_OLD,
        }
        _create_fbdi_workbook(old_path, sheets)
        _create_fbdi_workbook(new_path, sheets)

        rows = compare_fbdi_pair(old_path, new_path)
        tabs = {r.fbdi_tab for r in rows}
        assert "Instructions and CSV Generation" not in tabs
        assert "Data" in tabs

    def test_tab_only_in_old_skipped(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        _create_fbdi_workbook(old_path, {"Data": HEADERS_OLD, "ExtraTab": ["COL_A", "COL_B", "COL_C"]})
        _create_fbdi_workbook(new_path, {"Data": HEADERS_NEW})

        rows = compare_fbdi_pair(old_path, new_path)
        tabs = {r.fbdi_tab for r in rows}
        assert "ExtraTab" not in tabs

    def test_column_letters_correct(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        _create_fbdi_workbook(old_path, {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_path, {"Data": HEADERS_NEW})

        rows = compare_fbdi_pair(old_path, new_path)
        letters = [r.column_letter for r in rows]
        assert letters == ["A", "B", "C", "D", "E"]

    def test_column_numbers_1_based(self, tmp_path):
        old_path = tmp_path / "Template.xlsm"
        new_path = tmp_path / "Template_new.xlsm"
        _create_fbdi_workbook(old_path, {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_path, {"Data": HEADERS_NEW})

        rows = compare_fbdi_pair(old_path, new_path)
        numbers = [r.column_number for r in rows]
        assert numbers == [1, 2, 3, 4, 5]


class TestCompareAll:
    def test_writes_output_file(self, tmp_path):
        old_dir = tmp_path / "old"
        new_dir = tmp_path / "new"
        old_dir.mkdir()
        new_dir.mkdir()

        modified = HEADERS_NEW.copy()
        modified[0] = "INVOICE_NUM"  # Changed
        _create_fbdi_workbook(old_dir / "Template.xlsm", {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_dir / "Template.xlsm", {"Data": modified})

        output = tmp_path / "report.xlsx"
        result, _ = compare_all(old_dir, new_dir, output, changes_only=True)

        assert result.exists()
        wb = load_workbook(result)
        ws = wb.active
        # Header row
        assert ws.cell(row=1, column=1).value == "FBDI File"
        # Should have at least the header + 1 change row
        assert ws.max_row >= 2
        wb.close()

    def test_changes_only_filters_no_rows(self, tmp_path):
        old_dir = tmp_path / "old"
        new_dir = tmp_path / "new"
        old_dir.mkdir()
        new_dir.mkdir()

        modified = HEADERS_NEW.copy()
        modified[2] = "SUPPLIER_NAME"
        _create_fbdi_workbook(old_dir / "Template.xlsm", {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_dir / "Template.xlsm", {"Data": modified})

        output = tmp_path / "report.xlsx"
        compare_all(old_dir, new_dir, output, changes_only=True)

        wb = load_workbook(output)
        ws = wb.active
        # Row 1 = headers, Row 2+ = data
        # Only the changed column should appear
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 1
        assert data_rows[0][6] == "YES"  # Difference column
        wb.close()

    def test_all_rows_mode(self, tmp_path):
        old_dir = tmp_path / "old"
        new_dir = tmp_path / "new"
        old_dir.mkdir()
        new_dir.mkdir()

        modified = HEADERS_NEW.copy()
        modified[2] = "SUPPLIER_NAME"
        _create_fbdi_workbook(old_dir / "Template.xlsm", {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_dir / "Template.xlsm", {"Data": modified})

        output = tmp_path / "report.xlsx"
        compare_all(old_dir, new_dir, output, changes_only=False)

        wb = load_workbook(output)
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 5  # All 5 columns
        wb.close()


class TestCompareAllReturnSignature:
    def test_compare_all_returns_empty_timed_out_on_fast_files(self, tmp_path):
        """compare_all returns (output_path, timed_out) where timed_out is empty for normal files."""
        old_dir = tmp_path / "old"
        new_dir = tmp_path / "new"
        old_dir.mkdir()
        new_dir.mkdir()

        _create_fbdi_workbook(old_dir / "Template.xlsm", {"Data": HEADERS_OLD})
        _create_fbdi_workbook(new_dir / "Template.xlsm", {"Data": HEADERS_NEW})

        output = tmp_path / "report.xlsx"
        result_path, timed_out = compare_all(old_dir, new_dir, output)

        assert result_path.exists()
        assert timed_out == []
