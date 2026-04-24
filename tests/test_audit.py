import json
import tempfile
from pathlib import Path
from openpyxl import Workbook

from fbdi.audit import (
    SnapshotField, SnapshotKeySeq, SnapshotTable, ApplaudSnapshot,
    Candidate, EvidenceBundle, PriorRow, AuditRow,
    load_snapshot, load_catalog, load_prior_mapping, CatalogIndex,
)

def test_data_classes_importable():
    field = SnapshotField(
        name="TA4INVOICE_ID",
        bare_name="INVOICE_ID",
        is_legacy_tracking=False,
        data_type="N",
        length=15,
    )
    assert field.bare_name == "INVOICE_ID"
    assert not field.is_legacy_tracking

def test_legacy_tracking_field():
    field = SnapshotField(
        name="@TA4SITE",
        bare_name="SITE",
        is_legacy_tracking=True,
        data_type="X",
        length=10,
    )
    assert field.is_legacy_tracking
    assert field.bare_name == "SITE"


def _make_snapshot_dict(tables=None, missing=None) -> dict:
    return {
        "mdb_path": "C:/test/AP0STE.mdb",
        "extracted_at": "2026-04-21T12:00:00Z",
        "extractor_version": "1",
        "tables": tables or [],
        "missing_tables": missing or [],
    }


def test_load_snapshot_basic(tmp_path):
    snap_data = _make_snapshot_dict(tables=[{
        "name": "T_RA_INTERFACE_LINES_ALL",
        "prefix": "TA4",
        "description": "T_RA_INTERFACE_LINES_ALL (TA4)",
        "type": "1",
        "key_sequences": [{"seq": "1", "keys": ["TA4INVOICE_ID"]}],
        "fields": [
            {"name": "TA4INVOICE_ID", "bare_name": "INVOICE_ID",
             "is_legacy_tracking": False, "data_type": "N", "length": 15},
            {"name": "@TA4SITE", "bare_name": "SITE",
             "is_legacy_tracking": True, "data_type": "X", "length": 10},
        ],
    }])
    snap_file = tmp_path / "applaud_snapshot.json"
    snap_file.write_text(json.dumps(snap_data))
    snap = load_snapshot(snap_file)
    assert isinstance(snap, ApplaudSnapshot)
    assert len(snap.tables) == 1
    t = snap.tables[0]
    assert t.name == "T_RA_INTERFACE_LINES_ALL"
    assert t.prefix == "TA4"
    assert len(t.fields) == 2
    biz = t.business_fields()
    assert len(biz) == 1
    assert biz[0].bare_name == "INVOICE_ID"


def test_load_snapshot_missing_file(tmp_path):
    import pytest
    with pytest.raises(FileNotFoundError):
        load_snapshot(tmp_path / "missing.json")


def _make_catalog_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "26B"
    ws.append(["release", "file_name", "tab_name", "position",
                "column_label", "column_technical",
                "data_type", "length", "scale", "data_type_raw", "required"])
    ws.append(["26B", "AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL",
                1, "Invoice ID", "INVOICE_ID", "N", 15, None, "NUMBER(15)", "TRUE"])
    ws.append(["26B", "AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL",
                2, "Batch Source Name", "BATCH_SOURCE_NAME", "X", 50, None, "VARCHAR2(50)", "FALSE"])
    path = tmp_path / "FBDI_Master_Catalog.xlsx"
    wb.save(path)
    return path


def test_load_catalog_basic(tmp_path):
    path = _make_catalog_xlsx(tmp_path)
    index = load_catalog(path, release="26B")
    key = ("AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL")
    assert key in index
    assert "INVOICE_ID" in index[key]
    assert "BATCH_SOURCE_NAME" in index[key]


def test_load_catalog_missing_release(tmp_path):
    import pytest
    path = _make_catalog_xlsx(tmp_path)
    with pytest.raises(ValueError, match="25D"):
        load_catalog(path, release="25D")


def _make_thin_catalog_xlsx(tmp_path: Path) -> Path:
    """Catalog with a thin tab — rows have column_label but no column_technical."""
    wb = Workbook()
    ws = wb.active
    ws.title = "26B"
    ws.append(["release", "file_name", "tab_name", "position",
                "column_label", "column_technical",
                "data_type", "length", "scale", "data_type_raw", "required"])
    # Thin tab — only labels populated
    ws.append(["26B", "AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL",
                1, "Business Unit Identifier", None, "N", 15, None, "NUMBER(15)", "TRUE"])
    ws.append(["26B", "AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL",
                2, "*Business Unit Name", None, "X", 50, None, "VARCHAR2(50)", "FALSE"])
    ws.append(["26B", "AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL",
                3, "Payment Terms", None, "X", 80, None, "VARCHAR2(80)", "FALSE"])
    path = tmp_path / "FBDI_Master_Catalog.xlsx"
    wb.save(path)
    return path


def test_load_catalog_thin_tab_label_fallback(tmp_path):
    """When column_technical is missing, normalized column_label is indexed."""
    path = _make_thin_catalog_xlsx(tmp_path)
    index = load_catalog(path, release="26B")
    key = ("AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL")
    assert key in index
    # Labels should be normalized to UPPER_SNAKE_CASE, with * stripped
    assert "BUSINESS_UNIT_IDENTIFIER" in index[key]
    assert "BUSINESS_UNIT_NAME" in index[key]
    assert "PAYMENT_TERMS" in index[key]


def test_load_catalog_prefers_technical_over_label(tmp_path):
    """When both column_technical and column_label are present, technical wins."""
    wb = Workbook()
    ws = wb.active
    ws.title = "26B"
    ws.append(["release", "file_name", "tab_name", "position",
                "column_label", "column_technical",
                "data_type", "length", "scale", "data_type_raw", "required"])
    ws.append(["26B", "F", "T", 1, "Invoice Identifier", "INVOICE_ID",
               "N", 15, None, "NUMBER(15)", "TRUE"])
    path = tmp_path / "FBDI_Master_Catalog.xlsx"
    wb.save(path)
    index = load_catalog(path, release="26B")
    assert index[("F", "T")] == {"INVOICE_ID"}
    # The normalized label should NOT pollute the index when tech is present
    assert "INVOICE_IDENTIFIER" not in index[("F", "T")]


def _make_prior_mapping_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "FBDI Mapping"
    ws2 = wb.create_sheet("Applaud Tables")
    ws2.append(["#", "applaud_table", "status", "prefix",
                 "fbdi_template_mappings", "module", "notes"])
    ws2.append([1, "T_RA_INTERFACE_LINES_ALL", "YES", "TA4",
                 "AutoInvoiceImportTemplate / RA_INTERFACE_LINES_ALL", "Financials", ""])
    ws2.append([2, "T_GHOST_TABLE", "UNMAPPED", "", "", "HR", "no match found"])
    path = tmp_path / "fbdi_applaud_mapping.xlsx"
    wb.save(path)
    return path


def test_load_prior_mapping_basic(tmp_path):
    path = _make_prior_mapping_xlsx(tmp_path)
    mapping = load_prior_mapping(path)
    assert "T_RA_INTERFACE_LINES_ALL" in mapping
    row = mapping["T_RA_INTERFACE_LINES_ALL"]
    assert row.prior_status == "YES"
    assert row.prefix == "TA4"
    assert "AutoInvoiceImportTemplate" in row.mapping_text
    assert "T_GHOST_TABLE" in mapping
    assert mapping["T_GHOST_TABLE"].prior_status == "UNMAPPED"


from fbdi.audit import extract_prefix, derive_bare_name

def test_extract_prefix_standard():
    assert extract_prefix("T_RA_INTERFACE_LINES_ALL (TA4)") == "TA4"

def test_extract_prefix_alphanumeric():
    assert extract_prefix("T_EGP_COMPONENTS_INTERFACE (T91)") == "T91"

def test_extract_prefix_no_parens():
    assert extract_prefix("T_GHOST_TABLE") is None

def test_derive_bare_name_regular():
    bare, is_legacy = derive_bare_name("TA4INVOICE_ID", "TA4")
    assert bare == "INVOICE_ID"
    assert not is_legacy

def test_derive_bare_name_at_prefix():
    bare, is_legacy = derive_bare_name("@TA4SITE", "TA4")
    assert bare == "SITE"
    assert is_legacy

def test_derive_bare_name_no_prefix_match():
    bare, is_legacy = derive_bare_name("SOMETHING_ELSE", "TA4")
    assert bare == "SOMETHING_ELSE"
    assert not is_legacy


from fbdi.audit import (
    compute_name_alignment, compute_key_coverage,
    compute_column_overlap, check_prefix_conformance,
    SnapshotField,
)

# --- name_alignment ---

def test_name_alignment_exact():
    assert compute_name_alignment("T_RA_INTERFACE_LINES_ALL", "RA_INTERFACE_LINES_ALL") == "EXACT"

def test_name_alignment_partial_strip_all():
    assert compute_name_alignment("T_RA_INTERFACE_LINES_ALL", "RA_INTERFACE_LINES") == "PARTIAL"

def test_name_alignment_partial_strip_interface():
    assert compute_name_alignment("T_RCV_HEADERS_INTERFACE", "RCV_HEADERS") == "PARTIAL"

def test_name_alignment_none():
    assert compute_name_alignment("T_RA_INTERFACE_LINES_ALL", "TOTALLY_DIFFERENT_TAB") == "NONE"

def test_name_alignment_case_insensitive():
    assert compute_name_alignment("T_RA_INTERFACE_LINES_ALL", "ra_interface_lines_all") == "EXACT"

def test_name_alignment_human_readable_tab_exact():
    # Some Oracle templates (e.g. ImportAwards) use space-separated tab names.
    # These should align with the corresponding Applaud table after
    # normalizing spaces/underscores.
    assert compute_name_alignment("T_AWARD_BUDGET_PERIODS", "Award Budget Periods") == "EXACT"

def test_name_alignment_human_readable_partial():
    assert compute_name_alignment("T_RA_INTERFACE_LINES_ALL", "RA Interface Lines") == "PARTIAL"

def test_name_alignment_loose_oracle_T_suffix():
    # Oracle temp-table convention: Applaud has trailing _T, FBDI doesn't.
    assert compute_name_alignment("T_HZ_IMP_ACCOUNTRELS_T", "HZ_IMP_ACCOUNTRELS") == "PARTIAL"

def test_name_alignment_loose_embedded_separator_diff():
    # PARTY_SITES (underscored) vs PARTYSITES (glued) — same word, different punctuation.
    assert compute_name_alignment("T_HZ_IMP_PARTY_SITES_T", "HZ_IMP_PARTYSITES_T") == "PARTIAL"

def test_name_alignment_loose_singular_plural():
    # Applaud singular GL_BUDGETS vs FBDI singular GL_BUDGET.
    assert compute_name_alignment("T_GL_BUDGETS_INTERFACE", "GL_BUDGET_INTERFACE") == "PARTIAL"

def test_name_alignment_loose_glued_vs_spaced():
    # Glued Applaud name vs spaced FBDI label.
    assert compute_name_alignment("T_ADDITIONALTRANSFERORDERCOST",
                                    "Additional Transfer Order Costs") == "PARTIAL"

def test_name_alignment_loose_does_not_over_match():
    # Legitimately different acronyms must stay NONE.
    assert compute_name_alignment("T_PROJ_RES_REQ_INTERFACE", "PJR_RES_REQ_INTERFACE") == "NONE"
    # Different words in the middle: SUP vs SUPPLIER.
    assert compute_name_alignment("T_POZ_SUP_ADDRESSES_INT", "POZ_SUPPLIER_ADDRESSES_INT") == "NONE"

# --- key_coverage ---

def test_key_coverage_full():
    keys = {"INVOICE_ID", "LINE_NUMBER"}
    fbdi_cols = {"INVOICE_ID", "LINE_NUMBER", "BATCH_SOURCE_NAME"}
    assert compute_key_coverage(keys, fbdi_cols) == 1.0

def test_key_coverage_partial():
    keys = {"INVOICE_ID", "LINE_NUMBER"}
    fbdi_cols = {"INVOICE_ID", "BATCH_SOURCE_NAME"}
    assert compute_key_coverage(keys, fbdi_cols) == 0.5

def test_key_coverage_empty_keys():
    assert compute_key_coverage(set(), {"INVOICE_ID"}) == 0.0

# --- column_overlap ---

def _biz_field(bare_name: str) -> SnapshotField:
    return SnapshotField(name=f"TA4{bare_name}", bare_name=bare_name,
                         is_legacy_tracking=False, data_type="X", length=30)

def _legacy_field(bare_name: str) -> SnapshotField:
    return SnapshotField(name=f"@TA4{bare_name}", bare_name=bare_name,
                         is_legacy_tracking=True, data_type="X", length=10)

def test_column_overlap_excludes_legacy():
    fields = [_biz_field("INVOICE_ID"), _biz_field("LINE_NUM"), _legacy_field("SITE")]
    fbdi_cols = {"INVOICE_ID", "LINE_NUM"}
    assert compute_column_overlap(fields, fbdi_cols) == 1.0

def test_column_overlap_partial():
    fields = [_biz_field("INVOICE_ID"), _biz_field("LINE_NUM"), _biz_field("BATCH_NAME")]
    fbdi_cols = {"INVOICE_ID", "LINE_NUM"}
    assert abs(compute_column_overlap(fields, fbdi_cols) - 2/3) < 0.001

def test_column_overlap_all_legacy():
    fields = [_legacy_field("SITE"), _legacy_field("LEGACY_HEADER")]
    fbdi_cols = {"INVOICE_ID"}
    assert compute_column_overlap(fields, fbdi_cols) == 0.0

def test_column_overlap_case_insensitive():
    fields = [_biz_field("INVOICE_ID")]
    fbdi_cols = {"invoice_id"}
    assert compute_column_overlap(fields, fbdi_cols) == 1.0

# --- prefix conformance ---

def test_prefix_conformance_true():
    assert check_prefix_conformance("T_RA_INTERFACE_LINES_ALL", "TA4", "RA_INTERFACE_LINES_ALL") is True

def test_prefix_conformance_false():
    assert check_prefix_conformance("T_RA_INTERFACE_LINES_ALL", "TA4", "RA_INTERFACE_LINES") is False


# --- Pass 1 candidate index ---

from fbdi.audit import build_candidate_index, ApplaudSnapshot, SnapshotTable, SnapshotField, SnapshotKeySeq

def _make_snapshot_table(
    name: str, prefix: str, fields: list[SnapshotField], key_fields: list[str]
) -> SnapshotTable:
    return SnapshotTable(
        name=name, prefix=prefix,
        description=f"{name} ({prefix})", type="1",
        key_sequences=[SnapshotKeySeq(seq="1", keys=key_fields)],
        fields=fields,
    )

def _make_snap(*tables: SnapshotTable) -> ApplaudSnapshot:
    return ApplaudSnapshot(
        mdb_path="", extracted_at="", extractor_version="1",
        tables=list(tables), missing_tables=[],
    )


def test_pass1_exact_name_match_kept():
    fields = [SnapshotField("TA4INVOICE_ID", "INVOICE_ID", False, "N", 15)]
    table = _make_snapshot_table("T_RA_INTERFACE_LINES_ALL", "TA4", fields, ["TA4INVOICE_ID"])
    catalog = {("AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL"): {"INVOICE_ID"}}
    snap = _make_snap(table)
    idx = build_candidate_index(snap, catalog)
    assert "T_RA_INTERFACE_LINES_ALL" in idx
    candidates = idx["T_RA_INTERFACE_LINES_ALL"]
    assert len(candidates) == 1
    c = candidates[0]
    assert c.fbdi_file == "AutoInvoiceImportTemplate"
    assert c.fbdi_tab == "RA_INTERFACE_LINES_ALL"
    assert c.name_alignment == "EXACT"


def test_pass1_below_threshold_dropped():
    fields = [SnapshotField("TA4INVOICE_ID", "INVOICE_ID", False, "N", 15)]
    table = _make_snapshot_table("T_RA_INTERFACE_LINES_ALL", "TA4", fields, ["TA4INVOICE_ID"])
    catalog = {("SomeTemplate", "UNRELATED_TAB"): {"UNRELATED_COL"}}
    snap = _make_snap(table)
    idx = build_candidate_index(snap, catalog)
    assert idx.get("T_RA_INTERFACE_LINES_ALL", []) == []


def test_pass1_high_column_overlap_kept():
    fields = [SnapshotField(f"TA4F{i}", f"F{i}", False, "X", 10) for i in range(5)]
    table = _make_snapshot_table("T_SOME_TABLE", "TA4", fields, [])
    fbdi_cols = {f"F{i}" for i in range(4)}
    catalog = {("SomeTemplate", "UNRELATED_TAB"): fbdi_cols}
    snap = _make_snap(table)
    idx = build_candidate_index(snap, catalog)
    candidates = idx.get("T_SOME_TABLE", [])
    assert len(candidates) == 1
    assert candidates[0].column_overlap >= 0.3


def test_pass1_legacy_fields_excluded_from_overlap():
    biz = SnapshotField("TA4INVOICE_ID", "INVOICE_ID", False, "N", 15)
    leg = SnapshotField("@TA4SITE", "SITE", True, "X", 10)
    table = _make_snapshot_table("T_RA", "TA4", [biz, leg], [])
    catalog = {("AnyTemplate", "RA"): {"INVOICE_ID"}}
    snap = _make_snap(table)
    idx = build_candidate_index(snap, catalog)
    candidates = idx.get("T_RA", [])
    assert candidates  # kept because name PARTIAL match or high overlap
    assert candidates[0].column_overlap == 1.0  # 1/1 biz field matched


def test_pass1_sorted_strongest_first():
    fields = [SnapshotField("TA4INVOICE_ID", "INVOICE_ID", False, "N", 15)]
    table = _make_snapshot_table("T_RA_INTERFACE_LINES_ALL", "TA4", fields, ["TA4INVOICE_ID"])
    catalog = {
        ("TemplateA", "RA_INTERFACE_LINES_ALL"): {"INVOICE_ID"},   # EXACT + key 1.0
        ("TemplateB", "RA_INTERFACE_LINES"): {"INVOICE_ID"},       # PARTIAL + key 1.0
    }
    snap = _make_snap(table)
    idx = build_candidate_index(snap, catalog)
    candidates = idx["T_RA_INTERFACE_LINES_ALL"]
    assert len(candidates) == 2
    assert candidates[0].name_alignment == "EXACT"


from fbdi.audit import parse_prior_mapping, evaluate_confidence, Candidate


def _cand(name_align: str, key_cov: float, col_ovlp: float,
           key_fields_matched: list[str] | None = None) -> Candidate:
    # When key_coverage == 1.0 and caller didn't specify which keys matched,
    # default to two matched keys so evaluate_confidence's discriminative
    # guard (>=2 keys required for key-alone promotion to M) treats the
    # candidate as strong. Callers exercising the single-key edge case
    # should pass an explicit list.
    if key_fields_matched is None:
        key_fields_matched = ["K1", "K2"] if key_cov == 1.0 else []
    return Candidate(
        fbdi_file="F", fbdi_tab="T",
        name_alignment=name_align,
        key_coverage=key_cov,
        column_overlap=col_ovlp,
        prefix_conformance=True,
        applaud_key_fields_matched=key_fields_matched,
        applaud_fields_matched=[],
        applaud_fields_missing=[],
    )


# --- parse_prior_mapping ---

def test_parse_prior_mapping_single():
    result = parse_prior_mapping("AutoInvoiceImportTemplate / RA_INTERFACE_LINES_ALL")
    assert result == [("AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL")]

def test_parse_prior_mapping_multi():
    result = parse_prior_mapping(
        "AutoInvoiceImportTemplate / RA_INTERFACE_LINES_ALL; "
        "ItemStructureImportTemplate / EGP_COMPONENTS_INTERFACE"
    )
    assert len(result) == 2
    assert ("AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL") in result
    assert ("ItemStructureImportTemplate", "EGP_COMPONENTS_INTERFACE") in result

def test_parse_prior_mapping_blank():
    assert parse_prior_mapping("") == []
    assert parse_prior_mapping("   ") == []

def test_parse_prior_mapping_malformed(caplog):
    import logging
    with caplog.at_level(logging.WARNING):
        result = parse_prior_mapping("NoSlashHere")
    assert result == []

# --- evaluate_confidence ---

def test_evaluate_confidence_high():
    c = _cand("EXACT", 1.0, 0.85)
    assert evaluate_confidence(c) == "H"

def test_evaluate_confidence_high_no_keys():
    c = _cand("EXACT", 0.0, 0.75)
    assert evaluate_confidence(c) == "H"

def test_evaluate_confidence_medium_partial():
    c = _cand("PARTIAL", 0.8, 0.5)
    assert evaluate_confidence(c) == "M"

def test_evaluate_confidence_medium_key_coverage():
    c = _cand("NONE", 0.6, 0.45)
    assert evaluate_confidence(c) == "M"

def test_evaluate_confidence_low():
    c = _cand("NONE", 0.0, 0.1)
    assert evaluate_confidence(c) == "L"

def test_evaluate_confidence_partial_low_overlap_still_medium():
    # PARTIAL name match alone → M regardless of overlap
    c = _cand("PARTIAL", 0.0, 0.05)
    assert evaluate_confidence(c) == "M"

def test_evaluate_confidence_exact_weak_signals_is_medium():
    # EXACT name with weak key/overlap signals (e.g. thin tab) must be
    # at least M, never L. Prevents the signal inversion where EXACT+weak
    # scored lower than PARTIAL+weak.
    c = _cand("EXACT", 0.0, 0.05)
    assert evaluate_confidence(c) == "M"

def test_evaluate_confidence_exact_zero_signals_is_medium():
    c = _cand("EXACT", 0.0, 0.0)
    assert evaluate_confidence(c) == "M"

def test_evaluate_confidence_full_key_coverage_is_medium_even_with_none_name():
    # 100% key coverage is strong semantic evidence even when the Applaud
    # table name doesn't align with the FBDI tab name (e.g. T_BANKS_BRANCHES
    # → "Bank Account" where all 4 keys match perfectly).
    c = _cand("NONE", 1.0, 0.95)
    assert evaluate_confidence(c) == "M"

def test_evaluate_confidence_high_overlap_alone_is_low():
    # High overlap without any other signal must stay L — Oracle's generic
    # DFF columns (ATTRIBUTE1..20, ATTRIBUTE_DATE*) create spurious high
    # overlap across unrelated tabs.
    c = _cand("NONE", 0.0, 0.9)
    assert evaluate_confidence(c) == "L"

def test_evaluate_confidence_single_key_full_coverage_is_low():
    # A single key at 100% is not discriminative — generic column names
    # like SEQUENCE_NUMBER appear in many interface tabs.
    c = _cand("NONE", 1.0, 0.06, key_fields_matched=["SEQUENCE_NUMBER"])
    assert evaluate_confidence(c) == "L"


from fbdi.audit import (
    adjudicate_table, AuditRow, Candidate, EvidenceBundle,
    SnapshotTable, SnapshotField, SnapshotKeySeq, PriorRow,
)


def _pr(status: str, mapping: str = "", prefix: str = "TA4",
        module: str = "Fin", notes: str = "") -> PriorRow:
    return PriorRow("T_TEST", status, prefix, mapping, module, notes)


def _cand7(file: str, tab: str, align: str, key: float, overlap: float) -> Candidate:
    return Candidate(
        fbdi_file=file, fbdi_tab=tab,
        name_alignment=align, key_coverage=key, column_overlap=overlap,
        prefix_conformance=True,
        applaud_key_fields_matched=[], applaud_fields_matched=[], applaud_fields_missing=[],
    )


# Branch 1: NOT_IN_APPLAUD → UNMAPPED High
def test_adjudicate_not_in_applaud():
    row = adjudicate_table("T_GHOST", None, [], _pr("UNMAPPED"))
    assert row.verdict == "UNMAPPED"
    assert row.confidence == "H"
    assert "not present" in row.rationale.lower()


# Branch 2: FILE_TOO_LARGE carry-through
def test_adjudicate_file_too_large_carrythrough():
    row = adjudicate_table("T_TEST", None, [], _pr("FILE_TOO_LARGE"))
    assert row.verdict == "FILE_TOO_LARGE"
    assert row.confidence == ""


# Branch 3: Single prior, High signals → YES High
def test_adjudicate_single_prior_high():
    c = _cand7("AutoInvoice", "RA_INTERFACE_LINES_ALL", "EXACT", 1.0, 0.85)
    prior = _pr("YES", "AutoInvoice / RA_INTERFACE_LINES_ALL")
    snap_table = SnapshotTable("T_RA", "TA4", "T_RA (TA4)", "1", [], [])
    row = adjudicate_table("T_RA", snap_table, [c], prior)
    assert row.verdict == "YES"
    assert row.confidence == "H"
    assert not row.changed


# Branch 4: Single prior, low signals → NEEDS_REVIEW
def test_adjudicate_single_prior_low():
    c = _cand7("AutoInvoice", "WRONG_TAB", "NONE", 0.0, 0.1)
    prior = _pr("YES", "AutoInvoice / WRONG_TAB")
    snap_table = SnapshotTable("T_RA", "TA4", "T_RA (TA4)", "1", [], [])
    row = adjudicate_table("T_RA", snap_table, [c], prior)
    assert row.verdict == "NEEDS_REVIEW"
    assert row.needs_deep_rationale


# Branch 5: Multi prior, both High → multi retained
def test_adjudicate_multi_both_high():
    c1 = _cand7("TemplA", "TAB_X", "EXACT", 1.0, 0.85)
    c2 = _cand7("TemplB", "TAB_X", "EXACT", 1.0, 0.90)
    prior = _pr("YES", "TemplA / TAB_X; TemplB / TAB_X")
    snap_table = SnapshotTable("T_TAB_X", "TXX", "T_TAB_X (TXX)", "1", [], [])
    row = adjudicate_table("T_TAB_X", snap_table, [c1, c2], prior)
    assert row.verdict == "YES"
    assert ";" in row.fbdi_mapping  # multi retained


# Branch 6: Multi prior, one High + one Low → collapsed
def test_adjudicate_multi_collapse():
    c1 = _cand7("TemplA", "TAB_X", "EXACT", 1.0, 0.85)
    # c2 not in candidates list (low signal, filtered by pass 1)
    prior = _pr("YES", "TemplA / TAB_X; TemplB / TAB_MISSING")
    snap_table = SnapshotTable("T_TAB_X", "TXX", "T_TAB_X (TXX)", "1", [], [])
    row = adjudicate_table("T_TAB_X", snap_table, [c1], prior)
    assert row.verdict == "YES"
    assert "TemplB" not in row.fbdi_mapping
    assert row.changed  # collapsed from multi


# Branch 7: UNMAPPED + High candidate → promoted to YES
def test_adjudicate_unmapped_promoted():
    c = _cand7("AutoInvoice", "RA_INTERFACE_LINES_ALL", "EXACT", 1.0, 0.85)
    prior = _pr("UNMAPPED")
    snap_table = SnapshotTable("T_RA", "TA4", "T_RA (TA4)", "1", [], [])
    row = adjudicate_table("T_RA", snap_table, [c], prior)
    assert row.verdict == "YES"
    assert row.confidence == "H"
    assert row.changed


# Branch 8: UNMAPPED + Medium candidate → NEEDS_REVIEW
def test_adjudicate_unmapped_medium_candidate():
    c = _cand7("AutoInvoice", "RA_INTERFACE_LINES", "PARTIAL", 0.5, 0.5)
    prior = _pr("UNMAPPED")
    snap_table = SnapshotTable("T_RA_INTERFACE_LINES_ALL", "TA4", "T_RA (TA4)", "1", [], [])
    row = adjudicate_table("T_RA_INTERFACE_LINES_ALL", snap_table, [c], prior)
    assert row.verdict == "NEEDS_REVIEW"
    assert row.confidence == "M"


# Branch 9: UNMAPPED + no viable candidate → stays UNMAPPED High
def test_adjudicate_unmapped_no_candidate():
    prior = _pr("UNMAPPED")
    snap_table = SnapshotTable("T_GHOST", "TGH", "T_GHOST (TGH)", "1", [], [])
    row = adjudicate_table("T_GHOST", snap_table, [], prior)
    assert row.verdict == "UNMAPPED"
    assert row.confidence == "H"
    assert not row.changed


# Branch 10: prefix mismatch surfaces in notes, doesn't change verdict
def test_adjudicate_prefix_mismatch_noted():
    c = _cand7("AutoInvoice", "RA_INTERFACE_LINES_ALL", "EXACT", 1.0, 0.85)
    c.prefix_conformance = False
    prior = _pr("YES", "AutoInvoice / RA_INTERFACE_LINES_ALL")
    snap_table = SnapshotTable("T_RA_INTERFACE_LINES_ALL", "TA4", "T_RA (TA4)", "1", [], [])
    row = adjudicate_table("T_RA_INTERFACE_LINES_ALL", snap_table, [c], prior)
    assert row.verdict == "YES"
    assert "prefix" in row.rationale.lower() or any("prefix" in n.lower() for n in row.evidence.notes)


# Branch 11: deep_rationale trigger — changed from prior
def test_adjudicate_deep_rationale_on_change():
    c = _cand7("AutoInvoice", "RA_INTERFACE_LINES_ALL", "EXACT", 1.0, 0.85)
    prior = _pr("UNMAPPED")  # was UNMAPPED, now promoted to YES
    snap_table = SnapshotTable("T_RA", "TA4", "T_RA (TA4)", "1", [], [])
    row = adjudicate_table("T_RA", snap_table, [c], prior)
    assert row.changed
    assert row.needs_deep_rationale


from fbdi.audit import write_output_xlsx, write_audit_md, AuditRow, EvidenceBundle, CatalogIndex
from openpyxl import load_workbook


def _make_audit_row(
    table: str, verdict: str, confidence: str, mapping: str,
    prior_verdict: str = "YES", changed: bool = False,
    needs_deep: bool = False, prefix: str = "TA4",
) -> AuditRow:
    return AuditRow(
        applaud_table=table, prefix=prefix,
        verdict=verdict, fbdi_mapping=mapping,
        confidence=confidence,
        rationale=f"{verdict} because signals",
        prior_verdict=prior_verdict, changed=changed,
        needs_deep_rationale=needs_deep,
        evidence=EvidenceBundle(),
    )


def test_write_output_xlsx_sheets(tmp_path):
    rows = [
        _make_audit_row("T_RA", "YES", "H", "AutoInvoice / RA_TAB"),
        _make_audit_row("T_GHOST", "UNMAPPED", "H", "", prior_verdict="UNMAPPED"),
        _make_audit_row("T_PROBLEM", "NEEDS_REVIEW", "M", "SomeFile / SomeTab",
                        prior_verdict="YES", changed=True, needs_deep=True),
    ]
    catalog: CatalogIndex = {
        ("AutoInvoice", "RA_TAB"): {"INVOICE_ID"},
    }
    out = tmp_path / "Claude_test.xlsx"
    write_output_xlsx(rows, catalog, out)
    wb = load_workbook(out, read_only=True)
    assert "FBDI Mapping" in wb.sheetnames
    assert "Applaud Tables" in wb.sheetnames
    assert "Needs Review" in wb.sheetnames
    wb.close()


def test_write_output_xlsx_sheet2_rows(tmp_path):
    rows = [
        _make_audit_row("T_RA", "YES", "H", "AutoInvoice / RA_TAB"),
        _make_audit_row("T_GHOST", "UNMAPPED", "H", "", prior_verdict="UNMAPPED"),
    ]
    catalog: CatalogIndex = {}
    out = tmp_path / "Claude_test.xlsx"
    write_output_xlsx(rows, catalog, out)
    wb = load_workbook(out, read_only=True, data_only=True)
    ws2 = wb["Applaud Tables"]
    data = list(ws2.iter_rows(values_only=True))
    assert len(data) == 3  # header + 2 data rows
    table_names = [r[1] for r in data[1:]]
    assert "T_RA" in table_names
    assert "T_GHOST" in table_names
    wb.close()


def test_write_output_xlsx_needs_review_sheet(tmp_path):
    rows = [
        _make_audit_row("T_NEEDS", "NEEDS_REVIEW", "M", "SomeFile / SomeTab",
                        needs_deep=True),
        _make_audit_row("T_OK", "YES", "H", "AutoInvoice / RA_TAB"),
    ]
    catalog: CatalogIndex = {}
    out = tmp_path / "Claude_test.xlsx"
    write_output_xlsx(rows, catalog, out)
    wb = load_workbook(out, read_only=True, data_only=True)
    ws3 = wb["Needs Review"]
    data = list(ws3.iter_rows(values_only=True))
    assert len(data) == 2  # header + 1 needs-review row
    wb.close()


def test_write_audit_md(tmp_path):
    rows = [
        _make_audit_row("T_NEEDS", "NEEDS_REVIEW", "M", "SomeFile / SomeTab",
                        needs_deep=True, changed=True),
        _make_audit_row("T_OK", "YES", "H", "AutoInvoice / RA_TAB"),
    ]
    out = tmp_path / "audit.md"
    write_audit_md(rows, {"extracted_at": "2026-04-21T12:00:00Z"}, out)
    content = out.read_text()
    assert "T_NEEDS" in content
    assert "T_OK" not in content  # High confidence unchanged → no entry
    assert "NEEDS_REVIEW" in content


from fbdi.audit import run_audit
import json
from pathlib import Path
from openpyxl import Workbook


def _make_e2e_snapshot(tmp_path: Path) -> Path:
    tables = [
        # 1. EXACT + High → YES H
        {
            "name": "T_RA_INTERFACE_LINES_ALL",
            "prefix": "TA4",
            "description": "T_RA_INTERFACE_LINES_ALL (TA4)",
            "type": "1",
            "key_sequences": [{"seq": "1", "keys": ["TA4INVOICE_ID"]}],
            "fields": [
                {"name": "TA4INVOICE_ID", "bare_name": "INVOICE_ID",
                 "is_legacy_tracking": False, "data_type": "N", "length": 15},
                {"name": "@TA4SITE", "bare_name": "SITE",
                 "is_legacy_tracking": True, "data_type": "X", "length": 10},
            ],
        },
        # 2. PARTIAL match, prior=UNMAPPED → NEEDS_REVIEW M
        {
            "name": "T_RCV_HEADERS_INTERFACE",
            "prefix": "TH7",
            "description": "T_RCV_HEADERS_INTERFACE (TH7)",
            "type": "1",
            "key_sequences": [],
            "fields": [
                {"name": "TH7RECEIPT_NUM", "bare_name": "RECEIPT_NUM",
                 "is_legacy_tracking": False, "data_type": "X", "length": 30},
            ],
        },
        # 3. No candidate → UNMAPPED H (re-confirmed)
        {
            "name": "T_GHOST_TABLE",
            "prefix": "TGG",
            "description": "T_GHOST_TABLE (TGG)",
            "type": "1",
            "key_sequences": [],
            "fields": [],
        },
        # 4. Multi prior, both High → multi retained
        {
            "name": "T_EGP_COMPONENTS_INTERFACE",
            "prefix": "T91",
            "description": "T_EGP_COMPONENTS_INTERFACE (T91)",
            "type": "1",
            "key_sequences": [],
            "fields": [
                {"name": "T91COMPONENT_ITEM_ID", "bare_name": "COMPONENT_ITEM_ID",
                 "is_legacy_tracking": False, "data_type": "N", "length": 18},
            ],
        },
        # 5. Multi prior, one leg missing → collapse
        {
            "name": "T_DOO_ORDER_HEADERS_ALL",
            "prefix": "TC4",
            "description": "T_DOO_ORDER_HEADERS_ALL (TC4)",
            "type": "1",
            "key_sequences": [],
            "fields": [
                {"name": "TC4ORDER_NUMBER", "bare_name": "ORDER_NUMBER",
                 "is_legacy_tracking": False, "data_type": "X", "length": 50},
            ],
        },
    ]
    data = {
        "mdb_path": "test", "extracted_at": "2026-04-21T12:00:00Z",
        "extractor_version": "1", "tables": tables, "missing_tables": [],
    }
    p = tmp_path / "applaud_snapshot.json"
    p.write_text(json.dumps(data))
    return p


def _make_e2e_catalog(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "26B"
    ws.append(["release", "file_name", "tab_name", "position", "column_label",
                "column_technical", "data_type", "length", "scale", "data_type_raw", "required"])
    rows = [
        ("26B", "AutoInvoiceImportTemplate", "RA_INTERFACE_LINES_ALL", 1, "Invoice ID", "INVOICE_ID", "N", 15, None, "NUMBER(15)", "TRUE"),
        ("26B", "ReceivingReceiptImportTemplate", "RCV_HEADERS", 1, "Receipt Num", "RECEIPT_NUM", "X", 30, None, "VARCHAR2(30)", "FALSE"),
        ("26B", "ChangeOrderImportTemplate", "EGP_COMPONENTS_INTERFACE", 1, "Comp Item", "COMPONENT_ITEM_ID", "N", 18, None, "NUMBER(18)", "FALSE"),
        ("26B", "ItemStructureImportTemplate", "EGP_COMPONENTS_INTERFACE", 1, "Comp Item", "COMPONENT_ITEM_ID", "N", 18, None, "NUMBER(18)", "FALSE"),
        ("26B", "SourceSalesOrderImportTemplate", "DOO_ORDER_HEADERS_ALL_INT", 1, "Order Num", "ORDER_NUMBER", "X", 50, None, "VARCHAR2(50)", "FALSE"),
    ]
    for r in rows:
        ws.append(r)
    p = tmp_path / "FBDI_Master_Catalog.xlsx"
    wb.save(p)
    return p


def _make_e2e_prior(tmp_path: Path) -> Path:
    wb = Workbook()
    wb.active.title = "FBDI Mapping"
    ws2 = wb.create_sheet("Applaud Tables")
    ws2.append(["#", "applaud_table", "status", "prefix",
                 "fbdi_template_mappings", "module", "notes"])
    rows = [
        (1, "T_RA_INTERFACE_LINES_ALL", "YES", "TA4",
         "AutoInvoiceImportTemplate / RA_INTERFACE_LINES_ALL", "Financials", ""),
        (2, "T_RCV_HEADERS_INTERFACE", "UNMAPPED", "TH7", "", "Procurement", ""),
        (3, "T_GHOST_TABLE", "UNMAPPED", "TGG", "", "Unknown", ""),
        (4, "T_EGP_COMPONENTS_INTERFACE", "YES", "T91",
         "ChangeOrderImportTemplate / EGP_COMPONENTS_INTERFACE; "
         "ItemStructureImportTemplate / EGP_COMPONENTS_INTERFACE",
         "SCM", ""),
        (5, "T_DOO_ORDER_HEADERS_ALL", "YES", "TC4",
         "SourceSalesOrderImportTemplate / DOO_ORDER_HEADERS_ALL_INT; "
         "ItemStructureImportTemplate / NONEXISTENT_TAB",
         "SCM", ""),
    ]
    for r in rows:
        ws2.append(r)
    p = tmp_path / "fbdi_applaud_mapping.xlsx"
    wb.save(p)
    return p


def test_audit_end_to_end(tmp_path):
    snap_path = _make_e2e_snapshot(tmp_path)
    cat_path = _make_e2e_catalog(tmp_path)
    prior_path = _make_e2e_prior(tmp_path)
    out_xlsx = tmp_path / "Claude_test.xlsx"
    out_md = tmp_path / "audit.md"

    audit_rows = run_audit(snap_path, cat_path, prior_path, out_xlsx, out_md)

    assert len(audit_rows) == 5

    by_table = {ar.applaud_table: ar for ar in audit_rows}

    # T_RA: EXACT + High → YES H, unchanged
    ra = by_table["T_RA_INTERFACE_LINES_ALL"]
    assert ra.verdict == "YES"
    assert ra.confidence == "H"
    assert not ra.changed

    # T_RCV: PARTIAL (RCV_HEADERS_INTERFACE strips to RCV_HEADERS, catalog has RCV_HEADERS) → NEEDS_REVIEW or YES
    rcv = by_table["T_RCV_HEADERS_INTERFACE"]
    assert rcv.verdict in ("NEEDS_REVIEW", "YES")

    # T_GHOST: no candidates → UNMAPPED H
    ghost = by_table["T_GHOST_TABLE"]
    assert ghost.verdict == "UNMAPPED"
    assert ghost.confidence == "H"

    # T_EGP: multi, both legs High → multi retained YES
    egp = by_table["T_EGP_COMPONENTS_INTERFACE"]
    assert egp.verdict == "YES"
    assert ";" in egp.fbdi_mapping

    # T_DOO: multi, one leg missing → collapsed YES, changed=True
    doo = by_table["T_DOO_ORDER_HEADERS_ALL"]
    assert doo.verdict == "YES"
    assert "NONEXISTENT_TAB" not in doo.fbdi_mapping
    assert doo.changed

    # Outputs exist
    assert out_xlsx.exists()
    assert out_md.exists()

    # Legacy tracking exclusion: T_RA overlap denominator = 1 (INVOICE_ID only, not SITE)
    ra_candidates = ra.evidence.candidates_evaluated
    if ra_candidates:
        assert ra_candidates[0].column_overlap == 1.0
