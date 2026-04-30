"""
Build FBDI_to_ApplaudTables_Mapping.xlsx
- Proofread fixes
- TEXTJOIN formula for Applaud Tables "FBDI Template Mappings" column
- XLOOKUP formula standardised in FBDI Mapping "In Base System?" column
- Enterprise design formatting
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# ── Colour palette ────────────────────────────────────────────────────────────
C_NAVY        = "1B2A4A"
C_BLUE2       = "2E5F8A"
C_AMBER       = "E8912D"
C_ROW_ALT     = "EDF2FA"
C_MAPPED_BG   = "D6EFD8"
C_MAPPED_FG   = "1A5C2A"
C_UNMAPPED_BG = "FDECD0"
C_UNMAPPED_FG = "7A3E00"

def solid(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_border():
    return Border(bottom=Side(style="medium", color=C_AMBER))

HDR_FONT       = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
HDR_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT      = Font(name="Segoe UI", size=9, color="1A1A2E")
DATA_FONT_MONO = Font(name="Consolas", size=9, color="2C3E6B")
DATA_FONT_SM   = Font(name="Segoe UI", size=8, color="555555", italic=True)
ALT_FILL       = solid(C_ROW_ALT)
WHITE_FILL     = solid("FFFFFF")
DATA_ALIGN     = Alignment(vertical="center", wrap_text=False)
CTR_ALIGN      = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN     = Alignment(vertical="center", wrap_text=True)

mapped_fill   = PatternFill("solid", bgColor=C_MAPPED_BG)
unmapped_fill = PatternFill("solid", bgColor=C_UNMAPPED_BG)
mapped_font   = Font(name="Segoe UI", size=9, bold=True, color=C_MAPPED_FG)
unmapped_font = Font(name="Segoe UI", size=9, bold=True, color=C_UNMAPPED_FG)

# ── Load source ───────────────────────────────────────────────────────────────
src = openpyxl.load_workbook("Claude_fbdi_applaud_mapping.xlsx", data_only=False)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 – FBDI Mapping
# ─────────────────────────────────────────────────────────────────────────────
src_ws = src["FBDI Mapping"]
ws = wb.create_sheet("FBDI Mapping")
ws.sheet_view.showGridLines = False

# Collect populated rows
rows_data = []
for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, values_only=False):
    vals = [c.value for c in row]
    # Header row always included; data rows only if col A (FBDI Template) is populated
    if vals[0] is not None:
        rows_data.append(vals)

last_data_row = len(rows_data)  # includes header at index 0

# Write values/formulas
for r_idx, row in enumerate(rows_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=val)

# ── Proofread fix 1: header typo ─────────────────────────────────────────────
ws["H1"] = "BJH Verified"

# ── Proofread fix 2: rename G1 to reflect actual content ─────────────────────
ws["G1"] = "In Base System?"

# ── Proofread fix 3: standardise XLOOKUP formula in col G for all data rows ──
# Keep the manual "Multiple mapping is possible..." note for PJC rows;
# replace everything else with the standard XLOOKUP formula.
MANUAL_NOTE_MARKER = "Multiple mapping is possible"
# LET uses sentinel "__NF__" to distinguish "not found" from "found, G is empty".
# not-found  → "Needs to be created in base system"
# found + G empty (0) → "" (blank)
# found + G has value → that value
XLOOKUP_TMPL = (
    '=IF(C{r}="","",LET(v,'
    'XLOOKUP(C{r},\'Applaud Tables Reference\'!$A:$A,'
    '\'Applaud Tables Reference\'!$G:$G,"__NF__",0),'
    'IF(v="__NF__","Needs to be created in base system",'
    'IF(v=0,"",v))))'
)

for r_idx in range(2, last_data_row + 1):
    existing_g = ws.cell(row=r_idx, column=7).value
    # Preserve rows with manual multi-mapping note
    if existing_g and MANUAL_NOTE_MARKER in str(existing_g):
        continue
    ws.cell(row=r_idx, column=7, value=XLOOKUP_TMPL.format(r=r_idx))

# ── Proofread fix 4: R199 missing BJH Verified ───────────────────────────────
# Find last data row where col H is blank and col E is not blank
for r_idx in range(2, last_data_row + 1):
    if ws.cell(row=r_idx, column=5).value and not ws.cell(row=r_idx, column=8).value:
        ws.cell(row=r_idx, column=8, value="Y")

# ── Column widths ─────────────────────────────────────────────────────────────
for col, w in [("A",38),("B",40),("C",34),("D",9),("E",12),("F",18),("G",30),("H",14)]:
    ws.column_dimensions[col].width = w

ws.row_dimensions[1].height = 30
for r in range(2, last_data_row + 1):
    ws.row_dimensions[r].height = 16

ws.freeze_panes = "A2"

# ── Header row style ──────────────────────────────────────────────────────────
for c in range(1, 9):
    cell = ws.cell(row=1, column=c)
    cell.font      = HDR_FONT
    cell.fill      = solid(C_NAVY)
    cell.alignment = HDR_ALIGN
    cell.border    = hdr_border()

# ── Data row styles ───────────────────────────────────────────────────────────
for r in range(2, last_data_row + 1):
    bg = ALT_FILL if r % 2 == 0 else WHITE_FILL
    for c in range(1, 9):
        cell = ws.cell(row=r, column=c)
        cell.fill   = bg
        cell.border = thin_border()
        if c in (3, 4):
            cell.font      = DATA_FONT_MONO
            cell.alignment = DATA_ALIGN
        elif c == 7:
            cell.font      = DATA_FONT_SM
            cell.alignment = WRAP_ALIGN
        elif c in (5, 8):
            cell.font      = DATA_FONT
            cell.alignment = CTR_ALIGN
        else:
            cell.font      = DATA_FONT
            cell.alignment = DATA_ALIGN

# ── Conditional formatting: Status (E) ───────────────────────────────────────
e_range = "E2:E{}".format(last_data_row)
ws.conditional_formatting.add(
    e_range,
    Rule(type="containsText", operator="containsText", text="MAPPED",
         dxf=DifferentialStyle(fill=mapped_fill, font=mapped_font))
)
ws.conditional_formatting.add(
    e_range,
    Rule(type="containsText", operator="containsText", text="UNMAPPED",
         dxf=DifferentialStyle(fill=unmapped_fill, font=unmapped_font))
)

ws.sheet_properties.tabColor = C_NAVY


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 – Applaud Tables Reference
# ─────────────────────────────────────────────────────────────────────────────
src_ws2 = src["Applaud Tables O Base System"]
ws2 = wb.create_sheet("Applaud Tables Reference")
ws2.sheet_view.showGridLines = False

rows_app = []
for row in src_ws2.iter_rows(min_row=1, max_row=src_ws2.max_row, values_only=False):
    vals = [c.value for c in row]
    if any(v is not None for v in vals):
        rows_app.append(vals)

last_app_row = len(rows_app)

for r_idx, row in enumerate(rows_app, 1):
    for c_idx, val in enumerate(row, 1):
        ws2.cell(row=r_idx, column=c_idx, value=val)

# ── Fill column D with TEXTJOIN formula (FBDI Template Mappings) ──────────────
# Returns comma-separated list of every FBDI Template that maps to this table.
# Uses TEXTJOIN+IF array — works natively in Excel 365; no Ctrl+Shift+Enter needed.
# Column D header stays "FBDI Template Mappings".
TEXTJOIN_TMPL = (
    "=IFERROR(TEXTJOIN(\", \",TRUE,"
    "IF('FBDI Mapping'!$C$2:$C${last}=A{r},"
    "'FBDI Mapping'!$A$2:$A${last},\"\")),"
    "\"\")"
)

for r_idx in range(2, last_app_row + 1):
    formula = TEXTJOIN_TMPL.format(last=last_data_row, r=r_idx)
    ws2.cell(row=r_idx, column=4, value=formula)

# ── Column widths ─────────────────────────────────────────────────────────────
for col, w in [("A",36),("B",18),("C",10),("D",55)]:
    ws2.column_dimensions[col].width = w

ws2.row_dimensions[1].height = 30
for r in range(2, last_app_row + 1):
    ws2.row_dimensions[r].height = 15

ws2.freeze_panes = "A2"

# ── Header style ──────────────────────────────────────────────────────────────
for c in range(1, 5):
    cell = ws2.cell(row=1, column=c)
    cell.font      = HDR_FONT
    cell.fill      = solid(C_BLUE2)
    cell.alignment = HDR_ALIGN
    cell.border    = hdr_border()

# ── Data styles ───────────────────────────────────────────────────────────────
for r in range(2, last_app_row + 1):
    bg = ALT_FILL if r % 2 == 0 else WHITE_FILL
    for c in range(1, 5):
        cell = ws2.cell(row=r, column=c)
        cell.fill   = bg
        cell.border = thin_border()
        if c == 1:
            cell.font      = DATA_FONT_MONO
            cell.alignment = DATA_ALIGN
        elif c == 4:
            cell.font      = DATA_FONT_SM
            cell.alignment = WRAP_ALIGN
        elif c in (2, 3):
            cell.font      = DATA_FONT
            cell.alignment = CTR_ALIGN
        else:
            cell.font      = DATA_FONT
            cell.alignment = DATA_ALIGN

b_range = "B2:B{}".format(last_app_row)
ws2.conditional_formatting.add(
    b_range,
    Rule(type="containsText", operator="containsText", text="MAPPED",
         dxf=DifferentialStyle(fill=mapped_fill, font=mapped_font))
)
ws2.conditional_formatting.add(
    b_range,
    Rule(type="containsText", operator="containsText", text="UNMAPPED",
         dxf=DifferentialStyle(fill=unmapped_fill, font=unmapped_font))
)

ws2.sheet_properties.tabColor = C_BLUE2

# ── Save ──────────────────────────────────────────────────────────────────────
out = "FBDI_to_ApplaudTables_Mapping.xlsx"
wb.save(out)
print("Saved:", out)
print("  FBDI Mapping rows:   ", last_data_row - 1)
print("  Applaud Tables rows: ", last_app_row - 1)
